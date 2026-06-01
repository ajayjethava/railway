# C:\Backup\newV8\frontend (3)\frontend\Circuitbuilding\app\routes.py

import os
import sys
import io
import logging
import time
import re
import hashlib
import subprocess
import traceback
from datetime import datetime, timezone, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, send_file, flash, session , jsonify,g,current_app
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from .models import (
    user_projects,db,GroupSummary,Approval,TerminalHeaderSummary,ChokeSummary, ResistorSummary,Project, StationDrawing, JunctionBox, Cable,CableRowConfig,TerminalSummary,CableSummary,CableLocationAddition,JunctionBoxSummary,
    Terminal, Group, TerminalHeader, ChokeTable,StatusMaster, ResistorTable,CTRApprovalHistory, get_ist_now, User,CableBox,Notification,StationMaster,RoleMaster,DesignationMaster,JunctionApproval,CTRUpload,CTRSummary,CTRDiagram,CTRRowDetail,CTRApproval,
)
from .models import GeneratedPDF
from .schemas import SHEETS, HEADER_HINTS
from .database import db
from sqlalchemy import or_ , text , desc , nullslast
from datetime import datetime
from sqlalchemy import or_, and_
from flask import abort
from sqlalchemy.orm import joinedload
# Add these imports at the top of routes.py
import ftplib
import paramiko  # For SFTP/SCP
from ftplib import FTP
import shutil
from pathlib import Path
import os
from sqlalchemy import func
import pytz
from io import BytesIO
from openpyxl import Workbook
from flask import send_file
import pandas as pd
from .ctr_pdf_generator import generate_ctr_pdf_from_excel
import threading
from flask import current_app
from flask import send_file
from PyPDF2 import PdfMerger
from io import BytesIO



bp = Blueprint("main", __name__)
def get_role_display(role):
    """Convert role number to display name"""
    role_map = {
        '0': 'Viewer',
        '1': 'Creator',
        '2': 'Approver L2',
        '3': 'Approver L3',
        '4': 'Admin'
    }
    return role_map.get(str(role), 'Unknown')

def _parse_project_ids(project_ids):
    """
    Convert project_ids from various formats to a list of integers.
    
    Handles:
    - ['14', '15'] -> [14, 15]
    - ['14,15'] -> [14, 15]
    - [14, 15] -> [14, 15]
    
    Returns empty list if conversion fails for any ID.
    """
    if not project_ids:
        return []
    
    result = []
    
    # Flatten if comma-separated strings are present
    flattened = []
    for item in project_ids:
        if isinstance(item, str) and ',' in item:
            # Split comma-separated values
            flattened.extend([x.strip() for x in item.split(',')])
        else:
            flattened.append(item)
    
    # Convert to integers
    for item in flattened:
        try:
            result.append(int(item))
        except (ValueError, TypeError):
            # Log or handle invalid IDs if needed
            continue
    
    return result

def get_next_approval_level(upload_id):
    approvals = CTRApproval.query.filter_by(
        ctr_upload_id=upload_id
    ).order_by(CTRApproval.approval_level.asc()).all()

    if not approvals:
        return 1

    highest = max(a.approval_level for a in approvals)

    if highest == 1:
        return 2
    elif highest == 2:
        return 3

    return None
def can_user_sign(upload, user_role):
   return upload.current_approval_level == user_role

@bp.route('/upload-signed-pdf/<int:upload_id>', methods=['POST'])
@login_required
def upload_signed_pdf(upload_id):

    file = request.files.get('signed_pdf')

    if not file:
        flash('PDF required', 'danger')
        return redirect(url_for('main.ctr_drawing'))

    # Get upload record
    upload = CTRUpload.query.get_or_404(upload_id)

    # =========================
    # SAVE PDF FILE
    # =========================
    filename = secure_filename(file.filename)

    # Optional: make unique filename
    unique_filename = f"{upload_id}_{current_user.id}_{filename}"
    
    
    
    upload_folder = r"C:\Railway\git\static\signed_pdfs"
    # Create folder if not exists
    os.makedirs(upload_folder, exist_ok=True)

    file_path = os.path.join(upload_folder, unique_filename)

    # Save file
    file.save(file_path)

       
   
    # Save filename/path in DB field
    upload.sign_document = unique_filename
    approval_level =  int(request.form.get('roleid'))
    upload.current_approval_level = approval_level
    ist = pytz.timezone('Asia/Kolkata')
    upload.fully_approved_at = datetime.now(ist)
    #upload.fully_approved_at = datetime.utcnow()
    FINAL_APPROVAL_LEVEL = 3

    if approval_level == FINAL_APPROVAL_LEVEL:
        upload.is_fully_approved = True
        ist = pytz.timezone('Asia/Kolkata')
        upload.fully_approved_at = datetime.now(ist)
        from PyPDF2 import PdfReader,PdfWriter
        from reportlab.pdfgen import canvas
        from io import BytesIO

        reader = PdfReader(file_path)

        metadata = reader.metadata

        pdf_version = metadata.get("/Subject")

        writer = PdfWriter()

        for page in reader.pages:

            packet = BytesIO()
            page_width = float(page.mediabox.width)
            page_height = float(page.mediabox.height) 
            can = canvas.Canvas(packet, pagesize=(page_width, page_height))

        # page = reader.pages[0]

            

            #print(page_width)

            footer_text = f"{pdf_version}"

            text_width = can.stringWidth(footer_text)

            right_margin = 50
            
            completion_x = 3150 # page_width - text_width - right_margin
            completion_y = 135

            # SAME POSITION
            footer_text = f"{pdf_version}"

            can.drawString(
                completion_x,
                completion_y,
                footer_text
            )

            can.save()

            packet.seek(0)

            overlay_pdf = PdfReader(packet)

            page.merge_page(overlay_pdf.pages[0])

            writer.add_page(page)

        # Keep metadata
        writer.add_metadata({
            "/Subject": pdf_version
        })

        # SAVE UPDATED FILE
        with open(file_path, "wb") as output_file:
            writer.write(output_file)


    else:
        upload.is_fully_approved = False
    
    # =========================
    # APPROVAL ENTRY
    # =========================
   

    approval = CTRApproval(
        ctr_upload_id=upload_id,
        approver_role_id=approval_level, #current_user.role_id,
        approver_user_id=current_user.id,
        approval_level=approval_level,
        approval_status='approved',
        comments=request.form.get('comments')
    )

    db.session.add(approval)

    # =========================
    # HISTORY ENTRY
    # =========================
    history = CTRApprovalHistory(
        ctr_upload_id=upload_id,
        action='approved',
        action_level=approval_level,
        action_details=f'Level {approval_level} signed PDF uploaded',
        action_by_user_id=current_user.id,
        action_by_role_id=current_user.role_id,
        previous_status_id=None,
        new_status_id=None,
        version_number=1
    )

    db.session.add(history)

    
    db.session.commit()

    flash('Signed PDF uploaded successfully', 'success')

    return redirect(url_for('main.ctr_drawing'))



@bp.route("/download-file/<filename>")
def download_file(filename):
    file_path = os.path.join(
        r"C:\Railway\git\Circuitbuilding\uploads_ctr",
        filename
    )

    if not os.path.exists(file_path):
        return abort(404, description="File not found")

    return send_file(file_path, as_attachment=True)

@bp.route('/view-signed-pdf/<int:upload_id>/<int:level>')
def view_signed_pdf(upload_id, level):
    upload = CTRUpload.query.get_or_404(upload_id)
    print("DEBUG sign_document:", upload.sign_document, type(upload.sign_document))
    if not upload.sign_document:
        abort(404, description="Signed PDF not found")

    file_path = os.path.join(
        r"C:\Railway\git\static\signed_pdfs",
        upload.sign_document
    )

    return send_file(
      file_path,
      as_attachment=True
    )
@bp.route('/view-signed-pdf1/<int:upload_id>/<int:level>')
def view_signed_pdf1(upload_id, level):
    upload = CTRUpload.query.get_or_404(upload_id)
    print("DEBUG sign_document:", upload.sign_document, type(upload.sign_document))
    if not upload.sign_document:
        abort(404, description="Signed PDF not found")

    file_path = os.path.join(
        r"C:\Railway\git\static\signed_pdfs",
        upload.sign_document
    )

    return send_file(
      file_path
    )


def sync_station_to_master(station_drawing):
    """
    Sync a StationDrawing record to StationMaster (simplified fields only)
    Returns True if successful, False otherwise
    """
    try:
        # Check if station already exists in master
        existing_master = StationMaster.query.filter_by(
            station_id=station_drawing.station_id,
            project_id=station_drawing.project_id
        ).first()
        
        if existing_master:
            # Update existing record with only the required fields
            existing_master.station_name = station_drawing.station_name
            existing_master.station_code = station_drawing.station_code
            print(f"✅ Updated StationMaster: {station_drawing.station_name} ({station_drawing.station_id})")
        else:
            # Create new master record with only required fields
            new_master = StationMaster(
                project_id=station_drawing.project_id,
                station_id=station_drawing.station_id,
                station_name=station_drawing.station_name,
                station_code=station_drawing.station_code
            )
            db.session.add(new_master)
            print(f"✅ Created StationMaster: {station_drawing.station_name} ({station_drawing.station_id})")
        
        db.session.commit()
        return True
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error syncing station to master: {str(e)}")
        return False

def sync_all_stations_to_master(project_id=None):
 
    try:
        query = StationDrawing.query
        if project_id:
            query = query.filter_by(project_id=project_id)
        
        stations = query.all()
        success_count = 0
        error_count = 0
        
        for station in stations:
            if sync_station_to_master(station):
                success_count += 1
            else:
                error_count += 1
        
        result = {
            'total': len(stations),
            'success': success_count,
            'error': error_count
        }
        
        print(f"🔍 Sync completed: {result}")
        return result
        
    except Exception as e:
        print(f"❌ Error in sync_all_stations_to_master: {str(e)}")
        return {
            'total': 0,
            'success': 0,
            'error': 0,
            'message': str(e)
        }



def ensure_junction_approvals_for_pdf(pdf):
    """
    Create pending JunctionApproval records for a given PDF
    for all junctions of its project that do not already have a record.
    """
    if not pdf or not pdf.project:
        return 0

    project = pdf.project
    created_count = 0

    for junction in project.junction_boxes:
        existing = JunctionApproval.query.filter_by(
            project_id=project.id,
            generated_pdf_id=pdf.id,
            junction_box_id=junction.id
        ).first()

        if not existing:
            approval = JunctionApproval(
                project_id=project.id,
                generated_pdf_id=pdf.id,
                junction_box_id=junction.id,
                level1_status='pending',
                level2_status='pending',
                level3_status='pending',
                created_at=get_ist_now()
            )
            db.session.add(approval)
            created_count += 1

    if created_count > 0:
        db.session.commit()
        current_app.logger.info(
            f"✅ Created {created_count} pending JunctionApproval records for PDF {pdf.id} (Project {project.id})"
        )
    return created_count

# ==================== AUTHENTICATION SETUP ====================

# Add a before_request handler to require login for all routes
@bp.before_app_request
def require_login():
    """Ensure user is logged in for all routes except login, static files, and logout"""
    # List of endpoints that don't require authentication
    exempt_endpoints = ['main.login', 'static']
    
    # Get the endpoint name
    endpoint = request.endpoint
    
    # If user is not authenticated and trying to access protected endpoint, redirect to login
    if not current_user.is_authenticated and endpoint not in exempt_endpoints:
        return redirect(url_for('main.login', next=request.url))



# Add this function to get user permissions based on role_name
def get_user_permissions(user):
    """
    Get user permissions based on role (0,1,2,3,4)
    """
    role = str(user.role)  # Ensure it's a string
    
    permissions = {
        'can_view_approvals': False,
        'can_create_drawing': False,
        'can_approve_level1': False,
        'can_approve_level2': False,
        'can_approve_level3': False,
        'can_see_all': False,
        'role': role,
        'role_display': get_role_display(role)
    }
    
    if role == '0':  # Viewer
        permissions['can_view_approvals'] = True
        
    elif role == '1':  # Creator
        permissions['can_view_approvals'] = True
        permissions['can_create_drawing'] = True
        permissions['can_approve_level1'] = True
        
    elif role == '2':  # Approver Level 2
        permissions['can_view_approvals'] = True
        permissions['can_approve_level2'] = True
        
    elif role == '3':  # Approver Level 3
        permissions['can_view_approvals'] = True
        permissions['can_approve_level3'] = True
        
    elif role == '4':  # Admin
        permissions['can_view_approvals'] = True
        permissions['can_create_drawing'] = True
        permissions['can_approve_level1'] = True
        permissions['can_approve_level2'] = True
        permissions['can_approve_level3'] = True
        permissions['can_see_all'] = True
    
    return permissions



def ensure_junction_approvals_for_project_and_pdf(project_id, pdf_id):
    
    pdf = GeneratedPDF.query.get(pdf_id)
    if not pdf:
        return
    for junction in JunctionBox.query.filter_by(project_id=project_id).all():
        existing = JunctionApproval.query.filter_by(
            project_id=project_id,
            generated_pdf_id=pdf_id,
            junction_box_id=junction.id
        ).first()
        if not existing:
            approval = JunctionApproval(
                project_id=project_id,
                generated_pdf_id=pdf_id,
                junction_box_id=junction.id,
                level1_status='pending',
                level2_status='pending',
                level3_status='pending',
                created_at=get_ist_now()
            )
            db.session.add(approval)
    db.session.commit()

# Add context processor to make permissions available in templates
@bp.context_processor
def inject_permissions():
    """Inject user permissions into all templates"""
    if current_user.is_authenticated:
        return {'permissions': get_user_permissions(current_user)}
    return {'permissions': None}

# User loader function (moved from __init__.py for consistency)
@bp.route("/login", methods=["GET", "POST"])
def login():
    """Handle user login - accepts both username OR mobile number"""
    # If user is already logged in, redirect to index
   
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))
    
    if request.method == "POST":
        login_input = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
      
        if not login_input or not password:
            flash("Please enter both username/mobile number and password")
            return redirect(url_for('main.login'))
      
        # Find user by either username OR mobile number
        user = User.query.filter(
            or_(
                User.username == login_input,
                User.mobile_number == login_input
            )
        ).first()
       
        # Check if user exists
        if not user:
            flash("Invalid username/mobile number or password")
            return redirect(url_for('main.login'))
        
        # Check if user is active
        if not user.is_active:
            flash("Your account is inactive. Please contact the administrator.")
            return redirect(url_for('main.login'))
        
        # Check if password is correct
        
        if not user.check_password(password):
            flash("Invalid username/mobile number or password")
            return redirect(url_for('main.login'))
        
        
        # If all checks pass, log in the user
        login_user(user, remember=False)
        
        # Mark session as non-permanent (will expire when browser closes)
        session.permanent = False
        #return "12345"
        # UPDATED: For ALL users (including admin), auto-set to latest/first project if available
        if user.projects:
            # Prioritize latest project for admins, first for others
            if user.role_name == '4':
                latest_project = sorted(user.projects, key=lambda p: p.id, reverse=True)[0]
                session['current_project_id'] = latest_project.id
                session['project_id'] = latest_project.id
            else:
                session['current_project_id'] = user.projects[0].id
                session['project_id'] = user.projects[0].id
        else:
            # No projects: clear for all
            session.pop('current_project_id', None)
            session.pop('project_id', None)
            
        # Redirect to next page or index
        next_page = request.args.get('next')
        return redirect(next_page or url_for('main.approval_tracking'))
  
    return render_template("login.html")

@bp.route("/logout")
@login_required
def logout():
    """Handle user logout"""
    logout_user()
    session.clear()  # Clear all session data
    return redirect(url_for('main.login'))

# ==================== EXISTING CODE CONTINUES ====================

# Initialize Login Manager
login_manager = LoginManager()
login_manager.login_view = 'main.login'
login_manager.login_message = 'Please log in to access this page.'



@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Add custom template filter for sorting cables
@bp.app_template_filter('sort_cables')
def sort_cables_filter(cables):
    """Sort cables by cable_id as integer"""
    if not cables:
        return []
   
    try:
        # Convert cable_id to integer for proper numerical sorting
        return sorted(cables, key=lambda x: int(x.get('cable_id', 0)) if x.get('cable_id') else 0)
    except (ValueError, TypeError):
        # Fallback to string sorting if conversion fails
        return sorted(cables, key=lambda x: str(x.get('cable_id', '')))

@bp.app_template_filter('sort_junction_boxes')
def sort_junction_boxes_filter(junction_boxes):
    """Sort junction boxes by junction_id as integer"""
    if not junction_boxes:
        return []
   
    try:
        # Convert junction_id to integer for proper numerical sorting
        return sorted(junction_boxes, key=lambda x: int(x.get('junction_id', 0)) if x.get('junction_id') else 0)
    except (ValueError, TypeError):
        # Fallback to string sorting if conversion fails
        return sorted(junction_boxes, key=lambda x: str(x.get('junction_id', '')))
   
   
MODEL_MAP = {
    "StationDrawing": StationDrawing,
    "junction_box": JunctionBox,
    "cable": Cable,
    "cable_box": CableBox, # NEW: Add CableBox model
    "terminal": Terminal,
    "group": Group,
    "terminal_header": TerminalHeader,
    "choketable": ChokeTable,
    "resistortable": ResistorTable,
}

TABLE_WORKFLOW = [
    "StationDrawing",
    "junction_box",
    "cable",
    "terminal",
    "terminal_header", # Step 5 - Headers
    "group", # Step 6 - Groups
    "choketable",
    "resistortable",
    "cable_box",
]

IST = timezone(timedelta(hours=5, minutes=30))

@bp.app_template_filter('sort_terminals')
def sort_terminals_filter(terminals):
    """Sort terminals by cable_id first, then terminal_id, both as integers"""
    if not terminals:
        return []
   
    try:
        # Sort by cable_id first, then terminal_id, both as integers
        return sorted(terminals, key=lambda x: (
            int(x.get('cable_id', 0)) if x.get('cable_id') and str(x.get('cable_id')).isdigit() else 0,
            int(x.get('terminal_id', 0)) if x.get('terminal_id') and str(x.get('terminal_id')).isdigit() else 0
        ))
    except (ValueError, TypeError):
        # Fallback to string sorting if conversion fails
        return sorted(terminals, key=lambda x: (
            str(x.get('cable_id', '')),
            str(x.get('terminal_id', ''))
        ))

def format_location_name(name, size):
    """Format location name by appending (F) for Full and (H) for Half."""
    if not name or not size:
        return name
   
    # Remove any existing suffixes
    name = re.sub(r'\s*\([FH]\)$', '', name)
   
    if size == 'Full':
        return f"{name} (F)"
    elif size == 'Half':
        return f"{name} (H)"
    else:
        return name

def get_current_project():
    """Get the currently selected project ID from session with access control"""
    # Check both possible session keys for backward compatibility
    project_id = session.get('current_project_id') or session.get('project_id')
   
    print(f"🔍 DEBUG get_current_project: current_project_id={session.get('current_project_id')}, project_id={session.get('project_id')}, final={project_id}")
   
    if not project_id:
        print("❌ DEBUG: No project ID found in session")
        return None
   
    # For non-admin users, verify they have access to this project
    if current_user.is_authenticated and current_user.role != 'admin':
        user_project_ids = [p.id for p in current_user.projects]
        if project_id not in user_project_ids:
            print(f"❌ DEBUG: User {current_user.id} not allowed to access project {project_id}")
            # If user doesn't have access to current project, switch to their first project
            if current_user.projects:
                project_id = current_user.projects[0].id
                session['current_project_id'] = project_id
                session['project_id'] = project_id
                print(f"🔄 DEBUG: Auto-switched to user's first project: {project_id}")
            else:
                print("❌ DEBUG: User has no assigned projects")
                return None
   
    # Verify project exists in database
    project = Project.query.get(project_id)
    if not project:
        print(f"❌ DEBUG: Project {project_id} not found in database")
        # Clean up invalid session data
        session.pop('current_project_id', None)
        session.pop('project_id', None)
        return None
   
    print(f"✅ DEBUG: Returning valid project ID: {project_id}")
    return project_id

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'

def _md5_of_file(path, chunk_size=8192):
    hasher = hashlib.md5()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(chunk_size), b''):
            hasher.update(chunk)
    return hasher.hexdigest()

META_PATTERNS = {
    "metadata_checksum": re.compile(
        r"(?:Metadata\s+Checksum|Initial\s+checksum\s+with\s+initial\s+file\s+size)\s*:\s*([0-9a-fA-F]{32})",
        re.IGNORECASE,
    ),
    "metadata_data": re.compile(r"(?:Metadata\s+Data\s+string|Checksum\s+data\s+string)\s*:\s*(.+)", re.IGNORECASE),
    "initial_size_bytes": re.compile(r"Initial\s+file\s+size\s*:\s*(\d+)\s*bytes", re.IGNORECASE),
    "final_size_bytes": re.compile(r"(?:Final\s+file\s+size|Final\s+file\s+size\s+after\s+enhancement)\s*:\s*(\d+)\s*bytes", re.IGNORECASE),
    "metadata_ts_ist": re.compile(r"Timestamp\s*\(IST\)\s*:\s*([0-9:\-\s]+)", re.IGNORECASE),
    "station_code": re.compile(r"Station\s+code\s*:\s*([A-Za-z0-9\-_]+)", re.IGNORECASE),
    "source_pdf_name": re.compile(r"File\s+name\s*:\s*(.+?\.pdf)\s*$", re.IGNORECASE | re.MULTILINE),
    "full_file_md5": re.compile(r"Full\s+file\s+MD5(?:\s*hash)?\s*:\s*([0-9a-fA-F]{32})", re.IGNORECASE),
}

def parse_converter_stdout(stdout_text: str) -> dict:
    out = {}
    text = stdout_text or ""
    for key, pat in META_PATTERNS.items():
        m = pat.search(text)
        out[key] = m.group(1).strip() if m else None
    for k in ("initial_size_bytes", "final_size_bytes"):
        if out[k] is not None:
            try:
                out[k] = int(out[k])
            except Exception:
                out[k] = None
    if out["metadata_ts_ist"]:
        try:
            dt = datetime.strptime(out["metadata_ts_ist"], "%Y-%m-%d %H:%M:%S")
            out["metadata_ts_ist"] = dt.replace(tzinfo=IST)
        except Exception:
            out["metadata_ts_ist"] = None
    return out

def get_project_pdfs(project_id):
    upload_dir = os.path.join(os.getcwd(), 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    prefix = f"railway_project_{project_id}_"
    pdfs = []
    for name in os.listdir(upload_dir):
        if name.endswith(".pdf") and name.startswith(prefix):
            path = os.path.join(upload_dir, name)
            try:
                stat = os.stat(path)
                mtime = stat.st_mtime
                size_kb = max(1, stat.st_size // 1024)
                rec = GeneratedPDF.query.filter_by(
                    project_id=project_id, pdf_filename=name
                ).order_by(GeneratedPDF.id.desc()).first()
                checksum = rec.checksum_md5 if rec else None

                # ✅ Use inline_pdf for direct viewing, download_pdf for download
                inline_url = url_for('main.inline_pdf', filename=name)
                download_url = url_for('main.download_pdf', filename=name)

                pdfs.append({
                    "filename": name,
                    "inline_url": inline_url,       # "View" button will use this
                    "download_url": download_url,   # "Download" button will use this
                    "created": datetime.fromtimestamp(mtime).strftime("%d %b %Y %I:%M %p"),
                    "size_kb": size_kb,
                    "checksum_md5": checksum,
                    "mtime": mtime,
                })
            except Exception:
                continue
    pdfs.sort(key=lambda x: x["mtime"], reverse=True)
    for p in pdfs:
        p.pop("mtime", None)
    return pdfs

# ==============================================
# HELPER: Check if user has access to a project
# ==============================================
def user_has_project_access(project_id):
    """Return True if current user (admin/user) has access to this project"""
    if current_user.role == 'admin':
        return True
    if current_user.role == 'viewer':
        return False
    # For 'user' role: check if project is assigned
    project = Project.query.get(project_id)
    if project and project in current_user.projects:
        return True
    return False

# ==================== ADMIN USER MANAGEMENT ROUTES ====================

@bp.route("/admin/users")
@login_required
def admin_users():
    '''
    if current_user.role_name != '4':
        flash("Access denied. Admin privileges required.", "danger")
        return redirect(url_for('main.index'))
    '''
    from sqlalchemy import cast, Integer
    if current_user.role_name == '4':
        users = User.query.all()
        projects = Project.query.all()
        roles = RoleMaster.query.filter_by(is_active=True).all()
        designations = DesignationMaster.query.filter_by(is_active=True).all()
    else:
        users = User.query.filter(
            cast(User.role, Integer) < int(current_user.role_name)
        ).all()


        projects = Project.query.all()
        
        roles = RoleMaster.query.filter(
            RoleMaster.is_active == True,
            cast(RoleMaster.role_name, Integer) < int(current_user.role_name)
        ).all()
        designations = DesignationMaster.query.filter_by(is_active=True).all()
        

    # DEBUG: Log to console (check server logs)
    print(f"DEBUG: Fetched {len(designations)} active designations: {[d.designation_name.strip() for d in designations]}")
    for user in users[:3]:  # Sample first 3 users
        print(f"DEBUG: User {user.username} designation: '{user.designation}' (type: {type(user.designation)})")

    return render_template(
        "admin_users.html",
        users=users,
        projects=projects,
        roles=roles,
        designations=designations
    )

@bp.route("/admin/users/add", methods=['POST'])
@login_required
def admin_add_user():
    """Add new user"""
    '''
    if current_user.role_name != '4':
        flash("Access denied.", "danger")
        return redirect(url_for('main.index'))
    '''
    try:
        username = request.form.get('username')
        mobile_number = request.form.get('mobile_number') or None
        email = request.form.get('email') or None
        password = request.form.get('password')
        role = request.form.get('role')
        designation = request.form.get('designation')
        project_ids = request.form.getlist('project_ids')
        
        # Validate required fields
        if not all([username, password, role, designation]):
            flash("All required fields must be filled", "danger")
            return redirect(url_for('main.admin_users'))
        
        # Validate role exists
        role_exists = RoleMaster.query.filter_by(role_name=role, is_active=True).first()
        if not role_exists:
            flash("Invalid role selected", "danger")
            return redirect(url_for('main.admin_users'))
        
        # Validate designation exists
        desig_exists = DesignationMaster.query.filter_by(designation_name=designation, is_active=True).first()
        if not desig_exists:
            flash("Invalid designation selected", "danger")
            return redirect(url_for('main.admin_users'))
        
        # Check if username already exists
        if User.query.filter_by(username=username).first():
            flash("Username already registered", "danger")
            return redirect(url_for('main.admin_users'))
        
        # Check if mobile number already exists (if provided)
        if mobile_number:
            existing_mobile = User.query.filter_by(mobile_number=mobile_number).first()
            if existing_mobile:
                flash("Mobile number already registered to another user", "danger")
                return redirect(url_for('main.admin_users'))
        
        # Create new user
        new_user = User(
            username=username,
            mobile_number=mobile_number,
            email=email,
            role=role,
            designation=designation,
            is_active=True
        )
        new_user.set_password(password)
        
        # Assign projects - FIXED: Handle both comma-separated and list formats
        if project_ids:
            project_ids = _parse_project_ids(project_ids)
            if project_ids:  # Only query if we have valid IDs
                projects = Project.query.filter(Project.id.in_(project_ids)).all()
                new_user.projects = projects
        
        db.session.add(new_user)
        db.session.commit()
        
        flash(f"User {username} created successfully!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error creating user: {str(e)}", "danger")
    
    return redirect(url_for('main.admin_users'))


@bp.route("/admin/users/edit/<int:user_id>", methods=['POST'])
@login_required
def admin_edit_user(user_id):
    """Edit existing user"""
    '''
    if current_user.role_name != '4':
        flash("Access denied.", "danger")
        return redirect(url_for('main.index'))
    '''
    user = User.query.get(user_id)
    if not user:
        flash("User not found", "danger")
        return redirect(url_for('main.admin_users'))
    
    try:
        username = request.form.get('username')
        mobile_number = request.form.get('mobile_number') or None
        email = request.form.get('email') or None
        password = request.form.get('password')
        role = request.form.get('role')
        designation = request.form.get('designation')
        is_active = request.form.get('is_active', '1')
        remarks = request.form.get('remarks')
        project_ids = request.form.getlist('project_ids')
        
        # Validate required fields
        if not all([username, role, designation]):
            flash("All required fields must be filled", "danger")
            return redirect(url_for('main.admin_users'))
        
        # Validate role exists
        role_exists = RoleMaster.query.filter_by(role_name=role, is_active=True).first()
        if not role_exists:
            flash("Invalid role selected", "danger")
            return redirect(url_for('main.admin_users'))
        
        # Validate designation exists
        desig_exists = DesignationMaster.query.filter_by(designation_name=designation, is_active=True).first()
        if not desig_exists:
            flash("Invalid designation selected", "danger")
            return redirect(url_for('main.admin_users'))
        
        # Check if username is already taken by another user
        existing_user = User.query.filter_by(username=username).first()
        if existing_user and existing_user.id != user_id:
            flash("Username already registered to another user", "danger")
            return redirect(url_for('main.admin_users'))
        
        # Check if mobile number is already taken by another user (if provided)
        if mobile_number:
            existing_mobile = User.query.filter_by(mobile_number=mobile_number).first()
            if existing_mobile and existing_mobile.id != user_id:
                flash("Mobile number already registered to another user", "danger")
                return redirect(url_for('main.admin_users'))
        
        # Update user fields
        user.username = username
        user.mobile_number = mobile_number
        user.email = email
        user.role = role
        user.designation = designation
        user.is_active = int(is_active) == 1
        user.remarks = remarks
        
        # Update password if provided
        if password:
            user.set_password(password)
        
        # Update projects - FIXED: Handle both comma-separated and list formats
        if project_ids:
            project_ids = _parse_project_ids(project_ids)
            if project_ids:  # Only query if we have valid IDs
                projects = Project.query.filter(Project.id.in_(project_ids)).all()
                user.projects = projects
            else:
                user.projects = []
        else:
            user.projects = []
        
        db.session.commit()
        flash(f"User {username} updated successfully!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error updating user: {str(e)}", "danger")
    
    return redirect(url_for('main.admin_users'))


# HELPER FUNCTION - Add this at module level or in a utils file


@bp.route("/admin/users/delete/<int:user_id>", methods=['POST'])
@login_required
def admin_delete_user(user_id):
    """Delete user"""
    '''
    if current_user.role_name != '4':
        flash("Access denied.", "danger")
        return redirect(url_for('main.index'))
    '''
    user = User.query.get(user_id)
    if not user:
        flash("User not found", "danger")
        return redirect(url_for('main.admin_users'))
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        flash("You cannot delete your own account", "danger")
        return redirect(url_for('main.admin_users'))
    
    try:
        
        GeneratedPDF.query.filter(
            or_(
                GeneratedPDF.level1_approver_id == user.id,
                GeneratedPDF.level2_approver_id == user.id,
                GeneratedPDF.level3_approver_id == user.id
            )
        ).update({
            "level1_approver_id": None,
            "level2_approver_id": None,
            "level3_approver_id": None
        }, synchronize_session=False)
        
        Approval.query.filter(
            Approval.approver_id == user.id
        ).update({
            "approver_id": None
        }, synchronize_session=False)

        db.session.commit()

        db.session.commit()
        username = user.username
        db.session.delete(user)
        db.session.commit()
        flash(f"User {username} deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting user: {str(e)}", "danger")
    
    return redirect(url_for('main.admin_users'))

@bp.route('/change-password', methods=['GET', 'POST'])
def change_password():

    if request.method == 'POST':

        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        # Check current password
        if not check_password_hash(current_user.password_hash, current_password):
            flash('Current password is incorrect.', 'danger')
            return redirect(url_for('main.change_password'))

        # Check confirm password
        if new_password != confirm_password:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('main.approval_tracking'))

        # Update password
        current_user.password_hash = generate_password_hash(new_password)

        db.session.commit()

        flash('Password changed successfully.', 'success')

        return redirect(url_for('main.index'))

    return render_template('change_password.html')
# ==================== EXISTING ROUTES (UPDATED WITH @login_required) ====================

@bp.route("/a", methods=["GET"])
@login_required
def index():
    # Get user permissions
    permissions = get_user_permissions(current_user)
    
    # If user cannot create drawings and is not admin, redirect to approval tracking
    if not permissions['can_create_drawing'] and not permissions['can_see_all']:
        return redirect(url_for('main.approval_tracking'))
    
    # For non-admin users, only show their assigned projects
    if not permissions['can_see_all']:  # Not admin
        projects = current_user.projects
        print(f"🔍 DEBUG index: Non-admin - Fetching {len(projects)} assigned projects")
    else:
        projects = Project.query.order_by(Project.id.desc()).all()
        print(f"🔍 DEBUG index: ADMIN - Fetching ALL {len(projects)} projects from DB")
        for p in projects:
            print(f" - Project ID={p.id}, Name='{p.name}'")
   
    def _count_rows(project_id: int) -> int:
        total = 0
        try:
            for _sheet, Model in MODEL_MAP.items():
                if hasattr(Model, "project_id"):
                    total += Model.query.filter_by(project_id=project_id).count()
        except Exception:
            pass
        return total

    def _get_drawn_by(project_id: int) -> str:
        """Fetch drawn_by from StationDrawing table for this project"""
        try:
            station = StationDrawing.query.filter_by(project_id=project_id).first()
            if station and hasattr(station, 'drawn_by') and station.drawn_by:
                return str(station.drawn_by).strip()
        except Exception as e:
            print(f"Error fetching drawn_by for project {project_id}: {e}")
        return "-"
   
    projects_data = []
    for p in projects:
        # Get drawn_by for this project
        drawn_by = _get_drawn_by(p.id)
      
        pdfs = []
        current_pdf = None
        try:
            # Fetch PDFs ordered by creation time (oldest first for version assignment)
            pdf_records = GeneratedPDF.query.filter_by(project_id=p.id)\
                            .order_by(GeneratedPDF.created_at.asc()).all()
          
            # Assign sequential version numbers
            total_versions = len(pdf_records)
            for idx, rec in enumerate(pdf_records, start=1):
                pdf_data = {
                    "id": rec.id,
                    "filename": rec.pdf_filename,
                    "view_url": url_for("main.pdf_view", filename=rec.pdf_filename),
                    "inline_url": url_for("main.inline_pdf", filename=rec.pdf_filename),
                    "download_url": url_for("main.download_pdf", filename=rec.pdf_filename),
                    "size_kb": (rec.file_size // 1024) if getattr(rec, "file_size", None) else None,
                    "version": idx,
                    "is_current": (idx == total_versions),
                    "created_by": drawn_by,
                    "remarks": getattr(rec, "remarks", "-"),
                    "created_at": rec.created_at.strftime("%d %b %Y %I:%M %p") if getattr(rec, "created_at", None) else "-",
                    "level1_status": rec.level1_status,
                    "level2_status": rec.level2_status,
                    "level3_status": rec.level3_status,
                    "level1_approver": rec.level1_approver.username if rec.level1_approver else None,
                    "level2_approver": rec.level2_approver.username if rec.level2_approver else None,
                    "level3_approver": rec.level3_approver.username if rec.level3_approver else None,
                    "level1_approval_date": rec.level1_approval_date.strftime("%d %b %Y %I:%M %p") if rec.level1_approval_date else None,
                    "level2_approval_date": rec.level2_approval_date.strftime("%d %b %Y %I:%M %p") if rec.level2_approval_date else None,
                    "level3_approval_date": rec.level3_approval_date.strftime("%d %b %Y %I:%M %p") if rec.level3_approval_date else None,
                    "rejection_reason": rec.rejection_reason
                }
                pdfs.append(pdf_data)
              
                # Set current PDF (latest version)
                if pdf_data["is_current"]:
                    current_pdf = pdf_data
          
            # Reverse to show newest first in UI
            pdfs.reverse()
        except Exception as e:
            print(f"Error loading PDFs for project {p.id}: {e}")
            pdfs = []
        try:
            total_rows = _count_rows(p.id)
        except Exception:
            total_rows = 0
        projects_data.append({
            "project": p,
            "total_rows": total_rows,
            "pdfs": pdfs,
            "current_pdf": current_pdf
        })
   
    return render_template("index.html", projects_data=projects_data, permissions=permissions)
   
@bp.route('/check_terminal_duplicate')
@login_required
def check_terminal_duplicate():
    cable_id = request.args.get('cable_id')
    terminal_id = request.args.get('terminal_id')
   
    # Convert to strings for consistent comparison
    cable_id_str = str(cable_id) if cable_id else None
    terminal_id_str = str(terminal_id) if terminal_id else None
   
    # Query your database to check if a terminal with these IDs exists
    existing_terminal = Terminal.query.filter_by(
        cable_id=cable_id_str,
        terminal_id=terminal_id_str
    ).first()
   
    return jsonify({'exists': existing_terminal is not None})

@bp.route("/project_selection")
@login_required
def project_selection():
    return redirect(url_for("main.index"))

@bp.route("/upload/<sheet_name>", methods=["GET", "POST"])
@login_required
def upload_sheet(sheet_name):
    if sheet_name not in SHEETS:
        flash(f"Unknown sheet: {sheet_name}")
        return redirect(url_for("main.index"))
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    current_project = Project.query.get(project_id)
    model = MODEL_MAP[sheet_name]
    expected_headers = SHEETS[sheet_name]
    if request.method == "POST":
        if 'file' not in request.files:
            flash('No file uploaded')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No file selected')
            return redirect(request.url)
        if not allowed_file(file.filename):
            flash('Only XLSX files are allowed')
            return redirect(request.url)
        try:
            wb = load_workbook(file, data_only=True)
            sheet_found = None
            for ws_name in wb.sheetnames:
                if ws_name.lower() == sheet_name.lower():
                    sheet_found = ws_name
                    break
            if not sheet_found:
                flash(f'No "{sheet_name}" sheet found. Available: {", ".join(wb.sheetnames)}')
                return redirect(request.url)
            ws = wb[sheet_found]
            headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
            missing_headers = []
            header_mapping = {}
            for required_header in expected_headers:
                found = False
                for i, file_header in enumerate(headers):
                    if file_header == required_header.lower():
                        header_mapping[required_header] = i
                        found = True
                        break
                if not found:
                    missing_headers.append(required_header)
            if missing_headers:
                flash(f'Missing required columns: {", ".join(missing_headers)}')
                return redirect(request.url)
            # Clear existing rows for this project in this sheet to keep only latest Excel
            try:
                deleted = model.query.filter_by(project_id=project_id).delete(synchronize_session=False)
                if deleted:
                    db.session.flush()
                else:
                    # If no rows were deleted, check if there are any foreign key constraints
                    # that might prevent deletion
                    print(f"No rows to delete for {sheet_name} in project {project_id}")
            except Exception as e:
                db.session.rollback()
                flash(f'Error clearing old {sheet_name} data: {str(e)}')
                return redirect(request.url)
            imported_count = 0
            error_count = 0
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data = {'project_id': project_id}
                    has_data = False
                    for header in expected_headers:
                        if header in header_mapping:
                            col_index = header_mapping[header]
                            cell_value = row[col_index] if col_index < len(row) else None
                            if cell_value is not None:
                                text = str(cell_value).strip()
                               
                                # ENFORCE LOCATION NAMING CONVENTION FOR JUNCTION BOXES
                                if sheet_name == 'junction_box' and header == 'junction_name':
                                    # Get the junction_size for formatting
                                    size_header_index = None
                                    for h in expected_headers:
                                        if h == 'junction_size' and h in header_mapping:
                                            size_header_index = header_mapping[h]
                                            break
                                   
                                    if size_header_index is not None and size_header_index < len(row):
                                        size_value = row[size_header_index]
                                        if size_value is not None:
                                            size_text = str(size_value).strip()
                                            # Format the location name
                                            text = format_location_name(text, size_text)
                               
                                data[header] = text if text else None
                                if text:
                                    has_data = True
                            else:
                                data[header] = None
                        else:
                            data[header] = None
                    if has_data:
                        db.session.add(model(**data))
                        imported_count += 1
                except Exception as e:
                    error_count += 1
                    print(f"[UPLOAD:{sheet_name}] Row {row_num} error: {e}")
                    continue
            if imported_count > 0:
                db.session.commit()
                flash(f'Imported {imported_count} {sheet_name} records to Project ID {project_id} (latest data only)')
                if error_count > 0:
                    flash(f'Warning: {error_count} rows had errors and were skipped')
                return redirect(url_for("main.sheet_form", name=sheet_name))
            else:
                db.session.rollback()
                flash('No valid data found to import')
                return redirect(request.url)
        except Exception as e:
            db.session.rollback()
            flash(f'Error processing file: {str(e)}')
            return redirect(request.url)
    return render_template(
        "upload_sheet.html",
        current_project=current_project,
        sheet_name=sheet_name,
        expected_headers=expected_headers,
        sheet_display_name=sheet_name.replace('_', ' ').title(),
        hint=HEADER_HINTS.get(sheet_name, f"Upload {sheet_name} data from XLSX file")
    )

@bp.route("/workflow")
@login_required
def workflow_root():
    """
    Workflow entry point - redirects to step 2 by default
    since station drawing is already seeded with default data.
    Users can still navigate back to step 1 to edit if needed.
    """
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
   
    # Seed default station drawing data
    seed_default_station_drawing(project_id)
   
    # Redirect to step 2 (Junction Box Count) by default
    return redirect(url_for("main.workflow_step", step=2))

def can_edit_step(step, user):
    """Check if user can edit this step - viewers cannot edit any step"""
    if user.role == 'viewer':
        return False
    # Admin and regular users can edit all steps in their assigned projects
    return True



@bp.route('/set_project_and_continue/<int:project_id>', methods=['POST'])
@login_required
def set_project_and_continue(project_id):
    """Set the project ID in session and continue with existing data."""
    try:
        # Check if project exists and user has access
        project = Project.query.get_or_404(project_id)
        
        # Check user permissions
        has_permission = False
        
        # Admin (role 4) and creators (role 1) can access any project
        if current_user.role_name in ['1', '4']:
            has_permission = True
        # Regular users can only access if assigned to project
        elif current_user.id in [u.id for u in project.users]:
            has_permission = True
        
        if not has_permission:
            return jsonify({
                'success': False,
                'message': 'You do not have permission to access this project'
            }), 403
        
        # Set project ID in session - using the same session keys as workflow_step expects
        session['current_project_id'] = project_id
        session['project_id'] = project_id
        
        # Set continue flag
        session['is_continue_drawing'] = True
        
        return jsonify({
            'success': True,
            'redirect_url': url_for('main.workflow_step', step=2)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@bp.route("/workflow/step/<int:step>", methods=["GET", "POST"])
@login_required
def workflow_step(step):
    """
    Main workflow step handler with grid-first Add More junction boxes flow
    Uses junction_row field to determine number of cable popups
    """
    # Check if user can edit this step - viewers cannot edit
    can_edit = can_edit_step(step, current_user)
    
    # ==================== PROJECT SELECTION LOGIC ====================
    # Always handle project_id from query parameter (from "Continue with Existing Data")
    if request.args.get('project_id'):
        try:
            project_id_from_args = int(request.args.get('project_id'))
            project = Project.query.get(project_id_from_args)
            
            if project:
                # Check if user has permission to access this project
                has_permission = False
                
                # Admin (role 4) and creators (role 1) can access any project
                if current_user.role_name in ['1', '4']:
                    has_permission = True
                # Regular users can only access if assigned to project
                elif current_user.id in [u.id for u in project.users]:
                    has_permission = True
                
                if has_permission:
                    session['current_project_id'] = project_id_from_args
                    session['project_id'] = project_id_from_args
                    flash(f"Project '{project.name}' selected.", 'info')
                    print(f"✅ DEBUG: Set project from URL param: {project_id_from_args}")
                    
                    # Set continue drawing flag
                    session['is_continue_drawing'] = True
                else:
                    flash("You do not have permission to access this project.", 'danger')
                    return redirect(url_for("main.approval_tracking"))
            else:
                flash("Project not found.", 'danger')
                return redirect(url_for("main.approval_tracking"))
        except ValueError:
            flash("Invalid project ID.", 'danger')
            return redirect(url_for("main.approval_tracking"))
    
    project_id = get_current_project()
    if not project_id:
        print("❌ DEBUG: No project ID, redirecting to project selection")
        return redirect(url_for("main.new_drawing_selection"))
    
    print(f"✅ DEBUG: Using project ID: {project_id}")
    current_project = Project.query.get(project_id)
    if not current_project:
        print(f"❌ DEBUG: Project {project_id} not found in DB")
        session.pop('current_project_id', None)
        session.pop('project_id', None)
        return redirect(url_for("main.new_drawing_selection"))
    
    # Update to 9 steps total
    if step < 1 or step > 9:
        flash("Invalid workflow step")
        return redirect(url_for("main.approval_tracking"))
    
    # Corrected step mapping
    workflow_steps = {
        1: {"name": "StationDrawing", "display": "Station Information"},
        2: {"name": "junction_box", "display": "Location Box or CTR Setup"},
        3: {"name": "cable", "display": "Cable Information"},
        4: {"name": "terminal", "display": "Terminal Details"},
        5: {"name": "terminal_header", "display": "Terminal Headers"},
        6: {"name": "group", "display": "Group Information"},
        7: {"name": "choketable", "display": "Choke Table"},
        8: {"name": "resistortable", "display": "Resistor Table"},
        9: {"name": "cable_box", "display": "Relay Box"},
    }
    
    current_step = workflow_steps[step]
    
    sheet_name = current_step["name"]
    display_name = current_step["display"]
    model = MODEL_MAP.get(sheet_name)
    columns = SHEETS.get(sheet_name, [])
    
    
    # For step 1, ensure station drawing exists with project ID as station ID and auto-version
    if step == 1 and sheet_name == "StationDrawing":
        current_project = Project.query.get(project_id)
        project_name = current_project.name if current_project else "KHEDBRAHMA"
       
        # Check if StationDrawing already exists
        existing_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
        if not existing_drawing:
            # Create new with auto-version starting at 0
            seed_default_station_drawing(project_id, project_name)
        else:
            # Update version automatically for existing drawings
            pdf_count = GeneratedPDF.query.filter_by(project_id=project_id).count()
            current_version = str(pdf_count)
           
            if existing_drawing.version != current_version:
                existing_drawing.version = current_version
                db.session.commit()
                print(f"✅ Updated StationDrawing version to {current_version} for project {project_id} (PDF count: {pdf_count})")
           
            # Update checksum from latest PDF
            latest_pdf = GeneratedPDF.query.filter_by(project_id=project_id).order_by(GeneratedPDF.id.desc()).first()
            if latest_pdf and latest_pdf.checksum_md5:
                new_checksum = latest_pdf.checksum_md5
            else:
                new_checksum = "NO PDF GENERATED"
           
            if existing_drawing.checksum != new_checksum:
                existing_drawing.checksum = new_checksum
                db.session.commit()
                print(f"✅ Updated StationDrawing checksum to {new_checksum} for project {project_id}")
    
    start_junction_number = 1
    if step == 2:
        try:
            existing_boxes = JunctionBox.query.filter_by(project_id=project_id).all()
            if existing_boxes:
                max_num = 0
                for box in existing_boxes:
                    try:
                        val = box.junction_id
                        if val:
                            num = int(str(val).strip())
                            max_num = max(max_num, num)
                    except (ValueError, TypeError, AttributeError):
                        continue
                if max_num > 0:
                    start_junction_number = max_num + 1
        except Exception as e:
            print(f"Error calculating start_junction_number: {e}")
            start_junction_number = 1
    
    junction_count = session.get('junction_count', 0)
   
    # Initialize default_junction_count for session-based pre-fill
    default_junction_count = 0
   
    # Check if user has existing junctions
    has_existing_junctions = False
    is_continue_drawing = session.get('is_continue_drawing', False)
   
    if step == 2:
        existing_junction_count = JunctionBox.query.filter_by(project_id=project_id).count()
        has_existing_junctions = existing_junction_count > 0
        print(f"🔍 DEBUG Step 2: Project ID = {project_id}")
        print(f"🔍 DEBUG Step 2: Found {existing_junction_count} existing junctions for project {project_id}")
       
        # Load existing junctions for display
        existing_junctions = JunctionBox.query.filter_by(project_id=project_id).all()
        print(f"🔍 DEBUG Step 2: Existing junctions: {[(j.id, j.junction_id, j.junction_name) for j in existing_junctions]}")
    
    # Build data for step 3 (cable) grid
    junction_cables = []
    if step == 3:
        try:
            jboxes = JunctionBox.query.filter_by(project_id=project_id).all()
            for jbox in jboxes:
                jn = None
                for attr in ['junction_name', 'Junction Name', 'JunctionName', 'junction_id', 'Junction Id']:
                    try:
                        val = getattr(jbox, attr, None)
                        if val and str(val).strip() and str(val).strip() != '-':
                            jn = str(val).strip()
                            break
                    except:
                        continue
                if not jn:
                    jn = f"Junction_{jbox.id}"
                
                js = None
                for attr in ['junction_size', 'Junction Size', 'JunctionSize', 'junction_Size']:
                    try:
                        val = getattr(jbox, attr, None)
                        if val and str(val).strip() and str(val).strip() != '-':
                            js = val
                            break
                    except:
                        continue
                
                # Get junction_row count (number of cable popups to show)
                row_count = None
                for attr in ['junction_row', 'Junction Row', 'JunctionRow']:
                    try:
                        val = getattr(jbox, attr, None)
                        if val:
                            row_count = int(str(val).strip())
                            break
                    except (ValueError, TypeError):
                        continue
                
                if row_count and row_count > 0:
                    junction_cables.append({
                        'junction_name': jn,
                        'junction_size': js,
                        'row_count': row_count
                    })
        except Exception as e:
            print(f"Error loading junction cables: {e}")
    
    def _is_meaningful(value):
        """Check if a value is meaningful (not empty, not dash, not zero)"""
        if value is None:
            return False
        s = str(value).strip()
        if s in ("", "-"):
            return False
        try:
            if float(s) == 0.0:
                return False
        except Exception:
            pass
        return True
    
    def _norm(colname: str) -> str:
        """Normalize column name"""
        return str(colname).strip().lower().replace(" ", "_")
    
    pdf_history = []
    try:
        if project_id:
            pdf_records = GeneratedPDF.query.filter_by(project_id=project_id)\
                .order_by(GeneratedPDF.id.desc())\
                .limit(10).all()
            for rec in pdf_records:
                pdf_history.append({
                    "filename": rec.pdf_filename,
                    "view_url": url_for("main.pdf_view", filename=rec.pdf_filename),
                    "inline_url": url_for("main.inline_pdf", filename=rec.pdf_filename),
                    "download_url": url_for("main.download_pdf", filename=rec.pdf_filename),
                    "created": rec.created_at.strftime("%d %b %Y %I:%M %p") if getattr(rec, "created_at", None) else "",
                    "size_kb": (rec.file_size // 1024) if getattr(rec, "file_size", None) else None
                })
    except Exception:
        pdf_history = []
    
    # ==================== POST REQUEST HANDLING ====================
    if request.method == "POST":
        action = request.form.get('action')
        
        # Check if user can perform this action - viewers cannot perform any actions
        if not can_edit:
            flash("You are not allowed to modify this step.")
            return redirect(url_for('main.workflow_step', step=step))
        
        if action == 'prev':
            # Navigate to previous step
            if step > 1:
                return redirect(url_for('main.workflow_step', step=step - 1))
        
        elif action == 'save_draft':
            # Handle Save Draft for step 2
            if step == 2:
                try:
                    # Get project information
                    project_id = get_current_project()
                    
                    # Get junction count (only save this to the model)
                    junction_count = int(request.form.get('junction_count', 0))
                    
                    # Store start_junction_number in session only (not in database)
                    if 'start_junction_number' in request.form:
                        start_junction_number = int(request.form.get('start_junction_number', 1))
                        session['draft_start_junction_number'] = start_junction_number
                    
                    # Check for existing draft
                    draft = CableLocationAddition.query.filter_by(
                        project_id=project_id,
                        is_draft=True
                    ).first()
                    
                    if draft:
                        # Update existing draft - only junction_count
                        draft.junction_count = junction_count
                        draft.updated_at = get_ist_now()
                        print(f"✅ Updated existing draft for project {project_id}")
                    else:
                        # Create new draft - only junction_count
                        draft = CableLocationAddition(
                            project_id=project_id,
                            junction_count=junction_count,
                            is_draft=True
                        )
                        db.session.add(draft)
                        print(f"✅ Created new draft for project {project_id}")
                    
                    db.session.commit()
                    flash('✅ Draft saved successfully! You can continue later.', 'success')
                    
                    # Set flag to show success message in template
                    session['draft_saved'] = True
                    
                    # Also store junction_count in session for immediate use
                    session['junction_count'] = junction_count
                    
                except Exception as e:
                    db.session.rollback()
                    flash(f'❌ Error saving draft: {str(e)}', 'error')
                
                # Redirect back to the same page
                return redirect(url_for('main.workflow_step', step=2))
        
        elif action == 'add_more_junctions':
            # Handle adding more junctions - SHOW GRID FIRST, then popups after save
            if step == 2:
                try:
                    additional_count = int(request.form.get('junction_count', 0))
                except Exception:
                    additional_count = 0
                
                if additional_count > 0:
                    session['junction_count'] = additional_count
                    session['start_junction_number'] = start_junction_number
                    # NEW: Store the count in session for pre-fill on "Previous"
                    session['last_add_junction_count'] = additional_count
                    # Set flag to show grid first (not popups yet)
                    session['show_more_junction_grid'] = True
                    # Ensure we do NOT trigger fresh_step_2 (which shows popups immediately)
                    session.pop('fresh_step_2', None)
                    session.pop('is_continue_drawing', None)
                    flash(f"✅ Adding {additional_count} junction boxes starting from #{start_junction_number}")
                    return redirect(url_for('main.workflow_step', step=2))
                else:
                    flash("⚠️ Please enter a number greater than 0 to add junction boxes.")
                    return redirect(url_for('main.workflow_step', step=2))
        
        elif action == 'show_add_more':
            # Handle "Previous" button from grid - retrieve last count for pre-fill
            if step == 2:
                # Retrieve last count from session for pre-fill
                default_junction_count = session.get('last_add_junction_count', 0)
                # Set junction_count=0 to render the prompt block
                junction_count = 0
                # Clear grid flags to show the prompt
                session.pop('show_more_junction_grid', None)
                session.pop('fresh_step_2', None)
                flash("Returned to location count selection")
                # Stay on same page to show prompt with pre-filled value
                return redirect(url_for('main.workflow_step', step=2))
        
        elif action == 'next':
            if step == 2:
                # New Drawing flow: user submitted count from initial prompt
                if 'junction_count' in request.form:
                    try:
                        junction_count = int(request.form.get('junction_count', 0))
                    except Exception:
                        junction_count = 0
                    session['junction_count'] = junction_count
                    session['start_junction_number'] = start_junction_number
                    # Show popups immediately for new drawing
                    session['fresh_step_2'] = True
                    # Clear any add-more grid state
                    session.pop('show_more_junction_grid', None)
                    # Stay on same page to show grid and popups
                    return redirect(url_for('main.workflow_step', step=2))
                else:
                    # Continue Drawing flow - skip to next step
                    session['fresh_step_3'] = True
                    session.pop('is_continue_drawing', None)
                    return redirect(url_for('main.workflow_step', step=3))
            
            # Navigate to next step
            if step < 9:
                if step == 3:
                    return redirect(url_for('main.workflow_step', step=4))
                return redirect(url_for('main.workflow_step', step=step + 1))
            else:
                flash("🎉 All workflow steps completed")
                return redirect(url_for("main.approval_tracking"))
        
        elif action == 'add':
            if step == 1 and model:
                # Station Drawing step - update or create
                data = {'project_id': project_id}
                has_meaningful = False
                for col in columns:
                    raw = request.form.get(col, None)
                    value = (raw or "").strip() if raw is not None else None
                    data[col] = value if value else None
                    if _is_meaningful(value):
                        has_meaningful = True
                
                if has_meaningful:
                    try:
                        existing = model.query.filter_by(project_id=project_id).first()
                        if existing:
                            for col in columns:
                                setattr(existing, col, data[col])
                            db.session.commit()
                            flash(f"✅ Updated {display_name}")
                        else:
                            db.session.add(model(**data))
                            db.session.commit()
                            flash(f"✅ Added 1 record to {display_name}")
                    except Exception as e:
                        db.session.rollback()
                        flash(f"❌ Error adding record: {str(e)}")
                else:
                    flash("No changes to save (all fields blank or zero).")
            
            elif model and step == 2: # JunctionBox addition with naming convention
                data = {'project_id': project_id}
                has_meaningful = False
                
                for col in columns:
                    raw = request.form.get(col, None)
                    value = (raw or "").strip() if raw is not None else None
                    
                    # ENFORCE LOCATION NAMING CONVENTION FOR JUNCTION BOXES
                    if col == 'junction_name' and 'junction_size' in request.form:
                        location_name = value
                        location_size = request.form.get('junction_size', '')
                        
                        # Format the location name with (F) or (H) suffix
                        formatted_name = format_location_name(location_name, location_size)
                        data[col] = formatted_name
                        print(f"🔍 DEBUG: Formatted location name: '{location_name}' -> '{formatted_name}' for size '{location_size}'")
                    else:
                        data[col] = value if value else None
                    
                    if _is_meaningful(value):
                        has_meaningful = True
                
                if has_meaningful:
                    try:
                        db.session.add(model(**data))
                        db.session.commit()
                        flash(f"✅ Added 1 record to {display_name}")
                        
                        # ✅ UPDATE PROJECT'S JUNCTION DATA
                        try:
                            current_project = Project.query.get(project_id)
                            if current_project:
                                current_project.update_junction_data()
                                print(f"✅ Updated junction_data for project {project_id}")
                        except Exception as update_error:
                            print(f"⚠️ Warning: Could not update junction_data: {update_error}")
                        
                        # Clear draft after successful save
                        draft = CableLocationAddition.query.filter_by(
                            project_id=project_id,
                            is_draft=True
                        ).first()
                        if draft:
                            db.session.delete(draft)
                            db.session.commit()
                            print(f"✅ Cleared draft after saving locations")
                    except Exception as e:
                        db.session.rollback()
                        flash(f"❌ Error adding record: {str(e)}")
                else:
                    flash("No changes to save (all fields blank or zero).")
            
            elif model:
                # Generic add for other steps
                data = {'project_id': project_id}
                has_meaningful = False
                for col in columns:
                    raw = request.form.get(col, None)
                    value = (raw or "").strip() if raw is not None else None
                    data[col] = value if value else None
                    if _is_meaningful(value):
                        has_meaningful = True
                
                if has_meaningful:
                    try:
                        db.session.add(model(**data))
                        db.session.commit()
                        flash(f"✅ Added 1 record to {display_name}")
                    except Exception as e:
                        db.session.rollback()
                        flash(f"❌ Error adding record: {str(e)}")
                else:
                    flash("No changes to save (all fields blank or zero).")
            return redirect(url_for('main.workflow_step', step=step))
        
        # Default redirect for any other POST action
        return redirect(url_for('main.workflow_step', step=step))
    
    # ==================== GET REQUEST RENDERING ====================
   
    # Load current sheet rows
    rows = []
    cable_name_map = {} # Initialize empty cable name map
   
    if model:
        records = model.query.filter_by(project_id=project_id).order_by(model.id).all()
        rows = [{"id": r.id, **{col: getattr(r, col, '') for col in columns}} for r in records]
       
        # For steps 4-8 that use cable_id, get cable names mapping
        if (step == 4 and sheet_name == "terminal") or (step == 5 and sheet_name == "terminal_header") or (step == 6 and sheet_name == "group") or (step == 7 and sheet_name == "choketable") or (step == 8 and sheet_name == "resistortable"):
            try:
                # Get all cables for this project
                cables = Cable.query.filter_by(project_id=project_id).all()
                # Create mapping from cable_id to cable_name
                cable_name_map = {str(cable.cable_id): cable.cable_name for cable in cables}
                print(f"🔍 DEBUG Step {step}: Loaded {len(cable_name_map)} cable names for project {project_id}")
               
                # FOR RESISTOR TABLE: Also check if resistors have cable_name in their own table
                if step == 8 and sheet_name == "resistortable":
                    # If ResistorTable model has cable_name field, use it
                    try:
                        for row in rows:
                            cable_id = row.get('cable_id')
                            if cable_id:
                                # Check if resistor has cable_name stored
                                resistor_record = ResistorTable.query.filter_by(
                                    project_id=project_id,
                                    id=row['id']
                                ).first()
                                if resistor_record and hasattr(resistor_record, 'cable_name') and resistor_record.cable_name:
                                    # Use resistor's cable_name if available
                                    cable_name_map[str(cable_id)] = resistor_record.cable_name
                    except Exception as e:
                        print(f"⚠️ WARNING: Could not check resistor cable_name: {str(e)}")
            except Exception as e:
                print(f"❌ ERROR loading cable names for step {step}: {str(e)}")
                cable_name_map = {}
    
    # Update progress calculation for 9 steps
    progress_pct = int(step / 9 * 100)
    
    # Check for fresh navigation flags
    fresh_navigation = False
    if step == 2:
        fresh_navigation = session.pop('fresh_step_2', False)
    elif step == 3:
        fresh_navigation = session.pop('fresh_step_3', False)
   
    # Determine if we should show the "Add More" grid first (only step 2)
    show_more_junction_grid = False
    if step == 2:
        # Pop so it's one-time; page load consumes it and renders the grid
        show_more_junction_grid = session.pop('show_more_junction_grid', False)
   
    # Determine if we should show popups (highest priority for new drawing)
    # but NOT when add-more grid is active
    show_popups = False
    if step == 2:
        # Only show popups if:
        # 1. Fresh navigation (new drawing)
        # 2. Junction count > 0
        # 3. NOT showing add-more grid
        if fresh_navigation and junction_count > 0 and not show_more_junction_grid:
            show_popups = True
            print(f"🔔 DEBUG: Showing popups - junction_count={junction_count}, fresh_navigation={fresh_navigation}")
    
    # NEW: For step 2, check for existing draft and load data
    draft_data = None
    if step == 2:
        # Check for existing draft
        draft = CableLocationAddition.query.filter_by(
            project_id=project_id,
            is_draft=True
        ).first()
        
        if draft:
            draft_data = {
                'junction_count': draft.junction_count,
                'created_at': draft.created_at,
                'updated_at': draft.updated_at
            }
            
            # Load draft values into session if not already set
            if not session.get('junction_count'):
                session['junction_count'] = draft.junction_count
                junction_count = draft.junction_count
            
            # Get start_junction_number from session or calculate it
            if 'draft_start_junction_number' in session:
                start_junction_number = session['draft_start_junction_number']
            else:
                # Calculate it based on existing junctions
                try:
                    existing_boxes = JunctionBox.query.filter_by(project_id=project_id).all()
                    if existing_boxes:
                        max_num = 0
                        for box in existing_boxes:
                            try:
                                val = box.junction_id
                                if val:
                                    num = int(str(val).strip())
                                    max_num = max(max_num, num)
                            except (ValueError, TypeError, AttributeError):
                                continue
                        if max_num > 0:
                            start_junction_number = max_num + 1
                except Exception as e:
                    print(f"Error calculating start_junction_number: {e}")
                    start_junction_number = 1
            
            print(f"📋 Loaded draft: count={draft.junction_count}")
            
            # If draft exists and no other flags, show popups
            if draft.junction_count > 0 and not show_more_junction_grid and not fresh_navigation:
                show_popups = True
                print(f"🔔 DEBUG: Showing popups from draft - junction_count={draft.junction_count}")
    
    # Set default_junction_count from session if not already set
    if step == 2 and not default_junction_count:
        default_junction_count = session.get('last_add_junction_count', 0)
    
    # In the GET request rendering section, add:
    draft_saved = session.pop('draft_saved', False)

    return render_template(
        "workflow_step.html",
        current_project=current_project,
        step=step,
        total_steps=9,
        sheet_name=sheet_name,
        display_name=display_name,
        columns=columns,
        rows=rows,
        junction_count=junction_count,
        start_junction_number=start_junction_number,
        junction_cables=junction_cables if step == 3 else [],
        has_existing_junctions=has_existing_junctions if step == 2 else False,
        is_continue_drawing=is_continue_drawing if step == 2 else False,
        show_popups=show_popups if step == 2 else False,
        show_more_junction_grid=show_more_junction_grid if step == 2 else False,
        has_prev=step > 1,
        has_next=step < 9,
        progress_pct=progress_pct,
        hint=HEADER_HINTS.get(sheet_name, ""),
        pdf_history=pdf_history,
        fresh_navigation=fresh_navigation,
        default_junction_count=default_junction_count,
        cable_name_map=cable_name_map,
        can_edit=can_edit,
        draft_saved=draft_saved,
        draft_data=draft_data  # NEW: Pass draft data to template
    )



    
# New route: Set continue drawing flag (called from index page)
@bp.route("/set_continue_drawing", methods=["POST"])
@login_required
def set_continue_drawing():
    """Set flag when user clicks Continue Drawing button"""
    session['is_continue_drawing'] = True
    return jsonify({"success": True})

# New route: Set new drawing flag (called from index page)
@bp.route("/set_new_drawing", methods=["POST"])
@login_required
def set_new_drawing():
    """Set flag when user clicks New Drawing button"""
    session['is_continue_drawing'] = False
    return jsonify({"success": True})

def insert_row(sheet_name, row_data):
    """
    Simple insert_row function - replace with your actual Google Sheets logic
    """
    try:
        # TODO: Replace with your actual Google Sheets insertion code
        print(f"Would insert into {sheet_name}: {row_data}")
        return True
    except Exception as e:
        print(f"Error in insert_row: {str(e)}")
        return False

# NEW: Group AJAX route - ADD THIS ROUTE
@bp.route('/add_group_ajax', methods=['POST'])
@login_required
def add_group_ajax():
    """Handle AJAX submission for adding group configuration"""
    # Check if user can edit step 6 - viewers cannot edit
    if not can_edit_step(6, current_user):
        return jsonify({'success': False, 'message': 'You are not allowed to update this step.'}), 403
       
    try:
        data = request.get_json()
        project_id = get_current_project()
       
        print(f"🔍 DEBUG add_group_ajax: project_id={project_id}, data={data}")
       
        if not project_id:
            return jsonify({'success': False, 'message': 'No active project selected'}), 400
        # Required fields validation
        required_fields = ['cable_id', 'group_id', 'terminal_no', 'input_output']
        missing_fields = [field for field in required_fields if field not in data or not data[field]]
       
        if missing_fields:
            return jsonify({
                'success': False,
                'message': f'Missing required fields: {", ".join(missing_fields)}'
            }), 400
       
        # FIX: Convert cable_id to string for consistent database query
        cable_id_str = str(data['cable_id'])
       
        # Check if cable already exists in Cable table
        cable_exists = Cable.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).first()
       
        # Handle cable_name if provided
        cable_name = data.get('cable_name', '').strip()
        if cable_name and not cable_exists:
            # Create a new cable entry
            new_cable = Cable(
                project_id=project_id,
                cable_id=cable_id_str, # Use string version
                cable_name=cable_name,
                created_date=get_ist_now()
            )
            db.session.add(new_cable)
            print(f"✅ DEBUG: Created new cable entry: '{cable_id_str}' - {cable_name}")
        elif cable_name and cable_exists and cable_exists.cable_name != cable_name:
            # Update existing cable name
            cable_exists.cable_name = cable_name
            print(f"✅ DEBUG: Updated existing cable name: '{cable_id_str}' - {cable_name}")
       
        # Clean and validate terminal numbers
        terminal_no = data['terminal_no']
       
        # Remove any spaces and ensure proper comma format
        terminal_no = terminal_no.replace(' ', '').replace(',,', ',').strip(',')
       
        # Validate that it contains only numbers and commas
        if not re.match(r'^[\d,]+$', terminal_no):
            return jsonify({
                'success': False,
                'message': 'Invalid terminal numbers format. Please use only numbers and commas (e.g., 1,2,4,54).'
            }), 400
       
        # Create new group record
        new_group = Group(
            project_id=project_id,
            cable_id=cable_id_str, # Use string version
            group_id=data['group_id'],
            terminal_no=terminal_no, # Store as comma-separated string
            input_output=data['input_output'],
            text=data.get('text', ''),
            created_date=get_ist_now()
        )
       
        db.session.add(new_group)
        db.session.commit()
       
        print(f"✅ DEBUG: Saved group {new_group.id} to project {project_id}")
       
        return jsonify({
            'success': True,
            'message': 'Group configuration added successfully',
            'group_id': new_group.id
        })
           
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR saving group: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error saving group configuration: {str(e)}'
        }), 500

@bp.route('/add_choke_ajax', methods=['POST'])
@login_required
def add_choke_ajax():
    # Check if user can edit step 7 - viewers cannot edit
    if not can_edit_step(7, current_user):
        return jsonify({'success': False, 'message': 'You are not allowed to update this step.'}), 403
       
    try:
        data = request.get_json()
        project_id = get_current_project()
       
        print(f"🔍 DEBUG add_choke_ajax: project_id={project_id}, data={data}")
       
        if not project_id:
            return jsonify({'success': False, 'message': 'No active project selected'}), 400
        required_fields = ['cable_id', 'choke_id']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'success': False, 'message': f'Missing required field: {field}'}), 400
       
        # FIX: Convert cable_id to string for consistent database query
        cable_id_str = str(data['cable_id'])
       
        # Check if cable already exists in Cable table
        cable_exists = Cable.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).first()
       
        # Handle cable_name if provided
        cable_name = data.get('cable_name', '').strip()
        if cable_name and not cable_exists:
            # Create a new cable entry
            new_cable = Cable(
                project_id=project_id,
                cable_id=cable_id_str, # Use string version
                cable_name=cable_name,
                created_date=get_ist_now()
            )
            db.session.add(new_cable)
            print(f"✅ DEBUG: Created new cable entry: '{cable_id_str}' - {cable_name}")
        elif cable_name and cable_exists and cable_exists.cable_name != cable_name:
            # Update existing cable name
            cable_exists.cable_name = cable_name
            print(f"✅ DEBUG: Updated existing cable name: '{cable_id_str}' - {cable_name}")
       
        new_choke = ChokeTable(
            project_id=project_id,
            cable_id=cable_id_str, # Use string version
            choke_id=data['choke_id'],
            input_terminal=data.get('input_terminal', ''),
            output_terminal=data.get('output_terminal', ''),
            terminal_name=data.get('terminal_name', ''),
            output_type=data.get('output_type', ''),
            output_text=data.get('output_text', ''),
            output_connected=data.get('output_connected', ''),
            created_date=get_ist_now()
        )
       
        db.session.add(new_choke)
        db.session.commit()
       
        print(f"✅ DEBUG: Saved choke {new_choke.id} to project {project_id}")
       
        return jsonify({
            'success': True,
            'message': 'Choke data saved successfully',
            'choke_id': new_choke.id
        })
           
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR saving choke data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
   
@bp.route('/get_chokes_for_cable')
@login_required
def get_chokes_for_cable():
    """Fetch existing chokes for a cable"""
    cable_id = request.args.get('cable_id')
    project_id = get_current_project()
   
    print(f"🔍 DEBUG get_chokes_for_cable: cable_id={cable_id}, project_id={project_id}")
   
    if not project_id:
        return jsonify({'error': 'No project selected'}), 400
   
    if not cable_id:
        return jsonify([])
   
    try:
        # Convert cable_id to string for consistent query
        cable_id_str = str(cable_id)
       
        # Query chokes for the current project and cable
        chokes = ChokeTable.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).all()
       
        choke_data = []
        for choke in chokes:
            choke_data.append({
                'id': choke.id,
                'cable_id': choke.cable_id,
                'choke_id': choke.choke_id,
                'input_terminal': choke.input_terminal,
                'output_terminal': choke.output_terminal,
                'terminal_name': choke.terminal_name
            })
       
        print(f"✅ DEBUG: Found {len(choke_data)} chokes for cable {cable_id}")
       
        return jsonify(choke_data)
   
    except Exception as e:
        print(f"❌ ERROR fetching chokes for cable {cable_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to fetch chokes'}), 500

@bp.route('/get_resistors_for_cable')
@login_required
def get_resistors_for_cable():
    """Fetch existing resistors for a cable"""
    cable_id = request.args.get('cable_id')
    project_id = get_current_project()
   
    print(f"🔍 DEBUG get_resistors_for_cable: cable_id={cable_id}, project_id={project_id}")
   
    if not project_id:
        return jsonify({'error': 'No project selected'}), 400
   
    if not cable_id:
        return jsonify([])
   
    try:
        # Convert cable_id to string for consistent query
        cable_id_str = str(cable_id)
       
        # Query resistors for the current project and cable
        resistors = ResistorTable.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).all()
       
        resistor_data = []
        for resistor in resistors:
            resistor_data.append({
                'id': resistor.id,
                'cable_id': resistor.cable_id,
                'resistor_id': resistor.resistor_id,
                'input_terminal': resistor.input_terminal,
                'output_terminal': resistor.output_terminal,
                'resistor_name': resistor.resistor_name
            })
       
        print(f"✅ DEBUG: Found {len(resistor_data)} resistors for cable {cable_id}")
       
        return jsonify(resistor_data)
   
    except Exception as e:
        print(f"❌ ERROR fetching resistors for cable {cable_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to fetch resistors'}), 500

@bp.route('/check_terminals_for_cable')
@login_required
def check_terminals_for_cable():
    """Check if terminals exist for a cable"""
    cable_id = request.args.get('cable_id')
    project_id = get_current_project()
   
    if not project_id or not cable_id:
        return jsonify({'exists': False})
   
    try:
        # Convert cable_id to string for consistent query
        cable_id_str = str(cable_id)
       
        terminal_count = Terminal.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).count()
       
        return jsonify({
            'exists': terminal_count > 0,
            'count': terminal_count
        })
   
    except Exception as e:
        print(f"Error checking terminals for cable {cable_id}: {str(e)}")
        return jsonify({'exists': False})

@bp.route('/get_groups_for_cable')
@login_required
def get_groups_for_cable():
    """Fetch existing groups for a cable"""
    cable_id = request.args.get('cable_id')
    project_id = get_current_project()
   
    print(f"🔍 DEBUG get_groups_for_cable: cable_id={cable_id}, project_id={project_id}")
   
    if not project_id:
        return jsonify({'error': 'No project selected'}), 400
   
    if not cable_id:
        return jsonify([])
   
    try:
        # Convert cable_id to string for consistent query
        cable_id_str = str(cable_id)
       
        # Query groups for the current project and cable
        groups = Group.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).all()
       
        group_data = []
        for group in groups:
            group_data.append({
                'id': group.id,
                'cable_id': group.cable_id,
                'group_id': group.group_id,
                'terminal_no': group.terminal_no,
                'input_output': group.input_output,
                'text': group.text
            })
       
        print(f"✅ DEBUG: Found {len(group_data)} groups for cable {cable_id}")
       
        return jsonify(group_data)
   
    except Exception as e:
        print(f"❌ ERROR fetching groups for cable {cable_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to fetch groups'}), 500

# In the imports section, add this at the top if not already there:
import json

# Update the add_resistor_ajax route to handle cable_name:
@bp.route('/add_resistor_ajax', methods=['POST'])
@login_required
def add_resistor_ajax():
    # Check if user can edit step 8 - viewers cannot edit
    if not can_edit_step(8, current_user):
        return jsonify({'success': False, 'message': 'You are not allowed to update this step.'}), 403
       
    try:
        data = request.get_json()
        project_id = get_current_project()
       
        print(f"🔍 DEBUG add_resistor_ajax: project_id={project_id}, data={data}")
       
        if not project_id:
            return jsonify({'success': False, 'message': 'No active project selected'}), 400
        # Required fields validation
        required_fields = ['cable_id', 'resistor_id']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'success': False, 'message': f'Missing required field: {field}'}), 400
       
        # FIX: Convert cable_id to string for consistent database query
        cable_id_str = str(data['cable_id'])
       
        # Check if cable already exists in Cable table
        cable_exists = Cable.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).first()
       
        # Create new resistor record
        new_resistor = ResistorTable(
            project_id=project_id,
            cable_id=cable_id_str, # Use string version
            resistor_id=data['resistor_id'],
            input_terminal=data.get('input_terminal', ''),
            output_terminal=data.get('output_terminal', ''),
            resistor_name=data.get('resistor_name', ''),
            created_date=get_ist_now()
        )
       
        # If cable_name is provided and cable doesn't exist, create/update it in Cable table
        cable_name = data.get('cable_name', '').strip()
        if cable_name and not cable_exists:
            # Create a new cable entry
            new_cable = Cable(
                project_id=project_id,
                cable_id=cable_id_str, # Use string version
                cable_name=cable_name,
                created_date=get_ist_now()
            )
            db.session.add(new_cable)
            print(f"✅ DEBUG: Created new cable entry: '{cable_id_str}' - {cable_name}")
        elif cable_name and cable_exists and cable_exists.cable_name != cable_name:
            # Update existing cable name
            cable_exists.cable_name = cable_name
            print(f"✅ DEBUG: Updated existing cable name: '{cable_id_str}' - {cable_name}")
       
        db.session.add(new_resistor)
        db.session.commit()
       
        print(f"✅ DEBUG: Saved resistor {new_resistor.id} to project {project_id}")
       
        return jsonify({
            'success': True,
            'message': 'Resistor data saved successfully',
            'resistor_id': new_resistor.id
        })
           
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR saving resistor data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@bp.route("/add_junctions_ajax", methods=["POST"])
@login_required
def add_junctions_ajax():
    """Handle AJAX submission for adding multiple junction boxes with naming convention enforcement"""
    # Check if user can edit step 2 - viewers cannot edit
    if not can_edit_step(2, current_user):
        return jsonify({"success": False, "message": "You are not allowed to update this step."}), 403
       
    try:
        data = request.get_json()
        project_id = get_current_project()
       
        print(f"🔍 DEBUG add_junctions_ajax: project_id={project_id}, data={data}")
       
        if not project_id:
            return jsonify({"success": False, "message": "No active project selected"}), 400
        if not data or 'junctions' not in data or not isinstance(data['junctions'], list):
            return jsonify({"success": False, "message": "Invalid or missing 'junctions' array in request"}), 400
        junctions = data.get('junctions', [])
        if not junctions:
            return jsonify({"success": False, "message": "Junctions array is empty"}), 400
        added_junctions = []
        for junction_data in junctions:
            # Validate required fields
            if not junction_data.get('junction_size'):
                return jsonify({"success": False, "message": "Junction size is required for all junctions"}), 400
            if not junction_data.get('station_id'):
                return jsonify({"success": False, "message": "Station ID is required for all junctions"}), 400
            # ENFORCE NAMING CONVENTION: Format location name with (F) or (H) suffix
            location_name = junction_data.get('junction_name', '')
            location_size = junction_data.get('junction_size', '')
           
            # Format the location name with proper suffix
            formatted_name = format_location_name(location_name, location_size)
           
            print(f"🔍 DEBUG: Formatted location name: '{location_name}' -> '{formatted_name}' for size '{location_size}'")
           
            # Map 'row' to 'junction_row' for model compatibility
            junction = JunctionBox(
                project_id=project_id,
                station_id=junction_data.get('station_id'),
                junction_id=junction_data.get('junction_id'),
                junction_name=formatted_name, # Use the formatted name
                junction_size=junction_data.get('junction_size'),
                latitude=junction_data.get('latitude'),
                longitude=junction_data.get('longitude'),
                junction_row=junction_data.get('row') or junction_data.get('junction_row', 'A'),
                created_date=get_ist_now()
            )
            db.session.add(junction)
            added_junctions.append(junction)
        db.session.commit()
        # Prepare response with saved junction data
        response_junctions = [{
            "id": junction.id,
            "station_id": junction.station_id,
            "junction_id": junction.junction_id,
            "junction_name": junction.junction_name, # This will have (F)/(H) suffix
            "junction_size": junction.junction_size,
            "latitude": junction.latitude,
            "longitude": junction.longitude,
            "row": junction.junction_row # Return 'row' to match frontend
        } for junction in added_junctions]
        print(f"✅ DEBUG: Saved {len(added_junctions)} junctions to project {project_id}")
       
        return jsonify({
            "success": True,
            "message": f"Successfully added {len(added_junctions)} junction boxes",
            "junctions": response_junctions
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR adding junction boxes: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    
@bp.route("/test_approval/<int:pdf_id>")
@login_required
def test_approval(pdf_id):
    """Test route to check approval logic"""
    pdf = GeneratedPDF.query.get_or_404(pdf_id)
    
    print(f"PDF ID: {pdf.id}")
    print(f"Level1 Status: {pdf.level1_status}")
    print(f"Level2 Status: {pdf.level2_status}")
    print(f"Level3 Status: {pdf.level3_status}")
    print(f"User Designation: {current_user.designation}")
    print(f"Can Level1 Approve: {pdf.can_level1_approve()}")
    
    return jsonify({
        'pdf_id': pdf.id,
        'level1_status': pdf.level1_status,
        'level2_status': pdf.level2_status,
        'level3_status': pdf.level3_status,
        'user_designation': current_user.designation,
        'can_level1_approve': pdf.can_level1_approve(),
        'can_level2_approve': pdf.can_level2_approve(),
        'can_level3_approve': pdf.can_level3_approve()
    })

@bp.route('/get_terminals_for_cable')
@login_required
def get_terminals_for_cable():
    cable_id = request.args.get('cable_id')
    project_id = get_current_project()
    print(f"Fetching terminals for cable_id: {cable_id}")
   
    if not cable_id:
        return jsonify([])
   
    try:
        # Convert cable_id to string for consistent query
        cable_id_str = str(cable_id)
       
        # Query your database - adjust based on your actual model
        terminals = Terminal.query.filter_by(
            project_id=project_id, # ADD PROJECT SCOPE
            cable_id=cable_id_str # Use string version
        ).all()
       
        terminal_data = []
        for terminal in terminals:
            terminal_data.append({
                'id': terminal.id,
                'cable_id': terminal.cable_id,
                'terminal_id': terminal.terminal_id,
                'terminal_no': terminal.terminal_no,
                'symbol': terminal.symbol,
                'input_left': terminal.input_left,
                'input_right': terminal.input_right,
                'spare': terminal.spare,
                'input_connected': terminal.input_connected,
                'output_connected': terminal.output_connected,
                'input_connected_extra': terminal.input_connected_extra,
                'output_connected_extra': terminal.output_connected_extra,
                'output_left': terminal.output_left,
                'output_right': terminal.output_right
            })
       
        print(f"Found {len(terminal_data)} terminals for cable {cable_id}")
        for term in terminal_data:
            print(f"Terminal: id={term['terminal_id']}, name={term['terminal_no']}")
       
        return jsonify(terminal_data)
   
    except Exception as e:
        print(f"Error fetching terminals for cable {cable_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify([])

@bp.app_template_filter('sort_terminal_headers')
def sort_terminal_headers_filter(headers):
    """Sort terminal headers by cable_id as integer"""
    if not headers:
        return []
   
    try:
        # Sort by cable_id as integer first, then by id
        return sorted(headers, key=lambda x: (
            int(x.get('cable_id', 0)) if x.get('cable_id') and str(x.get('cable_id')).isdigit() else float('inf'),
            x.get('id', 0)
        ))
    except (ValueError, TypeError):
        # Fallback to string sorting if conversion fails
        return sorted(headers, key=lambda x: (
            str(x.get('cable_id', '')),
            x.get('id', 0)
        ))

def get_next_cable_id(project_id=None):
    """Get the next available cable ID for a project"""
    try:
        if project_id:
            # Use SQLAlchemy query instead of raw SQL
            max_id = db.session.query(db.func.max(Cable.cable_id)).filter(
                Cable.project_id == project_id
            ).scalar()
        else:
            max_id = db.session.query(db.func.max(Cable.cable_id)).scalar()
       
        # Handle case where max_id is None (no cables yet)
        if max_id is None:
            return "1"
       
        # Convert to integer, increment, then back to string
        try:
            next_id = int(max_id) + 1
        except (ValueError, TypeError):
            # If cable_id is not a number, start from 1
            next_id = 1
           
        return str(next_id)
       
    except Exception as e:
        print(f"Error getting next cable ID: {str(e)}")
        return "1" # Default to "1" if error

@bp.route('/get_existing_cables')
@login_required
def get_existing_cables():
    try:
        project_id = get_current_project()
        if not project_id:
            print("❌ ERROR: No project_id found in session")
            return jsonify({'error': 'No project selected'}), 400
       
        print(f"🔍 DEBUG: Fetching cables and cable_boxes for project_id={project_id}")
       
        # Fetch all cables for the current project
        cables = Cable.query.filter_by(project_id=project_id).all()
        cable_list = []
       
        for cable in cables:
            cable_list.append({
                'id': cable.id,
                'cable_id': cable.cable_id,
                'cable_name': getattr(cable, 'cable_name', ''),
                'junction_name': cable.junction_name,
                'row': cable.row,
                'junction_box': cable.junction_box,
                'position': getattr(cable, 'position', '1'),
                'terminal': cable.terminal,
                'start_no': cable.start_no,
                'cable_type': 'cable' # Default type
            })
       
        # Fetch all cable_boxes for the current project
        cable_boxes = CableBox.query.filter_by(project_id=project_id).all()
        for cable_box in cable_boxes:
            cable_list.append({
                'id': cable_box.id,
                'cable_id': cable_box.cable_id,
                'cable_name': getattr(cable_box, 'cable_name', ''),
                'junction_name': cable_box.junction_name,
                'row': cable_box.row,
                'junction_box': cable_box.junction_box,
                'position': getattr(cable_box, 'position', '1'),
                'terminal': cable_box.terminal,
                'start_no': cable_box.start_no,
                'cable_type': 'cable_box' # Specific type
            })
       
        print(f"✅ DEBUG: Found {len(cable_list)} cables and cable_boxes for project {project_id}")
       
        if cable_list:
            for cable in cable_list:
                print(f" - {cable['cable_type']} ID: {cable['cable_id']}, Name: {cable['cable_name']}, Junction: {cable['junction_name']}")
        else:
            print("ℹ️ INFO: No cables or cable_boxes found for this project")
       
        return jsonify(cable_list)
   
    except Exception as e:
        print(f"❌ ERROR: Error fetching existing cables and cable_boxes: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Failed to fetch existing cables and cable_boxes'}), 500

@bp.route("/add_cable_ajax", methods=["POST"])
@login_required
def add_cable_ajax():
    """Handle AJAX submission for adding or updating cable with cable_name and cable_type"""
    # Check if user can edit step 3 - viewers cannot edit
    if not can_edit_step(3, current_user):
        return jsonify({"success": False, "message": "You are not allowed to update this step."}), 403
       
    try:
        data = request.get_json()
        
        # Try to get project_id from session first, then from JSON data
        project_id = get_current_project()
        if not project_id and 'project_id' in data:
            project_id = data.get('project_id')
       
        print(f"🔍 DEBUG add_cable_ajax: project_id={project_id}, data={data}")
       
        if not project_id:
            return jsonify({"success": False, "message": "No active project selected"})
       
        # Convert cable_id to string for consistent database query
        cable_id = str(data.get('cable_id'))
        if not cable_id:
            return jsonify({"success": False, "message": "Cable ID is required"})
       
        # Get cable_type from frontend, default to 'cable'
        cable_type = data.get('cable_type', 'cable')
       
        # Calculate cable name if not provided
        cable_name = data.get('cable_name')
        
        # Define box types that should use the CableBox model
        box_types = ['cable_box', 'relay_box']
        
        if not cable_name and cable_type not in box_types:
            start_no = int(data.get('start_no', 1))
            terminal = int(data.get('terminal', 12))
            end_no = start_no + terminal - 1
            row = data.get('row', 'A')
            cable_name = f"{row} T{start_no}-{end_no}"
        elif not cable_name:
            cable_name = ""
       
        # Check if cable already exists
        if cable_type in box_types:
            existing_cable = CableBox.query.filter_by(
                project_id=project_id,
                cable_id=cable_id
            ).first()
        else:
            existing_cable = Cable.query.filter_by(
                project_id=project_id,
                cable_id=cable_id
            ).first()
        
        if existing_cable:
            # UPDATE existing cable
            print(f"🔄 Updating existing cable: project_id={project_id}, cable_id={cable_id}")
            
            # Update fields
            existing_cable.cable_name = cable_name
            existing_cable.junction_box = data.get('junction_box', existing_cable.junction_box)
            existing_cable.junction_name = data.get('junction_name', existing_cable.junction_name)
            existing_cable.row = data.get('row', existing_cable.row)
            existing_cable.position = str(data.get('position', existing_cable.position))
            existing_cable.terminal = str(data.get('terminal', existing_cable.terminal))
            existing_cable.start_no = str(data.get('start_no', existing_cable.start_no))
            
            # Update cable_type for box types
            if cable_type in box_types:
                existing_cable.cable_type = cable_type
            elif hasattr(existing_cable, 'cable_type'):
                existing_cable.cable_type = cable_type
            
            db.session.commit()
            
            print(f"✅ Updated cable: id={existing_cable.id}, cable_id={existing_cable.cable_id}, type={cable_type}")
            
            return jsonify({
                "success": True,
                "message": f"Cable {cable_id} updated successfully!",
                "cable_id": cable_id,
                "cable_name": cable_name,
                "cable_type": cable_type,
                "row_id": existing_cable.id
            })
        else:
            # CREATE new cable
            print(f"🆕 Creating new cable: project_id={project_id}, cable_id={cable_id}, type={cable_type}")
            
            # Determine which model to use based on cable_type
            if cable_type in box_types:
                cable_data = CableBox(
                    project_id=project_id,
                    junction_name=data.get('junction_name'),
                    row=data.get('row'),
                    position=str(data.get('position', '1')),
                    junction_box=data.get('junction_box'),
                    cable_id=cable_id,
                    cable_name=cable_name,
                    terminal=str(data.get('terminal')),
                    start_no=str(data.get('start_no')),
                    cable_type=cable_type,  # Use the actual cable_type from request
                    created_date=get_ist_now()
                )
            else:
                cable_data = Cable(
                    project_id=project_id,
                    junction_name=data.get('junction_name'),
                    row=data.get('row'),
                    position=str(data.get('position', '1')),
                    junction_box=data.get('junction_box'),
                    cable_id=cable_id,
                    cable_name=cable_name,
                    terminal=str(data.get('terminal')),
                    start_no=str(data.get('start_no')),
                    created_date=get_ist_now()
                )
            
            db.session.add(cable_data)
            db.session.commit()
            
            print(f"✅ Created new cable: id={cable_data.id}, cable_id={cable_id}, type={cable_type}")
            
            return jsonify({
                "success": True,
                "message": f"{cable_type.replace('_', ' ').title()} added successfully!",
                "cable_id": cable_id,
                "cable_name": cable_name,
                "cable_type": cable_type,
                "row_id": cable_data.id
            })
       
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR in add_cable_ajax: {str(e)}")
        import traceback
        traceback.print_exc()
       
        # Check if it's a unique constraint violation
        if "duplicate key value violates unique constraint" in str(e) or "uq_cable_project_cable" in str(e):
            return jsonify({
                "success": False,
                "message": f"Cable ID {cable_id} already exists for this project. Please use a different cable ID."
            })
       
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

@bp.route('/add_terminal_ajax', methods=['POST'])
@login_required
def add_terminal_ajax():
    # Check if user can edit step 4 - viewers cannot edit
    if not can_edit_step(4, current_user):
        return jsonify({'success': False, 'message': 'You are not allowed to update this step.'}), 403
       
    try:
        data = request.get_json()
        project_id = get_current_project()
       
        print(f"🔍 DEBUG add_terminal_ajax: project_id={project_id}, data={data}")
       
        if not project_id:
            return jsonify({'success': False, 'message': 'No active project selected'}), 400
        # Required fields validation
        required_fields = ['cable_id', 'terminal_id', 'terminal_no']
        missing_fields = [field for field in required_fields if field not in data or not data[field]]
       
        if missing_fields:
            return jsonify({
                'success': False,
                'message': f'Missing required fields: {", ".join(missing_fields)}'
            }), 400
       
        # FIX: Convert ALL IDs to strings consistently
        cable_id_str = str(data['cable_id'])
        terminal_id_str = str(data['terminal_id'])
       
        print(f"🔍 DEBUG: Converted cable_id: {data['cable_id']} -> '{cable_id_str}'")
        print(f"🔍 DEBUG: Converted terminal_id: {data['terminal_id']} -> '{terminal_id_str}'")
       
        # Check if cable already exists in Cable table
        cable_exists = Cable.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).first()
       
        print(f"🔍 DEBUG: Cable exists check for '{cable_id_str}': {cable_exists is not None}")
       
        # Handle cable_name if provided
        cable_name = data.get('cable_name', '').strip()
        if cable_name and not cable_exists:
            # Create a new cable entry
            new_cable = Cable(
                project_id=project_id,
                cable_id=cable_id_str, # Use string version
                cable_name=cable_name,
                created_date=get_ist_now()
            )
            db.session.add(new_cable)
            print(f"✅ DEBUG: Created new cable entry: '{cable_id_str}' - {cable_name}")
        elif cable_name and cable_exists and cable_exists.cable_name != cable_name:
            # Update existing cable name
            cable_exists.cable_name = cable_name
            print(f"✅ DEBUG: Updated existing cable name: '{cable_id_str}' - {cable_name}")
       
        # **SMART UPSERT: Check if terminal exists and update OR create new**
        existing_terminal = Terminal.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str, # Use string version
            terminal_id=terminal_id_str # Use string version
        ).first()
       
        if existing_terminal:
            # **UPDATE existing terminal**
            print(f"🔄 DEBUG: Updating existing terminal: cable_id='{cable_id_str}', terminal_id='{terminal_id_str}'")
           
            existing_terminal.terminal_no = data['terminal_no']
            existing_terminal.symbol = data.get('symbol', 'ara/wago')
            existing_terminal.input_left = data.get('input_left', '')
            existing_terminal.input_right = data.get('input_right', '')
            existing_terminal.spare = data.get('spare', 'N')
            existing_terminal.input_connected = data.get('input_connected', 'Y')
            existing_terminal.output_connected = data.get('output_connected', 'Y')
            # UPDATE THE TWO NEW FIELDS - AS REGULAR TEXT (NO DEFAULTS)
            existing_terminal.input_connected_extra = data.get('input_connected_extra', '')
            existing_terminal.output_connected_extra = data.get('output_connected_extra', '')
            existing_terminal.output_left = data.get('output_left', '')
            existing_terminal.output_right = data.get('output_right', '')
            existing_terminal.created_date = get_ist_now()
           
            db.session.commit()
           
            return jsonify({
                'success': True,
                'message': 'Terminal updated successfully',
                'terminal_id': existing_terminal.id,
                'action': 'updated'
            })
        else:
            # **CREATE new terminal**
            new_terminal = Terminal(
                project_id=project_id,
                cable_id=cable_id_str, # Use string version
                terminal_id=terminal_id_str, # Use string version
                terminal_no=data['terminal_no'],
                symbol=data.get('symbol', 'ara/wago'),
                input_left=data.get('input_left', ''),
                input_right=data.get('input_right', ''),
                spare=data.get('spare', 'N'),
                input_connected=data.get('input_connected', 'Y'),
                output_connected=data.get('output_connected', 'Y'),
                # ADD THE TWO NEW FIELDS - AS REGULAR TEXT (NO DEFAULTS)
                input_connected_extra=data.get('input_connected_extra', ''),
                output_connected_extra=data.get('output_connected_extra', ''),
                output_left=data.get('output_left', ''),
                output_right=data.get('output_right', ''),
                created_date=get_ist_now()
            )
           
            db.session.add(new_terminal)
            db.session.commit()
           
            print(f"✅ DEBUG: Created new terminal {new_terminal.id} for project {project_id}")
           
            return jsonify({
                'success': True,
                'message': 'Terminal added successfully',
                'terminal_id': new_terminal.id,
                'action': 'created'
            })
           
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR saving terminal: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error saving terminal: {str(e)}'
        }), 500

@bp.route('/add_header_ajax', methods=['POST'])
@login_required
def add_header_ajax():
    """Handle AJAX submission for adding header configuration"""
    # Check if user can edit step 5 - viewers cannot edit
    if not can_edit_step(5, current_user):
        return jsonify({'success': False, 'message': 'You are not allowed to update this step.'}), 403
       
    try:
        data = request.get_json()
        project_id = get_current_project()
       
        print(f"🔍 DEBUG add_header_ajax: project_id={project_id}, data={data}")
       
        if not project_id:
            return jsonify({'success': False, 'message': 'No active project selected'}), 400
        # Required fields validation
        required_fields = ['cable_id', 'header_type', 'terminal_start', 'terminal_end', 'input_output']
        missing_fields = [field for field in required_fields if field not in data or not data[field]]
       
        if missing_fields:
            return jsonify({
                'success': False,
                'message': f'Missing required fields: {", ".join(missing_fields)}'
            }), 400
       
        # FIX: Convert cable_id to string for consistent database query
        cable_id_str = str(data['cable_id'])
       
        # Check if cable already exists in Cable table
        cable_exists = Cable.query.filter_by(
            project_id=project_id,
            cable_id=cable_id_str # Use string version
        ).first()
       
        # Handle cable_name if provided
        cable_name = data.get('cable_name', '').strip()
        if cable_name and not cable_exists:
            # Create a new cable entry
            new_cable = Cable(
                project_id=project_id,
                cable_id=cable_id_str, # Use string version
                cable_name=cable_name,
                created_date=get_ist_now()
            )
            db.session.add(new_cable)
            print(f"✅ DEBUG: Created new cable entry: '{cable_id_str}' - {cable_name}")
        elif cable_name and cable_exists and cable_exists.cable_name != cable_name:
            # Update existing cable name
            cable_exists.cable_name = cable_name
            print(f"✅ DEBUG: Updated existing cable name: '{cable_id_str}' - {cable_name}")
       
        # Create new header record
        new_header = TerminalHeader(
            project_id=project_id,
            cable_id=cable_id_str, # Use string version
            header_type=data['header_type'],
            terminal_start=str(data['terminal_start']), # Convert to string as per model
            terminal_end=str(data['terminal_end']), # Convert to string as per model
            input_output=data['input_output'],
            text=data.get('text', ''),
            created_date=get_ist_now()
        )
       
        db.session.add(new_header)
        db.session.commit()
       
        print(f"✅ DEBUG: Saved header {new_header.id} to project {project_id}")
       
        return jsonify({
            'success': True,
            'message': 'Header configuration added successfully',
            'header_id': new_header.id
        })
           
    except Exception as e:
        db.session.rollback()
        print(f"❌ ERROR saving header: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error saving header configuration: {str(e)}'
        }), 500

@bp.route('/check_cable_terminals/<cable_id>')
@login_required
def check_cable_terminals(cable_id):
    """Check if terminals exist for a cable"""
    project_id = get_current_project()
    if not project_id:
        return jsonify({'exists': False, 'count': 0})
   
    # Convert cable_id to string for consistent query
    cable_id_str = str(cable_id)
   
    terminal_count = Terminal.query.filter_by(
        project_id=project_id,
        cable_id=cable_id_str # Use string version
    ).count()
   
    return jsonify({
        'exists': terminal_count > 0,
        'count': terminal_count
    })
   
# NEW: Workflow edit route
@bp.route("/workflow/<sheet_name>/edit/<int:row_id>/<int:step>", methods=["GET", "POST"])
@login_required
def workflow_edit_row(sheet_name, row_id, step):
    # Check if user can edit this step
    if not can_edit_step(step, current_user):
        flash("You are not allowed to update this step.")
        return redirect(url_for("main.workflow_step", step=step))
       
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    if sheet_name not in MODEL_MAP:
        flash("Invalid sheet")
        return redirect(url_for("main.workflow_step", step=step))
   
    model = MODEL_MAP[sheet_name]
    columns = SHEETS[sheet_name]
    current_project = Project.query.get(project_id)
    display_name = sheet_name.replace('_', ' ').title()
   
    row = model.query.filter_by(id=row_id, project_id=project_id).first()
    if not row:
        flash("Row not found")
        return redirect(url_for("main.workflow_step", step=step))
   
    if request.method == "POST":
        data = {}
        for col in columns:
            val = request.form.get(col, "").strip()
           
            # SPECIAL HANDLING FOR LOCATION NAMES IN STEP 2
            if sheet_name == 'junction_box' and col == 'junction_name' and 'junction_size' in request.form:
                location_name = val
                location_size = request.form.get('junction_size', '')
               
                # Format the location name with proper suffix
                formatted_name = format_location_name(location_name, location_size)
                data[col] = formatted_name
                print(f"🔍 DEBUG: Formatted location name in edit: '{location_name}' -> '{formatted_name}' for size '{location_size}'")
            else:
                data[col] = val if val else None
       
        if any(data.values()):
            try:
                for col, val in data.items():
                    if hasattr(row, col):
                        setattr(row, col, val)
                db.session.commit()
                
                # ✅ UPDATE PROJECT'S JUNCTION DATA FOR JUNCTION BOX EDITS
                if sheet_name == 'junction_box':
                    try:
                        current_project = Project.query.get(project_id)
                        if current_project:
                            current_project.update_junction_data()
                            print(f"✅ Updated junction_data for project {project_id} after editing")
                    except Exception as update_error:
                        print(f"⚠️ Warning: Could not update junction_data: {update_error}")
                
                flash(f"✅ Updated row in {display_name}")
                return redirect(url_for("main.workflow_step", step=step, project_id=project_id))
            except Exception as e:
                db.session.rollback()
                flash(f"❌ Error updating row: {str(e)}")
        else:
            flash("Please fill at least one field")
   
    row_data = {}
    for col in columns:
        row_data[col] = getattr(row, col, '')
   
    return render_template(
        "workflow_edit.html",
        current_project=current_project,
        step=step,
        sheet_name=sheet_name,
        display_name=display_name,
        columns=columns,
        row=row_data,
        row_id=row_id,
    )

# NEW: Workflow delete route
@bp.route("/workflow/<sheet_name>/delete/<int:row_id>/<int:step>", methods=["POST"])
@login_required
def workflow_delete_row(sheet_name, row_id, step):
    # Check if user can edit this step
    if not can_edit_step(step, current_user):
        flash("You are not allowed to update this step.")
        return redirect(url_for("main.workflow_step", step=step))
       
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    if sheet_name not in MODEL_MAP:
        flash("Invalid sheet")
        return redirect(url_for("main.workflow_step", step=step))
   
    model = MODEL_MAP[sheet_name]
    row = model.query.filter_by(id=row_id, project_id=project_id).first()
    if row:
        try:
            db.session.delete(row)
            db.session.commit()
            flash("✅ Row deleted")
        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error deleting row: {str(e)}")
    else:
        flash("Row not found")
   
    return redirect(url_for("main.workflow_step", step=step, project_id=project_id))

@bp.route("/excel_to_pdf", methods=["GET", "POST"])
@login_required
def excel_to_pdf():
    # Read force flag from query and (on POST) from form to persist across requests
    force_new_qs = request.args.get('new') == '1'
    force_new_form = request.form.get('force_new') == '1' if request.method == 'POST' else False
    force_new = force_new_qs or force_new_form
    # If forcing new, ignore any active project
    project_id = None if force_new else get_current_project()
    current_project = Project.query.get(project_id) if project_id else None
    auto_created_project = False
    
    if request.method == "POST":
        # Get remarks from form
        remarks = request.form.get('remarks', '').strip()
       
        if 'file' not in request.files:
            flash('❌ No file uploaded')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('❌ No file selected')
            return redirect(request.url)
        
        if not allowed_file(file.filename):
            flash('❌ Only XLSX files are allowed')
            return redirect(request.url)
        
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        upload_dir = os.path.join(BASE_DIR, 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        
        filename = secure_filename(file.filename)
        timestamp = get_ist_now().strftime('%Y%m%d_%H%M%S')
        
        # Save XLSX so we can load it for import
        temp_path = None
        if not project_id:
            # No project (forced new or none selected) -> will create from StationDrawing
            try:
                temp_path = os.path.join(upload_dir, f"temp_{timestamp}_{filename}")
                file.save(temp_path)
                wb = load_workbook(temp_path, data_only=True)
                
                # Extract station_name from StationDrawing
                station_name = None
                station_id = None
                station_code = None
                
                for ws_name in wb.sheetnames:
                    if ws_name.lower() == "stationdrawing":
                        ws = wb[ws_name]
                        headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
                       
                        # Find column indices
                        station_name_idx = None
                        station_id_idx = None
                        station_code_idx = None
                       
                        for i, h in enumerate(headers):
                            if h == 'station_name':
                                station_name_idx = i
                            elif h == 'station_id':
                                station_id_idx = i
                            elif h == 'station_code':
                                station_code_idx = i
                       
                        # Get values from first data row
                        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                            if row:
                                if station_name_idx is not None and len(row) > station_name_idx:
                                    station_name = row[station_name_idx]
                                    if station_name:
                                        station_name = str(station_name).strip()
                               
                                if station_id_idx is not None and len(row) > station_id_idx:
                                    station_id = row[station_id_idx]
                                    if station_id:
                                        station_id = str(station_id).strip()
                               
                                if station_code_idx is not None and len(row) > station_code_idx:
                                    station_code = row[station_code_idx]
                                    if station_code:
                                        station_code = str(station_code).strip()
                            break
                
                if not station_name:
                    station_name = f"Project_{timestamp}"
                
                # Create project
                new_project = Project(
                    name=station_name,
                    description=f"Auto-created from {filename}",
                    created_date=get_ist_now(),
                    updated_date=get_ist_now()
                )
                db.session.add(new_project)
                db.session.commit()
                
                project_id = new_project.id
                session['current_project_id'] = project_id
                session['project_id'] = project_id
                current_project = new_project
                auto_created_project = True
                
                flash(f"✅ Created new project: {station_name} (ID: {project_id})")
                
                # Import all sheets into this project (keep only latest -> clear then import)
                total_imported = 0
                for sheet_name in SHEETS.keys():
                    try:
                        model = MODEL_MAP[sheet_name]
                        expected_headers = SHEETS[sheet_name]
                        
                        # Find sheet
                        sheet_found = None
                        for ws_name in wb.sheetnames:
                            if ws_name.lower() == sheet_name.lower():
                                sheet_found = ws_name
                                break
                        
                        if not sheet_found:
                            continue
                        
                        ws = wb[sheet_found]
                        
                        # Map headers
                        headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
                        header_mapping = {}
                        for required_header in expected_headers:
                            for i, file_header in enumerate(headers):
                                if file_header == required_header.lower():
                                    header_mapping[required_header] = i
                                    break
                        
                        # Clear existing rows for this project in this sheet
                        model.query.filter_by(project_id=project_id).delete(synchronize_session=False)
                        
                        # Import rows
                        sheet_count = 0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            data = {'project_id': project_id}
                            has_data = False
                            
                            for header in expected_headers:
                                if header in header_mapping:
                                    col_index = header_mapping[header]
                                    cell_value = row[col_index] if col_index < len(row) else None
                                    
                                    if cell_value is not None:
                                        text = str(cell_value).strip()
                                       
                                        # ENFORCE LOCATION NAMING CONVENTION FOR JUNCTION BOXES
                                        if sheet_name == 'junction_box' and header == 'junction_name':
                                            # Get the junction_size for formatting
                                            size_header_index = None
                                            for h in expected_headers:
                                                if h == 'junction_size' and h in header_mapping:
                                                    size_header_index = header_mapping[h]
                                                    break
                                           
                                            if size_header_index is not None and size_header_index < len(row):
                                                size_value = row[size_header_index]
                                                if size_value is not None:
                                                    size_text = str(size_value).strip()
                                                    # Format the location name
                                                    text = format_location_name(text, size_text)
                                       
                                        data[header] = text if text else None
                                        if text:
                                            has_data = True
                                    else:
                                        data[header] = None
                                else:
                                    data[header] = None
                            
                            if has_data:
                                try:
                                    db.session.add(model(**data))
                                    sheet_count += 1
                                except Exception as e:
                                    print(f"Error importing {sheet_name} row: {e}")
                                    continue
                        
                        if sheet_count > 0:
                            db.session.commit()
                            total_imported += sheet_count
                    
                    except Exception as e:
                        print(f"Error importing sheet {sheet_name}: {e}")
                        db.session.rollback()
                        continue
                
                if total_imported > 0:
                    flash(f"✅ Imported {total_imported} records from XLSX into project (latest data only)")
                
                # Move temp to final path
                xlsx_filename = f"railway_project_{project_id}_{timestamp}_{filename}"
                xlsx_path = os.path.join(upload_dir, xlsx_filename)
                try:
                    os.rename(temp_path, xlsx_path)
                except Exception:
                    import shutil
                    shutil.copy2(temp_path, xlsx_path)
                    try:
                        os.remove(temp_path)
                    except:
                        pass
            
            except Exception as e:
                flash(f"❌ Error creating project from XLSX: {str(e)}")
                if temp_path and os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except:
                        pass
                return redirect(request.url)
        
        else:
            # Existing project -> we also import Excel as latest project data
            xlsx_filename = f"railway_project_{project_id}_{timestamp}_{filename}"
            xlsx_path = os.path.join(upload_dir, xlsx_filename)
            file.save(xlsx_path)
            
            try:
                wb = load_workbook(xlsx_path, data_only=True)
                total_imported = 0
               
                # NEW: Get current project details before import
                current_project = Project.query.get(project_id)
                project_name_before = current_project.name if current_project else None
               
                # For each sheet: clear then import so only latest Excel stays
                for sheet_name in SHEETS.keys():
                    try:
                        model = MODEL_MAP[sheet_name]
                        expected_headers = SHEETS[sheet_name]
                        
                        # Find sheet
                        sheet_found = None
                        for ws_name in wb.sheetnames:
                            if ws_name.lower() == sheet_name.lower():
                                sheet_found = ws_name
                                break
                        
                        if not sheet_found:
                            continue
                        
                        ws = wb[sheet_found]
                        
                        # Map headers
                        headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
                        header_mapping = {}
                        for required_header in expected_headers:
                            for i, file_header in enumerate(headers):
                                if file_header == required_header.lower():
                                    header_mapping[required_header] = i
                                    break
                        
                        # Clear existing rows for this project in this sheet
                        model.query.filter_by(project_id=project_id).delete(synchronize_session=False)
                        
                        # Import rows
                        sheet_count = 0
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            data = {'project_id': project_id}
                            has_data = False
                            
                            for header in expected_headers:
                                if header in header_mapping:
                                    col_index = header_mapping[header]
                                    cell_value = row[col_index] if col_index < len(row) else None
                                    
                                    if cell_value is not None:
                                        text = str(cell_value).strip()
                                       
                                        # ENFORCE LOCATION NAMING CONVENTION FOR JUNCTION BOXES
                                        if sheet_name == 'junction_box' and header == 'junction_name':
                                            # Get the junction_size for formatting
                                            size_header_index = None
                                            for h in expected_headers:
                                                if h == 'junction_size' and h in header_mapping:
                                                    size_header_index = header_mapping[h]
                                                    break
                                           
                                            if size_header_index is not None and size_header_index < len(row):
                                                size_value = row[size_header_index]
                                                if size_value is not None:
                                                    size_text = str(size_value).strip()
                                                    # Format the location name
                                                    text = format_location_name(text, size_text)
                                       
                                        data[header] = text if text else None
                                        if text:
                                            has_data = True
                                    else:
                                        data[header] = None
                                else:
                                    data[header] = None
                            
                            if has_data:
                                try:
                                    db.session.add(model(**data))
                                    sheet_count += 1
                                except Exception as e:
                                    print(f"Error importing {sheet_name} row: {e}")
                                    continue
                        
                        if sheet_count > 0:
                            db.session.commit()
                            total_imported += sheet_count
                           
                            # NEW: Update project name if StationDrawing was imported
                            # SPECIAL HANDLING FOR STATIONDRAWING SHEET
                            if sheet_name == 'StationDrawing':
                                # For StationDrawing, we should update the existing record instead of deleting and recreating
                                existing_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
                               
                                # Process only the first row (there should be only one StationDrawing per project)
                                row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
                                if row:
                                    data = {'project_id': project_id}
                                    has_data = False
                                   
                                    for header in expected_headers:
                                        if header in header_mapping:
                                            col_index = header_mapping[header]
                                            cell_value = row[col_index] if col_index < len(row) else None
                                            if cell_value is not None:
                                                text = str(cell_value).strip()
                                                data[header] = text if text else None
                                                if text and text != "NO PDF GENERATED": # Don't count "NO PDF GENERATED" as meaningful data
                                                    has_data = True
                                            else:
                                                data[header] = None
                                        else:
                                            data[header] = None
                                   
                                    if has_data:
                                        try:
                                            if existing_drawing:
                                                # UPDATE existing StationDrawing
                                                print(f"🔄 Updating existing StationDrawing for project {project_id}")
                                                for col, val in data.items():
                                                    if col != 'project_id': # Don't update project_id
                                                        setattr(existing_drawing, col, val)
                                                db.session.commit()
                                                print(f"✅ Updated StationDrawing: station_id={data.get('station_id')}, station_name={data.get('station_name')}, station_code={data.get('station_code')}")
                                            else:
                                                # CREATE new StationDrawing
                                                print(f"🆕 Creating new StationDrawing for project {project_id}")
                                                new_drawing = StationDrawing(**data)
                                                db.session.add(new_drawing)
                                                db.session.commit()
                                                print(f"✅ Created StationDrawing: station_id={data.get('station_id')}, station_name={data.get('station_name')}, station_code={data.get('station_code')}")
                                           
                                            sheet_count = 1
                                           
                                            # Update project name to match station name
                                            station_name = data.get('station_name')
                                            if station_name and current_project and current_project.name != station_name:
                                                current_project.name = station_name
                                                current_project.updated_date = get_ist_now()
                                                db.session.commit()
                                                print(f"✅ Updated project name to: {station_name}")
                                               
                                        except Exception as e:
                                            print(f"❌ Error processing StationDrawing: {str(e)}")
                                            db.session.rollback()
                    
                    except Exception as e:
                        print(f"Error importing sheet {sheet_name}: {e}")
                        db.session.rollback()
                        continue
                
                if total_imported > 0:
                    flash(f"✅ Replaced project data with latest Excel: imported {total_imported} records")
            
            except Exception as e:
                flash(f"❌ Error importing XLSX data: {str(e)}")
                return redirect(request.url)
        
        # Generate PDF
        
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))      # /app
        PROJECT_ROOT = os.path.dirname(BASE_DIR)                   # /Circuitbuilding
        GIT_ROOT = os.path.dirname(PROJECT_ROOT)                   # /git

        # ✅ uploads inside git folder
        upload_dir = os.path.join(GIT_ROOT, 'uploads')
        pdf_filename = xlsx_filename.replace('.xlsx', '.pdf')
        pdf_path = os.path.join(upload_dir, pdf_filename)
        #converter_script = os.path.join(os.getcwd(), 'excel_to_pdf_converter.py')
        
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))          # /app
        PROJECT_ROOT = os.path.dirname(BASE_DIR)                       # /Circuitbuilding
        GIT_ROOT = os.path.dirname(PROJECT_ROOT)                       # /git

        converter_script = os.path.join(GIT_ROOT, 'excel_to_pdf_converter.py')
        
        if not os.path.exists(converter_script):
            flash(f'❌ Converter script not found: {converter_script}')
            return redirect(request.url)
        
        try:
            python_exe = sys.executable
            start_ts = time.time()
            result = subprocess.run(
                [python_exe, converter_script, xlsx_path, pdf_path],
                capture_output=True,
                text=True,
                timeout=300,
                cwd=upload_dir
            )
            
            print("[PDF] Return code:", result.returncode)
            print("[PDF] STDOUT:", result.stdout)
            print("[PDF] STDERR:", result.stderr)
            print("[PDF] Expected PDF:", pdf_path, "exists:", os.path.exists(pdf_path))
            
            if result.returncode == 0 and not os.path.exists(pdf_path):
                newest = None
                newest_mtime = 0.0
                for name in os.listdir(upload_dir):
                    if name.lower().endswith(".pdf"):
                        p = os.path.join(upload_dir, name)
                        try:
                            m = os.path.getmtime(p)
                            if m >= start_ts - 5 and m > newest_mtime:
                                newest = p
                                newest_mtime = m
                        except Exception:
                            continue
                
                if newest and newest != pdf_path:
                    try:
                        os.replace(newest, pdf_path)
                        print(f"[PDF] Moved {newest} -> {pdf_path}")
                    except Exception as e:
                        print(f"[PDF] Rename fallback failed: {e}")
            
            if result.returncode == 0 and os.path.exists(pdf_path):
                file_md5 = _md5_of_file(pdf_path)
                file_size = os.path.getsize(pdf_path)
                meta = parse_converter_stdout(result.stdout)
                db_checksum = meta.get("metadata_checksum") or meta.get("full_file_md5") or file_md5
                max_version_record = GeneratedPDF.query.filter_by(project_id=project_id).order_by(GeneratedPDF.version.desc()).first()
                next_version = max_version_record.version + 1 if max_version_record else 1
                # Get junction data from the database at this moment
                junction_boxes = JunctionBox.query.filter_by(project_id=project_id).all()
                junction_data_list = []
                for jb in junction_boxes:
                    junction_data_list.append({
                        'junction_id': jb.junction_id,
                        'junction_name': jb.junction_name,
                        'junction_size': jb.junction_size,
                        'station_id': jb.station_id,
                        'latitude': jb.latitude,
                        'longitude': jb.longitude,
                        'junction_row': jb.junction_row
                    })
                
                # Convert to JSON string
                import json
                junction_data_json = json.dumps(junction_data_list) if junction_data_list else None
                record = GeneratedPDF(
                    project_id=project_id,
                    pdf_filename=pdf_filename,
                    xlsx_filename=xlsx_filename,
                    checksum_md5=db_checksum,
                    file_size=file_size,
                    checksum_algo="md5",
                    metadata_checksum=meta.get("metadata_checksum"),
                    metadata_data=meta.get("metadata_data"),
                    initial_size_bytes=meta.get("initial_size_bytes"),
                    final_size_bytes=meta.get("final_size_bytes"),
                    metadata_ts_ist=meta.get("metadata_ts_ist"),
                    station_code=meta.get("station_code"),
                    source_pdf_name=meta.get("source_pdf_name"),
                    full_file_md5=meta.get("full_file_md5") or file_md5,
                    remarks=remarks if remarks else None,
                    created_at=get_ist_now(),
                    version=next_version,
                    junction_data=junction_data_json  # Store junction data
                )
                db.session.add(record)
                db.session.commit()
               
                # ✅ CRITICAL: SET PROJECT STAGE TO 10 (PDF Generated)
                try:
                    if current_project:
                        # Update project stage to 10 (PDF Generated)
                        current_project.stage = 10
                        current_project.updated_date = get_ist_now()
                        print(f"✅ Updated project {project_id} stage to 10 (PDF generated)")
                except Exception as stage_error:
                    print(f"⚠️ Warning: Could not update project stage: {stage_error}")
                
                # Update StationDrawing with latest checksum and version
                try:
                    # Refresh the StationDrawing data
                    station_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
                    if station_drawing:
                        # Ensure project name matches station name
                        if station_drawing.station_name and current_project.name != station_drawing.station_name:
                            current_project.name = station_drawing.station_name
                            current_project.updated_date = get_ist_now()
                            db.session.commit()
                            print(f"✅ Synchronized project name with station name: {station_drawing.station_name}")
                       
                        # Debug output to verify data
                        print(f"🔍 DEBUG StationDrawing after import:")
                        print(f" - station_id: {station_drawing.station_id}")
                        print(f" - station_name: {station_drawing.station_name}")
                        print(f" - station_code: {station_drawing.station_code}")
                        print(f" - version: {station_drawing.version}")
                        print(f" - checksum: {station_drawing.checksum}")
                       
                except Exception as e:
                    print(f"❌ Error synchronizing project and station data: {str(e)}")
                    db.session.rollback()
                
               
                admin_users = User.query.filter_by(role='admin').all()
                for admin in admin_users:
                    # Check if admin is assigned to this project
                    if current_project in admin.projects or admin.role == 'admin':  # Admin might have access to all
                        notification = Notification(
                            user_id=admin.id,
                            pdf_id=record.id,
                            project_id=project_id,
                            level='New_Drawing',
                            status='pending',
                            # UPDATED MESSAGE FORMAT
                            message=f'NEW DRAWING requires admin attention: {current_project.name if current_project else "Unknown"}'
                        )
                        db.session.add(notification)

                # Also create notification for level1 users assigned to this project
                level1_users = User.query.filter_by(designation='level1').all()
                for user in level1_users:
                    # Check if level1 user is assigned to this project
                    if current_project in user.projects:
                        notification = Notification(
                            user_id=user.id,
                            pdf_id=record.id,
                            project_id=project_id,
                            level='level1',
                            status='pending',
                            # UPDATED MESSAGE FORMAT
                            message=f'NEW DRAWING requires level1 approval by: {user.username}'
                        )
                        db.session.add(notification)
                
                db.session.commit()  # Commit the notifications
                
                flash(f'✅ Successfully converted {filename} to PDF! Project stage updated to 10 (PDF Generated).')
                from flask import current_app

                if current_app.config.get('FTP_ENABLED'):
                    try:
                        # Upload XLSX file
                        xlsx_success, xlsx_msg, xlsx_remote = upload_to_ftp(xlsx_path, xlsx_filename)
                        if xlsx_success:
                            flash(f'✅ Excel file uploaded to FTP: {xlsx_remote}')
                            # Archive local copy
                            archive_local_file(xlsx_path)
                        else:
                            flash(f'⚠️ Excel FTP upload failed: {xlsx_msg}')
                        
                        # Upload PDF file
                        pdf_success, pdf_msg, pdf_remote = upload_to_ftp(pdf_path, pdf_filename)
                        if pdf_success:
                            flash(f'✅ PDF file uploaded to FTP: {pdf_remote}')
                            # Archive local copy
                            archive_local_file(pdf_path)
                        else:
                            flash(f'⚠️ PDF FTP upload failed: {pdf_msg}')
                            
                    except Exception as e:
                        flash(f'⚠️ FTP upload error: {str(e)}')

                return redirect(url_for('main.pdf_view', filename=pdf_filename))
            else:
                err = (result.stderr or 'Unknown error during conversion').strip()
                flash(f'❌ Error converting file: {err}')
                return redirect(request.url)
        
        except subprocess.TimeoutExpired:
            flash('❌ Conversion timed out. File might be too large.')
            return redirect(request.url)
        
        except Exception as e:
            flash(f'❌ Error proce  ssing file: {str(e)}')
            return redirect(request.url)
    
    # Render upload form; pass force_new so template can persist via hidden input
    return render_template("excel_to_pdf.html", current_project=current_project, force_new=force_new)


@bp.route("/upload_pdf", methods=["GET", "POST"])
@login_required
def upload_pdf():
    if request.method == "POST":  
        upload_dir = "uploads"

        file = request.files.get("file")
        project_id = request.form.get('project_id')  # get from form

        signed_status = request.form.get('signed_status')

         # Clean filename
        filename = secure_filename(file.filename)

        # Timestamp
        timestamp = int(time.time())

        # Custom filename
        new_filename = f"Project_{project_id}_{timestamp}_{filename}"
        
        if file and file.filename.endswith(".pdf"):
            #pdf_path = os.path.join(upload_dir, file.filename)
            pdf_path = os.path.join(upload_dir, new_filename)
            file.save(pdf_path)
            if os.path.exists(pdf_path):

                file_md5 = _md5_of_file(pdf_path)
                file_size = os.path.getsize(pdf_path)

                # No converter → no stdout metadata
                meta = {}

                db_checksum = file_md5

                max_version_record = GeneratedPDF.query.filter_by(project_id=project_id)\
                    .order_by(GeneratedPDF.version.desc()).first()

                next_version = max_version_record.version + 1 if max_version_record else 1

                # Junction data
                junction_boxes = JunctionBox.query.filter_by(project_id=project_id).all()
                junction_data_list = []

                for jb in junction_boxes:
                    junction_data_list.append({
                        'junction_id': jb.junction_id,
                        'junction_name': jb.junction_name,
                        'junction_size': jb.junction_size,
                        'station_id': jb.station_id,
                        'latitude': jb.latitude,
                        'longitude': jb.longitude,
                        'junction_row': jb.junction_row
                    })

                import json
                remarks='';
                junction_data_json = json.dumps(junction_data_list) if junction_data_list else None

                record = GeneratedPDF(
                    project_id=project_id,
                    pdf_filename=new_filename,
                    xlsx_filename=None,  # ❗ no Excel now
                    checksum_md5=db_checksum,
                    file_size=file_size,
                    checksum_algo="md5",
                    metadata_checksum=None,
                    metadata_data=None,
                    initial_size_bytes=file_size,
                    final_size_bytes=file_size,
                    metadata_ts_ist=get_ist_now(),
                    station_code=None,
                    source_pdf_name=new_filename,
                    full_file_md5=file_md5,
                    remarks=remarks if remarks else None,
                    created_at=get_ist_now(),
                    version=next_version,
                    junction_data=junction_data_json,
                    signed_status=signed_status
                )

                db.session.add(record)
                db.session.commit()
            return redirect(url_for('main.approval_tracking'))  # make sure this route exists
        else:
            return "Invalid file. Please upload a PDF."



@bp.route("/new_drawing/<int:project_id>", methods=["GET", "POST"])
@login_required
def new_drawing(project_id):
    """Clear drawing data for a project and prepare for new workflow."""
    project = Project.query.get_or_404(project_id)
    
    # Check if user has access to clear this project's drawing
    # Admin can clear any drawing, users can only clear drawings for their assigned projects
    if current_user.role == 'user' and project not in current_user.projects:
        flash("Access denied. You can only clear drawings for your own projects.")
        return redirect(url_for('main.index'))
    
    # Rest of the function remains the same...
    if request.method == "POST":
        # Handle AJAX POST request - MUST return JSON
        try:
            # Clear all workflow-related tables for this project (keep PDFs intact)
            tables_to_clear = [
                StationDrawing, JunctionBox, Cable, Terminal, Group,
                TerminalHeader, ChokeTable, ResistorTable, CableBox
            ]
            
            for model in tables_to_clear:
                records = model.query.filter_by(project_id=project_id).all()
                for record in records:
                    db.session.delete(record)
            
            db.session.commit()
            
            # IMPORTANT: Return JSON, not redirect
            return jsonify({"success": True, "message": "Drawing cleared successfully!"})
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ Error clearing drawing: {str(e)}")
            return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500
    
    # Handle GET request (direct link from browser)
    try:
        tables_to_clear = [
            StationDrawing, JunctionBox, Cable, Terminal, Group,
            TerminalHeader, ChokeTable, ResistorTable
        ]
        
        for model in tables_to_clear:
            records = model.query.filter_by(project_id=project_id).all()
            for record in records:
                db.session.delete(record)
        
        db.session.commit()
        flash("✅ Drawing data cleared. Ready for new setup!", "success")
        return redirect(url_for("main.index"))
        
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error clearing drawing data: {str(e)}", "error")
        return redirect(url_for("main.index"))

@bp.route("/pdf/view/<filename>")
@login_required
def pdf_view(filename):
    """
    Unified PDF viewer:
    - If the PDF belongs to a known project and is a single‑junction PDF,
      redirect to /preview?location_id=...
    - If it's a multi‑junction (full station) PDF, redirect to /preview (no location_id)
    - Otherwise (PDF not found or error), fall back to the original pdf_view.html
    """
    # ------------------------------------------------------------
    # TRY TO REDIRECT TO THE UNIFIED PREVIEW PAGE
    # ------------------------------------------------------------
    pdf_record = GeneratedPDF.query.filter_by(pdf_filename=filename).first()
    
    if pdf_record:
        # Check if this is a single-junction PDF using junction_data
        if pdf_record.junction_data:
            try:
                data = json.loads(pdf_record.junction_data)
                if isinstance(data, list) and len(data) == 1:
                    location_id = data[0].get('junction_id')
                    # Redirect to preview with location_id
                    return redirect(url_for('main.preview', location_id=location_id))
            except Exception as e:
                # Log the error if needed (optional)
                print(f"[pdf_view] Error parsing junction_data for {filename}: {e}")
        
        # If it's not a single-junction PDF, or junction_data is missing/incorrect,
        # redirect to the full‑station preview (no location_id)
        return redirect(url_for('main.preview'))
    
    # ------------------------------------------------------------
    # FALLBACK – ORIGINAL pdf_view LOGIC
    # (only executed if no PDF record is found)
    # ------------------------------------------------------------
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))

    current_project = Project.query.get(project_id)
    upload_dir = os.path.join(os.getcwd(), 'uploads')
    pdf_path = os.path.join(upload_dir, filename)

    if not os.path.exists(pdf_path):
        flash('❌ PDF file not found')
        return redirect(url_for('main.approval_tracking'))

    inline_url = url_for('main.inline_pdf', filename=filename)
    download_url = url_for('main.download_pdf', filename=filename)

    # Try to get the PDF record again (should be None here, but safe)
    pdf_record = GeneratedPDF.query.filter_by(pdf_filename=filename).first()

    # Approval & edit permissions (same as your original code)
    can_edit = False
    if current_user.designation == 'level1':
        if pdf_record and (pdf_record.level2_status == 'rejected' or 
                           pdf_record.level3_status == 'rejected'):
            can_edit = True

    can_approve_level1 = can_approve_level2 = can_approve_level3 = False
    can_reject_level1 = can_reject_level2 = can_reject_level3 = False

    if pdf_record:
        # Level 1 permissions
        if current_user.designation == 'level1':
            can_approve_level1 = pdf_record.can_level1_approve()
            can_reject_level1 = pdf_record.level1_status == 'pending'
        # Level 2 permissions
        elif current_user.designation == 'level2':
            can_approve_level2 = pdf_record.can_level2_approve()
            can_reject_level2 = (pdf_record.level1_status == 'approved' and 
                                 pdf_record.level2_status == 'pending')
        # Level 3 permissions
        elif current_user.designation == 'level3':
            can_approve_level3 = pdf_record.can_level3_approve()
            can_reject_level3 = (pdf_record.level1_status == 'approved' and 
                                 pdf_record.level2_status == 'approved' and
                                 pdf_record.level3_status == 'pending')

    return render_template(
        "pdf_view.html",
        current_project=current_project,
        filename=filename,
        inline_url=inline_url,
        download_url=download_url,
        pdf=pdf_record,
        user_designation=current_user.designation,
        can_edit=can_edit,
        can_approve_level1=can_approve_level1,
        can_approve_level2=can_approve_level2,
        can_approve_level3=can_approve_level3,
        can_reject_level1=can_reject_level1,
        can_reject_level2=can_reject_level2,
        can_reject_level3=can_reject_level3
    )

@bp.route("/pdf/inline/<filename>")
@login_required
def inline_pdf(filename):
    
    #upload_dir = os.path.join(os.getcwd(), 'uploads')
    upload_dir=r"C:\Railway\git\uploads" #"/var/www/html/git/uploads"
    pdf_path = os.path.join(upload_dir, filename)
    if not os.path.exists(pdf_path):
        flash('❌ PDF file not found')
        return redirect(url_for('main.approval_tracking'))
    return send_file(
        pdf_path,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=filename
    )

@bp.route("/download_pdf/<filename>")
@login_required
def download_pdf(filename):
    upload_dir=r"C:\Railway\git\uploads" #"/var/www/html/git/uploads"
    #upload_dir = os.path.join(os.getcwd(), 'uploads')
    pdf_path = os.path.join(upload_dir, filename)
    if not os.path.exists(pdf_path):
        flash('❌ PDF file not found')
        return redirect(url_for('main.approval_tracking'))
    try:
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
    except Exception as e:
        flash(f'❌ Error downloading PDF: {str(e)}')
        return redirect(url_for('main.index'))

@bp.route("/sheet/<name>", methods=["GET", "POST"])
@login_required
def sheet_form(name):
    if name not in SHEETS:
        flash(f"Unknown sheet: {name}")
        return redirect(url_for("main.index"))
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    current_project = Project.query.get(project_id)
    model = MODEL_MAP[name]
    columns = SHEETS[name]
    if name == "StationDrawing":
        seed_default_station_drawing(project_id)
    edit_id = request.args.get("edit", type=int)
    edit_row = None
    if edit_id:
        edit_row = model.query.filter_by(id=edit_id, project_id=project_id).first()
        if not edit_row:
            flash("Row not found for editing")
            return redirect(url_for("main.sheet_form", name=name))
    if request.method == "POST":
        data = {}
        for col in columns:
            val = request.form.get(col, "").strip()
            data[col] = val if val else None
        if any(data.values()):
            data["project_id"] = project_id
            try:
                if edit_id and edit_row:
                    for col, val in data.items():
                        if hasattr(edit_row, col):
                            setattr(edit_row, col, val)
                    db.session.commit()
                    flash(f"✅ Updated row in {name.replace('_',' ').title()}")
                else:
                    db.session.add(model(**data))
                    db.session.commit()
                    flash(f"✅ Added row to {name.replace('_',' ').title()}")
            except Exception as e:
                db.session.rollback()
                flash(f"❌ Error saving row: {str(e)}")
        else:
            flash("Please fill at least one field")
        return redirect(url_for("main.sheet_form", name=name))
    rows = model.query.filter_by(project_id=project_id).order_by(model.id).all()
    rows_data = []
    for r in rows:
        rd = {"id": r.id}
        for c in columns:
            rd[c] = getattr(r, c, "")
        rows_data.append(rd)
    edit_row_dict = None
    if edit_row:
        edit_row_dict = {"id": edit_row.id}
        for c in columns:
            edit_row_dict[c] = getattr(edit_row, c, "")
    return render_template(
        "sheet_form.html",
        sheet=name,
        columns=columns,
        rows=rows_data,
        hint=HEADER_HINTS.get(name, ""),
        current_project=current_project,
        edit_id=edit_id,
        edit_row=edit_row_dict,
    )

@bp.route("/sheet/<name>/delete/<int:row_id>", methods=["POST"])
@login_required
def delete_row(name, row_id):
    if name not in MODEL_MAP:
        flash("Invalid sheet")
        return redirect(url_for("main.index"))
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    model = MODEL_MAP[name]
    rec = model.query.filter_by(id=row_id, project_id=project_id).first()
    if rec:
        try:
            db.session.delete(rec)
            db.session.commit()
            flash("✅ Row deleted")
        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error deleting row: {str(e)}")
    else:
        flash("Row not found")
    return redirect(url_for("main.sheet_form", name=name))

@bp.route("/sheet/<name>/edit/<int:row_id>")
@login_required
def edit_row(name, row_id):
    return redirect(url_for("main.sheet_form", name=name, edit=row_id))

@bp.route('/check_duplicate_terminal')
@login_required
def check_duplicate_terminal():
    cable_id = request.args.get('cable_id')
    terminal_id = request.args.get('terminal_id')
    project_id = get_current_project()
   
    if not project_id:
        return jsonify({'exists': False})
   
    # **FIX: Convert to strings for consistent comparison**
    cable_id_str = str(cable_id) if cable_id else None
    terminal_id_str = str(terminal_id) if terminal_id else None
   
    if not cable_id_str or not terminal_id_str:
        return jsonify({'exists': False})
   
    # Query with consistent data types
    existing_terminal = Terminal.query.filter_by(
        project_id=project_id,
        cable_id=cable_id_str, # Use string
        terminal_id=terminal_id_str # Use string
    ).first()
   
    return jsonify({'exists': existing_terminal is not None})

@bp.route("/preview")
@login_required
def preview():
    print(f"[DEBUG] Rendering preview.html – this should print ONCE")
    import traceback
    traceback.print_stack()
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))

    current_project = Project.query.get(project_id)
    location_id = request.args.get('location_id', type=int)

    # ------------------------------------------------------------
    # Get ALL PDFs for this project, newest first
    # ------------------------------------------------------------
    all_pdfs = GeneratedPDF.query.filter_by(
        project_id=project_id
    ).order_by(GeneratedPDF.created_at.desc()).all()

    # Helper to detect single-junction PDF and extract info
    def get_single_junction_info(pdf):
        if not pdf.junction_data:
            return None, None
        try:
            data = json.loads(pdf.junction_data)
            if isinstance(data, list) and len(data) == 1:
                j = data[0]
                return int(j.get('junction_id')), j.get('junction_name', '')
        except:
            pass
        return None, None

    # ------------------------------------------------------------
    # Select the PDF to display
    # ------------------------------------------------------------
    selected_pdf = None
    junction_name = None

    if location_id:
        # 1st choice: latest single‑junction PDF for this location
        for pdf in all_pdfs:
            j_id, j_name = get_single_junction_info(pdf)
            if j_id == location_id:
                selected_pdf = pdf
                junction_name = j_name
                break

        # 2nd choice (fallback): latest full‑project PDF (multi-junction)
        if not selected_pdf:
            for pdf in all_pdfs:
                if pdf.junction_data and len(json.loads(pdf.junction_data)) > 1:
                    selected_pdf = pdf
                    break
            if selected_pdf:
                flash("No single‑location PDF found for this junction. Showing latest full‑project PDF.", "warning")
    else:
        # No location: show latest full‑project PDF (multi-junction)
        for pdf in all_pdfs:
            if pdf.junction_data and len(json.loads(pdf.junction_data)) > 1:
                selected_pdf = pdf
                break

    # Final fallback: any PDF
    if not selected_pdf and all_pdfs:
        selected_pdf = all_pdfs[0]
        flash("Showing the most recent PDF (could be a location PDF).", "info")

    # ------------------------------------------------------------
    # Prepare URLs and context
    # ------------------------------------------------------------
    full_pdf_url = None
    download_pdf_url = None
    if selected_pdf and selected_pdf.pdf_filename:
        #full_pdf_url = url_for('main.inlinepdf', filename=selected_pdf.pdf_filename)
        full_pdf_url = f"/pdf/inline/{selected_pdf.pdf_filename}"
        download_pdf_url = url_for('main.downloadpdf', filename=selected_pdf.pdf_filename)

    can_edit = selected_pdf and (
        selected_pdf.level2_status == 'rejected' or 
        selected_pdf.level3_status == 'rejected'
    )

    junctions = JunctionBox.query.filter_by(project_id=project_id).all()
    user_designation = getattr(current_user, 'designation', 'User')

    return render_template(
        "preview.html",
        current_project=current_project,
        pdf=selected_pdf,
        full_pdf_url=full_pdf_url,
        download_url=download_pdf_url,
        can_edit=can_edit,
        user_designation=user_designation,
        junctions=junctions,
        all_pdfs=all_pdfs,
        location_id=location_id,
        junction_name=junction_name   # 👈 passed directly, no JSON in template
    )

@bp.route("/download/")
@login_required
def download():
    """Download XLSX file - can be for entire project or specific location (JunctionBox)"""
    project_id = request.args.get('project_id', type=int) or get_current_project()
    location_id = request.args.get('location_id', type=int)
    download_only = request.args.get('download_only', type=bool, default=False)

    if not project_id:
        #flash("Please select a project first", "error")
        print ("No Project Id")
        #return redirect(url_for("main.project_selection"))

    # Get project
    current_project = Project.query.get(project_id)
    if not current_project:
        #flash("Project not found", "error")
        #return redirect(url_for("main.index"))
        print ("No Current Project ")

    # ---- CASE A: Specific location (Preview This Location) ----
    
    if location_id:
        location = JunctionBox.query.filter_by(id=location_id, project_id=project_id).first()
        print (" In Location ")
        if not location:
            flash("Location (Junction Box) not found in this project", "error")
            print (" In No Location ")
            return redirect(request.referrer or url_for("main.index"))

        # Check if this junction box has Level 1 approval
        approved = JunctionApproval.query.filter_by(
            project_id=project_id,
            junction_box_id=location.id,
            level1_status='approved'
        ).first() is not None

        if approved:
            # ✅ XLSX not needed – directly show PDF
            print (" In No Location Approved ")
            return jsonify({
                'skip_xlsx': True,
                'message': f"Junction '{location.junction_name}' is already approved at Level 1. Showing PDF.",
                'project_id': project_id,
                'location_id': location_id
            })
   
    # ---- CASE B: Full project (Preview Full Project / Preview Station) ----
    else:
        all_junctions = JunctionBox.query.filter_by(project_id=project_id).all()
        print (" In Location Else ")
        if all_junctions:
            all_approved = all(
                JunctionApproval.query.filter_by(
                    project_id=project_id,
                    junction_box_id=j.id,
                    level1_status='approved'
                ).first() is not None
                for j in all_junctions
            )

            if all_approved:
                # ✅ All locations are Level 1 approved – skip XLSX
                print (" In Location Else Approve")
                return jsonify({
                    'skip_xlsx': True,
                    'message': "All locations in this project are approved at Level 1. Showing PDF.",
                    'project_id': project_id
                })
    
    # ------------------------------------------------------------
    # 2. NORMAL XLSX GENERATION (only if not skipped above)
    # ------------------------------------------------------------
    try:
        # ========== CREATE XLSX WORKBOOK ==========
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        total_records = 0
        print (" 4087 - Create XLSX - Start ")   
        # Get cable names mapping
        cable_name_map = {}
        try:
            cables = Cable.query.filter_by(project_id=project_id).all()
            cable_name_map = {str(cable.cable_id): cable.cable_name for cable in cables}
        except Exception as e:
            print(f"Warning loading cable names: {str(e)}")

        # Get location info if location_id is provided (for status update)
        location = None
        if location_id:
            location = JunctionBox.query.filter_by(id=location_id, project_id=project_id).first()
            if not location:
                flash("Location (Junction Box) not found in this project", "error")
                return redirect(request.referrer or url_for("main.index"))

            # Update project status to 'preview' for location‑specific download
            current_project.status = 'preview'
            current_project.updated_date = get_ist_now()
            db.session.commit()
            print(f"✅ Project {project_id} status updated to 'preview' (location-specific download)")

        # ========== CREATE SHEETS ==========
        for sheet_name, columns in SHEETS.items():
            ws = wb.create_sheet(title=sheet_name)

            # Get the model for this sheet
            model = MODEL_MAP.get(sheet_name)
            if not model:
                continue

            # ========== WRITE HEADERS ==========
            if sheet_name == 'terminal':
                template_headers = [
                    'cable_id', 'cable_name', 'terminal_id', 'terminal_no', 'symbol',
                    'input_left', 'input_right', 'spare', 'input_connected',
                    'output_connected', 'input_connected_extra', 'output_connected_extra',
                    'output_left', 'output_right'
                ]
                ws.append(template_headers)
            else:
                ws.append(columns)

            # ========== GET RECORDS WITH LOCATION FILTERING ==========
            records = []

            if location_id:
                # Filter by specific location (junction_box)
                if sheet_name == 'junction_box':
                    records = model.query.filter_by(
                        project_id=project_id,
                        id=location_id
                    ).all()

                elif sheet_name == 'StationDrawing':
                    records = model.query.filter_by(project_id=project_id).all()

                elif sheet_name == 'cable':
                    records = model.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()

                elif sheet_name == 'cable_box':
                    records = model.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()

                elif sheet_name == 'terminal':
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        for cable_id in cable_ids:
                            cable_terminals = model.query.filter_by(
                                project_id=project_id,
                                cable_id=cable_id
                            ).all()
                            records.extend(cable_terminals)

                elif sheet_name in ['group', 'terminal_header', 'choketable', 'resistortable']:
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        for cable_id in cable_ids:
                            sheet_records = model.query.filter_by(
                                project_id=project_id,
                                cable_id=cable_id
                            ).all()
                            records.extend(sheet_records)

            else:
                # Full project download
                records = model.query.filter_by(project_id=project_id).all()

            # ========== WRITE DATA ==========
            if records:
                if sheet_name == 'terminal':
                    # Sort terminal records
                    sorted_records = sorted(records, key=lambda x: (
                        int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0,
                        int(x.terminal_id) if x.terminal_id and str(x.terminal_id).isdigit() else 0
                    ))

                    for record in sorted_records:
                        row = [
                            record.cable_id,
                            cable_name_map.get(str(record.cable_id), ''),
                            record.terminal_id,
                            record.terminal_no,
                            record.symbol,
                            record.input_left if record.input_left and record.input_left.strip() else '',
                            record.input_right if record.input_right and record.input_right.strip() else '',
                            record.spare,
                            record.input_connected,
                            record.output_connected,
                            record.input_connected_extra if record.input_connected_extra and record.input_connected_extra.strip() else '',
                            record.output_connected_extra if record.output_connected_extra and record.output_connected_extra.strip() else '',
                            record.output_left if record.output_left and record.output_left.strip() else '',
                            record.output_right if record.output_right and record.output_right.strip() else ''
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'cable':
                    # Sort cable records
                    sorted_records = sorted(records, key=lambda x: (
                        int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0
                    ))

                    for record in sorted_records:
                        row = [
                            record.cable_id,
                            record.cable_name,
                            record.junction_box,
                            record.junction_name,
                            record.row,
                            record.position,
                            record.terminal,
                            record.start_no
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'cable_box':
                    # Sort cable_box records
                    sorted_records = sorted(records, key=lambda x: (
                        int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0
                    ))

                    for record in sorted_records:
                        row = [
                            record.cable_id,
                            record.cable_name,
                            record.junction_box,
                            record.junction_name,
                            record.row,
                            record.position,
                            record.terminal,
                            record.start_no,
                            record.cable_type,
                            record.output
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'junction_box':
                    for record in records:
                        row = [
                            record.station_id,
                            record.junction_id,
                            record.junction_name,
                            record.latitude,
                            record.longitude,
                            record.junction_size,
                            record.junction_row
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'StationDrawing':
                    for record in records:
                        row = [
                            record.checksum,
                            record.station_id,
                            record.diagram_name,
                            record.station_name,
                            record.station_code,
                            record.version,
                            record.date,
                            record.drawn_by,
                            record.checked_by,
                            record.division,
                            record.zone,
                            record.total_sheet,
                            record.designation1,
                            record.designation2,
                            record.designation3
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'terminal_header':
                    for record in records:
                        row = [
                            record.cable_id,
                            record.header_type,
                            record.terminal_start,
                            record.terminal_end,
                            record.input_output,
                            record.text
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'group':
                    for record in records:
                        row = [
                            record.cable_id,
                            record.group_id,
                            record.terminal_no,
                            record.input_output,
                            record.text
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'choketable':
                    for record in records:
                        row = [
                            record.cable_id,
                            record.choke_id,
                            record.input_terminal,
                            record.output_terminal,
                            record.terminal_name,
                            record.output_type,
                            record.output_text,
                            record.output_connected
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'resistortable':
                    for record in records:
                        row = [
                            record.cable_id,
                            record.resistor_id,
                            record.input_terminal,
                            record.output_terminal,
                            record.resistor_name
                        ]
                        ws.append(row)
                        total_records += 1

            # (If no records, sheet is left with only headers – no "No data" message)

        # ========== UPDATE PROJECT STATUS FOR FULL PROJECT DOWNLOAD ==========
        if not location_id:
            seed_default_station_drawing(project_id)
            current_project.stage = 10
            current_project.status = 'ready_for_pdf'
            current_project.updated_date = get_ist_now()
            db.session.commit()
            print(f"✅ Project {project_id} marked as stage 10 with status 'ready_for_pdf' (full project download)")

        # ========== SAVE XLSX TO DOWNLOAD FOLDER ==========
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        download_dir = os.path.join(base_dir, 'xlsx_download')
        os.makedirs(download_dir, exist_ok=True)

        timestamp = get_ist_now().strftime('%Y%m%d_%H%M%S')
        project_name_clean = re.sub(r'[^\w\-_\. ]', '', current_project.name).replace(' ', '_')

        if location_id and location:
            junction_name_clean = re.sub(r'[^\w\-_\. ]', '', location.junction_name).replace(' ', '_') if location.junction_name else f"junction_{location_id}"
            filename = f"PROJECT_{project_id}_{project_name_clean}_JUNCTION_{location.junction_id}_{junction_name_clean}_{timestamp}.xlsx"
            pdf_filename = f"PROJECT_{project_id}_{project_name_clean}_JUNCTION_{location.junction_id}_{junction_name_clean}_{timestamp}.pdf"
            flash_message = f"✅ XLSX for junction box '{location.junction_name}' downloaded successfully! Project status updated to 'preview'."
        else:
            filename = f"RAILWAYPROJECT_ID{project_id}_{project_name_clean}_{timestamp}.xlsx"
            pdf_filename = f"railway_project_{project_id}_{project_name_clean}_{timestamp}.pdf"
            flash_message = f"✅ Full project XLSX downloaded successfully! Project marked as stage 10 with status 'ready_for_pdf'"

        file_path = os.path.join(download_dir, filename)
        print(file_path)
        wb.save(file_path)
        #=================New CODE 2026-03-11-16-34 ================
        download_dir = r"C:\Railway\git\xlsx_download" #"/var/www/html/git/xlsx_download"
        upload_dir = r"C:\Railway\git\uploads" #"/var/www/html/git/uploads"

        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)

        # Excel file path
        #file_path = os.path.join(download_dir, filename)

        pdf_filename_2=pdf_filename
        pdf_path = os.path.join(upload_dir, pdf_filename)
        print(pdf_filename)
        print(pdf_filename_2)
        try:
            python_exe = sys.executable
           
            
            #python_exe = "/usr/bin/python3"
            converter_script = r"C:\Railway\git\excel_to_pdf_converter.py" #"/var/www/html/git/excel_to_pdf_converter.py"
            #file_path = "/root/srv/local/git/xlsx_download/RAILWAYPROJECT_ID72_Upendra_20260312_151833.xlsx"
            #pdf_path = "/root/srv/local/git/uploads/RAILWAYPROJECT_ID72_Upendra_20260312_151835.pdf"
           


            result = subprocess.run(
                [python_exe, converter_script, file_path, pdf_path],
                capture_output=True,
                text=True,
                encoding='utf-8',
                timeout=300
            )
            print("==== SUBPROCESS DEBUG ====", file=sys.stderr)
            print("RETURN CODE:", result.returncode, file=sys.stderr)
            print("STDOUT:", result.stdout, file=sys.stderr)
            print("STDERR:", result.stderr, file=sys.stderr)
            print("=========================", file=sys.stderr)
            if result.returncode != 0:
                print("PDF conversion error:", result.stderr)
                 # ========== FLASH MESSAGE ==========
                flash(flash_message, "success")

                # ========== RETURN RESPONSE ==========
                if location_id:
                    print(f"✅ File saved to: {file_path}")
                    return jsonify({
                        'success': True,
                        'message': flash_message,
                        'file_path': file_path,
                        'filename': filename
                    })
                else:
                    return send_file(
                        file_path,
                        as_attachment=True,
                        download_name=filename,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
               
            else:
                # ===== MOVE EXCEL FILE TO uploads =====
                new_excel_path = os.path.join(upload_dir, filename)
                shutil.copy(file_path, new_excel_path)
                if result.returncode == 0 and os.path.exists(pdf_path):
                    file_md5 = _md5_of_file(pdf_path)
                    file_size = os.path.getsize(pdf_path)
                    meta = parse_converter_stdout(result.stdout)
                    db_checksum = meta.get("metadata_checksum") or meta.get("full_file_md5") or file_md5
                    max_version_record = GeneratedPDF.query.filter_by(project_id=project_id).order_by(GeneratedPDF.version.desc()).first()
                    next_version = max_version_record.version + 1 if max_version_record else 1
                    # Get junction data from the database at this moment
                    junction_boxes = JunctionBox.query.filter_by(project_id=project_id).all()
                    junction_data_list = []
                    for jb in junction_boxes:
                        junction_data_list.append({
                            'junction_id': jb.junction_id,
                            'junction_name': jb.junction_name,
                            'junction_size': jb.junction_size,
                            'station_id': jb.station_id,
                            'latitude': jb.latitude,
                            'longitude': jb.longitude,
                            'junction_row': jb.junction_row
                        })
                    
                    # Convert to JSON string
                    import json
                    remarks=''
                    junction_data_json = json.dumps(junction_data_list) if junction_data_list else None
                    record = GeneratedPDF(
                        project_id=project_id,
                        pdf_filename=pdf_filename_2,
                        xlsx_filename=filename,
                        checksum_md5=db_checksum,
                        file_size=file_size,
                        checksum_algo="md5",
                        metadata_checksum=meta.get("metadata_checksum"),
                        metadata_data=meta.get("metadata_data"),
                        initial_size_bytes=meta.get("initial_size_bytes"),
                        final_size_bytes=meta.get("final_size_bytes"),
                        metadata_ts_ist=meta.get("metadata_ts_ist"),
                        station_code=meta.get("station_code"),
                        source_pdf_name=meta.get("source_pdf_name"),
                        full_file_md5=meta.get("full_file_md5") or file_md5,
                        remarks=remarks if remarks else None,
                        created_at=get_ist_now(),
                        version=next_version,
                        junction_data=junction_data_json  # Store junction data
                    )
                    db.session.add(record)
                    db.session.commit()
                    
                    junction_approval = JunctionApproval.query.filter_by(
                        project_id=project_id,
                        generated_pdf_id=record.id,
                        junction_box_id=location_id
                    ).first()
                    
                    if not junction_approval:
                        # Create new junction approval
                        junction_approval = JunctionApproval(
                            project_id=project_id,
                            generated_pdf_id=record.id,
                            junction_box_id=location_id,
                            level1_status='pending',
                            level1_approver_id=current_user.id,
                            level1_approval_date=get_ist_now(),
                            level2_status='pending',
                            level3_status='pending',
                            created_at=get_ist_now()
                        )
                    db.session.add(junction_approval)
                    db.session.commit()
                    # ✅ CRITICAL: SET PROJECT STAGE TO 10 (PDF Generated)
                    try:
                        if current_project:
                            # Update project stage to 10 (PDF Generated)
                            current_project.stage = 10
                            current_project.updated_date = get_ist_now()
                            print(f"✅ Updated project {project_id} stage to 10 (PDF generated)")
                    except Exception as stage_error:
                        print(f"⚠️ Warning: Could not update project stage: {stage_error}")
                    
                    # Update StationDrawing with latest checksum and version
                    try:
                        # Refresh the StationDrawing data
                        station_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
                        if station_drawing:
                            # Ensure project name matches station name
                            if station_drawing.station_name and current_project.name != station_drawing.station_name:
                                current_project.name = station_drawing.station_name
                                current_project.updated_date = get_ist_now()
                                db.session.commit()
                                print(f"✅ Synchronized project name with station name: {station_drawing.station_name}")
                        
                            # Debug output to verify data
                            print(f"🔍 DEBUG StationDrawing after import:")
                            print(f" - station_id: {station_drawing.station_id}")
                            print(f" - station_name: {station_drawing.station_name}")
                            print(f" - station_code: {station_drawing.station_code}")
                            print(f" - version: {station_drawing.version}")
                            print(f" - checksum: {station_drawing.checksum}")
                        
                    except Exception as e:
                        print(f"❌ Error synchronizing project and station data: {str(e)}")
                        db.session.rollback()
                    
                
                    admin_users = User.query.filter_by(role='admin').all()
                    for admin in admin_users:
                        # Check if admin is assigned to this project
                        if current_project in admin.projects or admin.role == 'admin':  # Admin might have access to all
                            notification = Notification(
                                user_id=admin.id,
                                pdf_id=record.id,
                                project_id=project_id,
                                level='New_Drawing',
                                status='pending',
                                # UPDATED MESSAGE FORMAT
                                message=f'NEW DRAWING requires admin attention: {current_project.name if current_project else "Unknown"}'
                            )
                            db.session.add(notification)

                    # Also create notification for level1 users assigned to this project
                    level1_users = User.query.filter_by(designation='level1').all()
                    for user in level1_users:
                        # Check if level1 user is assigned to this project
                        if current_project in user.projects:
                            notification = Notification(
                                user_id=user.id,
                                pdf_id=record.id,
                                project_id=project_id,
                                level='level1',
                                status='pending',
                                # UPDATED MESSAGE FORMAT
                                message=f'NEW DRAWING requires level1 approval by: {user.username}'
                            )
                            db.session.add(notification)
                    
                    db.session.commit()  # Commit the notifications
                    
                
                
                  
                
        except Exception as e:
            print("PDF conversion failed222:", str(e))
            # Make sure the folder exists
            
        
        ##========================================================

        # ========== FLASH MESSAGE ==========
        flash(flash_message, "success")

        # ========== RETURN RESPONSE ==========
        if location_id:
            print(f"✅ File saved to: {file_path}")
            return send_file(
                file_path,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            return jsonify({
                'success': True,
                'message': flash_message,
                'file_path': file_path,
                'filename': filename
            })
        else:
            return send_file(
                file_path,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error generating download: {str(e)}", "error")
        print(f"❌ Download error: {str(e)}")
        import traceback
        traceback.print_exc()
        #return redirect(request.referrer or url_for("main.index"))
        if location_id:
            print(f"✅ File saved to: {file_path}")
            return jsonify({
                'success': True,
                'message': flash_message,
                'file_path': file_path,
                'filename': filename
            })
        else:
            return send_file(
                file_path,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

@bp.route("/downloadajax/")
def downloadajax():
    """Download XLSX file - can be for entire project or specific location (JunctionBox)"""
    project_id = request.args.get('project_id', type=int) or get_current_project()
    location_id = request.args.get('location_id', type=int)
    download_only = request.args.get('download_only', type=bool, default=False)

    if not project_id:
        flash("Please select a project first", "error")
        #return redirect(url_for("main.project_selection"))

    # Get project
    current_project = Project.query.get(project_id)
    if not current_project:
        flash("Project not found", "error")
        #return redirect(url_for("main.index"))

    # ---- CASE A: Specific location (Preview This Location) ----
    if location_id:
        location = JunctionBox.query.filter_by(id=location_id, project_id=project_id).first()
        if not location:
            flash("Location (Junction Box) not found in this project", "error")
            #return redirect(request.referrer or url_for("main.index"))

        # Check if this junction box has Level 1 approval
        approved = JunctionApproval.query.filter_by(
            project_id=project_id,
            junction_box_id=location.id,
            level1_status='approved'
        ).first() is not None

        if approved:
            # ✅ XLSX not needed – directly show PDF
            return jsonify({
                'skip_xlsx': True,
                'message': f"Junction '{location.junction_name}' is already approved at Level 1. Showing PDF.",
                'project_id': project_id,
                'location_id': location_id
            })

    # ---- CASE B: Full project (Preview Full Project / Preview Station) ----
    else:
        all_junctions = JunctionBox.query.filter_by(project_id=project_id).all()
        if all_junctions:
            all_approved = all(
                JunctionApproval.query.filter_by(
                    project_id=project_id,
                    junction_box_id=j.id,
                    level1_status='approved'
                ).first() is not None
                for j in all_junctions
            )

            if all_approved:
                # ✅ All locations are Level 1 approved – skip XLSX
                return jsonify({
                    'skip_xlsx': True,
                    'message': "All locations in this project are approved at Level 1. Showing PDF.",
                    'project_id': project_id
                })

    # ------------------------------------------------------------
    # 2. NORMAL XLSX GENERATION (only if not skipped above)
    # ------------------------------------------------------------
    try:
        # ========== CREATE XLSX WORKBOOK ==========
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        total_records = 0

        # Get cable names mapping
        cable_name_map = {}
        try:
            cables = Cable.query.filter_by(project_id=project_id).all()
            cable_name_map = {str(cable.cable_id): cable.cable_name for cable in cables}
        except Exception as e:
            print(f"Warning loading cable names: {str(e)}")

        # Get location info if location_id is provided (for status update)
        location = None
        if location_id:
            location = JunctionBox.query.filter_by(id=location_id, project_id=project_id).first()
            if not location:
                flash("Location (Junction Box) not found in this project", "error")
                #return redirect(request.referrer or url_for("main.index"))

            # Update project status to 'preview' for location‑specific download
            current_project.status = 'preview'
            current_project.updated_date = get_ist_now()
            db.session.commit()
            print(f"✅ Project {project_id} status updated to 'preview' (location-specific download)")

        # ========== CREATE SHEETS ==========
        for sheet_name, columns in SHEETS.items():
            ws = wb.create_sheet(title=sheet_name)

            # Get the model for this sheet
            model = MODEL_MAP.get(sheet_name)
            if not model:
                continue

            # ========== WRITE HEADERS ==========
            if sheet_name == 'terminal':
                template_headers = [
                    'cable_id', 'cable_name', 'terminal_id', 'terminal_no', 'symbol',
                    'input_left', 'input_right', 'spare', 'input_connected',
                    'output_connected', 'input_connected_extra', 'output_connected_extra',
                    'output_left', 'output_right'
                ]
                ws.append(template_headers)
            else:
                ws.append(columns)

            # ========== GET RECORDS WITH LOCATION FILTERING ==========
            records = []

            if location_id:
                # Filter by specific location (junction_box)
                if sheet_name == 'junction_box':
                    records = model.query.filter_by(
                        project_id=project_id,
                        id=location_id
                    ).all()

                elif sheet_name == 'StationDrawing':
                    records = model.query.filter_by(project_id=project_id).all()

                elif sheet_name == 'cable':
                    records = model.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()

                elif sheet_name == 'cable_box':
                    records = model.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()

                elif sheet_name == 'terminal':
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        for cable_id in cable_ids:
                            cable_terminals = model.query.filter_by(
                                project_id=project_id,
                                cable_id=cable_id
                            ).all()
                            records.extend(cable_terminals)

                elif sheet_name in ['group', 'terminal_header', 'choketable', 'resistortable']:
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        for cable_id in cable_ids:
                            sheet_records = model.query.filter_by(
                                project_id=project_id,
                                cable_id=cable_id
                            ).all()
                            records.extend(sheet_records)

            else:
                # Full project download
                records = model.query.filter_by(project_id=project_id).all()

            # ========== WRITE DATA ==========
            if records:
                if sheet_name == 'terminal':
                    # Sort terminal records
                    sorted_records = sorted(records, key=lambda x: (
                        int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0,
                        int(x.terminal_id) if x.terminal_id and str(x.terminal_id).isdigit() else 0
                    ))

                    for record in sorted_records:
                        row = [
                            record.cable_id,
                            cable_name_map.get(str(record.cable_id), ''),
                            record.terminal_id,
                            record.terminal_no,
                            record.symbol,
                            record.input_left if record.input_left and record.input_left.strip() else '',
                            record.input_right if record.input_right and record.input_right.strip() else '',
                            record.spare,
                            record.input_connected,
                            record.output_connected,
                            record.input_connected_extra if record.input_connected_extra and record.input_connected_extra.strip() else '',
                            record.output_connected_extra if record.output_connected_extra and record.output_connected_extra.strip() else '',
                            record.output_left if record.output_left and record.output_left.strip() else '',
                            record.output_right if record.output_right and record.output_right.strip() else ''
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'cable':
                    # Sort cable records
                    sorted_records = sorted(records, key=lambda x: (
                        int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0
                    ))

                    for record in sorted_records:
                        row = [
                            record.cable_id,
                            record.cable_name,
                            record.junction_box,
                            record.junction_name,
                            record.row,
                            record.position,
                            record.terminal,
                            record.start_no
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'cable_box':
                    # Sort cable_box records
                    sorted_records = sorted(records, key=lambda x: (
                        int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0
                    ))

                    for record in sorted_records:
                        row = [
                            record.cable_id,
                            record.cable_name,
                            record.junction_box,
                            record.junction_name,
                            record.row,
                            record.position,
                            record.terminal,
                            record.start_no,
                            record.cable_type,
                            record.output
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'junction_box':
                    for record in records:
                        row = [
                            record.station_id,
                            record.junction_id,
                            record.junction_name,
                            record.latitude,
                            record.longitude,
                            record.junction_size,
                            record.junction_row
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'StationDrawing':
                    for record in records:
                        row = [
                            record.checksum,
                            record.station_id,
                            record.diagram_name,
                            record.station_name,
                            record.station_code,
                            record.version,
                            record.date,
                            record.drawn_by,
                            record.checked_by,
                            record.division,
                            record.zone,
                            record.total_sheet,
                            record.designation1,
                            record.designation2,
                            record.designation3
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'terminal_header':
                    for record in records:
                        row = [
                            record.cable_id,
                            record.header_type,
                            record.terminal_start,
                            record.terminal_end,
                            record.input_output,
                            record.text
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'group':
                    for record in records:
                        row = [
                            record.cable_id,
                            record.group_id,
                            record.terminal_no,
                            record.input_output,
                            record.text
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'choketable':
                    for record in records:
                        row = [
                            record.cable_id,
                            record.choke_id,
                            record.input_terminal,
                            record.output_terminal,
                            record.terminal_name,
                            record.output_type,
                            record.output_text,
                            record.output_connected
                        ]
                        ws.append(row)
                        total_records += 1

                elif sheet_name == 'resistortable':
                    for record in records:
                        row = [
                            record.cable_id,
                            record.resistor_id,
                            record.input_terminal,
                            record.output_terminal,
                            record.resistor_name
                        ]
                        ws.append(row)
                        total_records += 1

            # (If no records, sheet is left with only headers – no "No data" message)

        # ========== UPDATE PROJECT STATUS FOR FULL PROJECT DOWNLOAD ==========
        if not location_id:
            seed_default_station_drawing(project_id)
            current_project.stage = 10
            current_project.status = 'ready_for_pdf'
            current_project.updated_date = get_ist_now()
            db.session.commit()
            print(f"✅ Project {project_id} marked as stage 10 with status 'ready_for_pdf' (full project download)")

        # ========== SAVE XLSX TO DOWNLOAD FOLDER ==========
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        download_dir = os.path.join(base_dir, 'xlsx_download')
        os.makedirs(download_dir, exist_ok=True)

        timestamp = get_ist_now().strftime('%Y%m%d_%H%M%S')
        project_name_clean = re.sub(r'[^\w\-_\. ]', '', current_project.name).replace(' ', '_')

        if location_id and location:
            junction_name_clean = re.sub(r'[^\w\-_\. ]', '', location.junction_name).replace(' ', '_') if location.junction_name else f"junction_{location_id}"
            filename = f"PROJECT_{project_id}_{project_name_clean}_JUNCTION_{location.junction_id}_{junction_name_clean}_{timestamp}.xlsx"
            flash_message = f"✅ XLSX for junction box '{location.junction_name}' downloaded successfully! Project status updated to 'preview'."
        else:
            filename = f"RAILWAYPROJECT_ID{project_id}_{project_name_clean}_{timestamp}.xlsx"
            flash_message = f"✅ Full project XLSX downloaded successfully! Project marked as stage 10 with status 'ready_for_pdf'"

        file_path = os.path.join(download_dir, filename)
        wb.save(file_path)
        #=================New CODE 2026-03-11-16-34 ================
        download_dir = "/root/srv/local/git/xlsx_download"
        upload_dir = "/root/srv/local/git/uploads"

        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)

        # Excel file path
        #file_path = os.path.join(download_dir, filename)

        # Convert .xlsx → .pdf
        pdf_filename = filename.replace(".xlsx", ".pdf")
        
        
        # remove extension
        name = os.path.splitext(filename)[0]

        # extract parts
        m = re.match(r"RAILWAYPROJECT_ID(\d+)_(.*?)_(\d{8}_\d{6})", name)

        if m:
            project_id = m.group(1)
            user = m.group(2)
            timestamp = m.group(3)

        pdf_filename = f"railway_project_{project_id}_{timestamp}_{user}.pdf"
        pdf_path = os.path.join(upload_dir, pdf_filename)
       
        import threading
        from flask import current_app

        threading.Thread(
            target=process_pdf_background,
            args=(
                current_app._get_current_object(),
                project_id,
                file_path,
                pdf_path,
                filename,
                pdf_filename,
                upload_dir
            )
        ).start()    
        
        ##========================================================

        # ========== FLASH MESSAGE ==========
        flash(flash_message, "success")

        # ========== RETURN RESPONSE ==========
        if location_id:
            print(f"✅ File saved to: {file_path}")
            return jsonify({
                'success': True,
                'message': flash_message,
                'file_path': file_path,
                'filename': filename
            })
        else:
            return send_file(
                file_path,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error generating download: {str(e)}", "error")
        print(f"❌ Download error: {str(e)}")
        import traceback
        traceback.print_exc()
        #return redirect(request.referrer or url_for("main.index"))
        if location_id:
            print(f"✅ File saved to: {file_path}")
            return jsonify({
                'success': True,
                'message': flash_message,
                'file_path': file_path,
                'filename': filename
            })
        else:
            return send_file(
                file_path,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
     
@bp.route("/project/<int:project_id>/switch")
@login_required
def switch_project(project_id):
    """Switch to a different project - with access control for non-admin users"""
    # For non-admin users, verify they have access to this project
    if current_user.role == 'user':  # Only check for 'user' role, admin has full access
        # CHANGE THIS LINE: current_user.projects -> user_project_ids
        user_project_ids = [p.id for p in current_user.projects]
        if project_id not in user_project_ids:
            flash("Access denied. You are not assigned to this project.")
            return redirect(url_for('main.index'))
    
    project = Project.query.get_or_404(project_id)
    session['current_project_id'] = project.id
    session['project_id'] = project.id
    print(f"✅ DEBUG switch_project: Set session to project {project.id}")
    flash(f"Switched to Project: {project.name}")
    return redirect(url_for("main.index"))

@bp.route("/new_project", methods=["GET", "POST"])
@login_required
def new_project():
    # Allow both admin and user roles to create projects
    if current_user.role_name not in ['admin', '4']:
        flash("Access denied. Only admin and user roles can create new projects.", "warning")
        return redirect(url_for('main.index'))
       
    if request.method == "POST":
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
       
        if name:
            # Check if project/station name already exists
            existing_project = Project.query.filter(
                func.lower(Project.name) == func.lower(name)
            ).first()
            
            if existing_project:
                flash(f"Station name '{name}' already exists. Please choose a different name.", "danger")
                return render_template("new_project.html")
            
            # Create the project with station_id set to project id
            project = Project(name=name, description=description)
            db.session.add(project)
            db.session.flush()  # Get the ID without committing
            
            # Set station_id to be the same as project id
            project.station_id = project.id
            
            db.session.commit()
           
            # Auto-assign the project to the user who created it
            current_user.projects.append(project)
            db.session.commit()
           
            # Set both session keys
            session['current_project_id'] = project.id
            session['project_id'] = project.id
            print(f"✅ DEBUG new_project: Created and set session to project {project.id}")
           
            # Use project name as station name
            seed_default_station_drawing(project.id, name)
           
            flash(f"Created new Project ID {project.id}: {name}")
            return redirect(url_for("main.approval_tracking"))
        else:
            flash("Station name is required")
   
    return render_template("new_project.html")


@bp.route('/api/check_station_name')
@login_required
def check_station_name():
    """Check if a station name already exists"""
    station_name = request.args.get('name', '').strip()
    
    if not station_name:
        return jsonify({'exists': False})
    
    # Check if project name already exists (case-insensitive)
    existing_project = Project.query.filter(
        func.lower(Project.name) == func.lower(station_name)
    ).first()
    
    return jsonify({'exists': existing_project is not None})
    
@bp.route("/clear_current_project", methods=["POST"])
@login_required
def clear_current_project():
    project_id = get_current_project()
    if not project_id:
        return redirect(url_for("main.project_selection"))
    
    # Check if user has access to clear this project
    # Admin can clear any project, users can only clear their assigned projects
    project = Project.query.get(project_id)
    if current_user.role == 'user' and project not in current_user.projects:
        flash("Access denied. You can only clear your own projects.")
        return redirect(url_for('main.index'))
    
    try:
        total_deleted = 0
        for model in MODEL_MAP.values():
            count = model.query.filter_by(project_id=project_id).count()
            model.query.filter_by(project_id=project_id).delete()
            total_deleted += count
        db.session.commit()
        flash(f"All data cleared from Project ID {project_id} (Deleted {total_deleted} records)")
    except Exception as e:
        flash(f"Error clearing project data: {str(e)}")
        db.session.rollback()
    return redirect(url_for("main.index"))

@bp.route("/project/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)
   
    # Check permissions:
    # 1. Admin can edit any project
    # 2. Users can edit only projects assigned to them
    if current_user.role == 'user' and project not in current_user.projects:
        flash("Access denied. You are not assigned to this project.")
        return redirect(url_for('main.index'))
   
    if request.method == "POST":
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
       
        if name:
            # Check if new name already exists for another project
            existing_project = Project.query.filter(
                func.lower(Project.name) == func.lower(name),
                Project.id != project_id
            ).first()
            
            if existing_project:
                flash(f"Station name '{name}' already exists for another project. Please choose a different name.", "danger")
                return render_template("edit_project.html", project=project)
            
            project.name = name
            project.description = description if description else None
            project.updated_date = get_ist_now()
           
            try:
                db.session.commit()
                flash(f"✅ Project '{project.name}' updated successfully!")
                return redirect(url_for("main.approval_tracking"))
            except Exception as e:
                db.session.rollback()
                flash(f"❌ Error updating project: {str(e)}")
        else:
            flash("❌ Project name is required")
   
    return render_template("edit_project.html", project=project)

@bp.route("/project/<int:project_id>/delete", methods=["POST"])
@login_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    project_name = project.name
    
    # Check permissions:
    # 1. Admin can delete any project
    # 2. Users can delete only projects assigned to them
    if current_user.role == 'user' and project not in current_user.projects:
        flash("Access denied. You can only delete your own projects.")
        return redirect(url_for('main.index'))
    
    try:
        # Delete all related data first
        total_deleted = 0
        
        # 1. Delete data from all MODEL_MAP tables
        for model in MODEL_MAP.values():
            count = model.query.filter_by(project_id=project_id).count()
            model.query.filter_by(project_id=project_id).delete()
            total_deleted += count
        
        # 2. Delete GeneratedPDF records
        pdf_count = GeneratedPDF.query.filter_by(project_id=project_id).count()
        GeneratedPDF.query.filter_by(project_id=project_id).delete()
        total_deleted += pdf_count
        
        # 3. Remove project from all users' assigned projects
        # CHANGE THIS LINE: project.users -> project.assigned_users
        for user in project.assigned_users:  # FIXED: Changed from project.users
            user.projects.remove(project)
        
        # 4. Delete the project itself
        db.session.delete(project)
        db.session.commit()
        
        # 5. Clear session if this was the current project
        if session.get('project_id') == project_id:
            session.pop('project_id', None)
            session.pop('current_project_id', None)
        
        flash(f"✅ Project '{project_name}' deleted successfully! ({total_deleted} database records removed)")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error deleting project: {str(e)}")
    
    return redirect(url_for("main.index")) 

def seed_default_station_drawing(project_id: int, project_name: str = "KHEDBRAHMA"):
    """Seed default station drawing using project name as station name, project ID as station ID, and auto-version starting at 0"""
    exists = StationDrawing.query.filter_by(project_id=project_id).first()
    if exists:
        return
   
    # Calculate version based on existing PDFs for this project
    pdf_count = GeneratedPDF.query.filter_by(project_id=project_id).count()
    current_version = str(pdf_count)  # Version = current PDF count (0, 1, 2...)
   
    # Get checksum from latest PDF if available
    latest_pdf = GeneratedPDF.query.filter_by(project_id=project_id).order_by(GeneratedPDF.id.desc()).first()
    checksum_value = latest_pdf.checksum_md5 if latest_pdf else "NO PDF GENERATED"
   
    # Generate station code from project name (first 4 characters uppercase)
    station_code = project_name[:4].upper() if project_name else "KDBM"
   
    # Use current date in dd-mm-yyyy format
    current_date = datetime.now().strftime("%d-%m-%Y")
    
    # Use logged-in username for drawn_by; fallback to 'user' if not logged in
    drawn_by_user = current_user.username if current_user.is_authenticated else "user"
   
    sd = StationDrawing(
        project_id=project_id,
        checksum=checksum_value,
        station_id=str(project_id),
        diagram_name="railways",
        station_name=project_name,
        station_code=station_code,
        version=current_version,
        date=current_date,
        drawn_by=drawn_by_user,  # ✅ Updated to take logged-in username
        checked_by="supervisor",
        division="Ahemdabad",
        zone="WRLY",
        total_sheet="17",
        designation1="DY.CSTE/C-II/ADI",
        designation2="DSTE/C/ADI",
        designation3="SSE/SIG/C/ADI",
        created_date=get_ist_now(),
    )
    
    db.session.add(sd)
    db.session.commit()
    
    print(f"✅ Seeded StationDrawing for project {project_id} with station name: {project_name}, station ID: {project_id}, version: {current_version}, checksum: {checksum_value}, drawn_by: {drawn_by_user}")

@bp.route("/project/<int:project_id>/continue")
@login_required
def continue_project(project_id):
    # Only allow admin users to switch projects
    if current_user.role != 'admin' and project_id != current_user.station_id:
        flash("Access denied. You are not assigned to this project.")
        return redirect(url_for('main.approval_tracking'))
       
    project = Project.query.get_or_404(project_id)
    session['current_project_id'] = project.id
    session['project_id'] = project.id
    print(f"✅ DEBUG continue_project: Set session to project {project.id}")
    flash(f"Switched to Project: {project.name}")
    return redirect(url_for("main.workflow_step", step=2))

@bp.route("/continue_from_version/<int:pdf_id>", methods=["POST"])
@login_required
def continue_from_version(pdf_id):
    """Load XLSX from a specific GeneratedPDF version, import data, set continue mode, and redirect to workflow step 2."""
    try:
        # Get the PDF record
        pdf_record = GeneratedPDF.query.get_or_404(pdf_id)
        project_id = pdf_record.project_id
        xlsx_filename = pdf_record.xlsx_filename
        
        if not xlsx_filename:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                return jsonify({"success": False, "message": "❌ No XLSX file associated with this PDF version."}), 400
            flash("❌ No XLSX file associated with this PDF version.", "danger")
            return redirect(url_for("main.index"))
        
        # Verify user access to project
        if not user_has_project_access(project_id):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                return jsonify({"success": False, "message": "❌ Access denied. You are not assigned to this project."}), 403
            flash("❌ Access denied. You are not assigned to this project.", "danger")
            return redirect(url_for("main.index"))
        
        upload_dir = os.path.join(os.getcwd(), 'uploads')
        xlsx_path = os.path.join(upload_dir, xlsx_filename)
        
        if not os.path.exists(xlsx_path):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                return jsonify({"success": False, "message": f"❌ XLSX file not found: {xlsx_filename}"}), 404
            flash(f"❌ XLSX file not found: {xlsx_filename}", "danger")
            return redirect(url_for("main.index"))
        
        # Load workbook (similar to excel_to_pdf)
        wb = load_workbook(xlsx_path, data_only=True)
        total_imported = 0
        
        # Get current project
        current_project = Project.query.get(project_id)
        
        # For each sheet: APPEND data instead of clearing
        for sheet_name in SHEETS.keys():
            try:
                model = MODEL_MAP[sheet_name]
                expected_headers = SHEETS[sheet_name]
                
                # Find sheet
                sheet_found = None
                for ws_name in wb.sheetnames:
                    if ws_name.lower() == sheet_name.lower():
                        sheet_found = ws_name
                        break
                if not sheet_found:
                    continue
                
                ws = wb[sheet_found]
                
                # Map headers
                headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
                header_mapping = {}
                for required_header in expected_headers:
                    for i, file_header in enumerate(headers):
                        if file_header == required_header.lower():
                            header_mapping[required_header] = i
                            break
                
                # DON'T clear existing rows - just import new rows
                # model.query.filter_by(project_id=project_id).delete(synchronize_session=False)
                
                # Import rows
                sheet_count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = {'project_id': project_id}
                    has_data = False
                    
                    # Skip if this is StationDrawing and we already have one (only keep one)
                    if sheet_name == 'StationDrawing':
                        existing_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
                        if existing_drawing:
                            # Only update the existing one, don't create new
                            continue
                    
                    for header in expected_headers:
                        if header in header_mapping:
                            col_index = header_mapping[header]
                            cell_value = row[col_index] if col_index < len(row) else None
                            if cell_value is not None:
                                text = str(cell_value).strip()
                                
                                # ENFORCE LOCATION NAMING CONVENTION FOR JUNCTION BOXES
                                if sheet_name == 'junction_box' and header == 'junction_name':
                                    size_header_index = None
                                    for h in expected_headers:
                                        if h == 'junction_size' and h in header_mapping:
                                            size_header_index = header_mapping[h]
                                            break
                                    
                                    if size_header_index is not None and size_header_index < len(row):
                                        size_value = row[size_header_index]
                                        if size_value is not None:
                                            size_text = str(size_value).strip()
                                            text = format_location_name(text, size_text)
                                
                                data[header] = text if text else None
                                if text:
                                    has_data = True
                            else:
                                data[header] = None
                        else:
                            data[header] = None
                    
                    if has_data:
                        try:
                            # Check for duplicates before adding
                            if sheet_name == 'junction_box':
                                # For junction boxes, check if same junction_id exists
                                junction_id = data.get('junction_id')
                                if junction_id:
                                    existing = JunctionBox.query.filter_by(
                                        project_id=project_id,
                                        junction_id=junction_id
                                    ).first()
                                    if existing:
                                        print(f"⚠️ Junction box {junction_id} already exists, skipping")
                                        continue
                            
                            elif sheet_name == 'cable':
                                # For cables, check if same cable_id exists
                                cable_id = data.get('cable_id')
                                if cable_id:
                                    existing = Cable.query.filter_by(
                                        project_id=project_id,
                                        cable_id=cable_id
                                    ).first()
                                    if existing:
                                        print(f"⚠️ Cable {cable_id} already exists, skipping")
                                        continue
                            
                            # Add new record
                            db.session.add(model(**data))
                            sheet_count += 1
                        except Exception as e:
                            print(f"Error importing {sheet_name} row: {e}")
                            continue
                
                if sheet_count > 0:
                    db.session.commit()
                    total_imported += sheet_count
                    
                    # SPECIAL HANDLING FOR STATIONDRAWING SHEET
                    if sheet_name == 'StationDrawing':
                        existing_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
                        
                        # Process first row (single record)
                        row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
                        if row:
                            data = {'project_id': project_id}
                            has_data = False
                            
                            for header in expected_headers:
                                if header in header_mapping:
                                    col_index = header_mapping[header]
                                    cell_value = row[col_index] if col_index < len(row) else None
                                    if cell_value is not None:
                                        text = str(cell_value).strip()
                                        data[header] = text if text else None
                                        if text and text != "NO PDF GENERATED":
                                            has_data = True
                                    else:
                                        data[header] = None
                                else:
                                    data[header] = None
                            
                            if has_data:
                                try:
                                    if existing_drawing:
                                        # UPDATE existing
                                        for col, val in data.items():
                                            if col != 'project_id':
                                                setattr(existing_drawing, col, val)
                                    
                                    # Set version to this PDF's version (pdf_count at creation time)
                                    pdf_count_at_creation = GeneratedPDF.query.filter_by(project_id=project_id).filter(GeneratedPDF.id <= pdf_record.id).count()
                                    existing_drawing.version = str(pdf_count_at_creation - 1)  # Adjust to match historical version
                                    
                                    # Set checksum to this PDF's checksum
                                    existing_drawing.checksum = pdf_record.checksum_md5
                                    
                                    db.session.commit()
                                    
                                    # Sync project name with station_name
                                    station_name = data.get('station_name')
                                    if station_name and current_project.name != station_name:
                                        current_project.name = station_name
                                        current_project.updated_date = get_ist_now()
                                        db.session.commit()
                                except Exception as e:
                                    print(f"❌ Error processing StationDrawing: {str(e)}")
                                    db.session.rollback()
                        
                        sheet_count = 1
                        
            except Exception as e:
                print(f"Error importing sheet {sheet_name}: {e}")
                db.session.rollback()
                continue
        
        # NEW: Restore junction data from PDF record if available
        if pdf_record.junction_data:
            try:
                import json
                junction_list = json.loads(pdf_record.junction_data)
                
                # DON'T clear existing junction boxes
                # JunctionBox.query.filter_by(project_id=project_id).delete(synchronize_session=False)
                
                # Restore junction boxes from stored data
                restored_count = 0
                for jb_data in junction_list:
                    # Check if this junction already exists
                    existing_junction = JunctionBox.query.filter_by(
                        project_id=project_id,
                        junction_id=jb_data.get('junction_id')
                    ).first()
                    
                    if not existing_junction:
                        # Only add if it doesn't exist
                        junction_box = JunctionBox(
                            project_id=project_id,
                            station_id=jb_data.get('station_id'),
                            junction_id=jb_data.get('junction_id'),
                            junction_name=jb_data.get('junction_name'),
                            junction_size=jb_data.get('junction_size'),
                            latitude=jb_data.get('latitude'),
                            longitude=jb_data.get('longitude'),
                            junction_row=jb_data.get('junction_row'),
                            created_date=get_ist_now()
                        )
                        db.session.add(junction_box)
                        restored_count += 1
                
                db.session.commit()
                print(f"✅ Restored {restored_count} new junction boxes from PDF record (skipped existing ones)")
            except Exception as e:
                print(f"❌ Error restoring junction data: {str(e)}")
                db.session.rollback()
        
        if total_imported > 0:
            # Set session to this project and continue mode
            session['current_project_id'] = project_id
            session['project_id'] = project_id
            session['is_continue_drawing'] = True
            
            # Check if it's an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                return jsonify({
                    "success": True, 
                    "message": f"✅ Loaded {total_imported} new records from version {pdf_record.id} ({pdf_record.pdf_filename}) and merged with existing data. Continuing drawing...",
                    "project_id": project_id
                })
            
            flash(f"✅ Loaded {total_imported} new records from version {pdf_record.id} ({pdf_record.pdf_filename}) and merged with existing data. Continuing drawing...", "success")
            
            # Redirect to workflow step 2
            return redirect(url_for("main.workflow_step", step=2, project_id=project_id))
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                return jsonify({"success": False, "message": "❌ No new data found in the XLSX file (all data already exists)."}), 400
            
            flash("❌ No new data found in the XLSX file (all data already exists).", "danger")
            return redirect(url_for("main.index"))
    
    except Exception as e:
        db.session.rollback()
        error_msg = f"❌ Error loading version: {str(e)}"
        print(f"❌ Error in continue_from_version: {str(e)}")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
            return jsonify({"success": False, "message": error_msg}), 500
        
        flash(error_msg, "danger")
        return redirect(url_for("main.index"))
    
    
# ==================== APPROVAL ROUTES ====================

# Helper function to normalize designation
def normalize_designation(designation):
    """Helper function to normalize designation"""
    if not designation: 
        return ''
    # Remove spaces and normalize to lowercase for comparison
    return designation.replace(' ', '').lower()

@bp.route('/get_approval_stats', methods=['GET'])
@login_required
def get_approval_stats():
    """Get approval statistics for dashboard"""
    try:
        # Count PDFs by approval status
        total_pdfs = GeneratedPDF.query.count()
        approved_pdfs = GeneratedPDF.query.filter_by(level3_status='approved').count()
        rejected_pdfs = GeneratedPDF.query.filter(
            (GeneratedPDF.level1_status == 'rejected') |
            (GeneratedPDF.level2_status == 'rejected') |
            (GeneratedPDF.level3_status == 'rejected')
        ).count()
        pending_pdfs = total_pdfs - approved_pdfs - rejected_pdfs
        
        # Count by level
        level1_pending = GeneratedPDF.query.filter_by(level1_status='pending').count()
        level2_pending = GeneratedPDF.query.filter_by(level2_status='pending').count()
        level3_pending = GeneratedPDF.query.filter_by(level3_status='pending').count()
        
        # Recent approvals (last 7 days) - This should now work since we're creating Approval records
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        recent_approvals = Approval.query.filter(
            Approval.created_at >= seven_days_ago,
            Approval.status == 'approved'
        ).count()
        
        # User-specific stats (if user is approver)
        user_approvals = Approval.query.filter_by(approver_id=current_user.id).count()
        user_rejections = Approval.query.filter_by(
            approver_id=current_user.id,
            status='rejected'
        ).count()
        
        # Pending approvals for this user
        user_pending_approvals = 0
        if current_user.designation == 'level1':
            user_pending_approvals = GeneratedPDF.query.filter_by(level1_status='pending').count()
        elif current_user.designation == 'level2':
            user_pending_approvals = GeneratedPDF.query.filter(
                GeneratedPDF.level1_status == 'approved',
                GeneratedPDF.level2_status == 'pending'
            ).count()
        elif current_user.designation == 'level3':
            user_pending_approvals = GeneratedPDF.query.filter(
                GeneratedPDF.level1_status == 'approved',
                GeneratedPDF.level2_status == 'approved',
                GeneratedPDF.level3_status == 'pending'
            ).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total': total_pdfs,
                'approved': approved_pdfs,
                'rejected': rejected_pdfs,
                'pending': pending_pdfs,
                'level1_pending': level1_pending,
                'level2_pending': level2_pending,
                'level3_pending': level3_pending,
                'recent_approvals': recent_approvals,
                'user_approvals': user_approvals,
                'user_rejections': user_rejections,
                'user_pending_approvals': user_pending_approvals
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/get_approval_history/<int:pdf_id>', methods=['GET'])
@login_required
def get_approval_history(pdf_id):
    """Get approval history for a PDF"""
    pdf = GeneratedPDF.query.get_or_404(pdf_id)
    
    # Check if user has permission to view this PDF
    if not current_user.role_name in ['1', '4'] and current_user.id not in [
        pdf.level1_approver_id, pdf.level2_approver_id, pdf.level3_approver_id
    ] and current_user.id != pdf.created_by:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get approval history from Approval table
    approvals = Approval.query.filter_by(generated_pdf_id=pdf_id)\
        .order_by(desc(Approval.created_at)).all()
    
    # Prepare response
    history = []
    for approval in approvals:
        history.append(approval.to_dict())
    
    # Also include current approval status from GeneratedPDF
    current_status = {
        'level1': {
            'status': pdf.level1_status,
            'approver': pdf.level1_approver.username if pdf.level1_approver else None,
            'date': pdf.level1_approval_date.strftime('%Y-%m-%d %H:%M:%S') if pdf.level1_approval_date else None
        },
        'level2': {
            'status': pdf.level2_status,
            'approver': pdf.level2_approver.username if pdf.level2_approver else None,
            'date': pdf.level2_approval_date.strftime('%Y-%m-%d %H:%M:%S') if pdf.level2_approval_date else None
        },
        'level3': {
            'status': pdf.level3_status,
            'approver': pdf.level3_approver.username if pdf.level3_approver else None,
            'date': pdf.level3_approval_date.strftime('%Y-%m-%d %H:%M:%S') if pdf.level3_approval_date else None
        }
    }
    
    return jsonify({
        'success': True,
        'pdf_id': pdf.id,
        'pdf_filename': pdf.pdf_filename,
        'project_name': pdf.project.name if pdf.project else 'Unknown',
        'version': pdf.version,
        'created_at': pdf.created_at.strftime('%Y-%m-%d %H:%M:%S') if pdf.created_at else None,
        'current_status': current_status,
        'approval_history': history
    })


@bp.route("/approve_pdf/<int:pdf_id>/<level>", methods=["POST"])
@login_required
def approve_pdf(pdf_id, level):
    """Approve PDF at specified level - ROLE-BASED (0,1,2,3,4)"""
    try:
        pdf = GeneratedPDF.query.get_or_404(pdf_id)
        project = Project.query.get_or_404(pdf.project_id)
        
        # Get approval remarks from form
        approval_remarks = request.form.get('approval_remarks', '')
        
        # Get user permissions based on role
        permissions = get_user_permissions(current_user)
        
        if level == '1':
            # ===== LEVEL 1 APPROVAL =====
            if not permissions['can_approve_level1']:
                flash("Access denied. You don't have permission to approve at Level 1.", "danger")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            if pdf.level1_status != 'pending':
                flash("This PDF has already been processed at Level 1.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Update PDF status
            pdf.level1_status = 'approved'
            pdf.level1_approver_id = current_user.id
            pdf.level1_approval_date = get_ist_now()
            
            # ✅ ADD THIS: Record approval in Approval table
            approval = Approval(
                generated_pdf_id=pdf.id,
                level=1,
                status='approved',
                approver_id=current_user.id,
                remarks=approval_remarks,
                created_at=get_ist_now()
            )
            db.session.add(approval)
            
            # ✅ ADD THIS: Create JunctionApproval records for ALL junctions
            junctions = project.junction_boxes
            for junction in junctions:
                # Check if junction approval already exists
                junction_approval = JunctionApproval.query.filter_by(
                    project_id=project.id,
                    generated_pdf_id=pdf.id,
                    junction_box_id=junction.id
                ).first()
                
                if not junction_approval:
                    # Create new junction approval
                    junction_approval = JunctionApproval(
                        project_id=project.id,
                        generated_pdf_id=pdf.id,
                        junction_box_id=junction.id,
                        level1_status='approved',
                        level1_approver_id=current_user.id,
                        level1_approval_date=get_ist_now(),
                        level2_status='pending',
                        level3_status='pending',
                        created_at=get_ist_now()
                    )
                    db.session.add(junction_approval)
                else:
                    # Update existing junction approval
                    junction_approval.level1_status = 'approved'
                    junction_approval.level1_approver_id = current_user.id
                    junction_approval.level1_approval_date = get_ist_now()
                
                # Update junction box status to "Under Process" (status=2)
                junction.status = 2
            
            # Update existing level1 notifications to 'approved'
            Notification.query.filter_by(
                pdf_id=pdf.id,
                level='level1',
                status='pending'
            ).update({
                'status': 'approved',
                'is_read': True,
                'updated_at': get_ist_now()
            }, synchronize_session=False)
            
            # Create notification for Level 2 users (role='2') assigned to this project
            level2_users = User.query.filter(
                User.role == '2',  # Role 2 = Level 2 Approvers
                User.is_active == True
            ).all()
            
            for user in level2_users:
                # Check if user is assigned to this project
                if project in user.projects:
                    notification = Notification(
                        user_id=user.id,
                        pdf_id=pdf.id,
                        project_id=project.id,
                        level='level2',
                        status='pending',
                        message=f"Drawing in {project.name} approved by {current_user.username} ({get_role_display(current_user.role)}) at Level 1. Requires your Level 2 approval."
                    )
                    db.session.add(notification)
                    print(f"✅ Created Level 2 notification for Role 2 user: {user.username}")
            
            # Notify admins (role='4') assigned to this project
            admin_users = User.query.filter(
                User.role == '4',  # Role 4 = Admin
                User.is_active == True
            ).all()
            
            for admin in admin_users:
                if admin.id != current_user.id and project in admin.projects:
                    notification = Notification(
                        user_id=admin.id,
                        pdf_id=pdf.id,
                        project_id=project.id,
                        level='admin',
                        status='approved',
                        message=f"[ADMIN] Drawing in {project.name} approved by {current_user.username} ({get_role_display(current_user.role)}) at Level 1."
                    )
                    db.session.add(notification)
                    print(f"✅ Created admin notification for: {admin.username}")
            
            flash("✅ PDF approved at Level 1! Notification sent to Level 2 approvers.", "success")
            
        elif level == '2':
            # ===== LEVEL 2 APPROVAL =====
            if not permissions['can_approve_level2']:
                flash("Access denied. You don't have permission to approve at Level 2.", "danger")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            if pdf.level1_status != 'approved':
                flash("This PDF must be approved at Level 1 first.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            if pdf.level2_status != 'pending':
                flash("This PDF has already been processed at Level 2.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Update PDF status
            pdf.level2_status = 'approved'
            pdf.level2_approver_id = current_user.id
            pdf.level2_approval_date = get_ist_now()
            
            # ✅ ADD THIS: Record approval in Approval table
            approval = Approval(
                generated_pdf_id=pdf.id,
                level=2,
                status='approved',
                approver_id=current_user.id,
                remarks=approval_remarks,
                created_at=get_ist_now()
            )
            db.session.add(approval)
            
            # ✅ ADD THIS: Update JunctionApproval records for ALL junctions
            junctions = project.junction_boxes
            for junction in junctions:
                # Check if junction approval already exists
                junction_approval = JunctionApproval.query.filter_by(
                    project_id=project.id,
                    generated_pdf_id=pdf.id,
                    junction_box_id=junction.id
                ).first()
                
                if not junction_approval:
                    # Create new junction approval (shouldn't happen if Level 1 was approved, but handle it)
                    junction_approval = JunctionApproval(
                        project_id=project.id,
                        generated_pdf_id=pdf.id,
                        junction_box_id=junction.id,
                        level1_status='approved',  # Assume Level 1 was approved
                        level2_status='approved',
                        level2_approver_id=current_user.id,
                        level2_approval_date=get_ist_now(),
                        level3_status='pending',
                        created_at=get_ist_now()
                    )
                    db.session.add(junction_approval)
                else:
                    # Update existing junction approval
                    junction_approval.level2_status = 'approved'
                    junction_approval.level2_approver_id = current_user.id
                    junction_approval.level2_approval_date = get_ist_now()
            
            # Update existing level2 notifications to 'approved'
            Notification.query.filter_by(
                pdf_id=pdf.id,
                level='level2',
                status='pending'
            ).update({
                'status': 'approved',
                'is_read': True,
                'updated_at': get_ist_now()
            }, synchronize_session=False)
            
            # Create notification for Level 3 users (role='3') assigned to this project
            level3_users = User.query.filter(
                User.role == '3',  # Role 3 = Level 3 Approvers
                User.is_active == True
            ).all()
            
            for user in level3_users:
                if project in user.projects:
                    notification = Notification(
                        user_id=user.id,
                        pdf_id=pdf.id,
                        project_id=project.id,
                        level='level3',
                        status='pending',
                        message=f"Drawing in {project.name} approved by {current_user.username} ({get_role_display(current_user.role)}) at Level 2. Requires your Level 3 approval."
                    )
                    db.session.add(notification)
                    print(f"✅ Created Level 3 notification for Role 3 user: {user.username}")
            
            # Notify admins (role='4')
            admin_users = User.query.filter(
                User.role == '4',
                User.is_active == True
            ).all()
            
            for admin in admin_users:
                if admin.id != current_user.id and project in admin.projects:
                    notification = Notification(
                        user_id=admin.id,
                        pdf_id=pdf.id,
                        project_id=project.id,
                        level='admin',
                        status='approved',
                        message=f"[ADMIN] Drawing in {project.name} approved by {current_user.username} ({get_role_display(current_user.role)}) at Level 2."
                    )
                    db.session.add(notification)
                    print(f"✅ Created admin notification for: {admin.username}")
            
            flash("✅ PDF approved at Level 2! Notification sent to Level 3 approvers.", "success")
            
        elif level == '3':
            # ===== LEVEL 3 APPROVAL (FINAL) =====
            if not permissions['can_approve_level3']:
                flash("Access denied. You don't have permission to approve at Level 3.", "danger")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            if pdf.level2_status != 'approved':
                flash("This PDF must be approved at Level 2 first.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            if pdf.level3_status != 'pending':
                flash("This PDF has already been processed at Level 3.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Update PDF status
            pdf.level3_status = 'approved'
            pdf.level3_approver_id = current_user.id
            pdf.level3_approval_date = get_ist_now()
            
            # ✅ ADD THIS: Record approval in Approval table
            approval = Approval(
                generated_pdf_id=pdf.id,
                level=3,
                status='approved',
                approver_id=current_user.id,
                remarks=approval_remarks,
                created_at=get_ist_now()
            )
            db.session.add(approval)
            
            # ✅ ADD THIS: Update JunctionApproval records for ALL junctions
            junctions = project.junction_boxes
            for junction in junctions:
                # Check if junction approval already exists
                junction_approval = JunctionApproval.query.filter_by(
                    project_id=project.id,
                    generated_pdf_id=pdf.id,
                    junction_box_id=junction.id
                ).first()
                
                if not junction_approval:
                    # Create new junction approval (shouldn't happen if Level 1 and 2 were approved, but handle it)
                    junction_approval = JunctionApproval(
                        project_id=project.id,
                        generated_pdf_id=pdf.id,
                        junction_box_id=junction.id,
                        level1_status='approved',  # Assume previous levels were approved
                        level2_status='approved',
                        level3_status='approved',
                        level3_approver_id=current_user.id,
                        level3_approval_date=get_ist_now(),
                        created_at=get_ist_now()
                    )
                    db.session.add(junction_approval)
                else:
                    # Update existing junction approval
                    junction_approval.level3_status = 'approved'
                    junction_approval.level3_approver_id = current_user.id
                    junction_approval.level3_approval_date = get_ist_now()
                
                # Update junction box status to "Completed" (status=3)
                junction.status = 3
            
            # Update existing level3 notifications to 'approved'
            Notification.query.filter_by(
                pdf_id=pdf.id,
                level='level3',
                status='pending'
            ).update({
                'status': 'approved',
                'is_read': True,
                'updated_at': get_ist_now()
            }, synchronize_session=False)
            
            # Notify ALL users assigned to this project (PDF is FULLY APPROVED)
            all_project_users = User.query.filter(
                User.is_active == True
            ).all()
            
            for user in all_project_users:
                if project in user.projects and user.id != current_user.id:
                    notification = Notification(
                        user_id=user.id,
                        pdf_id=pdf.id,
                        project_id=project.id,
                        level='final',
                        status='approved',
                        message=f"🎉 Drawing in {project.name} has been FULLY APPROVED by {current_user.username} ({get_role_display(current_user.role)}) at Level 3!"
                    )
                    db.session.add(notification)
                    print(f"✅ FULLY APPROVED notification sent to {user.username}")
            
            flash("✅ PDF FULLY APPROVED at Level 3! All users notified.", "success")
        
        else:
            flash("Invalid approval level.", "danger")
            return redirect(request.referrer or url_for('main.approval_tracking'))
        
        db.session.commit()
        return redirect(request.referrer or url_for('main.approval_tracking'))
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error approving PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f"Error approving PDF: {str(e)}", "danger")
        return redirect(request.referrer or url_for('main.approval_tracking'))


@bp.route("/reject_pdf/<int:pdf_id>/<level>", methods=["POST"])
@login_required
def reject_pdf(pdf_id, level):
    """Reject PDF at specified level"""
    try:
        pdf = GeneratedPDF.query.get_or_404(pdf_id)
        project = Project.query.get_or_404(pdf.project_id)
        rejection_reason = request.form.get('rejection_reason', 'No reason provided')
        
        # Get user permissions based on role
        permissions = get_user_permissions(current_user)
        
        # Check if user has permission for this level
        level_key = f'can_approve_level{level}'
        if not permissions.get(level_key):
            flash(f"Access denied. You don't have permission to reject at Level {level}.", "danger")
            return redirect(request.referrer or url_for('main.approval_tracking'))
        
        # Update PDF status
        if level == '1':
            pdf.level1_status = 'rejected'
            pdf.level1_approver_id = current_user.id
            pdf.level1_approval_date = get_ist_now()
        elif level == '2':
            pdf.level2_status = 'rejected'
            pdf.level2_approver_id = current_user.id
            pdf.level2_approval_date = get_ist_now()
        elif level == '3':
            pdf.level3_status = 'rejected'
            pdf.level3_approver_id = current_user.id
            pdf.level3_approval_date = get_ist_now()
        
        pdf.rejection_reason = rejection_reason
        
        # ✅ ADD THIS: Record rejection in Approval table
        approval = Approval(
            generated_pdf_id=pdf.id,
            level=int(level),
            status='rejected',
            approver_id=current_user.id,
            remarks=rejection_reason,
            created_at=get_ist_now()
        )
        db.session.add(approval)
        
        # ✅ ADD THIS: Create/Update JunctionApproval records for ALL junctions
        junctions = project.junction_boxes
        for junction in junctions:
            # Check if junction approval already exists
            junction_approval = JunctionApproval.query.filter_by(
                project_id=project.id,
                generated_pdf_id=pdf.id,
                junction_box_id=junction.id
            ).first()
            
            if not junction_approval:
                # Create new junction approval
                junction_approval = JunctionApproval(
                    project_id=project.id,
                    generated_pdf_id=pdf.id,
                    junction_box_id=junction.id,
                    created_at=get_ist_now()
                )
                db.session.add(junction_approval)
            
            # Update the specific level
            if level == '1':
                junction_approval.level1_status = 'rejected'
                junction_approval.level1_approver_id = current_user.id
                junction_approval.level1_approval_date = get_ist_now()
            elif level == '2':
                junction_approval.level2_status = 'rejected'
                junction_approval.level2_approver_id = current_user.id
                junction_approval.level2_approval_date = get_ist_now()
            elif level == '3':
                junction_approval.level3_status = 'rejected'
                junction_approval.level3_approver_id = current_user.id
                junction_approval.level3_approval_date = get_ist_now()
            
            junction_approval.rejection_reason = rejection_reason
            
            # Update junction box status to "No Drawing Data" (status=1)
            junction.status = 1
        
        # Update current level notifications to 'rejected'
        Notification.query.filter_by(
            pdf_id=pdf.id,
            level=f'level{level}',
            status='pending'
        ).update({
            'status': 'rejected',
            'is_read': True,
            'updated_at': get_ist_now()
        }, synchronize_session=False)
        
        # Notify creators (role='1') about rejection
        creator_users = User.query.filter(
            User.role == '1',
            User.is_active == True
        ).all()
        
        for user in creator_users:
            if project in user.projects:
                notification = Notification(
                    user_id=user.id,
                    pdf_id=pdf.id,
                    project_id=project.id,
                    level=f'level{level}',
                    status='rejected',
                    message=f"Drawing in {project.name} was REJECTED at Level {level} by {current_user.username} ({get_role_display(current_user.role)}). Reason: {rejection_reason}"
                )
                db.session.add(notification)
                print(f"✅ Rejection notification sent to creator: {user.username}")
        
        # Notify admins
        admin_users = User.query.filter(
            User.role == '4',
            User.is_active == True
        ).all()
        
        for admin in admin_users:
            if admin.id != current_user.id and project in admin.projects:
                notification = Notification(
                    user_id=admin.id,
                    pdf_id=pdf.id,
                    project_id=project.id,
                    level='admin',
                    status='rejected',
                    message=f"[ADMIN] Drawing in {project.name} REJECTED at Level {level} by {current_user.username} ({get_role_display(current_user.role)})."
                )
                db.session.add(notification)
        
        db.session.commit()
        flash(f"✅ PDF rejected at Level {level}. Notification sent to creators.", "success")
        return redirect(request.referrer or url_for('main.approval_tracking'))
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error rejecting PDF: {str(e)}")
        flash(f"Error rejecting PDF: {str(e)}", "danger")
        return redirect(request.referrer or url_for('main.approval_tracking'))


@bp.route('/', methods=['GET', 'POST'])
@login_required
def approval_tracking():
    """Show approval tracking for ALL users but filter data by their assigned projects"""
    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    rows_per_page = request.args.get('rows_per_page', 10, type=int)
    
    # Get user permissions
    permissions = get_user_permissions(current_user)
    
    # Determine accessible projects based on user role
    if permissions['can_see_all']:  # Admin (role 4)
        # Admin can see all projects
        all_projects = Project.query.options(joinedload(Project.junction_boxes)).all()
        accessible_project_ids = [p.id for p in all_projects]
    else:
        # For other roles, show only their assigned projects
        if current_user.projects:
            accessible_project_ids = [p.id for p in current_user.projects]
            all_projects = Project.query.filter(Project.id.in_(accessible_project_ids)).options(joinedload(Project.junction_boxes)).all()
        else:
            all_projects = []
            accessible_project_ids = []
   
    # Get filter parameters
    project_id = request.args.get('project_id', 'all')
    junction_id = request.args.get('junction_id', 'all')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    approval_status = request.args.get('approval_status', 'all')
    # UPDATED: Always show stations without drawings by default
    show_without_drawings = request.args.get('show_without_drawings', 'on')
    
    # For role 0 (viewer), only show fully approved drawings
    if str(current_user.role) == '0':
        approval_status = 'approved'
    
    # Get latest_only parameter
    latest_only = request.args.get('latest_only', 'true')
    
    if latest_only == 'true':
        latest_only = 'true'
    else:
        latest_only = 'false'
    
    if show_without_drawings == 'on':

        pdf_query = db.session.query(Project, GeneratedPDF)\
            .outerjoin(
                GeneratedPDF,
                Project.id == GeneratedPDF.project_id
            ).options(joinedload(Project.junction_boxes))

    else:

        pdf_query = db.session.query(Project, GeneratedPDF)\
            .join(
                GeneratedPDF,
                Project.id == GeneratedPDF.project_id
            )\
            .options(joinedload(Project.junction_boxes))
      
    latest_subquery = db.session.query(
        GeneratedPDF.project_id,
        func.max(GeneratedPDF.created_at).label("max_created_at")
    ).group_by(GeneratedPDF.project_id).subquery()

    pdf_query = db.session.query(Project, GeneratedPDF)\
        .outerjoin(
            latest_subquery,
            Project.id == latest_subquery.c.project_id
        )\
        .outerjoin(
            GeneratedPDF,
            and_(
                GeneratedPDF.project_id == latest_subquery.c.project_id,
                GeneratedPDF.created_at == latest_subquery.c.max_created_at
            )
        )  
    # Base query for PDFs
    #pdf_query = GeneratedPDF.query.options(
    #    joinedload(GeneratedPDF.project).joinedload(Project.junction_boxes)
    #)
   
    # Apply accessible project filter for non-admin users
    if not permissions['can_see_all'] and accessible_project_ids:
        pdf_query = pdf_query.filter(GeneratedPDF.project_id.in_(accessible_project_ids))
   
    # For role 0 (viewer), only show fully approved (level3 approved) drawings
    if str(current_user.role) == '0':
        pdf_query = pdf_query.filter(GeneratedPDF.level3_status == 'approved')
   
    # Apply project filter if selected
    if project_id and project_id != 'all':
        try:
            project_id_int = int(project_id)
            # Check if user has access to this project
            if permissions['can_see_all'] or project_id_int in accessible_project_ids:
                pdf_query = pdf_query.filter(Project.id == project_id_int)
            else:
                flash("You don't have access to this project.", "warning")
                # Fallback to all accessible projects
                if accessible_project_ids:
                    pdf_query = pdf_query.filter(Project.id.in_(accessible_project_ids))
        except ValueError:
            pass
   
    # Date filters - IMPORTANT: Convert to IST for comparison
    if start_date:
        try:
            # Parse the date in IST timezone
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            # Make it timezone aware in IST
            start_datetime_ist = start_datetime.replace(
                hour=0, minute=0, second=0, microsecond=0,
                tzinfo=IST
            )
            # Convert to UTC for database comparison (assuming database stores UTC)
            start_datetime_utc = start_datetime_ist.astimezone(timezone.utc)
            pdf_query = pdf_query.filter(GeneratedPDF.created_at >= start_datetime_utc)
        except ValueError:
            pass
   
    if end_date:
        try:
            # Parse the date in IST timezone
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            # Make it timezone aware in IST (end of day)
            end_datetime_ist = end_datetime.replace(
                hour=23, minute=59, second=59, microsecond=999999,
                tzinfo=IST
            )
            # Convert to UTC for database comparison
            end_datetime_utc = end_datetime_ist.astimezone(timezone.utc)
            pdf_query = pdf_query.filter(GeneratedPDF.created_at <= end_datetime_utc)
        except ValueError:
            pass
   
    # Approval status filter
    if approval_status != 'all':
        if approval_status == 'approved':
            pdf_query = pdf_query.filter(GeneratedPDF.level3_status == 'approved')
        elif approval_status == 'rejected':
            pdf_query = pdf_query.filter(or_(
                GeneratedPDF.level1_status == 'rejected',
                GeneratedPDF.level2_status == 'rejected',
                GeneratedPDF.level3_status == 'rejected'
            ))
        elif approval_status == 'pending':
            pdf_query = pdf_query.filter(
                GeneratedPDF.level3_status != 'approved',
                or_(
                    GeneratedPDF.level1_status == 'pending',
                    GeneratedPDF.level2_status == 'pending',
                    GeneratedPDF.level3_status == 'pending',
                    GeneratedPDF.level1_status.is_(None),
                    GeneratedPDF.level2_status.is_(None),
                    GeneratedPDF.level3_status.is_(None)
                )
            )
   
    # Order by most recent first
    # pdf_query = pdf_query.order_by(GeneratedPDF.created_at.desc())
   
    # Get all PDFs with filter applied
    #pdfs = pdf_query.all()cl
    #from sqlalchemy import and_, or_

    if junction_id and junction_id != 'all':
        # Apply filter
        '''
        pdf_query = pdf_query.filter(
            or_(
                # Case 1: records WITH PDF and matching junction
                and_(
                    GeneratedPDF.id.isnot(None),
                    Project.junction_boxes.any(
                        JunctionBox.junction_id == junction_id
                    )
                ),
                # Case 2: records WITHOUT PDF, only include if showing projects without drawings
                and_(
                    GeneratedPDF.id.is_(None),
                    show_without_drawings == 'on'
                )
            ))
        '''
    '''    
    if latest_only == 'true':    
        latest_subquery = db.session.query(
            GeneratedPDF.project_id,
            func.max(GeneratedPDF.version).label("max_version")
        ).group_by(GeneratedPDF.project_id).subquery()
        
        pdf_query = pdf_query.outerjoin(
            latest_subquery,
            (GeneratedPDF.project_id == latest_subquery.c.project_id) &
            (GeneratedPDF.version == latest_subquery.c.max_version)
        )
    '''
    
    
    pdf_query = pdf_query.order_by(nullslast(desc(GeneratedPDF.created_at)))
    print(pdf_query)
    pagination_obj = pdf_query.paginate(page=page, per_page=rows_per_page, error_out=False)
    pdfs = pagination_obj.items
    total_records = pagination_obj.total
    total_pages = pagination_obj.pages
    all_records=[]
    for pp, pdff in pdfs:
        if pdff is  None:
           pdff=create_dummy_pdf(pp)
        all_records.append(pdff)
    
   
    
    total_records = len(all_records)
    paginated_records=all_records
    start_idx = (page - 1) * rows_per_page + 1
    end_idx = min(page * rows_per_page, pagination_obj.total)
    
    print(all_records)
    # Calculate statistics
    approved_pdfs = len([p for p in all_records if hasattr(p, 'level3_status') and p.level3_status == 'approved'])
    rejected_pdfs = len([p for p in all_records if hasattr(p, 'level1_status') and 
                        (p.level1_status == 'rejected' or p.level2_status == 'rejected' or p.level3_status == 'rejected')])
    pending_pdfs = len([p for p in all_records if hasattr(p, 'level1_status') and 
                       p.level1_status == 'pending'])
    projects_no_drawings = 0 #len([p for p in all_records if p.pdf_filename is None])
    
    # =========================================================================
    # 🔥 AUTO‑CREATE PENDING JUNCTIONAPPROVAL RECORDS FOR ALL DISPLAYED PDFs
    # =========================================================================
    #for record in paginated_records:
    #    # Only real PDFs (with an id) have junction approvals
    #    if hasattr(record, 'id') and record.id is not None:
    #        ensure_junction_approvals_for_pdf(record)
    # =========================================================================
   
    return render_template('approval_tracking.html',
                         pdfs=paginated_records,
                         all_projects=all_projects,
                         current_filters={
                             'project_id': project_id,
                             'junction_id': junction_id,
                             'start_date': start_date,
                             'end_date': end_date,
                             'approval_status': approval_status,
                             'show_without_drawings': show_without_drawings,
                             'latest_only': latest_only
                         },
                         stats={
                             'total': total_records,
                             'approved': approved_pdfs,
                             'rejected': rejected_pdfs,
                             'pending': pending_pdfs,
                             'no_drawings': projects_no_drawings
                         },
                         permissions=permissions,
                         pagination={
                             "page": pagination_obj.page,
                             "rows_per_page": rows_per_page,
                             "total_pages": pagination_obj.pages,
                             "total_records": pagination_obj.total,
                             "has_next": pagination_obj.has_next,
                             "has_prev": pagination_obj.has_prev,
                             "start_idx" : start_idx,
                             "end_idx"   : end_idx
                         })

def create_dummy_pdf(project=None):
    dummy_pdf = type('DummyPDF', (), {})()

    dummy_pdf.id = None
    dummy_pdf.project_id = project.id if project else None
    dummy_pdf.project = project
    dummy_pdf.pdf_filename = None
    dummy_pdf.created_at = None
    dummy_pdf.created_at_ist = None

    dummy_pdf.level1_status = None
    dummy_pdf.level2_status = None
    dummy_pdf.level3_status = None
    dummy_pdf.level1_approver = None
    dummy_pdf.level2_approver = None
    dummy_pdf.level3_approver = None
    dummy_pdf.level1_approval_date = None
    dummy_pdf.level1_approval_date_ist = None
    dummy_pdf.level2_approval_date = None
    dummy_pdf.level2_approval_date_ist = None
    dummy_pdf.level3_approval_date = None
    dummy_pdf.level3_approval_date_ist = None

    dummy_pdf.rejection_reason = None
    dummy_pdf.remarks = None
    dummy_pdf.version = "0"
    dummy_pdf.parsed_junction_data = []
    dummy_pdf.junction_data = None
    dummy_pdf.checksum_md5 = None
    dummy_pdf.file_size = None
    dummy_pdf.created_by = None
    dummy_pdf.xlsx_filename = None
    dummy_pdf.checksum_algo = None
    dummy_pdf.metadata_checksum = None
    dummy_pdf.metadata_data = None
    dummy_pdf.initial_size_bytes = None
    dummy_pdf.final_size_bytes = None
    dummy_pdf.metadata_ts_ist = None
    dummy_pdf.station_code = None
    dummy_pdf.source_pdf_name = None
    dummy_pdf.full_file_md5 = None

    return dummy_pdf
'''
def approval_tracking():
    """Show approval tracking for ALL users but filter data by their assigned projects"""
    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    rows_per_page = request.args.get('rows_per_page', 10, type=int)
    
    # Get user permissions
    permissions = get_user_permissions(current_user)
    
    # Determine accessible projects based on user role
    if permissions['can_see_all']:  # Admin (role 4)
        # Admin can see all projects
        all_projects = Project.query.options(joinedload(Project.junction_boxes)).all()
        accessible_project_ids = [p.id for p in all_projects]
    else:
        # For other roles, show only their assigned projects
        if current_user.projects:
            accessible_project_ids = [p.id for p in current_user.projects]
            all_projects = Project.query.filter(Project.id.in_(accessible_project_ids)).options(joinedload(Project.junction_boxes)).all()
        else:
            all_projects = []
            accessible_project_ids = []
   
    # Get filter parameters
    project_id = request.args.get('project_id', 'all')
    junction_id = request.args.get('junction_id', 'all')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    approval_status = request.args.get('approval_status', 'all')
    # UPDATED: Always show stations without drawings by default
    show_without_drawings = request.args.get('show_without_drawings', 'on')
    
    # For role 0 (viewer), only show fully approved drawings
    if str(current_user.role) == '0':
        approval_status = 'approved'
    
    # Get latest_only parameter
    latest_only = request.args.get('latest_only', 'true')
    
    if latest_only == 'true':
        latest_only = 'true'
    else:
        latest_only = 'false'
   
    # Base query for PDFs
    pdf_query = GeneratedPDF.query.options(
        joinedload(GeneratedPDF.project).joinedload(Project.junction_boxes)
    )
   
    # Apply accessible project filter for non-admin users
    if not permissions['can_see_all'] and accessible_project_ids:
        pdf_query = pdf_query.filter(GeneratedPDF.project_id.in_(accessible_project_ids))
   
    # For role 0 (viewer), only show fully approved (level3 approved) drawings
    if str(current_user.role) == '0':
        pdf_query = pdf_query.filter(GeneratedPDF.level3_status == 'approved')
   
    # Apply project filter if selected
    if project_id and project_id != 'all':
        try:
            project_id_int = int(project_id)
            # Check if user has access to this project
            if permissions['can_see_all'] or project_id_int in accessible_project_ids:
                pdf_query = pdf_query.filter(GeneratedPDF.project_id == project_id_int)
            else:
                flash("You don't have access to this project.", "warning")
                # Fallback to all accessible projects
                if accessible_project_ids:
                    pdf_query = pdf_query.filter(GeneratedPDF.project_id.in_(accessible_project_ids))
        except ValueError:
            pass
   
    # Date filters - IMPORTANT: Convert to IST for comparison
    if start_date:
        try:
            # Parse the date in IST timezone
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            # Make it timezone aware in IST
            start_datetime_ist = start_datetime.replace(
                hour=0, minute=0, second=0, microsecond=0,
                tzinfo=IST
            )
            # Convert to UTC for database comparison (assuming database stores UTC)
            start_datetime_utc = start_datetime_ist.astimezone(timezone.utc)
            pdf_query = pdf_query.filter(GeneratedPDF.created_at >= start_datetime_utc)
        except ValueError:
            pass
   
    if end_date:
        try:
            # Parse the date in IST timezone
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            # Make it timezone aware in IST (end of day)
            end_datetime_ist = end_datetime.replace(
                hour=23, minute=59, second=59, microsecond=999999,
                tzinfo=IST
            )
            # Convert to UTC for database comparison
            end_datetime_utc = end_datetime_ist.astimezone(timezone.utc)
            pdf_query = pdf_query.filter(GeneratedPDF.created_at <= end_datetime_utc)
        except ValueError:
            pass
   
    # Approval status filter
    if approval_status != 'all':
        if approval_status == 'approved':
            pdf_query = pdf_query.filter(GeneratedPDF.level3_status == 'approved')
        elif approval_status == 'rejected':
            pdf_query = pdf_query.filter(or_(
                GeneratedPDF.level1_status == 'rejected',
                GeneratedPDF.level2_status == 'rejected',
                GeneratedPDF.level3_status == 'rejected'
            ))
        elif approval_status == 'pending':
            pdf_query = pdf_query.filter(
                GeneratedPDF.level3_status != 'approved',
                or_(
                    GeneratedPDF.level1_status == 'pending',
                    GeneratedPDF.level2_status == 'pending',
                    GeneratedPDF.level3_status == 'pending',
                    GeneratedPDF.level1_status.is_(None),
                    GeneratedPDF.level2_status.is_(None),
                    GeneratedPDF.level3_status.is_(None)
                )
            )
   
    # Order by most recent first
    pdf_query = pdf_query.order_by(GeneratedPDF.created_at.desc())
   
    # Get all PDFs with filter applied
    pdfs = pdf_query.all()
    
    # Convert dates to IST for display
    for pdf in pdfs:
        # Convert created_at to IST
        if pdf.created_at:
            if pdf.created_at.tzinfo is None:
                # Assuming stored as UTC, convert to IST
                pdf.created_at_ist = pdf.created_at.replace(tzinfo=timezone.utc).astimezone(IST)
            else:
                pdf.created_at_ist = pdf.created_at.astimezone(IST)
        else:
            pdf.created_at_ist = None
            
        # Convert approval dates to IST
        for date_field in ['level1_approval_date', 'level2_approval_date', 'level3_approval_date']:
            date_value = getattr(pdf, date_field, None)
            if date_value:
                if date_value.tzinfo is None:
                    setattr(pdf, f"{date_field}_ist", date_value.replace(tzinfo=timezone.utc).astimezone(IST))
                else:
                    setattr(pdf, f"{date_field}_ist", date_value.astimezone(IST))
            else:
                setattr(pdf, f"{date_field}_ist", None)
   
    # NEW: Get projects without drawings
    projects_without_drawings = []
    
    # UPDATED: Always get projects without drawings (default is 'on')
    if show_without_drawings == 'on':
        # Apply project filter to projects without drawings
        if project_id and project_id != 'all':
            try:
                project_id_int = int(project_id)
                if permissions['can_see_all'] or project_id_int in accessible_project_ids:
                    # Get this specific project
                    filtered_projects = Project.query.filter(
                        Project.id == project_id_int
                    ).all()
                else:
                    filtered_projects = []
            except ValueError:
                filtered_projects = []
        else:
            # Get all accessible projects
            filtered_projects = all_projects
        
        # Get ALL projects that have PDFs
        all_projects_with_pdfs = set([pdf.project_id for pdf in GeneratedPDF.query.all()])
        
        for project in filtered_projects:
            if project.id not in all_projects_with_pdfs:
                # Create a dummy PDF-like object for projects without drawings
                dummy_pdf = type('DummyPDF', (), {})()
                dummy_pdf.id = None
                dummy_pdf.project_id = project.id
                dummy_pdf.project = project
                dummy_pdf.pdf_filename = None
                
                # Convert project creation date to IST
                if project.created_date:
                    if project.created_date.tzinfo is None:
                        # Store as naive datetime (for sorting compatibility)
                        dummy_pdf.created_at = project.created_date
                        # Create IST version for display
                        dummy_pdf.created_at_ist = project.created_date.replace(tzinfo=timezone.utc).astimezone(IST) if project.created_date else None
                    else:
                        # Convert to naive UTC for sorting
                        dummy_pdf.created_at = project.created_date.astimezone(timezone.utc).replace(tzinfo=None)
                        # Create IST version for display
                        dummy_pdf.created_at_ist = project.created_date.astimezone(IST)
                else:
                    dummy_pdf.created_at = None
                    dummy_pdf.created_at_ist = None
                    
                dummy_pdf.level1_status = None
                dummy_pdf.level2_status = None
                dummy_pdf.level3_status = None
                dummy_pdf.level1_approver = None
                dummy_pdf.level2_approver = None
                dummy_pdf.level3_approver = None
                dummy_pdf.level1_approval_date = None
                dummy_pdf.level1_approval_date_ist = None
                dummy_pdf.level2_approval_date = None
                dummy_pdf.level2_approval_date_ist = None
                dummy_pdf.level3_approval_date = None
                dummy_pdf.level3_approval_date_ist = None
                dummy_pdf.rejection_reason = None
                dummy_pdf.remarks = None
                dummy_pdf.version = "0"
                dummy_pdf.parsed_junction_data = []
                dummy_pdf.junction_data = None
                dummy_pdf.checksum_md5 = None
                dummy_pdf.file_size = None
                dummy_pdf.created_by = None
                dummy_pdf.xlsx_filename = None
                dummy_pdf.checksum_algo = None
                dummy_pdf.metadata_checksum = None
                dummy_pdf.metadata_data = None
                dummy_pdf.initial_size_bytes = None
                dummy_pdf.final_size_bytes = None
                dummy_pdf.metadata_ts_ist = None
                dummy_pdf.station_code = None
                dummy_pdf.source_pdf_name = None
                dummy_pdf.full_file_md5 = None
                
                projects_without_drawings.append(dummy_pdf)
    
    # UPDATED: Combine records based on show_without_drawings filter
    # Always show both PDFs and projects without drawings when show_without_drawings is 'on'
    # (which is the new default)
    if show_without_drawings == 'on':
        # Show BOTH PDFs and projects without drawings
        all_records = list(pdfs) + projects_without_drawings
    else:
        # Show ONLY PDFs with drawings (for backward compatibility)
        all_records = list(pdfs)
    
    # Filter by junction_id (only for records with PDFs)
    if junction_id and junction_id != 'all':
        # Split unique key: ['1', 'AC BOX NO.K4 (F)']
        junction_parts = junction_id.split('|', 1)
        target_id = junction_parts[0] if len(junction_parts) > 0 else ''
        target_name = junction_parts[1] if len(junction_parts) > 1 else ''
        
        filtered_records = []
        import json
        for record in all_records:
            # Skip projects without drawings when filtering by junction
            if record.pdf_filename is None:
                # Projects without drawings don't have junction data
                # Keep them in the list only if we're showing all stations
                if show_without_drawings == 'on':
                    # Parse junctions from project data
                    pdf_junctions = []
                    if record.project.junction_boxes:
                        pdf_junctions = [jb for jb in record.project.junction_boxes 
                                       if jb.junction_id == target_id and jb.junction_name == target_name]
                    
                    if pdf_junctions:
                        # Re-parse for template
                        record.parsed_junction_data = []
                        for jb in record.project.junction_boxes:
                            if jb.junction_name:
                                record.parsed_junction_data.append({
                                    'junction_id': jb.junction_id,
                                    'junction_name': jb.junction_name
                                })
                        filtered_records.append(record)
                    else:
                        # Still parse (won't be rendered)
                        record.parsed_junction_data = []
                continue
            
            match_found = False
            
            # Parse historical from PDF
            pdf_junctions = []
            if hasattr(record, 'junction_data') and record.junction_data:
                try:
                    parsed_data = json.loads(record.junction_data)
                    # Match if ID and name both match
                    pdf_junctions = [jb for jb in parsed_data 
                                   if jb.get('junction_id') == target_id and jb.get('junction_name') == target_name]
                except:
                    pass
            
            # If no historical match, check current project junctions
            if not pdf_junctions and record.project.junction_boxes:
                pdf_junctions = [jb for jb in record.project.junction_boxes 
                               if jb.junction_id == target_id and jb.junction_name == target_name]
            
            if pdf_junctions:
                match_found = True
            
            if match_found:
                # Re-parse for template
                if hasattr(record, 'junction_data') and record.junction_data:
                    try:
                        record.parsed_junction_data = json.loads(record.junction_data)
                    except:
                        record.parsed_junction_data = []
                else:
                    record.parsed_junction_data = []
                filtered_records.append(record)
            else:
                # Still parse (won't be rendered)
                if hasattr(record, 'junction_data') and record.junction_data:
                    try:
                        record.parsed_junction_data = json.loads(record.junction_data)
                    except:
                        record.parsed_junction_data = []
                else:
                    record.parsed_junction_data = []
        
        all_records = filtered_records
    else:
        # No junction filter, parse junction data for all records
        import json
        for record in all_records:
            if record.pdf_filename is None:
                # For projects without drawings, use project junction data
                if hasattr(record, 'project') and record.project.junction_boxes:
                    record.parsed_junction_data = []
                    for jb in record.project.junction_boxes:
                        if jb.junction_name:
                            record.parsed_junction_data.append({
                                'junction_id': jb.junction_id,
                                'junction_name': jb.junction_name
                            })
                else:
                    record.parsed_junction_data = []
            elif hasattr(record, 'junction_data') and record.junction_data:
                try:
                    record.parsed_junction_data = json.loads(record.junction_data)
                except:
                    record.parsed_junction_data = []
            else:
                record.parsed_junction_data = []
        
    # Filter for latest versions only if checkbox is checked (only applies to PDFs)
    if latest_only == 'true':
        # Group by project and get the latest version for each project
        latest_by_project = {}
        
        for record in all_records:
            project_key = str(record.project_id)
            
            # Skip projects without drawings for version filtering
            if record.pdf_filename is None:
                # Always include projects without drawings
                if project_key not in latest_by_project:
                    latest_by_project[project_key] = record
                else:
                    # If we already have a PDF for this project, keep the PDF
                    existing_record = latest_by_project[project_key]
                    if existing_record.pdf_filename is None:
                        # Replace project without drawing with PDF if we find one
                        latest_by_project[project_key] = record
                continue
            
            # For PDFs with drawings
            if project_key not in latest_by_project:
                latest_by_project[project_key] = record
            else:
                # Check if existing record is a PDF or a project without drawing
                existing_record = latest_by_project[project_key]
                
                if existing_record.pdf_filename is None:
                    # Existing is project without drawing, replace with PDF
                    latest_by_project[project_key] = record
                else:
                    # Both are PDFs, compare versions
                    try:
                        current_ver_num = int(str(existing_record.version).replace('v', '').replace('V', ''))
                        new_ver_num = int(str(record.version).replace('v', '').replace('V', ''))
                        
                        if new_ver_num > current_ver_num:
                            latest_by_project[project_key] = record
                    except ValueError:
                        # Fallback: string comparison if not numeric
                        if str(record.version) > str(existing_record.version):
                            latest_by_project[project_key] = record
        
        # Keep only the latest version for each project
        all_records = list(latest_by_project.values())
    
    # Sort by created date (descending) - FIXED: Handle timezone comparison
    from datetime import datetime as dt
    from datetime import timezone as tz
    
    # Create a timezone-aware datetime.min for comparison
    datetime_min_utc = dt.min.replace(tzinfo=tz.utc)
    
    all_records.sort(key=lambda x: x.created_at.replace(tzinfo=tz.utc) 
                     if x.created_at and x.created_at.tzinfo is None 
                     else (x.created_at if x.created_at else datetime_min_utc), 
                     reverse=True)
    
    # Apply pagination
    total_records = len(all_records)
    
    # Calculate pagination
    if rows_per_page > 0:
        start_idx = (page - 1) * rows_per_page
        end_idx = min(start_idx + rows_per_page, total_records)
        paginated_records = all_records[start_idx:end_idx]
    else:
        # Show all records if rows_per_page is 0 or negative
        paginated_records = all_records
        start_idx = 0
        end_idx = total_records
    
    # Calculate total pages
    if rows_per_page > 0:
        total_pages = (total_records + rows_per_page - 1) // rows_per_page
    else:
        total_pages = 1
    
    # Calculate statistics
    approved_pdfs = len([p for p in all_records if hasattr(p, 'level3_status') and p.level3_status == 'approved'])
    rejected_pdfs = len([p for p in all_records if hasattr(p, 'level1_status') and 
                        (p.level1_status == 'rejected' or p.level2_status == 'rejected' or p.level3_status == 'rejected')])
    pending_pdfs = len([p for p in all_records if hasattr(p, 'level1_status') and 
                       p.level1_status == 'pending'])
    projects_no_drawings = len([p for p in all_records if p.pdf_filename is None])
    
    # Generate pagination numbers
    pagination_numbers = []
    if total_pages > 0:
        # Always show first page
        pagination_numbers.append(1)
        
        # Calculate range around current page
        start_page = max(2, page - 2)
        end_page = min(total_pages - 1, page + 2)
        
        # Add ellipsis after first page if needed
        if start_page > 2:
            pagination_numbers.append('...')
        
        # Add middle pages
        for i in range(start_page, end_page + 1):
            pagination_numbers.append(i)
        
        # Add ellipsis before last page if needed
        if end_page < total_pages - 1:
            pagination_numbers.append('...')
        
        # Always show last page if there's more than 1 page
        if total_pages > 1:
            pagination_numbers.append(total_pages)

    # =========================================================================
    # 🔥 AUTO‑CREATE PENDING JUNCTIONAPPROVAL RECORDS FOR ALL DISPLAYED PDFs
    # =========================================================================
    for record in paginated_records:
        # Only real PDFs (with an id) have junction approvals
        if hasattr(record, 'id') and record.id is not None:
            ensure_junction_approvals_for_pdf(record)
    # =========================================================================

    return render_template('approval_tracking.html',
                         pdfs=paginated_records,
                         all_projects=all_projects,
                         current_filters={
                             'project_id': project_id,
                             'junction_id': junction_id,
                             'start_date': start_date,
                             'end_date': end_date,
                             'approval_status': approval_status,
                             'show_without_drawings': show_without_drawings,
                             'latest_only': latest_only
                         },
                         stats={
                             'total': total_records,
                             'approved': approved_pdfs,
                             'rejected': rejected_pdfs,
                             'pending': pending_pdfs,
                             'no_drawings': projects_no_drawings
                         },
                         permissions=permissions,
                         pagination={
                             'page': page,
                             'rows_per_page': rows_per_page,
                             'total_pages': total_pages,
                             'total_records': total_records,
                             'start_idx': start_idx + 1,
                             'end_idx': end_idx,
                             'pagination_numbers': pagination_numbers
                         })

'''

@bp.context_processor
def utility_processor():
    def build_pagination_url(page=None, rows_per_page=None):
        args = request.args.copy()
        if page:
            args['page'] = page
        if rows_per_page:
            args['rows_per_page'] = rows_per_page
        return url_for('main.approval_tracking', **args)
    return dict(build_pagination_url=build_pagination_url)

                         
@bp.route('/admin/approval-summary', methods=['GET'])
@login_required
def admin_approval_summary():
    """API endpoint for approval summary data (for charts)"""
    if current_user.role_name != '4':
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Get date range from query params (default: last 30 days)
    days = int(request.args.get('days', 30))
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    # Query for summary data
    pdfs = GeneratedPDF.query.filter(
        GeneratedPDF.created_at >= start_date,
        GeneratedPDF.created_at <= end_date
    ).all()
    
    # Group by project
    projects_data = {}
    for pdf in pdfs:
        if pdf.project.name not in projects_data:
            projects_data[pdf.project.name] = {
                'total': 0,
                'approved': 0,
                'rejected': 0,
                'pending': 0
            }
        
        projects_data[pdf.project.name]['total'] += 1
        if pdf.level3_status == 'approved':
            projects_data[pdf.project.name]['approved'] += 1
        elif pdf.level1_status == 'rejected' or pdf.level2_status == 'rejected' or pdf.level3_status == 'rejected':
            projects_data[pdf.project.name]['rejected'] += 1
        else:
            projects_data[pdf.project.name]['pending'] += 1
    
    # Convert to list for chart
    chart_data = []
    for project_name, stats in projects_data.items():
        chart_data.append({
            'project': project_name,
            'total': stats['total'],
            'approved': stats['approved'],
            'rejected': stats['rejected'],
            'pending': stats['pending']
        })
    
    # Sort by total
    chart_data.sort(key=lambda x: x['total'], reverse=True)
    
    return jsonify(chart_data)

@bp.route('/notifications')
@login_required
def notifications():
    """Get notifications for current user, excluding dismissed ones for recent list"""
    # Get ALL notifications for dropdown (remove the limit to get all)
    user_notifications = Notification.query.filter_by(
        user_id=current_user.id
    ).order_by(Notification.created_at.desc()).all()  # Removed .limit(50)
    
    # For recent list, exclude dismissed ones (remove the limit to get all non-dismissed)
    recent_notifications = Notification.query.filter_by(
        user_id=current_user.id,
        is_dismissed=False  # Only show non-dismissed in recent
    ).order_by(Notification.created_at.desc()).all()  # Removed .limit(10)
    
    all_notifications_list = []
    recent_notifications_list = []
    
    # Process all notifications for dropdown
    for notif in user_notifications:
        pdf_filename = notif.pdf.pdf_filename if notif.pdf else 'Unknown'
        project_name = notif.project.name if notif.project else 'Unknown'
        
        all_notifications_list.append({
            'id': notif.id,
            'pdf_id': notif.pdf_id,
            'project_id': notif.project_id,
            'project_name': project_name,
            'level': notif.level,
            'status': notif.status,
            'message': notif.message,
            'is_read': notif.is_read,
            'is_dismissed': notif.is_dismissed,  # Include dismissed status
            'created_at': notif.created_at.strftime('%Y-%m-%d %H:%M'),
            'pdf_filename': pdf_filename
        })
    
    # Process recent notifications (non-dismissed)
    for notif in recent_notifications:
        pdf_filename = notif.pdf.pdf_filename if notif.pdf else 'Unknown'
        project_name = notif.project.name if notif.project else 'Unknown'
        
        recent_notifications_list.append({
            'id': notif.id,
            'pdf_id': notif.pdf_id,
            'project_id': notif.project_id,
            'project_name': project_name,
            'level': notif.level,
            'status': notif.status,
            'message': notif.message,
            'is_read': notif.is_read,
            'is_dismissed': notif.is_dismissed,
            'created_at': notif.created_at.strftime('%Y-%m-%d %H:%M'),
            'pdf_filename': pdf_filename
        })
    
    return jsonify({
        'all_notifications': all_notifications_list,
        'recent_notifications': recent_notifications_list
    })

@bp.route('/mark_notification_read/<int:notification_id>', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    """Mark a notification as read"""
    notification = Notification.query.get_or_404(notification_id)
    
    # Check if notification belongs to current user
    if notification.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    notification.is_read = True
    notification.updated_at = get_ist_now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Notification marked as read'})

@bp.route('/mark_all_notifications_read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    """Mark all notifications as read for current user"""
    notifications = Notification.query.filter_by(user_id=current_user.id, is_read=False).all()
    
    for notification in notifications:
        notification.is_read = True
        notification.updated_at = get_ist_now()
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Marked {len(notifications)} notifications as read'})

@bp.route('/notification_count')
@login_required
def notification_count():
    """Get count of unread notifications for current user"""
    count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    
    return jsonify({'count': count})

def create_pdf_notifications(pdf_file, project):
    """
    Create notifications based on ROLE (0,1,2,3,4) not designation
    
    Called when a new PDF is generated
    Role 1 (Creator) users get level1 approval notifications
    """
    try:
        print(f"📢 Creating notifications for PDF {pdf_file.id} in project {project.id}")
        
        # Level 1: Find all Role 1 users (Creators) assigned to this project
        if pdf_file.level1_status == 'pending':
            # FIX: Use the projects relationship properly
            level1_users = User.query.filter(
                User.role == '1',  # Role 1 = Creator/Level1 Approver
                User.is_active == True
            ).filter(User.projects.any(id=project.id)).all()
            
            for user in level1_users:
                notification = Notification(
                    user_id=user.id,
                    pdf_id=pdf_file.id,
                    project_id=project.id,
                    level='level1',
                    status='pending',
                    message=f"NEW DRAWING in {project.name} requires your Level 1 approval"
                )
                db.session.add(notification)
                print(f"✅ Created Level 1 notification for Role 1 user: {user.username} (assigned to project)")
        
        # Notify Admins (Role 4) - Only those assigned to this project
        admin_users = User.query.filter(
            User.role == '4',
            User.is_active == True
        ).filter(User.projects.any(id=project.id)).all()
        
        for admin in admin_users:
            notification = Notification(
                user_id=admin.id,
                pdf_id=pdf_file.id,
                project_id=project.id,
                level='admin',
                status='pending',
                message=f"[ADMIN] NEW DRAWING in {project.name} submitted for approval"
            )
            db.session.add(notification)
            print(f"✅ Created admin notification for: {admin.username} (assigned to project)")
        
        db.session.commit()
        print(f"✅ All initial notifications created for PDF {pdf_file.id}")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error creating notifications: {str(e)}")
        import traceback
        traceback.print_exc()

def update_pdf_notifications_on_approval(pdf_file, project, level, new_status, approver_user):
    """
    Update notification when user approves/rejects at a level
    
    level = 'level1', 'level2', 'level3'
    new_status = 'approved' or 'rejected'
    approver_user = User object who approved/rejected
    """
    try:
        # Update notification for current level
        notif = Notification.query.filter_by(
            pdf_id=pdf_file.id,
            level=level
        ).first()
        
        if notif:
            notif.status = new_status
            notif.is_read = True
            notif.updated_at = get_ist_now()
            print(f"✅ Updated notification: PDF {pdf_file.id}, Level {level} = {new_status}")
        
        # If approved at this level, create notification for NEXT level
        if new_status == 'approved':
            if level == 'level1':
                # Create level 2 notifications - Only for Role 2 users assigned to this project
                level2_users = User.query.filter(
                    User.role == '2',
                    User.is_active == True
                ).filter(User.projects.any(id=project.id)).all()
                
                for user in level2_users:
                    new_notif = Notification(
                        user_id=user.id,
                        pdf_id=pdf_file.id,
                        project_id=project.id,
                        level='level2',
                        status='pending',
                        message=f"Drawing in {project.name} requires your Level 2 approval"
                    )
                    db.session.add(new_notif)
                    print(f"✅ Created Level 2 notification for {user.username} (assigned to project)")
            
            elif level == 'level2':
                # Create level 3 notifications - Only for Role 3 users assigned to this project
                level3_users = User.query.filter(
                    User.role == '3',
                    User.is_active == True
                ).filter(User.projects.any(id=project.id)).all()
                
                for user in level3_users:
                    new_notif = Notification(
                        user_id=user.id,
                        pdf_id=pdf_file.id,
                        project_id=project.id,
                        level='level3',
                        status='pending',
                        message=f"Drawing in {project.name} requires your Level 3 approval"
                    )
                    db.session.add(new_notif)
                    print(f"✅ Created Level 3 notification for {user.username} (assigned to project)")
            
            elif level == 'level3':
                # PDF is fully approved - notify all project users (all roles assigned to project)
                all_project_users = User.query.filter(
                    User.is_active == True
                ).filter(User.projects.any(id=project.id)).all()
                
                for user in all_project_users:
                    # Create a "FULLY APPROVED" notification
                    final_notif = Notification(
                        user_id=user.id,
                        pdf_id=pdf_file.id,
                        project_id=project.id,
                        level='final',
                        status='approved',
                        message=f"Drawing in {project.name} has been FULLY APPROVED"
                    )
                    db.session.add(final_notif)
                    print(f"✅ Created FINAL APPROVED notification for {user.username} (assigned to project)")
        
        # If rejected, notify the PDF creator (role 1) assigned to this project
        if new_status == 'rejected':
            creators = User.query.filter(
                User.role == '1',
                User.is_active == True
            ).filter(User.projects.any(id=project.id)).all()
            
            for user in creators:
                reject_notif = Notification(
                    user_id=user.id,
                    pdf_id=pdf_file.id,
                    project_id=project.id,
                    level=level,
                    status='rejected',
                    message=f"Drawing in {project.name} was REJECTED at {level} by {approver_user.username}"
                )
                db.session.add(reject_notif)
                print(f"✅ Created REJECTION notification for {user.username} (assigned to project)")
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error updating notifications: {str(e)}")
        
@bp.app_template_filter('sort_terminal_headers')
def sort_terminal_headers_filter(headers):
    """Sort terminal headers by cable_id as integer"""
    if not headers:
        return []
   
    try:
        # Sort by cable_id as integer first, then by id
        return sorted(headers, key=lambda x: (
            int(x.get('cable_id', 0)) if x.get('cable_id') and str(x.get('cable_id')).isdigit() else float('inf'),
            x.get('id', 0)
        ))
    except (ValueError, TypeError):
        # Fallback to string sorting if conversion fails
        return sorted(headers, key=lambda x: (
            str(x.get('cable_id', '')),
            x.get('id', 0)
        ))
    
@bp.route("/admin/sync_all_stations", methods=["POST"])
@login_required
def admin_sync_all_stations():
    """Admin route to sync all StationDrawing records to StationMaster"""
    if current_user.role_name != '4':
        return jsonify({"success": False, "message": "Access denied. Admin only."}), 403
    
    try:
        result = sync_all_stations_to_master()
        return jsonify({
            "success": True,
            "message": f"Sync completed: {result['success']} successful, {result['error']} errors out of {result['total']} stations",
            "data": result
        })
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500

@bp.route("/admin/station_master")
@login_required
def admin_station_master():
    """Admin view of all stations in StationMaster"""
    if current_user.role != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('main.approval_tracking'))
    
    # Get filter parameters
    project_id = request.args.get('project_id', type=int)
    station_name = request.args.get('station_name', '')
    station_code = request.args.get('station_code', '')
    
    # Build query
    query = StationMaster.query
    
    if project_id:
        query = query.filter_by(project_id=project_id)
    
    if station_name:
        query = query.filter(StationMaster.station_name.ilike(f'%{station_name}%'))
    
    if station_code:
        query = query.filter(StationMaster.station_code.ilike(f'%{station_code}%'))
    
    # Order by last updated
    stations = query.order_by(StationMaster.last_updated.desc()).all()
    
    # Get all projects for filter dropdown
    all_projects = Project.query.order_by(Project.name).all()
    
    return render_template(
        "admin_station_master.html",
        stations=stations,
        all_projects=all_projects,
        current_project_id=project_id,
        station_name_filter=station_name,
        station_code_filter=station_code
    )

@bp.route('/pdf/view/<filename>')
@login_required
def inlinepdf(filename):
    print(f"🔍 DEBUG inlinepdf: Looking for file: {filename}")
    print(f"🔍 DEBUG inlinepdf: Current working directory: {os.getcwd()}")
    
    # Security check to prevent directory traversal
    if '..' in filename or '/' in filename:
        flash('Invalid file!', 'danger')
        return redirect(url_for('main.approval_tracking'))
    
    # Get the uploads directory path
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    uploads_dir = os.path.join(base_dir, 'uploads')
    
    
    print(f"🔍 DEBUG inlinepdf: Uploads directory: {uploads_dir}")
    
    # Check if uploads directory exists
    if not os.path.exists(uploads_dir):
        print(f"❌ DEBUG inlinepdf: Uploads directory does not exist!")
        flash('Uploads directory not found!', 'danger')
        return redirect(url_for('main.approval_tracking'))
    
    # List files in uploads directory for debugging
    try:
        files_in_uploads = os.listdir(uploads_dir)
        print(f"🔍 DEBUG inlinepdf: Files in uploads directory: {files_in_uploads}")
        
        # Check for any PDF files
        pdf_files = [f for f in files_in_uploads if f.lower().endswith('.pdf')]
        print(f"🔍 DEBUG inlinepdf: PDF files found: {pdf_files}")
    except Exception as e:
        print(f"❌ DEBUG inlinepdf: Error listing uploads directory: {str(e)}")
    
    path = os.path.join(uploads_dir, filename)
    print(f"🔍 DEBUG inlinepdf: Full path to check: {path}")
    print(f"🔍 DEBUG inlinepdf: File exists: {os.path.exists(path)}")
    
    if not os.path.exists(path):
        # Check if file exists with different extensions
        import glob
        possible_files = glob.glob(os.path.join(uploads_dir, f"{os.path.splitext(filename)[0]}*"))
        print(f"🔍 DEBUG inlinepdf: Possible matching files: {possible_files}")
        
        flash(f'❌ PDF file not found: {filename}', 'danger')
        return redirect(url_for('main.approval_tracking'))
    
    print(f"✅ DEBUG inlinepdf: File found, sending...")
    # Display PDF in browser
    return send_file(path, mimetype='application/pdf', as_attachment=False)

@bp.route('/pdf/download/<filename>')
@login_required
def downloadpdf(filename):
    print(f"🔍 DEBUG downloadpdf: Looking for file: {filename}")
    
    if '..' in filename or '/' in filename:
        flash('Invalid file!', 'danger')
        return redirect(url_for('main.approval_tracking'))
    
    # Get the uploads directory path
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    uploads_dir = os.path.join(base_dir, 'uploads')
    path = os.path.join(uploads_dir, filename)
    
    print(f"🔍 DEBUG downloadpdf: Full path: {path}")
    print(f"🔍 DEBUG downloadpdf: File exists: {os.path.exists(path)}")
    
    if not os.path.exists(path):
        # List what PDFs are available
        try:
            import glob
            pdf_files = glob.glob(os.path.join(uploads_dir, "*.pdf"))
            print(f"🔍 DEBUG downloadpdf: Available PDF files: {pdf_files}")
        except Exception as e:
            print(f"❌ DEBUG downloadpdf: Error listing files: {str(e)}")
        
        flash(f'File not found: {filename}', 'danger')
        return redirect(url_for('main.approval_tracking'))
    
    # Download PDF to computer
    return send_file(path, mimetype='application/pdf', as_attachment=True, download_name=filename)

@bp.route("/admin/designations_by_role/<int:role_id>", methods=["GET"])
@login_required
def get_designations_by_role(role_id):
    """
    AJAX endpoint to get active designations
    Used for dynamic dropdown in JavaScript
    """
    if current_user.role_name != 'admin':
        return {"error": "Access denied"}, 403
    
    role = RoleMaster.query.get(role_id)
    
    # Only approvers need designations
    if not role or role.role_name != 'approver':
        return {"designations": []}
    
    # Get active designations
    designations = DesignationMaster.query.filter_by(is_active=True).all()
    
    return {
        "designations": [
            {
                "id": d.id,
                "name": d.designation_name,
                "level": d.approval_level
            } 
            for d in designations
        ] 
    }

@bp.route('/dismiss_notification/<int:notification_id>', methods=['POST'])
@login_required
def dismiss_notification(notification_id):
    """Dismiss a notification"""
    notification = Notification.query.get_or_404(notification_id)
    
    # Check if notification belongs to current user
    if notification.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    notification.is_dismissed = True
    notification.updated_at = get_ist_now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Notification dismissed'})

@bp.route('/dismiss_all_notifications', methods=['POST'])
@login_required
def dismiss_all_notifications():
    """Dismiss all recent notifications for current user"""
    notifications = Notification.query.filter_by(
        user_id=current_user.id,
        is_dismissed=False
    ).all()
    
    count = 0
    for notification in notifications:
        notification.is_dismissed = True
        notification.updated_at = get_ist_now()
        count += 1
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Dismissed {count} notifications'})

@bp.route('/clear_all_notifications', methods=['POST'])
@login_required
def clear_all_notifications():
    """Clear all notifications for current user"""
    notifications = Notification.query.filter_by(user_id=current_user.id).all()
    
    for notification in notifications:
        db.session.delete(notification)
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'All notifications cleared'})

@bp.route("/new_drawing_selection")
@login_required
def new_drawing_selection():
    """Page to select a station for new drawing - with dropdown"""
    # Check if user has permission (user or admin)
    if current_user.role.lower() not in ['user', '4','1']:
        flash('You do not have permission to create new drawings.', 'danger')
        return redirect(url_for('main.index'))
    
    # Get all projects for the current user
    if current_user.role_name == '4':
        projects = Project.query.order_by(Project.name).all()
    else:
        projects = current_user.projects
    
    # Create a list of projects with additional data
    projects_data = []
    for project in projects:
        # Count various record types
        station_count = StationDrawing.query.filter_by(project_id=project.id).count()
        junction_count = JunctionBox.query.filter_by(project_id=project.id).count()
        cable_count = Cable.query.filter_by(project_id=project.id).count()
        terminal_count = Terminal.query.filter_by(project_id=project.id).count()
        
        total_records = station_count + junction_count + cable_count + terminal_count
        
        # Add location_box_count (junction boxes)
        location_box_count = junction_count
        
        projects_data.append({
            'project': project,
            'total_records': total_records,
            'location_box_count': location_box_count,  # Add this line
            'has_data': total_records > 0
        })
    
    return render_template('new_drawing_selection.html', 
                         projects_data=projects_data,
                         title="Select Station for New Drawing")


@bp.route("/start_new_drawing_from_sidebar/<int:project_id>", methods=["POST"])
@login_required
def start_new_drawing_from_sidebar(project_id):
    """Start a new drawing for selected project from sidebar"""
    # Check if user has permission
    if current_user.role.lower() not in ['user', '4','1']:
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    project = Project.query.get_or_404(project_id)
    
    # Check if project belongs to user (for non-admin users)
    if current_user.role != 'admin' and project not in current_user.projects:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        # Clear existing drawing data (keep PDFs)
        tables_to_clear = [
            StationDrawing, JunctionBox, Cable, CableBox, Terminal, 
            Group, TerminalHeader, ChokeTable, ResistorTable
        ]
        
        for model in tables_to_clear:
            records = model.query.filter_by(project_id=project_id).all()
            for record in records:
                db.session.delete(record)
        
        # Update StationDrawing with default values
        station_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
        if station_drawing:
            # Reset version to 0 and checksum to "NO PDF GENERATED"
            station_drawing.version = "0"
            station_drawing.checksum = "NO PDF GENERATED"
            # Reset other fields to defaults
            station_drawing.drawn_by = "user"
            station_drawing.checked_by = "supervisor"
            station_drawing.designation1 = "DY.CSTE/C-II/ADI"
            station_drawing.designation2 = "DSTE/C/ADI"
            station_drawing.designation3 = "SSE/SIG/C/ADI"
        
        db.session.commit()
        
        # Set session for the selected project
        session['current_project_id'] = project_id
        session['project_id'] = project_id
        session['is_continue_drawing'] = False  # This is a new drawing
        
        return jsonify({
            'success': True,
            'message': f'New drawing started for {project.name}',
            'redirect_url': url_for('main.workflow_step', step=2)
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error starting new drawing: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500



# Add this route in routes.py
@bp.route("/get_checksum/<int:pdf_id>")
@login_required
def get_checksum(pdf_id):
    """Get checksum for a fully approved PDF"""
    pdf = GeneratedPDF.query.get_or_404(pdf_id)
    
    if pdf.is_fully_approved():
        return jsonify({
            'success': True,
            'checksum': pdf.checksum_md5,
            'filename': pdf.pdf_filename
        })
    else:
        return jsonify({
            'success': False,
            'message': 'PDF is not fully approved yet'
        }), 400

    
@bp.route('/save_junction_box_draft', methods=['POST'])
@login_required
def save_junction_box_draft():
    """Save junction box data as draft"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not data or 'junctions' not in data:
            return jsonify({'success': False, 'message': 'Invalid data format'}), 400
        
        junctions = data['junctions']
        
        print(f"🔍 DEBUG: Saving draft for {len(junctions)} junction boxes")
        
        # Track totals for summary
        saved_junction_ids = []
        
        # Save each junction box to JunctionBox table
        for junction_data in junctions:
            # Extract required fields
            station_id = junction_data.get('station_id', '')
            junction_id = junction_data.get('junction_id', '')
            junction_name = junction_data.get('junction_name', '')
            junction_size = junction_data.get('junction_size', 'Full')
            junction_row = junction_data.get('junction_row', '1')
            latitude = junction_data.get('latitude', '')
            longitude = junction_data.get('longitude', '')
            
            if not junction_id or not junction_name:
                continue
            
            # Check if junction already exists for this project
            existing_junction = JunctionBox.query.filter_by(
                project_id=project_id,
                junction_id=junction_id
            ).first()
            
            if existing_junction:
                # Update existing junction
                existing_junction.station_id = station_id
                existing_junction.junction_name = junction_name
                existing_junction.junction_size = junction_size
                existing_junction.junction_row = junction_row
                existing_junction.latitude = latitude
                existing_junction.longitude = longitude
                existing_junction.updated_date = get_ist_now()
                print(f"🔄 DEBUG: Updated existing junction {junction_id}")
            else:
                # Create new junction
                new_junction = JunctionBox(
                    project_id=project_id,
                    station_id=station_id,
                    junction_id=junction_id,
                    junction_name=junction_name,
                    junction_size=junction_size,
                    junction_row=junction_row,
                    latitude=latitude,
                    longitude=longitude
                )
                db.session.add(new_junction)
                print(f"🆕 DEBUG: Created new junction {junction_id}")
            
            saved_junction_ids.append(junction_id)
        
        # Update or create JunctionBoxSummary
        # Count total junction boxes for this project
        total_junction_boxes = JunctionBox.query.filter_by(project_id=project_id).count()
        
        summary = JunctionBoxSummary.query.filter_by(project_id=project_id).first()
        
        if summary:
            # Update existing summary
            summary.total_junction_boxes = total_junction_boxes
            summary.updated_at = get_ist_now()
            print(f"📊 DEBUG: Updated existing summary: {summary.total_junction_boxes} boxes")
        else:
            # Create new summary
            summary = JunctionBoxSummary(
                project_id=project_id,
                total_junction_boxes=total_junction_boxes
            )
            db.session.add(summary)
            print(f"📊 DEBUG: Created new summary: {total_junction_boxes} boxes")
        
        # ✅ CRITICAL: Update project's junction_data after saving junction boxes
        try:
            current_project = Project.query.get(project_id)
            if current_project:
                current_project.update_junction_data()
                print(f"✅ Updated junction_data for project {project_id} after saving draft")
        except Exception as update_error:
            print(f"⚠️ Warning: Could not update junction_data: {update_error}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Draft saved successfully ({len(saved_junction_ids)} junction boxes)',
            'total_junction_boxes': total_junction_boxes,
            'saved_junction_ids': saved_junction_ids,
            'project_id': project_id  # ADDED: Return project ID for PDF modal
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving junction box draft: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving draft: {str(e)}'}), 500

@bp.route('/get_cable_configuration', methods=['GET'])
def get_cable_configuration():
    """Get saved cable configuration for a junction box"""
    try:
        project_id = request.args.get('project_id')
        junction_box_id = request.args.get('junction_box_id')
        
        if not project_id or not junction_box_id:
            return jsonify({'success': False, 'message': 'Missing required parameters'}), 400
        
        # Get configuration (non-draft only)
        config_rows = CableRowConfig.query.filter_by(
            project_id=project_id,
            junction_box_id=junction_box_id,
            is_draft=False
        ).order_by(CableRowConfig.row_number).all()
        
        if config_rows:
            config_data = []
            for row in config_rows:
                config_data.append({
                    'row_number': row.row_number,
                    'location_row_name': row.location_row_name,
                    'cable_type': row.cable_type,
                    'number_of_cables': row.number_of_cables
                })
            
            return jsonify({
                'success': True,
                'config_rows': config_data
            })
        
        return jsonify({
            'success': False,
            'message': 'No configuration found'
        })
        
    except Exception as e:
        print(f"Error getting cable configuration: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error getting configuration: {str(e)}'
        }), 500



@bp.route('/save_cable_configuration', methods=['POST'])
def save_cable_configuration():
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        junction_box_id = data.get('junction_box_id')  # Keep as-is
        config_rows = data.get('config_rows', [])
        
        if not project_id or not junction_box_id:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        # DON'T clean the junction_box_id - keep it as provided
        # OR clean but preserve size indicator
        cleaned_junction_box_id = junction_box_id.strip()  # Just trim whitespace
        
        print(f"Saving cable configuration for junction: {cleaned_junction_box_id}")
        
        # Delete existing rows for THIS EXACT junction_box_id
        deleted = CableRowConfig.query.filter_by(
            project_id=project_id,
            junction_box_id=cleaned_junction_box_id  # Use the exact ID
        ).delete()
        
        # Save new rows
        for row_data in config_rows:
            cable_config = CableRowConfig(
                project_id=project_id,
                junction_box_id=cleaned_junction_box_id,  # Use exact ID
                row_number=row_data.get('row_number'),
                location_row_name=row_data.get('location_row_name'),
                cable_type=row_data.get('cable_type'),
                number_of_cables=row_data.get('number_of_cables'),
                is_draft=False,
                draft_version=0
            )
            db.session.add(cable_config)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cable configuration saved successfully',
            'junction_box_id': cleaned_junction_box_id
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error saving cable configuration: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error saving configuration: {str(e)}'
        }), 500

@bp.route('/get_cable_config_draft', methods=['GET'])
def get_cable_config_draft():
    """Get cable configuration draft for a junction box"""
    try:
        project_id = request.args.get('project_id')
        junction_box_id = request.args.get('junction_box_id')
        
        if not project_id or not junction_box_id:
            return jsonify({'success': False, 'message': 'Missing required parameters'}), 400
        
        # Get draft configuration
        config_rows = CableRowConfig.query.filter_by(
            project_id=project_id,
            junction_box_id=junction_box_id,
            is_draft=True
        ).order_by(CableRowConfig.row_number).all()
        
        if config_rows:
            config_data = []
            draft_version = config_rows[0].draft_version if config_rows else 1
            
            for row in config_rows:
                config_data.append({
                    'row_number': row.row_number,
                    'location_row_name': row.location_row_name,
                    'cable_type': row.cable_type,
                    'number_of_cables': row.number_of_cables
                })
            
            return jsonify({
                'success': True,
                'config_rows': config_data,
                'draft_version': draft_version
            })
        
        return jsonify({
            'success': False,
            'message': 'No draft found'
        })
        
    except Exception as e:
        print(f"Error getting cable configuration draft: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error getting draft: {str(e)}'
        }), 500

@bp.route('/clear_cable_config_draft', methods=['POST'])
def clear_cable_config_draft():
    """Clear cable configuration draft"""
    try:
        project_id = request.args.get('project_id')
        junction_box_id = request.args.get('junction_box_id')
        
        if not project_id or not junction_box_id:
            return jsonify({'success': False, 'message': 'Missing required parameters'}), 400
        
        # Delete draft configuration
        deleted_count = CableRowConfig.query.filter_by(
            project_id=project_id,
            junction_box_id=junction_box_id,
            is_draft=True
        ).delete()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Draft cleared successfully ({deleted_count} rows deleted)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error clearing cable configuration draft: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error clearing draft: {str(e)}'
        }), 500
        
@bp.route('/save_more_junctions', methods=['POST'])
@login_required
def save_more_junctions():
    """Save additional junction boxes (final save)"""
    try:
        project_id = get_current_project()
        
        if not project_id:
            flash('No project selected', 'error')
            return redirect(request.referrer or url_for('main.dashboard'))
        
        # Get form data - use getlist to get all values
        junctions = []
        
        # Loop through all junction indices
        i = 0
        while True:
            station_id = request.form.get(f'junctions[{i}][station_id]')
            if station_id is None:
                break
            
            junction_data = {
                'station_id': station_id,
                'junction_id': request.form.get(f'junctions[{i}][junction_id]', ''),
                'junction_name': request.form.get(f'junctions[{i}][junction_name]', ''),
                'latitude': request.form.get(f'junctions[{i}][latitude]', ''),
                'longitude': request.form.get(f'junctions[{i}][longitude]', ''),
                'junction_size': request.form.get(f'junctions[{i}][junction_size]', 'Full'),
                'junction_row': request.form.get(f'junctions[{i}][junction_row]', '1')
            }
            
            # Validate required fields
            if not junction_data['junction_id'] or not junction_data['junction_name']:
                i += 1
                continue
                
            junctions.append(junction_data)
            i += 1
        
        print(f"🔍 DEBUG: Saving {len(junctions)} junction boxes")
        
        saved_junction_ids = []
        
        # Save each junction box
        for junction_data in junctions:
            station_id = junction_data.get('station_id', '')
            junction_id = junction_data.get('junction_id', '')
            junction_name = junction_data.get('junction_name', '')
            junction_size = junction_data.get('junction_size', 'Full')
            junction_row = junction_data.get('junction_row', '1')
            latitude = junction_data.get('latitude', '')
            longitude = junction_data.get('longitude', '')
            
            if not junction_id or not junction_name:
                continue
            
            # Check if junction already exists
            existing_junction = JunctionBox.query.filter_by(
                project_id=project_id,
                junction_id=junction_id
            ).first()
            
            if existing_junction:
                # Update existing junction
                existing_junction.station_id = station_id
                existing_junction.junction_name = junction_name
                existing_junction.junction_size = junction_size
                existing_junction.junction_row = junction_row
                existing_junction.latitude = latitude
                existing_junction.longitude = longitude
                existing_junction.updated_date = get_ist_now()
                print(f"🔄 Updated junction {junction_id}")
            else:
                # Create new junction
                new_junction = JunctionBox(
                    project_id=project_id,
                    station_id=station_id,
                    junction_id=junction_id,
                    junction_name=junction_name,
                    junction_size=junction_size,
                    junction_row=junction_row,
                    latitude=latitude,
                    longitude=longitude
                )
                db.session.add(new_junction)
                print(f"🆕 Created new junction {junction_id}")
            
            saved_junction_ids.append(junction_id)
        
        # Update JunctionBoxSummary
        total_junction_boxes = JunctionBox.query.filter_by(project_id=project_id).count()
        
        summary = JunctionBoxSummary.query.filter_by(project_id=project_id).first()
        
        if summary:
            summary.total_junction_boxes = total_junction_boxes
            summary.updated_at = get_ist_now()
        else:
            summary = JunctionBoxSummary(
                project_id=project_id,
                total_junction_boxes=total_junction_boxes
            )
            db.session.add(summary)
        
        db.session.commit()
        
        flash(f'Successfully saved {len(saved_junction_ids)} junction boxes!', 'success')
        return redirect(url_for('main.junction_boxes'))
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving junction boxes: {str(e)}")
        flash(f'Error saving junction boxes: {str(e)}', 'error')
        return redirect(request.referrer or url_for('main.dashboard'))


@bp.route('/save_cable_table_draft', methods=['POST'])
@login_required
def save_cable_table_draft():
    """Save cable table configuration - UPDATE IF EXISTS, INSERT IF NEW"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        
        print(f"🔍 DEBUG save_cable_table_draft: project_id={project_id}, data={data}")
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not data or 'junction_box_id' not in data or 'cable_data' not in data:
            return jsonify({'success': False, 'message': 'Invalid data format'}), 400
        
        junction_box_id = data['junction_box_id']
        cable_data = data['cable_data']
        junction_box_name = data.get('junction_box_name', '')
        
        print(f"🔍 DEBUG: Saving/updating cable table for junction {junction_box_id} with {len(cable_data)} cables")
        
        # Track totals and unique rows
        total_cables = 0
        unique_rows = set()
        cables_saved = []
        
        # Process each cable item - UPDATE if exists, INSERT if new
        for cable_item in cable_data:
            try:
                # Extract data from cable_item
                row = cable_item.get('row', '')
                position = cable_item.get('position', '')
                terminal = cable_item.get('terminal', '')
                start_no = cable_item.get('start_no', '')
                cable_id = cable_item.get('cable_id', f'{junction_box_id}-{row}-{position}')
                
                # Generate a unique cable name if not provided
                cable_name = cable_item.get('cable_name', f'Cable {row}{position}')
                
                # Check if this cable already exists
                existing_cable = Cable.query.filter(
                    Cable.project_id == project_id,
                    Cable.cable_id == cable_id
                ).first()
                
                if existing_cable:
                    # UPDATE existing cable
                    existing_cable.cable_name = cable_name
                    existing_cable.junction_box = junction_box_id
                    existing_cable.junction_name = junction_box_name
                    existing_cable.row = row
                    existing_cable.position = position
                    existing_cable.terminal = terminal
                    existing_cable.start_no = start_no
                    existing_cable.created_date = get_ist_now()
                    
                    print(f"🔄 Updated existing cable: {cable_id}")
                else:
                    # INSERT new cable
                    cable = Cable(
                        project_id=project_id,
                        cable_id=cable_id,
                        cable_name=cable_name,
                        junction_box=junction_box_id,
                        junction_name=junction_box_name,
                        row=row,
                        position=position,
                        terminal=terminal,
                        start_no=start_no,
                        created_date=get_ist_now()
                    )
                    db.session.add(cable)
                    print(f"➕ Created new cable: {cable_id}")
                
                cables_saved.append(cable_id)
                total_cables += 1
                unique_rows.add(row)
                
            except Exception as e:
                print(f"⚠️ WARNING: Error saving cable {cable_item}: {str(e)}")
                continue
        
        # Now, delete any cables for this junction box that are no longer in the current data
        current_cable_ids = []
        for cable_item in cable_data:
            row = cable_item.get('row', '')
            position = cable_item.get('position', '')
            cable_id = cable_item.get('cable_id')
            
            if not cable_id:
                cable_id = f"{junction_box_id}-{row}-{position}"
            
            current_cable_ids.append(cable_id)
        
        # Delete cables that exist in DB but not in current data
        stale_cables = Cable.query.filter(
            Cable.project_id == project_id,
            Cable.junction_box == junction_box_id,
            ~Cable.cable_id.in_(current_cable_ids)
        ).all()
        
        for stale_cable in stale_cables:
            db.session.delete(stale_cable)
            print(f"🗑️ Deleted stale cable: {stale_cable.cable_id}")
        
        # Update or create CableSummary
        summary = CableSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Count all cables for this project
            all_cables = Cable.query.filter(
                Cable.project_id == project_id
            ).count()
            
            # Count unique rows from all cables
            all_rows_query = db.session.query(Cable.row).filter(
                Cable.project_id == project_id
            ).distinct()
            total_rows = all_rows_query.count()
            
            # Count unique junctions from all cables
            all_junctions_query = db.session.query(Cable.junction_box).filter(
                Cable.project_id == project_id
            ).distinct()
            total_junctions = all_junctions_query.count()
            
            # Update summary
            summary.total_cables = all_cables
            summary.total_rows = total_rows
            summary.total_junctions = total_junctions
            summary.updated_at = get_ist_now()
        else:
            # Create new summary
            summary = CableSummary(
                project_id=project_id,
                total_cables=total_cables,
                total_rows=len(unique_rows),
                total_junctions=1,
                updated_at=get_ist_now()
            )
            db.session.add(summary)
        
        db.session.commit()
        
        print(f"✅ DEBUG: Saved/Updated {total_cables} cables to Cable table")
        
        return jsonify({
            'success': True,
            'message': f'Cable table saved successfully ({total_cables} cables)',
            'total_cables': total_cables,
            'total_rows': len(unique_rows),
            'cables_saved': cables_saved
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving cable table: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving cable table: {str(e)}'}), 500


@bp.route('/finalize_cable_table', methods=['POST'])
@login_required
def finalize_cable_table():
    """Finalize cable table (now just updates summary)"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        junction_box_id = data.get('junction_box_id')
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        # Update CableSummary
        summary = CableSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Count all cables for this project
            all_cables = Cable.query.filter(
                Cable.project_id == project_id
            ).count()
            
            # Count unique rows
            all_rows_query = db.session.query(Cable.row).filter(
                Cable.project_id == project_id
            ).distinct()
            total_rows = all_rows_query.count()
            
            # Count unique junctions
            all_junctions_query = db.session.query(Cable.junction_box).filter(
                Cable.project_id == project_id
            ).distinct()
            total_junctions = all_junctions_query.count()
            
            summary.total_cables = all_cables
            summary.total_rows = total_rows
            summary.total_junctions = total_junctions
            summary.updated_at = get_ist_now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Cable table finalized',
            'total_cables': all_cables,
            'total_rows': total_rows,
            'total_junctions': total_junctions
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error finalizing cable table: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error finalizing cable table: {str(e)}'}), 500

@bp.route('/get_cable_table_draft', methods=['GET'])
@login_required
def get_cable_table_draft():
    """Load cable table draft from Cable table"""
    try:
        project_id = get_current_project()
        junction_box_id = request.args.get('junction_box_id')
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not junction_box_id:
            return jsonify({'success': False, 'message': 'Junction box ID required'}), 400
        
        # Load draft cables from Cable table
        draft_cables = Cable.query.filter(
            Cable.project_id == project_id,
            Cable.junction_box == junction_box_id,
            Cable.cable_id.like('DRAFT-%')
        ).order_by(Cable.row, Cable.position).all()
        
        # Reconstruct cable data from draft cables
        cable_data = []
        for cable in draft_cables:
            # Remove DRAFT- prefix for display
            display_cable_id = cable.cable_id[6:] if cable.cable_id.startswith('DRAFT-') else cable.cable_id
            
            cable_data.append({
                'cable_id': display_cable_id,
                'cable_name': cable.cable_name,
                'row': cable.row,
                'position': cable.position,
                'terminal': cable.terminal,
                'start_no': cable.start_no
            })
        
        # Also get rows configuration by grouping
        rows_config = []
        rows_data = {}
        
        for cable in draft_cables:
            row = cable.row
            if row not in rows_data:
                rows_data[row] = {
                    'row': row,
                    'cable_count': 0
                }
            rows_data[row]['cable_count'] += 1
        
        rows_config = list(rows_data.values())
        
        print(f"🔍 DEBUG: Found {len(cable_data)} draft cables in Cable table")
        
        return jsonify({
            'success': True,
            'cable_data': cable_data,
            'rows_config': rows_config,
            'has_draft': len(cable_data) > 0
        })
        
    except Exception as e:
        print(f"❌ Error loading cable table draft: {str(e)}")
        return jsonify({'success': False, 'message': f'Error loading cable table draft: {str(e)}'}), 500


@bp.route('/clear_cable_table_draft', methods=['POST'])
@login_required
def clear_cable_table_draft():
    """Clear cable table draft from Cable table"""
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        junction_box_id = data.get('junction_box_id')
        
        if not all([project_id, junction_box_id]):
            return jsonify({'success': False, 'message': 'Missing parameters'}), 400
        
        print(f"🔍 DEBUG clear_cable_table_draft: project_id={project_id}, junction_box_id={junction_box_id}")
        
        # Delete draft cables from Cable table
        deleted_cables_count = Cable.query.filter(
            Cable.project_id == project_id,
            Cable.junction_box == junction_box_id,
            Cable.cable_id.like('DRAFT-%')
        ).delete(synchronize_session=False)
        
        # Update CableSummary
        summary = CableSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Recalculate totals from final cables only
            final_cables = Cable.query.filter(
                Cable.project_id == project_id,
                ~Cable.cable_id.like('DRAFT-%')
            ).count()
            
            # Count unique rows from final cables
            final_rows_query = db.session.query(Cable.row).filter(
                Cable.project_id == project_id,
                ~Cable.cable_id.like('DRAFT-%')
            ).distinct()
            total_rows = final_rows_query.count()
            
            # Count unique junctions from final cables
            final_junctions_query = db.session.query(Cable.junction_box).filter(
                Cable.project_id == project_id,
                ~Cable.cable_id.like('DRAFT-%')
            ).distinct()
            total_junctions = final_junctions_query.count()
            
            summary.total_cables = final_cables
            summary.total_rows = total_rows
            summary.total_junctions = total_junctions
            summary.updated_at = get_ist_now()
        
        db.session.commit()
        
        print(f"✅ DEBUG: Cleared {deleted_cables_count} draft cables from Cable table")
        
        return jsonify({
            'success': True,
            'message': f'Cable table draft cleared successfully ({deleted_cables_count} cables removed)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error clearing cable table draft: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/get_cable_summary', methods=['GET'])
@login_required
def get_cable_summary():
    """Get current cable summary"""
    try:
        project_id = get_current_project()
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        summary = CableSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if not summary:
            return jsonify({
                'success': True,
                'total_cables': 0,
                'total_rows': 0,
                'total_junctions': 0,
                'updated_at': None
            })
        
        return jsonify({
            'success': True,
            'total_cables': summary.total_cables,
            'total_rows': summary.total_rows,
            'total_junctions': summary.total_junctions,
            'updated_at': summary.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"❌ Error getting cable summary: {str(e)}")
        return jsonify({'success': False, 'message': f'Error getting cable summary: {str(e)}'}), 500


@bp.route('/get_cable_stats', methods=['GET'])
@login_required
def get_cable_stats():
    """Get detailed cable statistics"""
    try:
        project_id = get_current_project()
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        # Get total cables (excluding drafts)
        total_cables = Cable.query.filter(
            Cable.project_id == project_id,
            ~Cable.cable_id.like('DRAFT-%')
        ).count()
        
        # Get total draft cables
        draft_cables = Cable.query.filter(
            Cable.project_id == project_id,
            Cable.cable_id.like('DRAFT-%')
        ).count()
        
        # Get unique rows
        rows_query = db.session.query(Cable.row).filter(
            Cable.project_id == project_id,
            ~Cable.cable_id.like('DRAFT-%')
        ).distinct()
        total_rows = rows_query.count()
        
        # Get unique junctions
        junctions_query = db.session.query(Cable.junction_box).filter(
            Cable.project_id == project_id,
            ~Cable.cable_id.like('DRAFT-%')
        ).distinct()
        total_junctions = junctions_query.count()
        
        return jsonify({
            'success': True,
            'total_cables': total_cables,
            'draft_cables': draft_cables,
            'total_rows': total_rows,
            'total_junctions': total_junctions
        })
        
    except Exception as e:
        print(f"❌ Error getting cable stats: {str(e)}")
        return jsonify({'success': False, 'message': f'Error getting cable stats: {str(e)}'}), 500
    
@bp.route('/save_terminal_draft', methods=['POST'])
@login_required
def save_terminal_draft():
    """Save terminal configuration directly to Terminal table (final version)"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        
        print(f"🔍 DEBUG save_terminal_draft: project_id={project_id}, data={data}")
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not data or 'cable_id' not in data or 'terminal_data' not in data:
            return jsonify({'success': False, 'message': 'Invalid data format'}), 400
        
        # Convert cable_id to string for consistent storage
        cable_id_str = str(data['cable_id'])
        terminal_data = data['terminal_data']
        junction_box_id = data.get('junction_box_id', '')
        cable_name = data.get('cable_name', '')
        
        print(f"🔍 DEBUG: Saving terminals for cable {cable_id_str} with {len(terminal_data)} terminals")
        
        # First, delete any existing terminals for this cable
        deleted_count = Terminal.query.filter(
            Terminal.project_id == project_id,
            Terminal.cable_id == cable_id_str
        ).delete()
        print(f"🔍 DEBUG: Deleted {deleted_count} existing terminals for cable {cable_id_str}")
        
        db.session.flush()
        
        # Track totals
        total_terminals = 0
        terminals_saved = []
        
        # Save to Terminal table (directly without draft prefix)
        for i, terminal_item in enumerate(terminal_data):
            try:
                # Extract terminal fields
                terminal_id = terminal_item.get('terminal_id', f'{cable_id_str}-T{i+1}')
                terminal_no = terminal_item.get('terminal_no', str(i+1))
                symbol = terminal_item.get('symbol', 'ara/wago')
                input_left = terminal_item.get('input_left', '')
                input_right = terminal_item.get('input_right', '')
                spare = terminal_item.get('spare', 'N')
                input_connected = terminal_item.get('input_connected', 'Y')
                output_connected = terminal_item.get('output_connected', 'Y')
                input_connected_extra = terminal_item.get('input_connected_extra', '')
                output_connected_extra = terminal_item.get('output_connected_extra', '')
                output_left = terminal_item.get('output_left', '')
                output_right = terminal_item.get('output_right', '')
                
                terminal = Terminal(
                    project_id=project_id,
                    cable_id=cable_id_str,
                    terminal_id=terminal_id,
                    terminal_no=terminal_no,
                    symbol=symbol,
                    input_left=input_left,
                    input_right=input_right,
                    spare=spare,
                    input_connected=input_connected,
                    output_connected=output_connected,
                    input_connected_extra=input_connected_extra,
                    output_connected_extra=output_connected_extra,
                    output_left=output_left,
                    output_right=output_right,
                    created_date=get_ist_now()
                )
                db.session.add(terminal)
                terminals_saved.append(terminal.terminal_id)
                total_terminals += 1
                
                print(f"✅ Saved terminal {terminal_id} for cable {cable_id_str}")
                
            except Exception as e:
                print(f"⚠️ WARNING: Error saving terminal {terminal_item}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        # Update TerminalSummary with total count
        summary = TerminalSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Get count of all terminals for this project
            total_project_terminals = Terminal.query.filter(
                Terminal.project_id == project_id
            ).count()
            
            summary.total_terminals = total_project_terminals
            summary.updated_at = get_ist_now()
            print(f"📊 Updated existing summary: {total_project_terminals} terminals")
        else:
            # Create new summary
            total_project_terminals = Terminal.query.filter(
                Terminal.project_id == project_id
            ).count()
            
            summary = TerminalSummary(
                project_id=project_id,
                total_terminals=total_project_terminals,
                updated_at=get_ist_now()
            )
            db.session.add(summary)
            print(f"📊 Created new summary: {total_project_terminals} terminals")
        
        db.session.commit()
        
        print(f"✅ DEBUG: Successfully saved {total_terminals} terminals to Terminal table")
        print(f"📊 DEBUG: Total terminals in project: {total_project_terminals}")
        
        return jsonify({
            'success': True,
            'message': f'Terminal configuration saved successfully ({total_terminals} terminals)',
            'total_terminals': total_project_terminals,
            'terminals_saved': terminals_saved,
            'cable_id': cable_id_str
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving terminals: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving terminals: {str(e)}'}), 500

@bp.route('/finalize_terminal_config', methods=['POST'])
@login_required
def finalize_terminal_config():
    """Finalize terminal configuration (simplified - just returns success)"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        cable_id = data.get('cable_id')
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        # Just return success since terminals are already saved
        return jsonify({
            'success': True,
            'message': 'Terminal configuration completed successfully'
        })
        
    except Exception as e:
        print(f"❌ Error finalizing terminal config: {str(e)}")
        return jsonify({'success': False, 'message': f'Error finalizing terminal config: {str(e)}'}), 500

@bp.route('/get_terminal_draft', methods=['GET'])
@login_required
def get_terminal_draft():
    """Load existing terminal configuration from Terminal table"""
    try:
        project_id = get_current_project()
        cable_id = request.args.get('cable_id')
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not cable_id:
            return jsonify({'success': False, 'message': 'Cable ID required'}), 400
        
        cable_id_str = str(cable_id)
        
        # Load existing terminals from Terminal table
        existing_terminals = Terminal.query.filter(
            Terminal.project_id == project_id,
            Terminal.cable_id == cable_id_str
        ).order_by(Terminal.terminal_no).all()
        
        # Reconstruct terminal data
        terminal_data = []
        for terminal in existing_terminals:
            terminal_data.append({
                'terminal_id': terminal.terminal_id,
                'terminal_no': terminal.terminal_no,
                'symbol': terminal.symbol,
                'input_left': terminal.input_left,
                'input_right': terminal.input_right,
                'spare': terminal.spare,
                'input_connected': terminal.input_connected,
                'output_connected': terminal.output_connected,
                'input_connected_extra': terminal.input_connected_extra,
                'output_connected_extra': terminal.output_connected_extra,
                'output_left': terminal.output_left,
                'output_right': terminal.output_right
            })
        
        print(f"🔍 DEBUG: Found {len(terminal_data)} existing terminals for cable {cable_id_str}")
        
        return jsonify({
            'success': True,
            'terminal_data': terminal_data,
            'has_draft': len(terminal_data) > 0
        })
        
    except Exception as e:
        print(f"❌ Error loading terminal configuration: {str(e)}")
        return jsonify({'success': False, 'message': f'Error loading terminal configuration: {str(e)}'}), 500
  
@bp.route('/save_terminal_header_draft', methods=['POST'])
@login_required
def save_terminal_header_draft():
    """Save terminal header configuration - UPDATE IF EXISTS, INSERT IF NEW"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        
        print(f"🔍 DEBUG save_terminal_header_draft: project_id={project_id}, data={data}")
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not data or 'cable_id' not in data or 'header_data' not in data:
            return jsonify({'success': False, 'message': 'Invalid data format'}), 400
        
        cable_id = data['cable_id']
        header_data = data['header_data']
        
        print(f"🔍 DEBUG: Saving/updating terminal header for cable {cable_id} with {len(header_data)} headers")
        print(f"🔍 DEBUG: cable_id type from request: {type(cable_id)}, value: {cable_id}")
        
        # Convert cable_id to string for consistency with database
        cable_id_str = str(cable_id).strip()
        
        # Track totals
        total_headers = 0
        headers_saved = []
        
        # Get ALL existing headers for this cable in one query
        existing_headers = TerminalHeader.query.filter(
            TerminalHeader.project_id == project_id,
            TerminalHeader.cable_id == cable_id_str
        ).all()
        
        print(f"🔍 Found {len(existing_headers)} existing headers for cable {cable_id_str}")
        for h in existing_headers:
            print(f"   - Existing: {h.header_type} ({h.terminal_start}-{h.terminal_end})")
        
        # Create a dictionary for fast lookup
        existing_headers_dict = {}
        for h in existing_headers:
            # Use tuple as key: (header_type, terminal_start, terminal_end)
            header_type_clean = (h.header_type or '').strip().upper()
            terminal_start_clean = str(h.terminal_start or '').strip()
            terminal_end_clean = str(h.terminal_end or '').strip()
            text_clean = str(h.text).strip()
            key = (header_type_clean, terminal_start_clean, terminal_end_clean,text_clean)
            existing_headers_dict[key] = h
            print(f"🔍 Added to lookup dict: {key}")
        
        # Track which headers we've processed
        processed_keys = set()
        
        # Process each header item - UPDATE if exists, INSERT if new
        for idx, header_item in enumerate(header_data):
            try:
                print(f"\n🔍 Processing header item {idx}: {header_item}")
                
                # Extract data from header_item
                header_type = header_item.get('header_type', '')
                terminal_start = header_item.get('terminal_start', '')
                terminal_end = header_item.get('terminal_end', '')
                input_output = header_item.get('input_output', '')
                text = header_item.get('text', '')
                
                print(f"   Raw data: type={header_type}, start={terminal_start}, end={terminal_end}, io={input_output}, text={text}")
                
                # Skip if required fields are missing
                if not header_type or not terminal_start or not terminal_end:
                    print(f"   ⚠️ Skipping - missing required fields")
                    continue
                
                # Clean the data for comparison
                header_type_clean = str(header_type).strip().upper()
                terminal_start_clean = str(terminal_start).strip()
                terminal_end_clean = str(terminal_end).strip()
                text_clean = str(text).strip()
                key = (header_type_clean, terminal_start_clean, terminal_end_clean,text_clean)
                processed_keys.add(key)
                
                print(f"   Cleaned key: {key}")
                print(f"   Looking for key in existing_headers_dict: {key in existing_headers_dict}")
                
                # Check if this header already exists in our dictionary
                existing_header = existing_headers_dict.get(key)
                
                if existing_header:
                    # Check if data actually changed
                    if (existing_header.input_output != input_output or 
                        existing_header.text != text):
                        # UPDATE existing header
                        existing_header.input_output = input_output
                        existing_header.text = text
                        existing_header.created_date = get_ist_now()
                        
                        print(f"🔄 Updated existing header: {header_type_clean} ({terminal_start_clean}-{terminal_end_clean})")
                        headers_saved.append(f"UPDATED: {header_type_clean}")
                    else:
                        print(f"✅ No changes for existing header: {header_type_clean} ({terminal_start_clean}-{terminal_end_clean})")
                        headers_saved.append(f"NO CHANGE: {header_type_clean}")
                else:
                    # INSERT new header (no DRAFT- prefix added)
                    header = TerminalHeader(
                        project_id=project_id,
                        cable_id=cable_id_str,
                        header_type=header_type_clean,
                        terminal_start=terminal_start_clean,
                        terminal_end=terminal_end_clean,
                        input_output=input_output,
                        text=text,
                        created_date=get_ist_now()
                    )
                    db.session.add(header)
                    print(f"➕ Created new header: {header_type_clean} ({terminal_start_clean}-{terminal_end_clean})")
                    headers_saved.append(f"NEW: {header_type_clean}")
                
                total_headers += 1
                
            except Exception as e:
                print(f"⚠️ WARNING: Error saving header {header_item}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        # Delete headers that are no longer in the current data
        deleted_count = 0
        for key, header in existing_headers_dict.items():
            if key not in processed_keys:
                db.session.delete(header)
                deleted_count += 1
                print(f"🗑️ Deleted stale header: {key[0]} ({key[1]}-{key[2]})")
        
        # Update or create TerminalHeaderSummary
        summary = TerminalHeaderSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        # Count all headers for this project (no draft distinction)
        all_headers_count = TerminalHeader.query.filter(
            TerminalHeader.project_id == project_id
        ).count()
        
        if summary:
            summary.total_terminal_headers = all_headers_count
            summary.updated_at = get_ist_now()
        else:
            summary = TerminalHeaderSummary(
                project_id=project_id,
                total_terminal_headers=all_headers_count,
                updated_at=get_ist_now()
            )
            db.session.add(summary)
        
        db.session.commit()
        
        print(f"✅ DEBUG: Saved/Updated {total_headers} headers, Deleted {deleted_count} headers")
        
        return jsonify({
            'success': True,
            'message': f'Terminal header saved successfully ({total_headers} headers, {deleted_count} deleted)',
            'total_headers': total_headers,
            'deleted_count': deleted_count,
            'headers_saved': headers_saved
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving terminal header: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving terminal header: {str(e)}'}), 500


# 🗑️ REMOVED finalize_terminal_header endpoint – no longer needed

@bp.route('/get_terminal_header_draft', methods=['GET'])
@login_required
def get_terminal_header_draft():
    """Load terminal header configuration (all headers, no draft prefix)"""
    try:
        project_id = get_current_project()
        cable_id = request.args.get('cable_id')
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not cable_id:
            return jsonify({'success': False, 'message': 'Cable ID required'}), 400
        
        # Load all headers for this cable (no draft filtering)
        headers = TerminalHeader.query.filter(
            TerminalHeader.project_id == project_id,
            TerminalHeader.cable_id == cable_id
        ).order_by(TerminalHeader.terminal_start).all()
        
        header_data = []
        for header in headers:
            header_data.append({
                'cable_id': header.cable_id,
                'header_type': header.header_type,  # stored without prefix
                'terminal_start': header.terminal_start,
                'terminal_end': header.terminal_end,
                'input_output': header.input_output,
                'text': header.text
            })
        
        print(f"🔍 DEBUG: Found {len(header_data)} headers for cable {cable_id}")
        
        return jsonify({
            'success': True,
            'header_data': header_data,
            'has_draft': len(header_data) > 0  # kept for compatibility; always true if any exist
        })
        
    except Exception as e:
        print(f"❌ Error loading terminal header draft: {str(e)}")
        return jsonify({'success': False, 'message': f'Error loading terminal header draft: {str(e)}'}), 500


@bp.route('/clear_terminal_header_draft', methods=['POST'])
@login_required
def clear_terminal_header_draft():
    """Clear all terminal headers for a cable (delete all)"""
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        cable_id = data.get('cable_id')
        
        if not all([project_id, cable_id]):
            return jsonify({'success': False, 'message': 'Missing parameters'}), 400
        
        print(f"🔍 DEBUG clear_terminal_header_draft: project_id={project_id}, cable_id={cable_id}")
        
        # Delete ALL headers for this cable (no draft filter)
        deleted_headers_count = TerminalHeader.query.filter(
            TerminalHeader.project_id == project_id,
            TerminalHeader.cable_id == cable_id
        ).delete(synchronize_session=False)
        
        # Update TerminalHeaderSummary
        summary = TerminalHeaderSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Recalculate total headers
            total_headers = TerminalHeader.query.filter(
                TerminalHeader.project_id == project_id
            ).count()
            summary.total_terminal_headers = total_headers
            summary.updated_at = get_ist_now()
        
        db.session.commit()
        
        print(f"✅ DEBUG: Cleared {deleted_headers_count} headers for cable {cable_id}")
        
        return jsonify({
            'success': True,
            'message': f'Terminal headers cleared successfully ({deleted_headers_count} headers removed)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error clearing terminal header draft: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/get_terminal_header_summary', methods=['GET'])
@login_required
def get_terminal_header_summary():
    """Get current terminal header summary"""
    try:
        project_id = get_current_project()
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        summary = TerminalHeaderSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if not summary:
            return jsonify({
                'success': True,
                'total_terminal_headers': 0,
                'updated_at': None
            })
        
        return jsonify({
            'success': True,
            'total_terminal_headers': summary.total_terminal_headers,
            'updated_at': summary.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"❌ Error getting terminal header summary: {str(e)}")
        return jsonify({'success': False, 'message': f'Error getting terminal header summary: {str(e)}'}), 500
    
@bp.route('/save_group_table_draft', methods=['POST'])
@login_required
def save_group_table_draft():
    """Save group table configuration - UPDATE IF EXISTS, INSERT IF NEW"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        
        print(f"🔍 DEBUG save_group_table_draft: project_id={project_id}, data={data}")
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not data or 'cable_id' not in data or 'group_data' not in data:
            return jsonify({'success': False, 'message': 'Invalid data format'}), 400
        
        cable_id = data['cable_id']
        group_data = data['group_data']
        
        print(f"🔍 DEBUG: Saving/updating group table for cable {cable_id} with {len(group_data)} groups")
        
        # Track totals
        total_groups = 0
        groups_saved = []
        
        # Process each group item - UPDATE if exists, INSERT if new
        for group_item in group_data:
            try:
                # Extract data from group_item
                group_id = group_item.get('group_id', '')
                terminal_no = group_item.get('terminal_no', '')
                input_output = group_item.get('input_output', '')
                text = group_item.get('text', '')
                
                # Generate unique group ID if not provided
                if not group_id:
                    # Find the next available group ID
                    existing_groups = Group.query.filter(
                        Group.project_id == project_id,
                        Group.cable_id == cable_id
                    ).count()
                    group_id = f'GR{existing_groups + 1:03d}'
                
                # Check if this group already exists
                existing_group = Group.query.filter(
                    Group.project_id == project_id,
                    Group.cable_id == cable_id,
                    Group.group_id == group_id
                ).first()
                
                if existing_group:
                    # UPDATE existing group
                    existing_group.terminal_no = terminal_no
                    existing_group.input_output = input_output
                    existing_group.text = text
                    existing_group.created_date = get_ist_now()
                    
                    print(f"🔄 Updated existing group: {group_id}")
                else:
                    # INSERT new group
                    group = Group(
                        project_id=project_id,
                        cable_id=cable_id,
                        group_id=group_id,
                        terminal_no=terminal_no,
                        input_output=input_output,
                        text=text,
                        created_date=get_ist_now()
                    )
                    db.session.add(group)
                    print(f"➕ Created new group: {group_id}")
                
                groups_saved.append(group_id)
                total_groups += 1
                
            except Exception as e:
                print(f"⚠️ WARNING: Error saving group {group_item}: {str(e)}")
                continue
        
        # Delete groups that are no longer in the current data
        current_group_ids = []
        for idx, group_item in enumerate(group_data):
            group_id = group_item.get('group_id', '')
            if not group_id:
                group_id = f"GR{idx + 1:03d}"
            current_group_ids.append(group_id)
        
        stale_groups = Group.query.filter(
            Group.project_id == project_id,
            Group.cable_id == cable_id,
            ~Group.group_id.in_(current_group_ids)
        ).all()
        
        for stale_group in stale_groups:
            db.session.delete(stale_group)
            print(f"🗑️ Deleted stale group: {stale_group.group_id}")
        
        # Update or create GroupSummary
        summary = GroupSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Count all groups for this project
            all_groups = Group.query.filter(
                Group.project_id == project_id
            ).count()
            
            summary.total_groups = all_groups
            summary.updated_at = get_ist_now()
        else:
            # Create new summary with groups
            summary = GroupSummary(
                project_id=project_id,
                total_groups=total_groups,
                updated_at=get_ist_now()
            )
            db.session.add(summary)
        
        db.session.commit()
        
        print(f"✅ DEBUG: Saved/Updated {total_groups} groups to Group table")
        
        return jsonify({
            'success': True,
            'message': f'Group table saved successfully ({total_groups} groups)',
            'total_groups': total_groups,
            'groups_saved': groups_saved
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving group table: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving group table: {str(e)}'}), 500


@bp.route('/finalize_group_table', methods=['POST'])
@login_required
def finalize_group_table():
    """Convert draft groups to final (remove DRAFT- prefix)"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        cable_id = data.get('cable_id')
        
        if not project_id or not cable_id:
            return jsonify({'success': False, 'message': 'Missing parameters'}), 400
        
        # Get all draft groups for this cable
        draft_groups = Group.query.filter(
            Group.project_id == project_id,
            Group.cable_id == cable_id,
            Group.group_id.like('DRAFT-%')
        ).all()
        
        if not draft_groups:
            return jsonify({'success': False, 'message': 'No draft groups found'}), 400
        
        # Update each group to remove DRAFT- prefix
        for group in draft_groups:
            # Remove DRAFT- prefix from group_id
            if group.group_id.startswith('DRAFT-'):
                group.group_id = group.group_id[6:]  # Remove 'DRAFT-' prefix
        
        # Update GroupSummary with final counts (excluding drafts)
        summary = GroupSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Count final groups (excluding drafts)
            final_groups_count = Group.query.filter(
                Group.project_id == project_id,
                ~Group.group_id.like('DRAFT-%')
            ).count()
            
            summary.total_groups = final_groups_count
            summary.updated_at = get_ist_now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Group table finalized ({len(draft_groups)} groups)',
            'groups_finalized': len(draft_groups),
            'total_groups': final_groups_count
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error finalizing group table: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error finalizing group table: {str(e)}'}), 500


@bp.route('/get_group_table_draft', methods=['GET'])
@login_required
def get_group_table_draft():
    """Load group table draft from Group table"""
    try:
        project_id = get_current_project()
        cable_id = request.args.get('cable_id')
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not cable_id:
            return jsonify({'success': False, 'message': 'Cable ID required'}), 400
        
        # Load draft groups from Group table
        draft_groups = Group.query.filter(
            Group.project_id == project_id,
            Group.cable_id == cable_id,
            Group.group_id.like('DRAFT-%')
        ).order_by(Group.group_id).all()
        
        # Reconstruct group data from draft groups
        group_data = []
        for group in draft_groups:
            # Remove DRAFT- prefix for display
            display_group_id = group.group_id[6:] if group.group_id.startswith('DRAFT-') else group.group_id
            
            group_data.append({
                'cable_id': group.cable_id,
                'group_id': display_group_id,
                'terminal_no': group.terminal_no,
                'input_output': group.input_output,
                'text': group.text
            })
        
        print(f"🔍 DEBUG: Found {len(group_data)} draft groups in Group table")
        
        return jsonify({
            'success': True,
            'group_data': group_data,
            'has_draft': len(group_data) > 0
        })
        
    except Exception as e:
        print(f"❌ Error loading group table draft: {str(e)}")
        return jsonify({'success': False, 'message': f'Error loading group table draft: {str(e)}'}), 500


@bp.route('/clear_group_table_draft', methods=['POST'])
@login_required
def clear_group_table_draft():
    """Clear group table draft from Group table"""
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        cable_id = data.get('cable_id')
        
        if not all([project_id, cable_id]):
            return jsonify({'success': False, 'message': 'Missing parameters'}), 400
        
        print(f"🔍 DEBUG clear_group_table_draft: project_id={project_id}, cable_id={cable_id}")
        
        # Delete draft groups from Group table
        deleted_groups_count = Group.query.filter(
            Group.project_id == project_id,
            Group.cable_id == cable_id,
            Group.group_id.like('DRAFT-%')
        ).delete(synchronize_session=False)
        
        # Update GroupSummary
        summary = GroupSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Recalculate total by counting only non-draft groups
            total_groups = Group.query.filter(
                Group.project_id == project_id,
                ~Group.group_id.like('DRAFT-%')
            ).count()
            
            summary.total_groups = total_groups
            summary.updated_at = get_ist_now()
        
        db.session.commit()
        
        print(f"✅ DEBUG: Cleared {deleted_groups_count} draft groups from Group table")
        
        return jsonify({
            'success': True,
            'message': f'Group table draft cleared successfully ({deleted_groups_count} groups removed)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error clearing group table draft: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/get_group_summary', methods=['GET'])
@login_required
def get_group_summary():
    """Get current group summary"""
    try:
        project_id = get_current_project()
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        summary = GroupSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if not summary:
            return jsonify({
                'success': True,
                'total_groups': 0,
                'updated_at': None
            })
        
        return jsonify({
            'success': True,
            'total_groups': summary.total_groups,
            'updated_at': summary.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"❌ Error getting group summary: {str(e)}")
        return jsonify({'success': False, 'message': f'Error getting group summary: {str(e)}'}), 500
    
@bp.route('/save_choke_table_draft', methods=['POST'])
@login_required
def save_choke_table_draft():
    """Save choke table configuration - UPDATE IF EXISTS, INSERT IF NEW"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        
        print(f"🔍 DEBUG save_choke_table_draft: project_id={project_id}, data={data}")
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not data or 'cable_id' not in data or 'choke_data' not in data:
            return jsonify({'success': False, 'message': 'Invalid data format'}), 400
        
        cable_id = data['cable_id']
        choke_data = data['choke_data']
        
        print(f"🔍 DEBUG: Saving/updating choke table for cable {cable_id} with {len(choke_data)} chokes")
        
        # Track totals
        total_chokes = 0
        chokes_saved = []
        validation_errors = []
        
        # Process each choke item - UPDATE if exists, INSERT if new
        for idx, choke_item in enumerate(choke_data):
            try:
                # Extract data from choke_item
                choke_id = choke_item.get('choke_id', '')
                input_terminal = choke_item.get('input_terminal', '')
                output_terminal = choke_item.get('output_terminal', '')
                terminal_name = choke_item.get('terminal_name', 'CHOKE')
                output_type = choke_item.get('output_type', '')
                output_text = choke_item.get('output_text', '')
                output_connected = choke_item.get('output_connected', '')
                
                # Validate output_connected format (optional)
                if output_connected:
                    # Check if it's comma-separated numbers
                    parts = output_connected.split(',')
                    for part in parts:
                        part = part.strip()
                        if part and not part.isdigit():
                            validation_errors.append(f"Row {idx + 1}: Invalid terminal number '{part}' in output_connected")
                
                # Generate unique choke ID if not provided
                if not choke_id:
                    # Find the next available choke ID
                    existing_chokes = ChokeTable.query.filter(
                        ChokeTable.project_id == project_id,
                        ChokeTable.cable_id == cable_id
                    ).count()
                    choke_id = f'CH{existing_chokes + 1:03d}'
                
                # Check if this choke already exists
                existing_choke = ChokeTable.query.filter(
                    ChokeTable.project_id == project_id,
                    ChokeTable.cable_id == cable_id,
                    ChokeTable.choke_id == choke_id
                ).first()
                
                if existing_choke:
                    # UPDATE existing choke
                    existing_choke.input_terminal = input_terminal
                    existing_choke.output_terminal = output_terminal
                    existing_choke.terminal_name = terminal_name
                    existing_choke.output_type = output_type
                    existing_choke.output_text = output_text
                    existing_choke.output_connected = output_connected
                    existing_choke.created_date = get_ist_now()
                    
                    print(f"🔄 Updated existing choke: {choke_id}")
                else:
                    # INSERT new choke
                    choke = ChokeTable(
                        project_id=project_id,
                        cable_id=cable_id,
                        choke_id=choke_id,
                        input_terminal=input_terminal,
                        output_terminal=output_terminal,
                        terminal_name=terminal_name,
                        output_type=output_type,
                        output_text=output_text,
                        output_connected=output_connected,
                        created_date=get_ist_now()
                    )
                    db.session.add(choke)
                    print(f"➕ Created new choke: {choke_id}")
                
                chokes_saved.append(choke_id)
                total_chokes += 1
                
            except Exception as e:
                print(f"⚠️ WARNING: Error saving choke {choke_item}: {str(e)}")
                validation_errors.append(f"Row {idx + 1}: {str(e)}")
                continue
        
        if validation_errors:
            return jsonify({
                'success': False,
                'message': 'Validation errors',
                'errors': validation_errors
            }), 400
        
        # Delete chokes that are no longer in the current data
        current_choke_ids = []
        for idx, choke_item in enumerate(choke_data):
            choke_id = choke_item.get('choke_id', '')
            if not choke_id:
                choke_id = f"CH{idx + 1:03d}"
            current_choke_ids.append(choke_id)
        
        stale_chokes = ChokeTable.query.filter(
            ChokeTable.project_id == project_id,
            ChokeTable.cable_id == cable_id,
            ~ChokeTable.choke_id.in_(current_choke_ids)
        ).all()
        
        for stale_choke in stale_chokes:
            db.session.delete(stale_choke)
            print(f"🗑️ Deleted stale choke: {stale_choke.choke_id}")
        
        # Update or create ChokeSummary
        summary = ChokeSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Count all chokes for this project
            all_chokes = ChokeTable.query.filter(
                ChokeTable.project_id == project_id
            ).count()
            
            summary.total_chokes = all_chokes
            summary.updated_at = get_ist_now()
        else:
            # Create new summary with chokes
            summary = ChokeSummary(
                project_id=project_id,
                total_chokes=total_chokes,
                updated_at=get_ist_now()
            )
            db.session.add(summary)
        
        db.session.commit()
        
        print(f"✅ DEBUG: Saved/Updated {total_chokes} chokes to ChokeTable")
        
        return jsonify({
            'success': True,
            'message': f'Choke table saved successfully ({total_chokes} chokes)',
            'total_chokes': total_chokes,
            'chokes_saved': chokes_saved
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving choke table: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving choke table: {str(e)}'}), 500
    
@bp.route('/finalize_choke_table', methods=['POST'])
@login_required
def finalize_choke_table():
    """Convert draft chokes to final (remove DRAFT- prefix)"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        cable_id = data.get('cable_id')
        
        if not project_id or not cable_id:
            return jsonify({'success': False, 'message': 'Missing parameters'}), 400
        
        # Get all draft chokes for this cable
        draft_chokes = ChokeTable.query.filter(
            ChokeTable.project_id == project_id,
            ChokeTable.cable_id == cable_id,
            ChokeTable.choke_id.like('DRAFT-%')
        ).all()
        
        if not draft_chokes:
            return jsonify({'success': False, 'message': 'No draft chokes found'}), 400
        
        # Update each choke to remove DRAFT- prefix
        for choke in draft_chokes:
            # Remove DRAFT- prefix from choke_id
            if choke.choke_id.startswith('DRAFT-'):
                choke.choke_id = choke.choke_id[6:]  # Remove 'DRAFT-' prefix
        
        # Update ChokeSummary with final counts (excluding drafts)
        summary = ChokeSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Count final chokes (excluding drafts)
            final_chokes_count = ChokeTable.query.filter(
                ChokeTable.project_id == project_id,
                ~ChokeTable.choke_id.like('DRAFT-%')
            ).count()
            
            summary.total_chokes = final_chokes_count
            summary.updated_at = get_ist_now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Choke table finalized ({len(draft_chokes)} chokes)',
            'chokes_finalized': len(draft_chokes),
            'total_chokes': final_chokes_count
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error finalizing choke table: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error finalizing choke table: {str(e)}'}), 500


@bp.route('/get_choke_table_draft', methods=['GET'])
@login_required
def get_choke_table_draft():
    """Load choke table draft from ChokeTable"""
    try:
        project_id = get_current_project()
        cable_id = request.args.get('cable_id')
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not cable_id:
            return jsonify({'success': False, 'message': 'Cable ID required'}), 400
        
        # Load draft chokes from ChokeTable
        draft_chokes = ChokeTable.query.filter(
            ChokeTable.project_id == project_id,
            ChokeTable.cable_id == cable_id,
            ChokeTable.choke_id.like('DRAFT-%')
        ).order_by(ChokeTable.choke_id).all()
        
        # Reconstruct choke data from draft chokes
        choke_data = []
        for choke in draft_chokes:
            # Remove DRAFT- prefix for display
            display_choke_id = choke.choke_id[6:] if choke.choke_id.startswith('DRAFT-') else choke.choke_id
            
            choke_data.append({
                'cable_id': choke.cable_id,
                'choke_id': display_choke_id,
                'input_terminal': choke.input_terminal,
                'output_terminal': choke.output_terminal,
                'terminal_name': choke.terminal_name,
                'output_type': choke.output_type,
                'output_text': choke.output_text,
                'output_connected': choke.output_connected
            })
        
        print(f"🔍 DEBUG: Found {len(choke_data)} draft chokes in ChokeTable")
        
        return jsonify({
            'success': True,
            'choke_data': choke_data,
            'has_draft': len(choke_data) > 0
        })
        
    except Exception as e:
        print(f"❌ Error loading choke table draft: {str(e)}")
        return jsonify({'success': False, 'message': f'Error loading choke table draft: {str(e)}'}), 500


@bp.route('/clear_choke_table_draft', methods=['POST'])
@login_required
def clear_choke_table_draft():
    """Clear choke table draft from ChokeTable"""
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        cable_id = data.get('cable_id')
        
        if not all([project_id, cable_id]):
            return jsonify({'success': False, 'message': 'Missing parameters'}), 400
        
        print(f"🔍 DEBUG clear_choke_table_draft: project_id={project_id}, cable_id={cable_id}")
        
        # Delete draft chokes from ChokeTable
        deleted_chokes_count = ChokeTable.query.filter(
            ChokeTable.project_id == project_id,
            ChokeTable.cable_id == cable_id,
            ChokeTable.choke_id.like('DRAFT-%')
        ).delete(synchronize_session=False)
        
        # Update ChokeSummary
        summary = ChokeSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Recalculate total by counting only non-draft chokes
            total_chokes = ChokeTable.query.filter(
                ChokeTable.project_id == project_id,
                ~ChokeTable.choke_id.like('DRAFT-%')
            ).count()
            
            summary.total_chokes = total_chokes
            summary.updated_at = get_ist_now()
        
        db.session.commit()
        
        print(f"✅ DEBUG: Cleared {deleted_chokes_count} draft chokes from ChokeTable")
        
        return jsonify({
            'success': True,
            'message': f'Choke table draft cleared successfully ({deleted_chokes_count} chokes removed)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error clearing choke table draft: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/get_choke_summary', methods=['GET'])
@login_required
def get_choke_summary():
    """Get current choke summary"""
    try:
        project_id = get_current_project()
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        summary = ChokeSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if not summary:
            return jsonify({
                'success': True,
                'total_chokes': 0,
                'updated_at': None
            })
        
        return jsonify({
            'success': True,
            'total_chokes': summary.total_chokes,
            'updated_at': summary.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"❌ Error getting choke summary: {str(e)}")
        return jsonify({'success': False, 'message': f'Error getting choke summary: {str(e)}'}), 500
    
@bp.route('/save_resistor_table_draft', methods=['POST'])
@login_required
def save_resistor_table_draft():
    """Save resistor table configuration - UPDATE IF EXISTS, INSERT IF NEW"""
    try:
        data = request.get_json()
        project_id = get_current_project()
        
        print(f"🔍 DEBUG save_resistor_table_draft: project_id={project_id}, data={data}")
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not data or 'cable_id' not in data or 'resistor_data' not in data:
            return jsonify({'success': False, 'message': 'Invalid data format'}), 400
        
        cable_id = data['cable_id']
        resistor_data = data['resistor_data']
        
        print(f"🔍 DEBUG: Saving/updating resistor table for cable {cable_id} with {len(resistor_data)} resistors")
        
        # Track totals
        total_resistors = 0
        resistors_saved = []
        
        # Process each resistor item - UPDATE if exists, INSERT if new
        for resistor_item in resistor_data:
            try:
                # Extract data from resistor_item
                resistor_id = resistor_item.get('resistor_id', '')
                input_terminal = resistor_item.get('input_terminal', '')
                output_terminal = resistor_item.get('output_terminal', '')
                resistor_name = resistor_item.get('resistor_name', 'R')
                
                # Generate unique resistor ID if not provided
                if not resistor_id:
                    # Find the next available resistor ID
                    existing_resistors = ResistorTable.query.filter(
                        ResistorTable.project_id == project_id,
                        ResistorTable.cable_id == cable_id
                    ).count()
                    resistor_id = f'R{existing_resistors + 1:03d}'
                
                # Check if this resistor already exists
                existing_resistor = ResistorTable.query.filter(
                    ResistorTable.project_id == project_id,
                    ResistorTable.cable_id == cable_id,
                    ResistorTable.resistor_id == resistor_id
                ).first()
                
                if existing_resistor:
                    # UPDATE existing resistor
                    existing_resistor.input_terminal = input_terminal
                    existing_resistor.output_terminal = output_terminal
                    existing_resistor.resistor_name = resistor_name
                    existing_resistor.created_date = get_ist_now()
                    
                    print(f"🔄 Updated existing resistor: {resistor_id}")
                else:
                    # INSERT new resistor
                    resistor = ResistorTable(
                        project_id=project_id,
                        cable_id=cable_id,
                        resistor_id=resistor_id,
                        input_terminal=input_terminal,
                        output_terminal=output_terminal,
                        resistor_name=resistor_name,
                        created_date=get_ist_now()
                    )
                    db.session.add(resistor)
                    print(f"➕ Created new resistor: {resistor_id}")
                
                resistors_saved.append(resistor_id)
                total_resistors += 1
                
            except Exception as e:
                print(f"⚠️ WARNING: Error saving resistor {resistor_item}: {str(e)}")
                continue
        
        # Delete resistors that are no longer in the current data
        current_resistor_ids = []
        for idx, resistor_item in enumerate(resistor_data):
            resistor_id = resistor_item.get('resistor_id', '')
            if not resistor_id:
                resistor_id = f"R{idx + 1:03d}"
            current_resistor_ids.append(resistor_id)
        
        stale_resistors = ResistorTable.query.filter(
            ResistorTable.project_id == project_id,
            ResistorTable.cable_id == cable_id,
            ~ResistorTable.resistor_id.in_(current_resistor_ids)
        ).all()
        
        for stale_resistor in stale_resistors:
            db.session.delete(stale_resistor)
            print(f"🗑️ Deleted stale resistor: {stale_resistor.resistor_id}")
        
        # Update or create ResistorSummary
        summary = ResistorSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Count all resistors for this project
            all_resistors = ResistorTable.query.filter(
                ResistorTable.project_id == project_id
            ).count()
            
            summary.total_resistors = all_resistors
            summary.updated_at = get_ist_now()
        else:
            # Create new summary with resistors
            summary = ResistorSummary(
                project_id=project_id,
                total_resistors=total_resistors,
                updated_at=get_ist_now()
            )
            db.session.add(summary)
        
        db.session.commit()
        
        print(f"✅ DEBUG: Saved/Updated {total_resistors} resistors to ResistorTable")
        
        return jsonify({
            'success': True,
            'message': f'Resistor table saved successfully ({total_resistors} resistors)',
            'total_resistors': total_resistors,
            'resistors_saved': resistors_saved
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error saving resistor table: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saving resistor table: {str(e)}'}), 500


@bp.route('/get_resistor_table_draft', methods=['GET'])
@login_required
def get_resistor_table_draft():
    """Load resistor table data"""
    try:
        project_id = get_current_project()
        cable_id = request.args.get('cable_id')
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        if not cable_id:
            return jsonify({'success': False, 'message': 'Cable ID required'}), 400
        
        # Load resistors for this cable
        resistors = ResistorTable.query.filter(
            ResistorTable.project_id == project_id,
            ResistorTable.cable_id == cable_id
        ).order_by(ResistorTable.resistor_id).all()
        
        # Reconstruct resistor data
        resistor_data = []
        for resistor in resistors:
            resistor_data.append({
                'cable_id': resistor.cable_id,
                'resistor_id': resistor.resistor_id,
                'input_terminal': resistor.input_terminal,
                'output_terminal': resistor.output_terminal,
                'resistor_name': resistor.resistor_name
            })
        
        print(f"🔍 DEBUG: Found {len(resistor_data)} resistors in ResistorTable")
        
        return jsonify({
            'success': True,
            'resistor_data': resistor_data,
            'has_data': len(resistor_data) > 0
        })
        
    except Exception as e:
        print(f"❌ Error loading resistor table data: {str(e)}")
        return jsonify({'success': False, 'message': f'Error loading resistor table data: {str(e)}'}), 500


@bp.route('/get_resistor_summary', methods=['GET'])
@login_required
def get_resistor_summary():
    """Get current resistor summary"""
    try:
        project_id = get_current_project()
        
        if not project_id:
            return jsonify({'success': False, 'message': 'No project selected'}), 400
        
        summary = ResistorSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if not summary:
            return jsonify({
                'success': True,
                'total_resistors': 0,
                'updated_at': None
            })
        
        return jsonify({
            'success': True,
            'total_resistors': summary.total_resistors,
            'updated_at': summary.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"❌ Error getting resistor summary: {str(e)}")
        return jsonify({'success': False, 'message': f'Error getting resistor summary: {str(e)}'}), 500


@bp.route('/clear_resistor_table_draft', methods=['POST'])
@login_required
def clear_resistor_table_draft():
    """Clear resistor table data for a cable"""
    try:
        data = request.get_json()
        project_id = data.get('project_id')
        cable_id = data.get('cable_id')
        
        if not all([project_id, cable_id]):
            return jsonify({'success': False, 'message': 'Missing parameters'}), 400
        
        print(f"🔍 DEBUG clear_resistor_table_draft: project_id={project_id}, cable_id={cable_id}")
        
        # Delete resistors for this cable
        deleted_resistors_count = ResistorTable.query.filter(
            ResistorTable.project_id == project_id,
            ResistorTable.cable_id == cable_id
        ).delete(synchronize_session=False)
        
        # Update ResistorSummary
        summary = ResistorSummary.query.filter_by(
            project_id=project_id
        ).first()
        
        if summary:
            # Count remaining resistors
            total_resistors = ResistorTable.query.filter(
                ResistorTable.project_id == project_id
            ).count()
            
            summary.total_resistors = total_resistors
            summary.updated_at = get_ist_now()
        
        db.session.commit()
        
        print(f"✅ DEBUG: Cleared {deleted_resistors_count} resistors from ResistorTable")
        
        return jsonify({
            'success': True,
            'message': f'Resistor table cleared successfully ({deleted_resistors_count} resistors removed)'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error clearing resistor table: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500
    
# ============================================================================
# FIXED API ROUTES WITH DEBUGGING
# ============================================================================

# Add this for debugging requests
@bp.before_app_request
def log_request_info():
    """Log all API requests for debugging"""
    if request.path.startswith('/api/'):
        try:
            current_app.logger.info(f"\n{'='*60}")
            current_app.logger.info(f"API Request: {request.method} {request.path}")
            current_app.logger.info(f"Headers: {dict(request.headers)}")
            
            if request.is_json:
                data = request.get_json(silent=True) or {}
                current_app.logger.info(f"JSON Data: {json.dumps(data, indent=2)}")
            elif request.form:
                current_app.logger.info(f"Form Data: {dict(request.form)}")
            elif request.data:
                try:
                    current_app.logger.info(f"Raw Data: {json.loads(request.data)}")
                except:
                    current_app.logger.info(f"Raw Data: {request.data[:500]}")
            
            current_app.logger.info(f"{'='*60}\n")
        except Exception as e:
            current_app.logger.error(f"Error logging request: {str(e)}")


@bp.route('/api/project/<int:project_id>/cables', methods=['GET'])
@login_required
def api_get_cables(project_id):
    """Get all cables for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        cables = Cable.query.filter_by(project_id=project_id).order_by(Cable.cable_id).all()
        
        cable_data = []
        for cable in cables:
            cable_data.append({
                'id': cable.id,  # Make sure we include the database ID
                'cable_id': cable.cable_id,
                'cable_name': cable.cable_name,
                'junction_name': cable.junction_name,
                'row': cable.row,
                'terminal': cable.terminal,
                'start_no': cable.start_no,
                'created_date': cable.created_date.isoformat() if cable.created_date else None
            })
        
        current_app.logger.info(f"API /cables: Returning {len(cable_data)} records")
        return jsonify(cable_data)
    
    except Exception as e:
        current_app.logger.error(f"API /cables ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/terminals', methods=['GET'])
@login_required
def api_get_terminals(project_id):
    """Get all terminals for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        terminals = Terminal.query.filter_by(project_id=project_id).order_by(
            Terminal.cable_id, Terminal.terminal_id
        ).all()
        
        terminal_data = []
        for terminal in terminals:
            terminal_data.append({
                'id': terminal.id,
                'cable_id': terminal.cable_id,
                'terminal_id': terminal.terminal_id,
                'terminal_no': terminal.terminal_no,
                'symbol': terminal.symbol,
                'spare': terminal.spare,
                'input_connected': terminal.input_connected,
                'output_connected': terminal.output_connected,
                'created_date': terminal.created_date.isoformat() if terminal.created_date else None
            })
        
        current_app.logger.info(f"API /terminals: Returning {len(terminal_data)} records")
        return jsonify(terminal_data)
    
    except Exception as e:
        current_app.logger.error(f"API /terminals ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/headers', methods=['GET'])
@login_required
def api_get_headers(project_id):
    """Get all terminal headers for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        headers = TerminalHeader.query.filter_by(project_id=project_id).order_by(
            TerminalHeader.cable_id
        ).all()
        
        header_data = []
        for header in headers:
            header_data.append({
                'id': header.id,
                'cable_id': header.cable_id,
                'header_type': header.header_type,
                'terminal_start': header.terminal_start,
                'terminal_end': header.terminal_end,
                'input_output': header.input_output,
                'text': header.text,
                'created_date': header.created_date.isoformat() if header.created_date else None
            })
        
        current_app.logger.info(f"API /headers: Returning {len(header_data)} records")
        return jsonify(header_data)
    
    except Exception as e:
        current_app.logger.error(f"API /headers ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/groups', methods=['GET'])
@login_required
def api_get_groups(project_id):
    """Get all groups for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        groups = Group.query.filter_by(project_id=project_id).order_by(
            Group.cable_id, Group.group_id
        ).all()
        
        group_data = []
        for group in groups:
            group_data.append({
                'id': group.id,
                'cable_id': group.cable_id,
                'group_id': group.group_id,
                'terminal_no': group.terminal_no,
                'input_output': group.input_output,
                'text': group.text,
                'created_date': group.created_date.isoformat() if group.created_date else None
            })
        
        current_app.logger.info(f"API /groups: Returning {len(group_data)} records")
        return jsonify(group_data)
    
    except Exception as e:
        current_app.logger.error(f"API /groups ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/chokes', methods=['GET'])
@login_required
def api_get_chokes(project_id):
    """Get all chokes for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        chokes = ChokeTable.query.filter_by(project_id=project_id).order_by(
            ChokeTable.cable_id, ChokeTable.choke_id
        ).all()
        
        choke_data = []
        for choke in chokes:
            choke_data.append({
                'id': choke.id,
                'cable_id': choke.cable_id,
                'choke_id': choke.choke_id,
                'input_terminal': choke.input_terminal,
                'output_terminal': choke.output_terminal,
                'terminal_name': choke.terminal_name,
                'output_type': choke.output_type,
                'created_date': choke.created_date.isoformat() if choke.created_date else None
            })
        
        current_app.logger.info(f"API /chokes: Returning {len(choke_data)} records")
        return jsonify(choke_data)
    
    except Exception as e:
        current_app.logger.error(f"API /chokes ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/resistors', methods=['GET'])
@login_required
def api_get_resistors(project_id):
    """Get all resistors for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        resistors = ResistorTable.query.filter_by(project_id=project_id).order_by(
            ResistorTable.cable_id, ResistorTable.resistor_id
        ).all()
        
        resistor_data = []
        for resistor in resistors:
            resistor_data.append({
                'id': resistor.id,
                'cable_id': resistor.cable_id,
                'resistor_id': resistor.resistor_id,
                'input_terminal': resistor.input_terminal,
                'output_terminal': resistor.output_terminal,
                'resistor_name': resistor.resistor_name,
                'created_date': resistor.created_date.isoformat() if resistor.created_date else None
            })
        
        current_app.logger.info(f"API /resistors: Returning {len(resistor_data)} records")
        return jsonify(resistor_data)
    
    except Exception as e:
        current_app.logger.error(f"API /resistors ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/junctions', methods=['GET'])
@login_required
def api_get_junctions(project_id):
    """Get all junction boxes for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        junctions = JunctionBox.query.filter_by(project_id=project_id).order_by(
            JunctionBox.junction_id
        ).all()
        
        junction_data = []
        for junction in junctions:
            junction_data.append({
                'id': junction.id,
                'junction_id': junction.junction_id,
                'junction_name': junction.junction_name,
                'junction_size': junction.junction_size,
                'junction_row': junction.junction_row,
                'latitude': str(junction.latitude) if junction.latitude else '',
                'longitude': str(junction.longitude) if junction.longitude else '',
                'created_date': junction.created_date.isoformat() if junction.created_date else None
            })
        
        current_app.logger.info(f"API /junctions: Returning {len(junction_data)} records")
        return jsonify(junction_data)
    
    except Exception as e:
        current_app.logger.error(f"API /junctions ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/station', methods=['GET'])
@login_required
def api_get_station(project_id):
    """Get station drawing information for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        station = StationDrawing.query.filter_by(project_id=project_id).first()
        
        if not station:
            current_app.logger.info(f"API /station: No station data found")
            return jsonify([])
        
        # Return as array of field-value pairs for display
        station_data = [
            {'field': 'Station ID', 'value': station.station_id or '-'},
            {'field': 'Station Name', 'value': station.station_name or '-'},
            {'field': 'Station Code', 'value': station.station_code or '-'},
            {'field': 'Version', 'value': station.version or '-'},
            {'field': 'Zone', 'value': station.zone or '-'},
            {'field': 'Division', 'value': station.division or '-'},
            {'field': 'Date', 'value': station.date or '-'},
            {'field': 'Drawn By', 'value': station.drawn_by or '-'},
            {'field': 'Checked By', 'value': station.checked_by or '-'}
        ]
        
        current_app.logger.info(f"API /station: Returning station info")
        return jsonify(station_data)
    
    except Exception as e:
        current_app.logger.error(f"API /station ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/station', methods=['POST', 'PUT'])
@login_required
def api_update_station(project_id):
    """Create or update station data"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Updating station for project {project_id}: {data}")
        
        # Check if station exists
        station = StationDrawing.query.filter_by(project_id=project_id).first()
        
        if station:
            # Update existing station
            station.station_id = data.get('station_id', station.station_id)
            station.station_name = data.get('station_name', station.station_name)
            station.station_code = data.get('station_code', station.station_code)
            station.version = data.get('version', station.version)
            station.zone = data.get('zone', station.zone)
            station.division = data.get('division', station.division)
            station.date = data.get('date', station.date)
            station.drawn_by = data.get('drawn_by', station.drawn_by)
            station.checked_by = data.get('checked_by', station.checked_by)
            station.updated_date = datetime.utcnow()
        else:
            # Create new station
            station = StationDrawing(
                project_id=project_id,
                station_id=data.get('station_id', ''),
                station_name=data.get('station_name', ''),
                station_code=data.get('station_code', ''),
                version=data.get('version', ''),
                zone=data.get('zone', ''),
                division=data.get('division', ''),
                date=data.get('date', ''),
                drawn_by=data.get('drawn_by', ''),
                checked_by=data.get('checked_by', '')
            )
            db.session.add(station)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Station data saved successfully',
            'id': station.id
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating station: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ============================================================================
# CREATE, UPDATE, DELETE ROUTES
# ============================================================================

# Cables routes
@bp.route('/api/project/<int:project_id>/cables', methods=['POST'])
@login_required
def api_create_cable(project_id):
    """Create a new cable"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Creating cable: {data}")
        
        # Validate required fields
        required_fields = ['cable_id', 'cable_name']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Check if cable_id already exists
        existing = Cable.query.filter_by(
            project_id=project_id, 
            cable_id=data['cable_id']
        ).first()
        if existing:
            return jsonify({'error': f'Cable with ID {data["cable_id"]} already exists'}), 400
        
        # Create new cable
        cable = Cable(
            project_id=project_id,
            cable_id=data['cable_id'],
            cable_name=data.get('cable_name', ''),
            junction_name=data.get('junction_name', ''),
            row=data.get('row', ''),
            terminal=data.get('terminal', ''),
            start_no=data.get('start_no')
        )
        
        db.session.add(cable)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cable created successfully',
            'id': cable.id,
            'cable_id': cable.cable_id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating cable: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/cables/<int:record_id>', methods=['PUT'])
@login_required
def api_update_cable(project_id, record_id):
    """Update a cable"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        cable = Cable.query.filter_by(project_id=project_id, id=record_id).first()
        if not cable:
            return jsonify({'error': 'Cable not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Updating cable {record_id}: {data}")
        
        # Update fields (don't update cable_id as it's an identifier)
        if 'cable_name' in data:
            cable.cable_name = data['cable_name']
        if 'junction_name' in data:
            cable.junction_name = data['junction_name']
        if 'row' in data:
            cable.row = data['row']
        if 'terminal' in data:
            cable.terminal = data['terminal']
        if 'start_no' in data:
            cable.start_no = data['start_no']
        
        cable.updated_date = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cable updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating cable: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/cables/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_cable(project_id, record_id):
    """Delete a cable"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        cable = Cable.query.filter_by(project_id=project_id, id=record_id).first()
        if not cable:
            return jsonify({'error': 'Cable not found'}), 404        
        current_app.logger.info(f"Deleting cable {record_id} ({cable.cable_id})")
        
        db.session.delete(cable)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cable deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting cable: {str(e)}")
        return jsonify({'error': str(e)}), 500


# Terminals routes
@bp.route('/api/project/<int:project_id>/terminals', methods=['POST'])
@login_required
def api_create_terminal(project_id):
    """Create a new terminal"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Creating terminal: {data}")
        
        # Validate required fields
        required_fields = ['cable_id', 'terminal_id', 'terminal_no']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Check if terminal_id already exists for this cable
        existing = Terminal.query.filter_by(
            project_id=project_id, 
            cable_id=data['cable_id'],
            terminal_id=data['terminal_id']
        ).first()
        if existing:
            return jsonify({'error': f'Terminal {data["terminal_id"]} already exists for cable {data["cable_id"]}'}), 400
        
        # Create new terminal
        terminal = Terminal(
            project_id=project_id,
            cable_id=data['cable_id'],
            terminal_id=data['terminal_id'],
            terminal_no=data.get('terminal_no', ''),
            symbol=data.get('symbol', ''),
            spare=data.get('spare', 'No'),
            input_connected=data.get('input_connected', ''),
            output_connected=data.get('output_connected', '')
        )
        
        db.session.add(terminal)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Terminal created successfully',
            'id': terminal.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating terminal: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/terminals/<int:record_id>', methods=['PUT'])
@login_required
def api_update_terminal(project_id, record_id):
    """Update a terminal"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        terminal = Terminal.query.filter_by(project_id=project_id, id=record_id).first()
        if not terminal:
            return jsonify({'error': 'Terminal not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Updating terminal {record_id}: {data}")
        
        # Update fields (don't update cable_id or terminal_id as they're identifiers)
        if 'terminal_no' in data:
            terminal.terminal_no = data['terminal_no']
        if 'symbol' in data:
            terminal.symbol = data['symbol']
        if 'spare' in data:
            terminal.spare = data['spare']
        if 'input_connected' in data:
            terminal.input_connected = data['input_connected']
        if 'output_connected' in data:
            terminal.output_connected = data['output_connected']
        
        terminal.updated_date = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Terminal updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating terminal: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/terminals/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_terminal(project_id, record_id):
    """Delete a terminal"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        terminal = Terminal.query.filter_by(project_id=project_id, id=record_id).first()
        if not terminal:
            return jsonify({'error': 'Terminal not found'}), 404
        
        current_app.logger.info(f"Deleting terminal {record_id} ({terminal.terminal_id})")
        
        db.session.delete(terminal)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Terminal deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting terminal: {str(e)}")
        return jsonify({'error': str(e)}), 500


# Headers routes
@bp.route('/api/project/<int:project_id>/headers', methods=['POST'])
@login_required
def api_create_header(project_id):
    """Create a new header"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Creating header: {data}")
        
        # Validate required fields
        required_fields = ['cable_id', 'header_type']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Create new header
        header = TerminalHeader(
            project_id=project_id,
            cable_id=data['cable_id'],
            header_type=data['header_type'],
            terminal_start=data.get('terminal_start', ''),
            terminal_end=data.get('terminal_end', ''),
            input_output=data.get('input_output', 'Input'),
            text=data.get('text', '')
        )
        
        db.session.add(header)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Header created successfully',
            'id': header.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating header: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/headers/<int:record_id>', methods=['PUT'])
@login_required
def api_update_header(project_id, record_id):
    """Update a header"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        header = TerminalHeader.query.filter_by(project_id=project_id, id=record_id).first()
        if not header:
            return jsonify({'error': 'Header not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Updating header {record_id}: {data}")
        
        # Update fields
        if 'header_type' in data:
            header.header_type = data['header_type']
        if 'terminal_start' in data:
            header.terminal_start = data['terminal_start']
        if 'terminal_end' in data:
            header.terminal_end = data['terminal_end']
        if 'input_output' in data:
            header.input_output = data['input_output']
        if 'text' in data:
            header.text = data['text']
        
        header.updated_date = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Header updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating header: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/headers/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_header(project_id, record_id):
    """Delete a header"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        header = TerminalHeader.query.filter_by(project_id=project_id, id=record_id).first()
        if not header:
            return jsonify({'error': 'Header not found'}), 404
        
        current_app.logger.info(f"Deleting header {record_id}")
        
        db.session.delete(header)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Header deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting header: {str(e)}")
        return jsonify({'error': str(e)}), 500


# Groups routes
@bp.route('/api/project/<int:project_id>/groups', methods=['POST'])
@login_required
def api_create_group(project_id):
    """Create a new group"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Creating group: {data}")
        
        # Validate required fields
        required_fields = ['cable_id', 'group_id']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Create new group
        group = Group(
            project_id=project_id,
            cable_id=data['cable_id'],
            group_id=data['group_id'],
            terminal_no=data.get('terminal_no', ''),
            input_output=data.get('input_output', 'Input'),
            text=data.get('text', '')
        )
        
        db.session.add(group)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Group created successfully',
            'id': group.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating group: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/groups/<int:record_id>', methods=['PUT'])
@login_required
def api_update_group(project_id, record_id):
    """Update a group"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        group = Group.query.filter_by(project_id=project_id, id=record_id).first()
        if not group:
            return jsonify({'error': 'Group not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Updating group {record_id}: {data}")
        
        # Update fields
        if 'group_id' in data:
            group.group_id = data['group_id']
        if 'terminal_no' in data:
            group.terminal_no = data['terminal_no']
        if 'input_output' in data:
            group.input_output = data['input_output']
        if 'text' in data:
            group.text = data['text']
        
        group.updated_date = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Group updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating group: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/groups/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_group(project_id, record_id):
    """Delete a group"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        group = Group.query.filter_by(project_id=project_id, id=record_id).first()
        if not group:
            return jsonify({'error': 'Group not found'}), 404
        
        current_app.logger.info(f"Deleting group {record_id} ({group.group_id})")
        
        db.session.delete(group)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Group deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting group: {str(e)}")
        return jsonify({'error': str(e)}), 500


# Chokes routes
@bp.route('/api/project/<int:project_id>/chokes', methods=['POST'])
@login_required
def api_create_choke(project_id):
    """Create a new choke"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Creating choke: {data}")
        
        # Validate required fields
        required_fields = ['cable_id', 'choke_id']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Create new choke
        choke = ChokeTable(
            project_id=project_id,
            cable_id=data['cable_id'],
            choke_id=data['choke_id'],
            input_terminal=data.get('input_terminal', ''),
            output_terminal=data.get('output_terminal', ''),
            terminal_name=data.get('terminal_name', ''),
            output_type=data.get('output_type', '')
        )
        
        db.session.add(choke)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Choke created successfully',
            'id': choke.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating choke: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/chokes/<int:record_id>', methods=['PUT'])
@login_required
def api_update_choke(project_id, record_id):
    """Update a choke"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        choke = ChokeTable.query.filter_by(project_id=project_id, id=record_id).first()
        if not choke:
            return jsonify({'error': 'Choke not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Updating choke {record_id}: {data}")
        
        # Update fields
        if 'choke_id' in data:
            choke.choke_id = data['choke_id']
        if 'input_terminal' in data:
            choke.input_terminal = data['input_terminal']
        if 'output_terminal' in data:
            choke.output_terminal = data['output_terminal']
        if 'terminal_name' in data:
            choke.terminal_name = data['terminal_name']
        if 'output_type' in data:
            choke.output_type = data['output_type']
        
        choke.updated_date = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Choke updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating choke: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/chokes/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_choke(project_id, record_id):
    """Delete a choke"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        choke = ChokeTable.query.filter_by(project_id=project_id, id=record_id).first()
        if not choke:
            return jsonify({'error': 'Choke not found'}), 404
        
        current_app.logger.info(f"Deleting choke {record_id} ({choke.choke_id})")
        
        db.session.delete(choke)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Choke deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting choke: {str(e)}")
        return jsonify({'error': str(e)}), 500


# Resistors routes
@bp.route('/api/project/<int:project_id>/resistors', methods=['POST'])
@login_required
def api_create_resistor(project_id):
    """Create a new resistor"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Creating resistor: {data}")
        
        # Validate required fields
        required_fields = ['cable_id', 'resistor_id']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Create new resistor
        resistor = ResistorTable(
            project_id=project_id,
            cable_id=data['cable_id'],
            resistor_id=data['resistor_id'],
            input_terminal=data.get('input_terminal', ''),
            output_terminal=data.get('output_terminal', ''),
            resistor_name=data.get('resistor_name', '')
        )
        
        db.session.add(resistor)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Resistor created successfully',
            'id': resistor.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating resistor: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/resistors/<int:record_id>', methods=['PUT'])
@login_required
def api_update_resistor(project_id, record_id):
    """Update a resistor"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        resistor = ResistorTable.query.filter_by(project_id=project_id, id=record_id).first()
        if not resistor:
            return jsonify({'error': 'Resistor not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Updating resistor {record_id}: {data}")
        
        # Update fields
        if 'resistor_id' in data:
            resistor.resistor_id = data['resistor_id']
        if 'input_terminal' in data:
            resistor.input_terminal = data['input_terminal']
        if 'output_terminal' in data:
            resistor.output_terminal = data['output_terminal']
        if 'resistor_name' in data:
            resistor.resistor_name = data['resistor_name']
        
        resistor.updated_date = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Resistor updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating resistor: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/resistors/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_resistor(project_id, record_id):
    """Delete a resistor"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        resistor = ResistorTable.query.filter_by(project_id=project_id, id=record_id).first()
        if not resistor:
            return jsonify({'error': 'Resistor not found'}), 404
        
        current_app.logger.info(f"Deleting resistor {record_id} ({resistor.resistor_id})")
        
        db.session.delete(resistor)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Resistor deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting resistor: {str(e)}")
        return jsonify({'error': str(e)}), 500


# Junctions routes
@bp.route('/api/project/<int:project_id>/junctions', methods=['POST'])
@login_required
def api_create_junction(project_id):
    """Create a new junction"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Creating junction: {data}")
        
        # Validate required fields
        required_fields = ['junction_id', 'junction_name']
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Check if junction_id already exists
        existing = JunctionBox.query.filter_by(
            project_id=project_id, 
            junction_id=data['junction_id']
        ).first()
        if existing:
            return jsonify({'error': f'Junction with ID {data["junction_id"]} already exists'}), 400
        
        # Create new junction
        junction = JunctionBox(
            project_id=project_id,
            junction_id=data['junction_id'],
            junction_name=data['junction_name'],
            junction_size=data.get('junction_size', ''),
            junction_row=data.get('junction_row', ''),
            latitude=data.get('latitude'),
            longitude=data.get('longitude')
        )
        
        db.session.add(junction)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Junction created successfully',
            'id': junction.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating junction: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/junctions/<int:record_id>', methods=['PUT'])
@login_required
def api_update_junction(project_id, record_id):
    """Update a junction"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        junction = JunctionBox.query.filter_by(project_id=project_id, id=record_id).first()
        if not junction:
            return jsonify({'error': 'Junction not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        current_app.logger.info(f"Updating junction {record_id}: {data}")
        
        # Update fields (don't update junction_id as it's an identifier)
        if 'junction_name' in data:
            junction.junction_name = data['junction_name']
        if 'junction_size' in data:
            junction.junction_size = data['junction_size']
        if 'junction_row' in data:
            junction.junction_row = data['junction_row']
        if 'latitude' in data:
            junction.latitude = data['latitude']
        if 'longitude' in data:
            junction.longitude = data['longitude']
        
        junction.updated_date = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Junction updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating junction: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/project/<int:project_id>/junctions/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_junction(project_id, record_id):
    """Delete a junction"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        junction = JunctionBox.query.filter_by(project_id=project_id, id=record_id).first()
        if not junction:
            return jsonify({'error': 'Junction not found'}), 404
        
        current_app.logger.info(f"Deleting junction {record_id} ({junction.junction_id})")
        
        db.session.delete(junction)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Junction deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting junction: {str(e)}")
        return jsonify({'error': str(e)}), 500


# ============================================================================
# DEBUG ENDPOINT
# ============================================================================

@bp.route('/api/debug/project/<int:project_id>', methods=['GET'])
@login_required
def api_debug_project(project_id):
    """Debug endpoint - shows all data for a project"""
    try:
        if not user_has_project_access(project_id):
            return jsonify({'error': 'Access denied'}), 403
        
        return jsonify({
            'project_id': project_id,
            'cables': Cable.query.filter_by(project_id=project_id).count(),
            'terminals': Terminal.query.filter_by(project_id=project_id).count(),
            'headers': TerminalHeader.query.filter_by(project_id=project_id).count(),
            'groups': Group.query.filter_by(project_id=project_id).count(),
            'chokes': ChokeTable.query.filter_by(project_id=project_id).count(),
            'resistors': ResistorTable.query.filter_by(project_id=project_id).count(),
            'junctions': JunctionBox.query.filter_by(project_id=project_id).count(),
            'station': StationDrawing.query.filter_by(project_id=project_id).count()
        })
    
    except Exception as e:
        current_app.logger.error(f"DEBUG ERROR: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
# Delete All Data

@bp.route('/delete_projects/<int:project_id>', methods=['POST'])
@login_required
def delete_projects(project_id):
    """Delete all related data for a project but keep the project itself"""
    # Only admin users can delete project data
    if current_user.role_name != '4':
        return jsonify({'success': False, 'message': 'Unauthorized: Only admin users can delete project data'}), 403
    
    try:
        # Get the project (but don't delete it)
        project = Project.query.get_or_404(project_id)
        project_name = project.name
        
        # Delete all related data in correct order to avoid foreign key constraints
        
        # IMPORTANT: First, delete JunctionApproval ,junction_box_summary records to avoid NULL constraint violation
        
        #JunctionApproval.query.filter_by(project_id=project_id).delete()
        #JunctionBoxSummary.query.filter_by(project_id=project_id).delete()
       
        # 1. Delete notifications for this project
        # Notification.query.filter_by(project_id=project_id).delete()
        
        # # 2. Delete generated PDFs
        GeneratedPDF.query.filter_by(project_id=project_id).delete()
        
        # 3. Delete from user_projects association table
        # db.session.execute(
        #     user_projects.delete().where(user_projects.c.project_id == project_id)
        # )
        
        # 4. Delete station master records
        #StationMaster.query.filter_by(project_id=project_id).delete()
        
        # 5. Delete cable box records
        CableBox.query.filter_by(project_id=project_id).delete()
        
        # 7. Delete junction boxes
        # JunctionBox.query.filter_by(project_id=project_id).delete()
        
        # 8. Delete cables
        Cable.query.filter_by(project_id=project_id).delete()
        
        # 9. Delete terminals
        Terminal.query.filter_by(project_id=project_id).delete()
        
        # 10. Delete groups
        Group.query.filter_by(project_id=project_id).delete()
        
        # 11. Delete terminal headers
        TerminalHeader.query.filter_by(project_id=project_id).delete()
        
        # 12. Delete choke table records
        ChokeTable.query.filter_by(project_id=project_id).delete()
        
        # 13. Delete resistor table records
        ResistorTable.query.filter_by(project_id=project_id).delete()
        
        # Delete station drawings first
        StationDrawing.query.filter_by(project_id=project.id).delete()
        
        # IMPORTANT: Do NOT delete the project itself
        db.session.delete(project)  # REMOVED THIS LINE
        
        # Commit all changes
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'All data for project "{project_name}" has been cleared successfully. The project record remains.'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting project data {project_id}: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error clearing project data: {str(e)}'
        }), 500
        
@bp.route('/project/<int:project_id>/continue-draft')
@login_required
def continue_draft(project_id):
    """Continue working on a project draft from its current stage"""
    project = Project.query.get_or_404(project_id)
    
    # Clear any existing location filter
    if 'current_location_id' in session:
        session.pop('current_location_id')
    
    # If no stage is set, start from stage 1
    if not project.stage or project.stage == 0:
        return redirect(url_for('main.stage_data_view', project_id=project_id, stage=1))
    
    # If stage is 9 or 10 (Cable Box or PDF generated), go to stage 9
    if project.stage >= 9:
        return redirect(url_for('main.stage_data_view', project_id=project_id, stage=9))
    
    # Otherwise, go to the current stage
    return redirect(url_for('main.stage_data_view', project_id=project_id, stage=project.stage))

@bp.route('/project/<int:project_id>/continue-draft/<int:location_id>')
@login_required
def continue_draft_with_location(project_id, location_id):
    """Continue working on a project draft from a specific location"""
    project = Project.query.get_or_404(project_id)
    
    # Get the specific location
    location = JunctionBox.query.filter_by(
        project_id=project_id, 
        id=location_id
    ).first_or_404()
    
    # Store location_id in session to filter data
    session['current_location_id'] = location_id
    session['current_location_name'] = location.junction_name
    
    # If no stage is set, start from stage 1
    if not project.stage or project.stage == 0:
        return redirect(url_for('main.stage_data_view', project_id=project_id, stage=1))
    
    # If stage is 9 or 10 (Cable Box or PDF generated), go to stage 9
    if project.stage >= 9:
        return redirect(url_for('main.stage_data_view', project_id=project_id, stage=9))
    
    # Otherwise, go to the current stage
    return redirect(url_for('main.stage_data_view', project_id=project_id, stage=project.stage))

@bp.route('/project/<int:project_id>/stage/<int:stage>/view')
@login_required
def stage_data_view(project_id, stage):
    """View and edit data for a specific stage"""
    project = Project.query.get_or_404(project_id)
    
    # Validate stage number
    if stage < 1 or stage > 9:
        flash('Invalid stage number', 'error')
        return redirect(url_for('main.approval_tracking'))
    
    # Fix: Update the session with the correct project_id for this view
    session['current_project_id'] = project_id
    
    # Get all cables for filtering dropdown
    cables = Cable.query.filter_by(project_id=project_id).all()
    
    # Check if we have a location_id in the query parameters (from popup click)
    location_id_from_query = request.args.get('location_id')
    
    # Variables to pass to template
    current_location_id = None
    current_location_name = None
    
    # If location_id is in query params, use it and store in session
    if location_id_from_query:
        try:
            # Store as string to handle both ID and junction_id
            current_location_id = str(location_id_from_query)
            session['current_location_id'] = current_location_id
            
            # Try to get location by ID (primary key)
            try:
                location_id_int = int(location_id_from_query)
                location = JunctionBox.query.filter_by(
                    project_id=project_id, 
                    id=location_id_int
                ).first()
            except ValueError:
                # If not a valid integer, try by junction_id
                location = JunctionBox.query.filter_by(
                    project_id=project_id,
                    junction_id=location_id_from_query
                ).first()
            
            if location:
                current_location_name = location.junction_name
                session['current_location_name'] = current_location_name
            else:
                current_location_name = f"Location {current_location_id}"
                session['current_location_name'] = current_location_name
        except Exception as e:
            print(f"Error processing location_id: {e}")
            # Clear location filters if there's an error
            if 'current_location_id' in session:
                session.pop('current_location_id')
            if 'current_location_name' in session:
                session.pop('current_location_name')
    # Otherwise, check if we have location_id in session
    else:
        if 'current_location_id' in session:
            current_location_id = session.get('current_location_id')
            current_location_name = session.get('current_location_name', '')
    
    return render_template('workflow/stage_data_view.html', 
                         project=project, 
                         current_stage=stage,
                         cables=cables,
                         current_location_id=current_location_id,
                         current_location_name=current_location_name)

@bp.route('/project/<int:project_id>/stage/<int:stage>/api-data')
@login_required
def get_stage_api_data(project_id, stage):
    """Get JSON data for a specific stage"""
    try:
        project = Project.query.get_or_404(project_id)
        cable_id = request.args.get('cable_id')
        junction_box = request.args.get('junction_box')  
        junction_name = request.args.get('junction_name')
        location_id = request.args.get('location_id')  # Get location_id from query params
        
        print(f"DEBUG: Loading stage {stage} data for project {project_id}, cable_id: {cable_id}, location_id: {location_id}")
        
        data = {}
        counts = {}
        location_name = None
        
        # Helper function to get junction details for location filtering
        def get_junction_for_location(loc_id):
            if loc_id:
                try:
                    # Convert to integer and get junction
                    junction = JunctionBox.query.filter_by(
                        project_id=project_id,
                        id=int(loc_id)
                    ).first()
                    return junction
                except (ValueError, TypeError):
                    return None
            return None
        
        # Get location name if location_id is provided
        if location_id:
            junction = get_junction_for_location(location_id)
            if junction:
                location_name = junction.junction_name
                print(f"DEBUG: Found location: {junction.junction_name} (ID: {junction.id})")
        
        # Stage 1: Station Info (no location filtering)
        if stage == 1:
            station_master = StationMaster.query.filter_by(project_id=project_id).first()
            station_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
            
            data = {
                'station_master': station_master.to_dict() if station_master else None,
                'station_drawing': station_drawing.to_dict() if station_drawing else None
            }
            counts = {
                'station_master': 1 if station_master else 0,
                'station_drawing': 1 if station_drawing else 0
            }
        
        # Stage 2: Locations (Junction Boxes)
        elif stage == 2:
            query = JunctionBox.query.filter_by(project_id=project_id)
            
            if location_id:
                # Get only the specific location by ID
                junction = get_junction_for_location(location_id)
                if junction:
                    query = query.filter_by(id=junction.id)
            
            junction_boxes = query.all()
            print(f"DEBUG: Found {len(junction_boxes)} junction boxes")
            
            data = {
                'junction_boxes': [jb.to_dict() for jb in junction_boxes]
            }
            counts = {
                'junction_boxes': len(junction_boxes)
            }
        
        # Stage 3: Cables
        elif stage == 3:
            query = Cable.query.filter_by(project_id=project_id)
            
            if location_id:
                # Get junction details for this location
                junction = get_junction_for_location(location_id)
                if junction:
                    # Filter by both junction_box AND junction_name for precise matching
                    query = query.filter_by(
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    )
            elif cable_id:
                query = query.filter_by(cable_id=cable_id)
            elif junction_box and junction_name:
                query = query.filter_by(junction_box=junction_box, junction_name=junction_name)
            elif junction_box:
                query = query.filter_by(junction_box=junction_box)
            
            cables = query.all()
            print(f"DEBUG: Found {len(cables)} cables")
            data = {
                'cables': [cable.to_dict() for cable in cables]
            }
            counts = {
                'cables': len(cables)
            }
        
        # Stage 4: Terminals
        elif stage == 4:
            if location_id:
                # Get junction details
                junction = get_junction_for_location(location_id)
                if junction:
                    # Get all cables for this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    ).all()
                    cable_ids = [cable.cable_id for cable in junction_cables]
                    
                    # Get terminals for these cables
                    terminals = Terminal.query.filter(
                        Terminal.project_id == project_id,
                        Terminal.cable_id.in_(cable_ids)
                    ).all()
                else:
                    terminals = []
            elif cable_id:
                terminals = Terminal.query.filter_by(project_id=project_id, cable_id=cable_id).all()
            else:
                terminals = Terminal.query.filter_by(project_id=project_id).all()
            
            print(f"DEBUG: Found {len(terminals)} terminals")
            data = {
                'terminals': [terminal.to_dict() for terminal in terminals]
            }
            counts = {
                'terminals': len(terminals)
            }
        
        # Stage 5-8: Similar filtering for other stages
        elif stage in [5, 6, 7, 8]:
            model_map = {
                5: (TerminalHeader, 'headers'),
                6: (Group, 'groups'),
                7: (ChokeTable, 'chokes'),
                8: (ResistorTable, 'resistors')
            }
            
            model, data_key = model_map[stage]
            query = model.query.filter_by(project_id=project_id)
            
            # Apply location filter if provided
            if location_id:
                junction = get_junction_for_location(location_id)
                if junction:
                    # Get all cables for this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    ).all()
                    cable_ids = [cable.cable_id for cable in junction_cables]
                    
                    if cable_ids:
                        # Filter by these cable IDs
                        query = query.filter(model.cable_id.in_(cable_ids))
                    else:
                        query = query.filter(model.id == -1)  # Return empty
                else:
                    query = query.filter(model.id == -1)  # Return empty
            elif cable_id:
                query = query.filter_by(cable_id=cable_id)
            
            items = query.all()
            print(f"DEBUG: Found {len(items)} {data_key}")
            data = {
                data_key: [item.to_dict() for item in items]
            }
            counts = {
                data_key: len(items)
            }
        
        # Stage 9: Cable Box (Relay Box) - FIXED
        # Stage 9: Cable Box (Relay Box) - FIXED
        elif stage == 9:
            query = CableBox.query.filter_by(project_id=project_id)
            
            # Apply location filter if provided
            if location_id:
                junction = get_junction_for_location(location_id)
                if junction:
                    print(f"🔍 DEBUG STAGE 9: Filtering for location_id={location_id}")
                    print(f"🔍 DEBUG STAGE 9: Junction found - id={junction.id}, junction_id='{junction.junction_id}', junction_name='{junction.junction_name}'")
                    
                    # Get count before filtering
                    all_count = CableBox.query.filter_by(project_id=project_id).count()
                    print(f"🔍 DEBUG STAGE 9: Total CableBox records in project: {all_count}")
                    
                    # Check what we're filtering for
                    print(f"🔍 DEBUG STAGE 9: Looking for CableBox with: junction_box='{junction.junction_id}', junction_name='{junction.junction_name}'")
                    
                    # Apply filter - IMPORTANT: junction.junction_id might be string or integer
                    query = query.filter(
                        CableBox.junction_box == str(junction.junction_id),
                        CableBox.junction_name == str(junction.junction_name)
                    )
                    
                    # Get SQL to debug
                    print(f"🔍 DEBUG STAGE 9: Filter SQL: {query}")
                else:
                    print(f"🔍 DEBUG STAGE 9: No junction found for location_id={location_id}")
                    query = query.filter(CableBox.id == -1)
            elif cable_id:
                query = query.filter_by(cable_id=cable_id)
            
            cable_boxes = query.all()
            print(f"🔍 DEBUG STAGE 9: Query returned {len(cable_boxes)} cable boxes")
            
            # Debug what we got
            for cb in cable_boxes:
                print(f"🔍 DEBUG STAGE 9: Result - CableBox {cb.id}: cable_id={cb.cable_id}, junction_box='{cb.junction_box}', junction_name='{cb.junction_name}'")
            
            data = {
                'cable_boxes': [cb.to_dict() for cb in cable_boxes]
            }
            counts = {
                'cable_boxes': len(cable_boxes)
            }
        else:
            return jsonify({
                'success': False,
                'error': f'Invalid stage number: {stage}. Stage must be between 1 and 9.'
            }), 400
        
        # Check if any data exists
        has_data = False
        if isinstance(data, dict):
            for value in data.values():
                if value and (isinstance(value, list) and len(value) > 0) or (value is not None):
                    has_data = True
                    break
        
        print(f"DEBUG: Has data: {has_data}, Location ID: {location_id}, Location Name: {location_name}")
        
        # Make sure location_id is defined for the response
        response_location_id = location_id
        
        return jsonify({
            'success': True,
            'stage': stage,
            'project_id': project_id,
            'project_name': project.name,
            'project_stage': project.stage,
            'cable_id': cable_id,
            'location_id': response_location_id,  # Use the defined variable
            'location_name': location_name,  # Include location_name in response
            'data': data,
            'counts': counts,
            'has_data': has_data
        })
        
    except Exception as e:
        print(f"ERROR in get_stage_api_data: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'error': f'Internal server error: {str(e)}',
            'traceback': traceback.format_exc() if current_app.debug else None
        }), 500



@bp.route('/project/<int:project_id>/clear-location-filter', methods=['POST'])
@login_required
def clear_location_filter(project_id):
    """Clear the location filter from session"""
    stage = request.args.get('stage', 1)
    
    if 'current_location_id' in session:
        session.pop('current_location_id')
    if 'current_location_name' in session:
        session.pop('current_location_name')
    
    return jsonify({'success': True, 'redirect_url': f'/project/{project_id}/stage/{stage}/view'})


# ==================== CRUD OPERATIONS ====================

@bp.route('/project/<int:project_id>/stage/1/station-master/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def station_master_crud(project_id, item_id):
    """CRUD operations for station master"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        # Get station master data
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        station_master = StationMaster.query.filter_by(id=item_id, project_id=project_id).first()
        if not station_master:
            return jsonify({'success': False, 'error': 'Station master not found'}), 404
        
        return jsonify({'success': True, 'data': station_master.to_dict()})
    
    elif request.method == 'POST':
        # Create or update station master
        data = request.form
        
        if item_id == 'new':
            station_master = StationMaster(
                project_id=project_id,
                station_id=data.get('station_id'),
                station_name=data.get('station_name'),
                station_code=data.get('station_code'),
                remarks=data.get('remarks')
            )
            db.session.add(station_master)
            message = 'Station master created successfully'
        else:
            station_master = StationMaster.query.filter_by(id=item_id, project_id=project_id).first()
            if not station_master:
                return jsonify({'success': False, 'error': 'Station master not found'}), 404
            
            station_master.station_id = data.get('station_id')
            station_master.station_name = data.get('station_name')
            station_master.station_code = data.get('station_code')
            station_master.remarks = data.get('remarks')
            message = 'Station master updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        station_master = StationMaster.query.filter_by(id=item_id, project_id=project_id).first()
        if not station_master:
            return jsonify({'success': False, 'error': 'Station master not found'}), 404
        
        db.session.delete(station_master)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Station master deleted successfully'})

@bp.route('/project/<int:project_id>/stage/1/station-drawing/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def station_drawing_crud(project_id, item_id):
    """CRUD operations for station drawing"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        station_drawing = StationDrawing.query.filter_by(id=item_id, project_id=project_id).first()
        if not station_drawing:
            return jsonify({'success': False, 'error': 'Station drawing not found'}), 404
        
        return jsonify({'success': True, 'data': station_drawing.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            station_drawing = StationDrawing(
                project_id=project_id,
                diagram_name=data.get('diagram_name'),
                station_name=data.get('station_name'),
                station_code=data.get('station_code'),
                version=data.get('version'),
                drawn_by=data.get('drawn_by'),
                checked_by=data.get('checked_by'),
                remarks=data.get('remarks')
            )
            db.session.add(station_drawing)
            message = 'Station drawing created successfully'
        else:
            station_drawing = StationDrawing.query.filter_by(id=item_id, project_id=project_id).first()
            if not station_drawing:
                return jsonify({'success': False, 'error': 'Station drawing not found'}), 404
            
            station_drawing.diagram_name = data.get('diagram_name')
            station_drawing.station_name = data.get('station_name')
            station_drawing.station_code = data.get('station_code')
            station_drawing.version = data.get('version')
            station_drawing.drawn_by = data.get('drawn_by')
            station_drawing.checked_by = data.get('checked_by')
            station_drawing.remarks = data.get('remarks')
            message = 'Station drawing updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        station_drawing = StationDrawing.query.filter_by(id=item_id, project_id=project_id).first()
        if not station_drawing:
            return jsonify({'success': False, 'error': 'Station drawing not found'}), 404
        
        db.session.delete(station_drawing)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Station drawing deleted successfully'})

@bp.route('/project/<int:project_id>/stage/2/junction-box/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def junction_box_crud(project_id, item_id):
    """CRUD operations for junction box"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        junction_box = JunctionBox.query.filter_by(id=item_id, project_id=project_id).first()
        if not junction_box:
            return jsonify({'success': False, 'error': 'Junction box not found'}), 404
        
        return jsonify({'success': True, 'data': junction_box.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            junction_box = JunctionBox(
                project_id=project_id,
                junction_id=data.get('junction_id'),
                junction_name=data.get('junction_name'),
                junction_size=data.get('junction_size'),
                junction_row=data.get('junction_row'),
                latitude=data.get('latitude'),
                longitude=data.get('longitude'),
                remarks=data.get('remarks')
            )
            db.session.add(junction_box)
            message = 'Junction box created successfully'
        else:
            junction_box = JunctionBox.query.filter_by(id=item_id, project_id=project_id).first()
            if not junction_box:
                return jsonify({'success': False, 'error': 'Junction box not found'}), 404
            
            junction_box.junction_id = data.get('junction_id')
            junction_box.junction_name = data.get('junction_name')
            junction_box.junction_size = data.get('junction_size')
            junction_box.junction_row = data.get('junction_row')
            junction_box.latitude = data.get('latitude')
            junction_box.longitude = data.get('longitude')
            junction_box.remarks = data.get('remarks')
            message = 'Junction box updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        junction_box = JunctionBox.query.filter_by(id=item_id, project_id=project_id).first()
        if not junction_box:
            return jsonify({'success': False, 'error': 'Junction box not found'}), 404
        
        db.session.delete(junction_box)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Junction box deleted successfully'})

@bp.route('/project/<int:project_id>/stage/3/cable/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def cable_crud(project_id, item_id):
    """CRUD operations for cable"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        cable = Cable.query.filter_by(id=item_id, project_id=project_id).first()
        if not cable:
            return jsonify({'success': False, 'error': 'Cable not found'}), 404
        
        return jsonify({'success': True, 'data': cable.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            cable = Cable(
                project_id=project_id,
                cable_id=data.get('cable_id'),
                cable_name=data.get('cable_name'),
                junction_box=data.get('junction_box'),
                row=data.get('row'),
                terminal=data.get('terminal'),
                start_no=data.get('start_no'),
                remarks=data.get('remarks')
            )
            db.session.add(cable)
            message = 'Cable created successfully'
        else:
            cable = Cable.query.filter_by(id=item_id, project_id=project_id).first()
            if not cable:
                return jsonify({'success': False, 'error': 'Cable not found'}), 404
            
            cable.cable_id = data.get('cable_id')
            cable.cable_name = data.get('cable_name')
            cable.junction_box = data.get('junction_box')
            cable.row = data.get('row')
            cable.terminal = data.get('terminal')
            cable.start_no = data.get('start_no')
            cable.remarks = data.get('remarks')
            message = 'Cable updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        cable = Cable.query.filter_by(id=item_id, project_id=project_id).first()
        if not cable:
            return jsonify({'success': False, 'error': 'Cable not found'}), 404
        
        db.session.delete(cable)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Cable deleted successfully'})

@bp.route('/project/<int:project_id>/stage/4/terminal/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def terminal_crud(project_id, item_id):
    """CRUD operations for terminal"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        terminal = Terminal.query.filter_by(id=item_id, project_id=project_id).first()
        if not terminal:
            return jsonify({'success': False, 'error': 'Terminal not found'}), 404
        
        return jsonify({'success': True, 'data': terminal.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            terminal = Terminal(
                project_id=project_id,
                cable_id=data.get('cable_id'),
                terminal_id=data.get('terminal_id'),
                terminal_no=data.get('terminal_no'),
                symbol=data.get('symbol'),
                spare=data.get('spare'),
                input_connected=data.get('input_connected'),
                output_connected=data.get('output_connected'),
                remarks=data.get('remarks')
            )
            db.session.add(terminal)
            message = 'Terminal created successfully'
        else:
            terminal = Terminal.query.filter_by(id=item_id, project_id=project_id).first()
            if not terminal:
                return jsonify({'success': False, 'error': 'Terminal not found'}), 404
            
            terminal.cable_id = data.get('cable_id')
            terminal.terminal_id = data.get('terminal_id')
            terminal.terminal_no = data.get('terminal_no')
            terminal.symbol = data.get('symbol')
            terminal.spare = data.get('spare')
            terminal.input_connected = data.get('input_connected')
            terminal.output_connected = data.get('output_connected')
            terminal.remarks = data.get('remarks')
            message = 'Terminal updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        terminal = Terminal.query.filter_by(id=item_id, project_id=project_id).first()
        if not terminal:
            return jsonify({'success': False, 'error': 'Terminal not found'}), 404
        
        db.session.delete(terminal)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Terminal deleted successfully'})

@bp.route('/project/<int:project_id>/stage/5/terminal-header/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def terminal_header_crud(project_id, item_id):
    """CRUD operations for terminal header"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        header = TerminalHeader.query.filter_by(id=item_id, project_id=project_id).first()
        if not header:
            return jsonify({'success': False, 'error': 'Terminal header not found'}), 404
        
        return jsonify({'success': True, 'data': header.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            header = TerminalHeader(
                project_id=project_id,
                cable_id=data.get('cable_id'),
                header_type=data.get('header_type'),
                terminal_start=data.get('terminal_start'),
                terminal_end=data.get('terminal_end'),
                input_output=data.get('input_output'),
                text=data.get('text')
            )
            db.session.add(header)
            message = 'Terminal header created successfully'
        else:
            header = TerminalHeader.query.filter_by(id=item_id, project_id=project_id).first()
            if not header:
                return jsonify({'success': False, 'error': 'Terminal header not found'}), 404
            
            header.cable_id = data.get('cable_id')
            header.header_type = data.get('header_type')
            header.terminal_start = data.get('terminal_start')
            header.terminal_end = data.get('terminal_end')
            header.input_output = data.get('input_output')
            header.text = data.get('text')
            message = 'Terminal header updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        header = TerminalHeader.query.filter_by(id=item_id, project_id=project_id).first()
        if not header:
            return jsonify({'success': False, 'error': 'Terminal header not found'}), 404
        
        db.session.delete(header)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Terminal header deleted successfully'})

@bp.route('/project/<int:project_id>/stage/6/group/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def group_crud(project_id, item_id):
    """CRUD operations for group"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        group = Group.query.filter_by(id=item_id, project_id=project_id).first()
        if not group:
            return jsonify({'success': False, 'error': 'Group not found'}), 404
        
        return jsonify({'success': True, 'data': group.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            group = Group(
                project_id=project_id,
                cable_id=data.get('cable_id'),
                group_id=data.get('group_id'),
                terminal_no=data.get('terminal_no'),
                input_output=data.get('input_output'),
                text=data.get('text')
            )
            db.session.add(group)
            message = 'Group created successfully'
        else:
            group = Group.query.filter_by(id=item_id, project_id=project_id).first()
            if not group:
                return jsonify({'success': False, 'error': 'Group not found'}), 404
            
            group.cable_id = data.get('cable_id')
            group.group_id = data.get('group_id')
            group.terminal_no = data.get('terminal_no')
            group.input_output = data.get('input_output')
            group.text = data.get('text')
            message = 'Group updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        group = Group.query.filter_by(id=item_id, project_id=project_id).first()
        if not group:
            return jsonify({'success': False, 'error': 'Group not found'}), 404
        
        db.session.delete(group)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Group deleted successfully'})

@bp.route('/project/<int:project_id>/stage/7/choke-table/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def choke_table_crud(project_id, item_id):
    """CRUD operations for choke table"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        choke = ChokeTable.query.filter_by(id=item_id, project_id=project_id).first()
        if not choke:
            return jsonify({'success': False, 'error': 'Choke not found'}), 404
        
        return jsonify({'success': True, 'data': choke.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            choke = ChokeTable(
                project_id=project_id,
                cable_id=data.get('cable_id'),
                choke_id=data.get('choke_id'),
                input_terminal=data.get('input_terminal'),
                output_terminal=data.get('output_terminal'),
                terminal_name=data.get('terminal_name'),
                output_type=data.get('output_type'),
                remarks=data.get('remarks')
            )
            db.session.add(choke)
            message = 'Choke created successfully'
        else:
            choke = ChokeTable.query.filter_by(id=item_id, project_id=project_id).first()
            if not choke:
                return jsonify({'success': False, 'error': 'Choke not found'}), 404
            
            choke.cable_id = data.get('cable_id')
            choke.choke_id = data.get('choke_id')
            choke.input_terminal = data.get('input_terminal')
            choke.output_terminal = data.get('output_terminal')
            choke.terminal_name = data.get('terminal_name')
            choke.output_type = data.get('output_type')
            choke.remarks = data.get('remarks')
            message = 'Choke updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        choke = ChokeTable.query.filter_by(id=item_id, project_id=project_id).first()
        if not choke:
            return jsonify({'success': False, 'error': 'Choke not found'}), 404
        
        db.session.delete(choke)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Choke deleted successfully'})

@bp.route('/project/<int:project_id>/stage/8/resistor-table/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def resistor_table_crud(project_id, item_id):
    """CRUD operations for resistor table"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        resistor = ResistorTable.query.filter_by(id=item_id, project_id=project_id).first()
        if not resistor:
            return jsonify({'success': False, 'error': 'Resistor not found'}), 404
        
        return jsonify({'success': True, 'data': resistor.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            resistor = ResistorTable(
                project_id=project_id,
                cable_id=data.get('cable_id'),
                resistor_id=data.get('resistor_id'),
                input_terminal=data.get('input_terminal'),
                output_terminal=data.get('output_terminal'),
                resistor_name=data.get('resistor_name'),
                resistance_value=data.get('resistance_value'),
                wattage=data.get('wattage'),
                tolerance=data.get('tolerance'),
                remarks=data.get('remarks')
            )
            db.session.add(resistor)
            message = 'Resistor created successfully'
        else:
            resistor = ResistorTable.query.filter_by(id=item_id, project_id=project_id).first()
            if not resistor:
                return jsonify({'success': False, 'error': 'Resistor not found'}), 404
            
            resistor.cable_id = data.get('cable_id')
            resistor.resistor_id = data.get('resistor_id')
            resistor.input_terminal = data.get('input_terminal')
            resistor.output_terminal = data.get('output_terminal')
            resistor.resistor_name = data.get('resistor_name')
            resistor.resistance_value = data.get('resistance_value')
            resistor.wattage = data.get('wattage')
            resistor.tolerance = data.get('tolerance')
            resistor.remarks = data.get('remarks')
            message = 'Resistor updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        resistor = ResistorTable.query.filter_by(id=item_id, project_id=project_id).first()
        if not resistor:
            return jsonify({'success': False, 'error': 'Resistor not found'}), 404
        
        db.session.delete(resistor)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Resistor deleted successfully'})

@bp.route('/project/<int:project_id>/stage/9/cable-box/<item_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def cable_box_crud(project_id, item_id):
    """CRUD operations for cable box (relay box)"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'GET':
        if item_id == 'new':
            return jsonify({'success': True, 'data': None})
        
        cable_box = CableBox.query.filter_by(id=item_id, project_id=project_id).first()
       
        if not cable_box:
            return jsonify({'success': False, 'error': 'Cable box not found'}), 404
        
        return jsonify({'success': True, 'data': cable_box.to_dict()})
    
    elif request.method == 'POST':
        data = request.form
        
        if item_id == 'new':
            cable_box = CableBox(
                project_id=project_id,
                cable_id=data.get('cable_id'),
                cable_name=data.get('cable_name'),
                junction_box=data.get('junction_box'),
                row=data.get('row'),
                terminal=data.get('terminal'),
                start_no=data.get('start_no'),
                cable_type=data.get('cable_type', 'relay_box'),
                junction_name=data.get('junction_name', '')
            )
            db.session.add(cable_box)
            message = 'Cable box created successfully'
        else:
            cable_box = CableBox.query.filter_by(id=item_id, project_id=project_id).first()
            if not cable_box:
                return jsonify({'success': False, 'error': 'Cable box not found'}), 404
            
            cable_box.cable_id = data.get('cable_id')
            cable_box.cable_name = data.get('cable_name')
            cable_box.junction_box = data.get('junction_box')
            cable_box.row = data.get('row')
            cable_box.terminal = data.get('terminal')
            cable_box.start_no = data.get('start_no')
            cable_box.cable_type = data.get('cable_type', 'relay_box')
            cable_box.junction_name = data.get('junction_name', '')
            cable_box.output = data.get('output', '')
            message = 'Cable box updated successfully'
        
        try:
            db.session.commit()
            return jsonify({'success': True, 'message': message})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        cable_box = CableBox.query.filter_by(id=item_id, project_id=project_id).first()
        if not cable_box:
            return jsonify({'success': False, 'error': 'Cable box not found'}), 404
        
        db.session.delete(cable_box)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Cable box deleted successfully'})
# ==================== HELPER ROUTES ====================

@bp.route('/project/<int:project_id>/stage/2/junction-boxes/json')
@login_required
def get_junction_boxes_json(project_id):
    """Get junction boxes for dropdowns"""
    junction_boxes = JunctionBox.query.filter_by(project_id=project_id).all()
    return jsonify({
        'success': True,
        'data': [{
            'junction_id': jb.junction_id,
            'junction_name': jb.junction_name
        } for jb in junction_boxes]
    })

@bp.route('/project/<int:project_id>/stage/3/cables/json')
@login_required
def get_cables_json(project_id):
    """Get cables for dropdowns"""
    cables = Cable.query.filter_by(project_id=project_id).all()
    return jsonify({
        'success': True,
        'data': [{
            'cable_id': cable.cable_id,
            'cable_name': cable.cable_name
        } for cable in cables]
    })

@bp.route('/project/<int:project_id>/stage/3/cables-by-junction/<junction_id>/json')
@login_required
def get_cables_by_junction_json(project_id, junction_id):
    """Get cables for a specific junction"""
    cables = Cable.query.filter_by(project_id=project_id, junction_box=junction_id).all()
    junction = JunctionBox.query.filter_by(project_id=project_id, junction_id=junction_id).first()
    
    return jsonify({
        'success': True,
        'junction_name': junction.junction_name if junction else junction_id,
        'cables': [{
            'cable_id': cable.cable_id,
            'cable_name': cable.cable_name,
            'row': cable.row,
            'terminal': cable.terminal,
            'start_no': cable.start_no
        } for cable in cables]
    })

@bp.route('/project/<int:project_id>/stage/3/cable/<cable_id>/details/json')
@login_required
def get_cable_details_json(project_id, cable_id):
    """Get detailed cable information with terminals"""
    cable = Cable.query.filter_by(project_id=project_id, cable_id=cable_id).first()
    if not cable:
        return jsonify({'success': False, 'error': 'Cable not found'}), 404
    
    terminals = Terminal.query.filter_by(project_id=project_id, cable_id=cable_id).all()
    
    return jsonify({
        'success': True,
        'cable': cable.to_dict() if cable else None,
        'terminals': [terminal.to_dict() for terminal in terminals]
    })

@bp.route('/project/<int:project_id>/all-stages-summary/json')
@login_required
def get_all_stages_summary(project_id):
    """Get summary of all stages"""
    project = Project.query.get_or_404(project_id)
    
    summary = {}
    
    # Stage 1
    station_master = StationMaster.query.filter_by(project_id=project_id).first()
    station_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
    summary['stage_1'] = {
        'count': (1 if station_master else 0) + (1 if station_drawing else 0),
        'last_updated': max(
            station_master.created_date if station_master else datetime.min,
            station_drawing.created_date if station_drawing else datetime.min
        ).strftime('%Y-%m-%d %H:%M:%S') if station_master or station_drawing else 'Not started'
    }
    
    # Stage 2
    junction_count = JunctionBox.query.filter_by(project_id=project_id).count()
    last_junction = JunctionBox.query.filter_by(project_id=project_id).order_by(JunctionBox.created_date.desc()).first()
    summary['stage_2'] = {
        'count': junction_count,
        'last_updated': last_junction.created_date.strftime('%Y-%m-%d %H:%M:%S') if last_junction else 'Not started'
    }
    
    # Stage 3
    cable_count = Cable.query.filter_by(project_id=project_id).count()
    last_cable = Cable.query.filter_by(project_id=project_id).order_by(Cable.created_date.desc()).first()
    summary['stage_3'] = {
        'count': cable_count,
        'last_updated': last_cable.created_date.strftime('%Y-%m-%d %H:%M:%S') if last_cable else 'Not started'
    }
    
    # Stage 4
    terminal_count = Terminal.query.filter_by(project_id=project_id).count()
    last_terminal = Terminal.query.filter_by(project_id=project_id).order_by(Terminal.created_date.desc()).first()
    summary['stage_4'] = {
        'count': terminal_count,
        'last_updated': last_terminal.created_date.strftime('%Y-%m-%d %H:%M:%S') if last_terminal else 'Not started'
    }
    
    # Stage 5
    header_count = TerminalHeader.query.filter_by(project_id=project_id).count()
    last_header = TerminalHeader.query.filter_by(project_id=project_id).order_by(TerminalHeader.created_date.desc()).first()
    summary['stage_5'] = {
        'count': header_count,
        'last_updated': last_header.created_date.strftime('%Y-%m-%d %H:%M:%S') if last_header else 'Not started'
    }
    
    # Stage 6
    group_count = Group.query.filter_by(project_id=project_id).count()
    last_group = Group.query.filter_by(project_id=project_id).order_by(Group.created_date.desc()).first()
    summary['stage_6'] = {
        'count': group_count,
        'last_updated': last_group.created_date.strftime('%Y-%m-%d %H:%M:%S') if last_group else 'Not started'
    }
    
    # Stage 7
    choke_count = ChokeTable.query.filter_by(project_id=project_id).count()
    last_choke = ChokeTable.query.filter_by(project_id=project_id).order_by(ChokeTable.created_date.desc()).first()
    summary['stage_7'] = {
        'count': choke_count,
        'last_updated': last_choke.created_date.strftime('%Y-%m-%d %H:%M:%S') if last_choke else 'Not started'
    }
    
    # Stage 8
    resistor_count = ResistorTable.query.filter_by(project_id=project_id).count()
    last_resistor = ResistorTable.query.filter_by(project_id=project_id).order_by(ResistorTable.created_date.desc()).first()
    summary['stage_8'] = {
        'count': resistor_count,
        'last_updated': last_resistor.created_date.strftime('%Y-%m-%d %H:%M:%S') if last_resistor else 'Not started'
    }
    
    # Stage 9
    cablebox_count = CableBox.query.filter_by(project_id=project_id).count()
    last_cablebox = CableBox.query.filter_by(project_id=project_id).order_by(CableBox.created_date.desc()).first()
    summary['stage_9'] = {
        'count': cablebox_count,
        'last_updated': last_cablebox.created_date.strftime('%Y-%m-%d %H:%M:%S') if last_cablebox else 'Not started'
    }
    
    return jsonify({
        'success': True,
        'project_id': project_id,
        'project_name': project.name,
        **summary
    })


@bp.route('/project/<int:project_id>/comprehensive-view')
@login_required
def comprehensive_view(project_id):
    """Comprehensive view with step-by-step workflow"""
    project = Project.query.get_or_404(project_id)
    cables = Cable.query.filter_by(project_id=project_id).all()
    
    return render_template('workflow/comprehensive_view.html',
                         project=project,
                         cables=cables,
                         current_stage=project.stage or 1)


# ============================================================================
# 📊 FLASK ROUTES FOR STAGE DATA LOADING AND CRUD OPERATIONS
# ============================================================================

@bp.route('/project/<int:project_id>/stage/<int:stage>/api-data', methods=['GET'])
def get_stage_data(project_id, stage):
    """
    Get data for a specific stage of a project
    Optional query parameter: cable_id (for stages 3-9)
    """
    try:
        # Check if project exists
        project = Project.query.get_or_404(project_id)
        
        # Get optional filters
        cable_id = request.args.get('cable_id')
        location_id = request.args.get('location_id')  # Add this line
        
        # Initialize response structure
        response = {
            'success': True,
            'has_data': False,
            'data': {},
            'counts': {}
        }
        
        # Stage-specific data fetching
        if stage == 1:
            # Stage 1: Station Information
            station_master = StationMaster.query.filter_by(project_id=project_id).first()
            station_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
            
            data = {
                'station_master': station_master.to_dict() if station_master else None,
                'station_drawing': station_drawing.to_dict() if station_drawing else None
            }
            
            counts = {
                'station_master': 1 if station_master else 0,
                'station_drawing': 1 if station_drawing else 0
            }
            
            response['data'] = data
            response['counts'] = counts
            response['has_data'] = station_master is not None or station_drawing is not None
            
        # Stage 2: Locations (Junction Boxes)
        elif stage == 2:
            query = JunctionBox.query.filter_by(project_id=project_id)
            
            if location_id:
                # Get only the specific location by ID
                query = query.filter_by(id=location_id)
            
            junction_boxes = query.all()
            print(f"DEBUG: Found {len(junction_boxes)} junction boxes for location_id: {location_id}")
            
            data = {
                'junction_boxes': [jb.to_dict() for jb in junction_boxes]
            }
            counts = {
                'junction_boxes': len(junction_boxes)
            }
            
            response['data'] = data
            response['counts'] = counts
            response['has_data'] = len(junction_boxes) > 0
            
        elif stage == 3:
            # Stage 3: Cables
            query = Cable.query.filter_by(project_id=project_id)
            
            # Apply location filter if provided
            if location_id:
                # Get the junction box for this location
                junction = JunctionBox.query.filter_by(
                    project_id=project_id,
                    id=location_id
                ).first()
                
                if junction:
                    # Filter by both junction_box AND junction_name for precise matching
                    query = query.filter_by(
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    )
            
            if cable_id:
                query = query.filter_by(cable_id=cable_id)
            
            cables = query.order_by(Cable.cable_id).all()
            
            data = {
                'cables': [cable.to_dict() for cable in cables]
            }
            
            counts = {
                'cables': len(cables)
            }
            
            response['data'] = data
            response['counts'] = counts
            response['has_data'] = len(cables) > 0
            
        elif stage == 4:
            # Stage 4: Terminals
            query = Terminal.query.filter_by(project_id=project_id)
            
            print(f"DEBUG STAGE 4: Initial query, location_id={location_id}, cable_id={cable_id}")
            
            # Apply location filter if provided
            if location_id:
                # Get the junction box for this location
                junction = JunctionBox.query.filter_by(
                    project_id=project_id,
                    id=location_id
                ).first()
                
                print(f"DEBUG STAGE 4: Junction found: {junction.junction_id if junction else 'None'}")
                
                if junction:
                    # Get all cables for this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    ).all()
                    
                    print(f"DEBUG STAGE 4: Found {len(junction_cables)} cables for junction")
                    
                    cable_ids = [cable.cable_id for cable in junction_cables]
                    
                    if cable_ids:
                        query = query.filter(Terminal.cable_id.in_(cable_ids))
                        print(f"DEBUG STAGE 4: Filtering by cable_ids: {cable_ids}")
                    else:
                        query = query.filter(Terminal.id == -1)  # Return empty
                        print(f"DEBUG STAGE 4: No cables found, returning empty")
            
            if cable_id:
                query = query.filter_by(cable_id=cable_id)
                print(f"DEBUG STAGE 4: Filtering by cable_id: {cable_id}")
            
            terminals = query.order_by(Terminal.terminal_id).all()
            
            print(f"DEBUG STAGE 4: Found {len(terminals)} terminals")
            for t in terminals:
                print(f"  Terminal: cable_id={t.cable_id}, terminal_id={t.terminal_id}, terminal_no={t.terminal_no}, input_left={t.input_left}")
            
        elif stage == 5:
            # Stage 5: Headers
            query = TerminalHeader.query.filter_by(project_id=project_id)
            
            # Apply location filter if provided
            if location_id:
                # Get the junction box for this location
                junction = JunctionBox.query.filter_by(
                    project_id=project_id,
                    id=location_id
                ).first()
                
                if junction:
                    # Get all cables for this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    ).all()
                    
                    cable_ids = [cable.cable_id for cable in junction_cables]
                    
                    if cable_ids:
                        query = query.filter(TerminalHeader.cable_id.in_(cable_ids))
                    else:
                        query = query.filter(TerminalHeader.id == -1)  # Return empty
            
            if cable_id:
                query = query.filter_by(cable_id=cable_id)
            
            headers = query.order_by(TerminalHeader.terminal_start).all()
            
            data = {
                'headers': [header.to_dict() for header in headers]
            }
            
            counts = {
                'headers': len(headers)
            }
            
            response['data'] = data
            response['counts'] = counts
            response['has_data'] = len(headers) > 0
            
        elif stage == 6:
            # Stage 6: Groups
            query = Group.query.filter_by(project_id=project_id)
            
            # Apply location filter if provided
            if location_id:
                # Get the junction box for this location
                junction = JunctionBox.query.filter_by(
                    project_id=project_id,
                    id=location_id
                ).first()
                
                if junction:
                    # Get all cables for this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    ).all()
                    
                    cable_ids = [cable.cable_id for cable in junction_cables]
                    
                    if cable_ids:
                        query = query.filter(Group.cable_id.in_(cable_ids))
                    else:
                        query = query.filter(Group.id == -1)  # Return empty
            
            if cable_id:
                query = query.filter_by(cable_id=cable_id)
            
            groups = query.order_by(Group.group_id).all()
            
            data = {
                'groups': [group.to_dict() for group in groups]
            }
            
            counts = {
                'groups': len(groups)
            }
            
            response['data'] = data
            response['counts'] = counts
            response['has_data'] = len(groups) > 0
            
        elif stage == 7:
            # Stage 7: Choke Table
            query = ChokeTable.query.filter_by(project_id=project_id)
            
            # Apply location filter if provided
            if location_id:
                # Get the junction box for this location
                junction = JunctionBox.query.filter_by(
                    project_id=project_id,
                    id=location_id
                ).first()
                
                if junction:
                    # Get all cables for this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    ).all()
                    
                    cable_ids = [cable.cable_id for cable in junction_cables]
                    
                    if cable_ids:
                        query = query.filter(ChokeTable.cable_id.in_(cable_ids))
                    else:
                        query = query.filter(ChokeTable.id == -1)  # Return empty
            
            if cable_id:
                query = query.filter_by(cable_id=cable_id)
            
            chokes = query.order_by(ChokeTable.choke_id).all()
            
            data = {
                'chokes': [choke.to_dict() for choke in chokes]
            }
            
            counts = {
                'chokes': len(chokes)
            }
            
            response['data'] = data
            response['counts'] = counts
            response['has_data'] = len(chokes) > 0
            
        elif stage == 8:
            # Stage 8: Resistor Table
            query = ResistorTable.query.filter_by(project_id=project_id)
            
            # Apply location filter if provided
            if location_id:
                # Get the junction box for this location
                junction = JunctionBox.query.filter_by(
                    project_id=project_id,
                    id=location_id
                ).first()
                
                if junction:
                    # Get all cables for this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    ).all()
                    
                    cable_ids = [cable.cable_id for cable in junction_cables]
                    
                    if cable_ids:
                        query = query.filter(ResistorTable.cable_id.in_(cable_ids))
                    else:
                        query = query.filter(ResistorTable.id == -1)  # Return empty
            
            if cable_id:
                query = query.filter_by(cable_id=cable_id)
            
            resistors = query.order_by(ResistorTable.resistor_id).all()
            
            data = {
                'resistors': [resistor.to_dict() for resistor in resistors]
            }
            
            counts = {
                'resistors': len(resistors)
            }
            
            response['data'] = data
            response['counts'] = counts
            response['has_data'] = len(resistors) > 0
            
        elif stage == 9:
            # Stage 9: Cable Box
            query = CableBox.query.filter_by(project_id=project_id)
            
            # Apply location filter if provided
            if location_id:
                # Get the junction box for this location
                junction = JunctionBox.query.filter_by(
                    project_id=project_id,
                    id=location_id
                ).first()
                
                if junction:
                    # Get all cables for this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=junction.junction_id,
                        junction_name=junction.junction_name
                    ).all()
                    
                    cable_ids = [cable.cable_id for cable in junction_cables]
                    
                    if cable_ids:
                        query = query.filter(CableBox.cable_id.in_(cable_ids))
                    else:
                        query = query.filter(CableBox.id == -1)  # Return empty
            
            if cable_id:
                query = query.filter_by(cable_id=cable_id)
            
            cable_boxes = query.order_by(CableBox.cable_id).all()
            
            data = {
                'cable_boxes': [cb.to_dict() for cb in cable_boxes]
            }
            
            counts = {
                'cable_boxes': len(cable_boxes)
            }
            
            response['data'] = data
            response['counts'] = counts
            response['has_data'] = len(cable_boxes) > 0
            
        else:
            return jsonify({
                'success': False,
                'error': f'Stage {stage} not found'
            }), 404
        
        # Add location info to response
        response['location_id'] = location_id
        if location_id:
            # Get location name for the response
            location = JunctionBox.query.filter_by(
                project_id=project_id,
                id=location_id
            ).first()
            if location:
                response['location_name'] = location.junction_name
            else:
                response['location_name'] = None
            
        return jsonify(response)
        
    except Exception as e:
        current_app.logger.error(f'Error getting stage {stage} data for project {project_id}: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error loading data: {str(e)}'
        }), 500


# ============================================================================
# 📋 STEP 5: HEADERS CRUD ROUTES
# ============================================================================

@bp.route('/project/<int:project_id>/stage/5/header/new', methods=['POST'])
def add_header(project_id):
    """Add a new header"""
    try:
        data = request.form.to_dict()
        data['project_id'] = project_id
        
        # Validate required fields
        if not data.get('cable_id') or not data.get('header_type'):
            return jsonify({
                'success': False,
                'error': 'Cable ID and Header Type are required'
            }), 400
        
        header = TerminalHeader(**data)
        db.session.add(header)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Header added successfully',
            'id': header.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error adding header: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error adding header: {str(e)}'
        }), 500

@bp.route('/project/<int:project_id>/stage/5/header/<int:header_id>', methods=['POST', 'DELETE'])
def manage_header(project_id, header_id):
    """Update or delete a header"""
    try:
        header = TerminalHeader.query.get_or_404(header_id)
        
        # Verify ownership
        if header.project_id != project_id:
            return jsonify({
                'success': False,
                'error': 'Header not found in this project'
            }), 404
        
        if request.method == 'POST':
            # Update header
            data = request.form.to_dict()
            for key, value in data.items():
                if hasattr(header, key):
                    setattr(header, key, value)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Header updated successfully'
            })
            
        elif request.method == 'DELETE':
            # Delete header
            db.session.delete(header)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Header deleted successfully'
            })
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error managing header: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error managing header: {str(e)}'
        }), 500


# ============================================================================
# 📋 STEP 6: GROUPS CRUD ROUTES
# ============================================================================

@bp.route('/project/<int:project_id>/stage/6/group/new', methods=['POST'])
def add_group(project_id):
    """Add a new group"""
    try:
        data = request.form.to_dict()
        data['project_id'] = project_id
        
        # Validate required fields
        if not data.get('cable_id') or not data.get('group_id'):
            return jsonify({
                'success': False,
                'error': 'Cable ID and Group ID are required'
            }), 400
        
        group = Group(**data)
        db.session.add(group)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Group added successfully',
            'id': group.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error adding group: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error adding group: {str(e)}'
        }), 500

@bp.route('/project/<int:project_id>/stage/6/group/<int:group_id>', methods=['POST', 'DELETE'])
def manage_group(project_id, group_id):
    """Update or delete a group"""
    try:
        group = Group.query.get_or_404(group_id)
        
        # Verify ownership
        if group.project_id != project_id:
            return jsonify({
                'success': False,
                'error': 'Group not found in this project'
            }), 404
        
        if request.method == 'POST':
            # Update group
            data = request.form.to_dict()
            for key, value in data.items():
                if hasattr(group, key):
                    setattr(group, key, value)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Group updated successfully'
            })
            
        elif request.method == 'DELETE':
            # Delete group
            db.session.delete(group)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Group deleted successfully'
            })
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error managing group: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error managing group: {str(e)}'
        }), 500


# ============================================================================
# 📋 STEP 7: CHOKES CRUD ROUTES
# ============================================================================

@bp.route('/project/<int:project_id>/stage/7/choke/new', methods=['POST'])
def add_choke(project_id):
    """Add a new choke"""
    try:
        data = request.form.to_dict()
        data['project_id'] = project_id
        
        # Validate required fields
        if not data.get('cable_id') or not data.get('choke_id'):
            return jsonify({
                'success': False,
                'error': 'Cable ID and Choke ID are required'
            }), 400
        
        choke = ChokeTable(**data)
        db.session.add(choke)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Choke added successfully',
            'id': choke.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error adding choke: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error adding choke: {str(e)}'
        }), 500

@bp.route('/project/<int:project_id>/stage/7/choke/<int:choke_id>', methods=['POST', 'DELETE'])
def manage_choke(project_id, choke_id):
    """Update or delete a choke"""
    try:
        choke = ChokeTable.query.get_or_404(choke_id)
        
        # Verify ownership
        if choke.project_id != project_id:
            return jsonify({
                'success': False,
                'error': 'Choke not found in this project'
            }), 404
        
        if request.method == 'POST':
            # Update choke
            data = request.form.to_dict()
            for key, value in data.items():
                if hasattr(choke, key):
                    setattr(choke, key, value)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Choke updated successfully'
            })
            
        elif request.method == 'DELETE':
            # Delete choke
            db.session.delete(choke)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Choke deleted successfully'
            })
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error managing choke: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error managing choke: {str(e)}'
        }), 500


# ============================================================================
# 📋 STEP 8: RESISTORS CRUD ROUTES
# ============================================================================

@bp.route('/project/<int:project_id>/stage/8/resistor/new', methods=['POST'])
def add_resistor(project_id):
    """Add a new resistor"""
    try:
        data = request.form.to_dict()
        data['project_id'] = project_id
        
        # Validate required fields
        if not data.get('cable_id') or not data.get('resistor_id'):
            return jsonify({
                'success': False,
                'error': 'Cable ID and Resistor ID are required'
            }), 400
        
        resistor = ResistorTable(**data)
        db.session.add(resistor)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Resistor added successfully',
            'id': resistor.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error adding resistor: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error adding resistor: {str(e)}'
        }), 500

@bp.route('/project/<int:project_id>/stage/8/resistor/<int:resistor_id>', methods=['POST', 'DELETE'])
def manage_resistor(project_id, resistor_id):
    """Update or delete a resistor"""
    try:
        resistor = ResistorTable.query.get_or_404(resistor_id)
        
        # Verify ownership
        if resistor.project_id != project_id:
            return jsonify({
                'success': False,
                'error': 'Resistor not found in this project'
            }), 404
        
        if request.method == 'POST':
            # Update resistor
            data = request.form.to_dict()
            for key, value in data.items():
                if hasattr(resistor, key):
                    setattr(resistor, key, value)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Resistor updated successfully'
            })
            
        elif request.method == 'DELETE':
            # Delete resistor
            db.session.delete(resistor)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Resistor deleted successfully'
            })
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error managing resistor: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error managing resistor: {str(e)}'
        }), 500


# ============================================================================
# 📋 JUNCTION BOX CRUD ROUTES
# ============================================================================

@bp.route('/project/<int:project_id>/stage/2/junction-box/new', methods=['POST'])
def add_junction_box(project_id):
    """Add a new junction box"""
    try:
        data = request.form.to_dict()
        data['project_id'] = project_id
        
        # Validate required fields
        if not data.get('junction_id') or not data.get('junction_name'):
            return jsonify({
                'success': False,
                'error': 'Junction ID and Name are required'
            }), 400
        
        junction_box = JunctionBox(**data)
        db.session.add(junction_box)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Junction box added successfully',
            'id': junction_box.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error adding junction box: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error adding junction box: {str(e)}'
        }), 500

@bp.route('/project/<int:project_id>/stage/2/junction-box/<int:junction_id>', methods=['POST', 'DELETE'])
def manage_junction_box(project_id, junction_id):
    """Update or delete a junction box"""
    try:
        junction_box = JunctionBox.query.get_or_404(junction_id)
        
        # Verify ownership
        if junction_box.project_id != project_id:
            return jsonify({
                'success': False,
                'error': 'Junction box not found in this project'
            }), 404
        
        if request.method == 'POST':
            # Update junction box
            data = request.form.to_dict()
            for key, value in data.items():
                if hasattr(junction_box, key):
                    setattr(junction_box, key, value)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Junction box updated successfully'
            })
            
        elif request.method == 'DELETE':
            # Delete junction box
            db.session.delete(junction_box)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Junction box deleted successfully'
            })
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error managing junction box: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error managing junction box: {str(e)}'
        }), 500


# ============================================================================
# 📋 MARK STEP COMPLETE ROUTE
# ============================================================================

@bp.route('/project/<int:project_id>/mark-step-complete', methods=['POST'])
def mark_step_complete(project_id):
    """Mark a step as complete"""
    try:
        data = request.get_json()
        step = data.get('step')
        
        if not step or step < 1 or step > 9:
            return jsonify({
                'success': False,
                'error': 'Invalid step number'
            }), 400
        
        project = Project.query.get_or_404(project_id)
        
        # Update project stage if the completed step is greater than current stage
        if not project.stage or step > project.stage:
            project.stage = step
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Step {step} marked complete',
                'new_stage': project.stage
            })
        else:
            return jsonify({
                'success': True,
                'message': f'Step {step} already completed',
                'current_stage': project.stage
            })
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error marking step complete: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Error marking step complete: {str(e)}'
        }), 500

@bp.route('/project/<int:project_id>/save-cable-row-config', methods=['POST'])
def save_cable_row_config(project_id):
    """Save cable row configuration (update existing rows in place)"""
    try:
        data = request.get_json()
        junction_box_id = data.get('junction_box_id')
        rows = data.get('rows', [])
        
        if not junction_box_id:
            return jsonify({'success': False, 'error': 'junction_box_id required'}), 400
        
        if not rows:
            return jsonify({'success': False, 'error': 'No rows to save'}), 400
        
        # FIXED: Update existing rows in place instead of marking as draft
        # First, get all existing rows for this junction box
        existing_rows = CableRowConfig.query.filter_by(
            project_id=project_id,
            junction_box_id=junction_box_id,
            is_draft=False
        ).all()
        
        # Create a dictionary of existing rows by row_number for easy lookup
        existing_rows_dict = {row.row_number: row for row in existing_rows}
        
        # Track which rows we've processed
        processed_row_numbers = set()
        updated_count = 0
        created_count = 0
        
        # Update or create rows
        for row_data in rows:
            row_number = row_data.get('row_number')
            if not row_number:
                continue  # Skip if no row number
            
            processed_row_numbers.add(row_number)
            
            if row_number in existing_rows_dict:
                # Update existing row
                existing_row = existing_rows_dict[row_number]
                existing_row.location_row_name = row_data.get('location_row_name')
                existing_row.cable_type = row_data.get('cable_type', 'cable')
                existing_row.number_of_cables = row_data.get('number_of_cables', 1)
                existing_row.updated_at = datetime.utcnow()
                updated_count += 1
            else:
                # Create new row
                row_config = CableRowConfig(
                    project_id=project_id,
                    junction_box_id=junction_box_id,
                    row_number=row_number,
                    location_row_name=row_data.get('location_row_name'),
                    cable_type=row_data.get('cable_type', 'cable'),
                    number_of_cables=row_data.get('number_of_cables', 1),
                    is_draft=False,
                    draft_version=0
                )
                db.session.add(row_config)
                created_count += 1
        
        # Delete rows that are no longer in the configuration
        deleted_count = 0
        for row_number, existing_row in existing_rows_dict.items():
            if row_number not in processed_row_numbers:
                db.session.delete(existing_row)
                deleted_count += 1
        
        db.session.commit()
        
        print(f"DEBUG: Updated {updated_count}, created {created_count}, deleted {deleted_count} rows")
        return jsonify({
            'success': True,
            'message': f'Configuration saved: {updated_count} updated, {created_count} created, {deleted_count} deleted',
            'updated_count': updated_count,
            'created_count': created_count,
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"ERROR in save_cable_row_config: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/project/<int:project_id>/cable-row-config')
def get_cable_row_config(project_id):
    """Get cable row configuration for a specific junction box"""
    
    junction_box_id = request.args.get('junction_box_id')
    
    # Clean the junction_box_id - remove any whitespace including newlines
    if junction_box_id:
        junction_box_id = junction_box_id.strip()
        print(f"🔍 DEBUG: Original junction_box_id: '{request.args.get('junction_box_id')}'")
        print(f"🔍 DEBUG: Cleaned junction_box_id: '{junction_box_id}'")
    
    if not junction_box_id:
        return jsonify({
            'success': False, 
            'error': 'junction_box_id parameter is required',
            'debug_info': {
                'received_value': request.args.get('junction_box_id'),
                'note': 'Value may be empty or only whitespace'
            }
        }), 400
    
    try:
        # Now query with cleaned value
        rows = CableRowConfig.query.filter_by(
            project_id=project_id,
            junction_box_id=junction_box_id,
            is_draft=False
        ).order_by(CableRowConfig.row_number.asc()).all()
        
        print(f"🔍 DEBUG: Querying for project_id={project_id}, junction_box_id='{junction_box_id}'")
        print(f"🔍 DEBUG: Found {len(rows)} rows")
        
        rows_data = []
        for row in rows:
            rows_data.append({
                'id': row.id,
                'project_id': row.project_id,
                'junction_box_id': row.junction_box_id,
                'row_number': row.row_number,
                'location_row_name': row.location_row_name,
                'cable_type': row.cable_type,
                'number_of_cables': row.number_of_cables,
                'is_draft': row.is_draft,
                'draft_version': row.draft_version,
                'created_at': row.created_at.isoformat() if row.created_at else None,
                'updated_at': row.updated_at.isoformat() if row.updated_at else None
            })
        
        return jsonify({
            'success': True,
            'rows': rows_data,
            'has_data': len(rows_data) > 0,
            'count': len(rows_data),
            'debug_info': {
                'project_id': project_id,
                'junction_box_id_original': request.args.get('junction_box_id'),
                'junction_box_id_cleaned': junction_box_id,
                'rows_found': len(rows_data),
                'query_used': f"project_id={project_id}, junction_box_id='{junction_box_id}'"
            }
        })
        
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'debug_info': {
                'project_id': project_id,
                'junction_box_id_original': request.args.get('junction_box_id'),
                'junction_box_id_cleaned': junction_box_id if junction_box_id else 'None'
            }
        }), 500
                
        
@bp.route('/project/<int:project_id>/stage/3/cable-config/finalize', methods=['POST'])
def finalize_cable_config(project_id):
    """Finalize cable configuration (mark as non-draft)"""
    try:
        data = request.json
        junction_box_id = data.get('junction_box_id')
        
        if not junction_box_id:
            return jsonify(success=False, error="Missing junction_box_id"), 400
        
        # Get the latest draft configuration
        draft_configs = CableRowConfig.query.filter_by(
            project_id=project_id,
            junction_box_id=junction_box_id,
            is_draft=True
        ).all()
        
        if not draft_configs:
            return jsonify(success=False, error="No draft configuration found"), 404
        
        # Mark all drafts as non-draft
        for config in draft_configs:
            config.is_draft = False
        
        db.session.commit()
        
        return jsonify(success=True, message="Configuration finalized")
        
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, error=str(e)), 500

@bp.route('/project/<int:project_id>/stage/3/generate-cables', methods=['POST'])
def generate_cables_from_config(project_id):
    """Generate cables based on cable row configuration"""
    try:
        data = request.json
        junction_box_id = data.get('junction_box_id')
        junction_name = data.get('junction_name')
        junction_size = data.get('junction_size')
        
        if not all([junction_box_id, junction_name]):
            return jsonify(success=False, error="Missing required data"), 400
        
        # Get the finalized configuration
        configs = CableRowConfig.query.filter_by(
            project_id=project_id,
            junction_box_id=junction_box_id,
            is_draft=False
        ).order_by(CableRowConfig.row_number).all()
        
        if not configs:
            return jsonify(success=False, error="No configuration found"), 404
        
        # Get existing cables for this junction box to find max cable_id
        existing_cables = Cable.query.filter_by(
            project_id=project_id,
            junction_box=junction_name
        ).all()
        
        max_cable_id = 0
        for cable in existing_cables:
            try:
                cable_id_int = int(cable.cable_id)
                if cable_id_int > max_cable_id:
                    max_cable_id = cable_id_int
            except (ValueError, TypeError):
                continue
        
        cables_created = []
        
        # Generate cables based on configuration
        for config in configs:
            for i in range(config.number_of_cables):
                max_cable_id += 1
                
                # Create cable object
                cable = Cable(
                    project_id=project_id,
                    cable_id=str(max_cable_id),
                    cable_name=f"{config.location_row_name} Cable {i+1}",
                    junction_box=junction_name,
                    junction_size=junction_size,
                    row=config.location_row_name,
                    cable_type=config.cable_type,
                    terminal=12 if config.cable_type == 'cable' else 0,
                    start_no=1,
                    is_draft=True
                )
                
                db.session.add(cable)
                cables_created.append({
                    'cable_id': str(max_cable_id),
                    'cable_name': cable.cable_name,
                    'row': config.location_row_name,
                    'cable_type': config.cable_type,
                    'terminal': cable.terminal
                })
        
        db.session.commit()
        
        return jsonify(
            success=True,
            message=f"Generated {len(cables_created)} cables",
            cables=cables_created
        )
        
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, error=str(e)), 500


@bp.route('/project/<int:project_id>/stage/1/station-drawing/<int:drawing_id>', methods=['GET', 'POST'])
def station_drawing_edit(project_id, drawing_id):
    if request.method == 'GET':
        # Load data for editing
        drawing = StationDrawing.query.get_or_404(drawing_id)
        return jsonify({
            'success': True,
            'data': drawing.to_dict()
        })
    elif request.method == 'POST':
        # Save data
        data = request.form.to_dict()
        
        if drawing_id:  # Update existing
            drawing = StationDrawing.query.get_or_404(drawing_id)
            for key, value in data.items():
                if hasattr(drawing, key):
                    setattr(drawing, key, value)
        else:  # Create new
            drawing = StationDrawing(project_id=project_id, **data)
            db.session.add(drawing)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Saved successfully'})


@bp.route('/update_cable_ajax/<int:cable_id>', methods=['POST'])
def update_cable_ajax(cable_id):
    try:
        data = request.get_json()
        print(f"Updating cable ID {cable_id} with data: {data}")
        
        # Find the cable by database ID
        cable = Cable.query.get(cable_id)
        
        if not cable:
            print(f"Cable with ID {cable_id} not found")
            return jsonify({
                'success': False,
                'error': f'Cable with ID {cable_id} not found'
            }), 404
        
        print(f"Found cable: ID={cable.id}, cable_id={cable.cable_id}, name={cable.cable_name}")
        
        # Update fields
        if 'cable_name' in data:
            cable.cable_name = data['cable_name']
        if 'junction_box' in data:
            cable.junction_box = data['junction_box']
        if 'junction_name' in data:
            cable.junction_name = data['junction_name']
        if 'row' in data:
            cable.row = data['row']
        if 'terminal' in data:
            cable.terminal = int(data['terminal']) if data['terminal'] else 0
        if 'start_no' in data:
            cable.start_no = int(data['start_no']) if data['start_no'] else 1
        if 'cable_type' in data:
            cable.cable_type = data['cable_type']
        
        db.session.commit()
        
        print(f"Cable updated successfully: ID={cable.id}, cable_id={cable.cable_id}")
        
        return jsonify({
            'success': True,
            'message': f'Cable {cable.cable_id} updated successfully',
            'cable_id': cable.cable_id
        })
        
    except Exception as e:
        print(f"Error updating cable {cable_id}: {str(e)}")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/project/<int:project_id>/stage/3/cable/<int:cable_id>', methods=['POST', 'DELETE'])
def handle_cable(project_id, cable_id):
    try:
        if request.method == 'POST':
            # Update existing cable
            cable = Cable.query.filter_by(id=cable_id, project_id=project_id).first()
            
            if not cable:
                return jsonify({'success': False, 'error': 'Cable not found'}), 404
            
            # Update fields
            cable.cable_name = request.form.get('cable_name', cable.cable_name)
            cable.junction_box = request.form.get('junction_box', cable.junction_box)
            cable.row = request.form.get('row', cable.row)
            
            terminal = request.form.get('terminal')
            if terminal is not None:
                cable.terminal = int(terminal) if terminal else 0
            
            start_no = request.form.get('start_no')
            if start_no is not None:
                cable.start_no = int(start_no) if start_no else 1
            
            cable.cable_type = request.form.get('cable_type', cable.cable_type)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Cable updated successfully',
                'cable_id': cable.cable_id
            })
            
        elif request.method == 'DELETE':
            # Delete cable and all associated data
            cable = Cable.query.filter_by(id=cable_id, project_id=project_id).first()
            
            if not cable:
                return jsonify({'success': False, 'error': 'Cable not found'}), 404
            
            # Delete associated data
            Terminal.query.filter_by(cable_id=cable.cable_id, project_id=project_id).delete()
            TerminalHeader.query.filter_by(cable_id=cable.cable_id, project_id=project_id).delete()
            Group.query.filter_by(cable_id=cable.cable_id, project_id=project_id).delete()
            ChokeTable.query.filter_by(cable_id=cable.cable_id, project_id=project_id).delete()
            ResistorTable.query.filter_by(cable_id=cable.cable_id, project_id=project_id).delete()
            
            # Delete the cable
            db.session.delete(cable)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Cable and all associated data deleted successfully'
            })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/view_list')
@login_required
def view_list():
    """View all projects with summary statistics"""
    try:
        # Get filter parameters
        page = request.args.get('page', 1, type=int)
        rows_per_page = request.args.get('rows_per_page', 20, type=int)
        project_id = request.args.get('project_id', 'all')
        approval_status = request.args.get('approval_status', 'all')
        
        # Base query for projects
        if current_user.role_name == '4' or current_user.role_name == '1':
            # Admin or Creator can see all projects
            base_query = Project.query
        else:
            # Others see only their assigned projects
            base_query = Project.query.filter(
                Project.assigned_users.any(id=current_user.id)
            )
        
        # Apply filters
        if project_id != 'all':
            base_query = base_query.filter(Project.id == project_id)
        
        # Apply approval status filter if not 'all'
        if approval_status != 'all':
            if approval_status == 'approved':
                base_query = base_query.filter(Project.status == 'drawing_approved')
            elif approval_status == 'rejected':
                base_query = base_query.filter(Project.status == 'rejected')
            elif approval_status == 'pending':
                base_query = base_query.filter(
                    Project.status.in_(['level1_pending', 'level2_pending', 'level3_pending'])
                )
            elif approval_status == 'drawing_in_progress':
                base_query = base_query.filter(Project.status == 'drawing_in_progress')
            elif approval_status == 'no_drawing':
                base_query = base_query.filter(
                    ~Project.generated_pdfs.any()
                )
        
        # Order by latest updated
        projects_query = base_query.order_by(Project.updated_date.desc())
        
        # Pagination
        total_projects = projects_query.count()
        total_pages = (total_projects + rows_per_page - 1) // rows_per_page
        start_idx = (page - 1) * rows_per_page + 1
        end_idx = min(start_idx + rows_per_page - 1, total_projects)
        
        projects = projects_query.paginate(
            page=page, 
            per_page=rows_per_page,
            error_out=False
        )
        
        # Get all projects for filter dropdown
        if current_user.role_name == '4' or current_user.role_name == '1':
            all_projects_for_filter = Project.query.order_by(Project.name).all()
        else:
            all_projects_for_filter = Project.query.filter(
                Project.assigned_users.any(id=current_user.id)
            ).order_by(Project.name).all()
        
        # Get summary statistics for each project
        project_summaries = []
        for project in projects.items:
            # Get latest PDF
            latest_pdf = GeneratedPDF.query.filter_by(
                project_id=project.id
            ).order_by(GeneratedPDF.created_at.desc()).first()
            
            # Get counts
            junction_count = JunctionBox.query.filter_by(project_id=project.id).count()
            cable_count = Cable.query.filter_by(project_id=project.id).count()
            terminal_count = Terminal.query.filter_by(project_id=project.id).count()
            
            # Get approval status
            if latest_pdf:
                approval_status_str = latest_pdf.get_approval_status()
                # Map to more readable status
                status_map = {
                    'approved': 'Approved',
                    'rejected': 'Rejected',
                    'level1_pending': 'Pending Level 1',
                    'level2_pending': 'Pending Level 2',
                    'level3_pending': 'Pending Level 3'
                }
                approval_status_display = status_map.get(approval_status_str, 'Unknown')
            else:
                approval_status_display = 'No Drawing'
            
            # Get stage information
            stage_display = 'Not Started'
            if project.stage:
                if project.stage == 10:
                    stage_display = 'PDF Generated'
                elif project.stage == 9:
                    stage_display = 'Drawing Complete'
                elif project.stage < 9:
                    stage_display = f'Stage {project.stage}/9'
            
            project_summaries.append({
                'project': project,
                'latest_pdf': latest_pdf,
                'junction_count': junction_count,
                'cable_count': cable_count,
                'terminal_count': terminal_count,
                'approval_status': approval_status_display,
                'stage': stage_display,
                'has_drawing': latest_pdf is not None,
                'last_updated': project.updated_date
            })
        
        # Calculate overall statistics
        total_junctions = JunctionBox.query.count()
        total_cables = Cable.query.count()
        total_terminals = Terminal.query.count()
        
        # Get approval status distribution
        status_counts = {
            'approved': GeneratedPDF.query.filter_by(level3_status='approved').count(),
            'rejected': GeneratedPDF.query.filter(
                db.or_(
                    GeneratedPDF.level1_status == 'rejected',
                    GeneratedPDF.level2_status == 'rejected',
                    GeneratedPDF.level3_status == 'rejected'
                )
            ).count(),
            'pending': GeneratedPDF.query.filter(
                db.and_(
                    GeneratedPDF.level3_status == 'pending',
                    GeneratedPDF.level2_status != 'rejected',
                    GeneratedPDF.level1_status != 'rejected'
                )
            ).count(),
            'no_drawing': Project.query.filter(~Project.generated_pdfs.any()).count()
        }
        
        # Pagination numbers for template
        pagination_numbers = []
        if total_pages <= 7:
            pagination_numbers = list(range(1, total_pages + 1))
        else:
            if page <= 4:
                pagination_numbers = list(range(1, 6)) + ['...', total_pages]
            elif page >= total_pages - 3:
                pagination_numbers = [1, '...'] + list(range(total_pages - 4, total_pages + 1))
            else:
                pagination_numbers = [1, '...'] + list(range(page - 1, page + 2)) + ['...', total_pages]
        
        return render_template('view_list.html',
                            projects=project_summaries,
                            all_projects=all_projects_for_filter,
                            current_filters={
                                'project_id': project_id,
                                'approval_status': approval_status
                            },
                            pagination={
                                'page': page,
                                'total_pages': total_pages,
                                'total_records': total_projects,
                                'rows_per_page': rows_per_page,
                                'start_idx': start_idx,
                                'end_idx': end_idx,
                                'pagination_numbers': pagination_numbers
                            },
                            stats={
                                'total_junctions': total_junctions,
                                'total_cables': total_cables,
                                'total_terminals': total_terminals,
                                'total_projects': total_projects,
                                'status_counts': status_counts
                            })
        
    except Exception as e:
        print(f"Error in view_list: {e}")
        flash('Error loading view list', 'danger')
        return redirect(url_for('main.approval_tracking'))


@bp.route('/dashboard')
@login_required
def dashboard():
    """View all projects with summary statistics"""
    try:
        # Get filter parameters
        page = request.args.get('page', 1, type=int)
        rows_per_page = request.args.get('rows_per_page', 20, type=int)
        project_id = request.args.get('project_id', 'all')
        approval_status = request.args.get('approval_status', 'all')
        
        # Base query for projects
        if current_user.role_name == '4' or current_user.role_name == '1':
            # Admin or Creator can see all projects
            base_query = Project.query
        else:
            # Others see only their assigned projects
            base_query = Project.query.filter(
                Project.assigned_users.any(id=current_user.id)
            )
        
        # Apply filters
        if project_id != 'all':
            base_query = base_query.filter(Project.id == project_id)
        
        # Apply approval status filter if not 'all'
        if approval_status != 'all':
            if approval_status == 'approved':
                base_query = base_query.filter(Project.status == 'drawing_approved')
            elif approval_status == 'rejected':
                base_query = base_query.filter(Project.status == 'rejected')
            elif approval_status == 'pending':
                base_query = base_query.filter(
                    Project.status.in_(['level1_pending', 'level2_pending', 'level3_pending'])
                )
            elif approval_status == 'drawing_in_progress':
                base_query = base_query.filter(Project.status == 'drawing_in_progress')
            elif approval_status == 'no_drawing':
                base_query = base_query.filter(
                    ~Project.generated_pdfs.any()
                )
        
        # Order by latest updated
        projects_query = base_query.order_by(Project.updated_date.desc())
        
        # Pagination
        total_projects = projects_query.count()
        total_pages = (total_projects + rows_per_page - 1) // rows_per_page
        start_idx = (page - 1) * rows_per_page + 1
        end_idx = min(start_idx + rows_per_page - 1, total_projects)
        
        projects = projects_query.paginate(
            page=page, 
            per_page=rows_per_page,
            error_out=False
        )
        
        # Get all projects for filter dropdown
        if current_user.role_name == '4' or current_user.role_name == '1':
            all_projects_for_filter = Project.query.order_by(Project.name).all()
        else:
            all_projects_for_filter = Project.query.filter(
                Project.assigned_users.any(id=current_user.id)
            ).order_by(Project.name).all()
        
        # Get summary statistics for each project
        project_summaries = []
        for project in projects.items:
            # Get latest PDF
            latest_pdf = GeneratedPDF.query.filter_by(
                project_id=project.id
            ).order_by(GeneratedPDF.created_at.desc()).first()
            
            # Get counts
            junction_count = JunctionBox.query.filter_by(project_id=project.id).count()
            cable_count = Cable.query.filter_by(project_id=project.id).count()
            terminal_count = Terminal.query.filter_by(project_id=project.id).count()
            
            # Get approval status
            if latest_pdf:
                approval_status_str = latest_pdf.get_approval_status()
                # Map to more readable status
                status_map = {
                    'approved': 'Approved',
                    'rejected': 'Rejected',
                    'level1_pending': 'Pending Level 1',
                    'level2_pending': 'Pending Level 2',
                    'level3_pending': 'Pending Level 3'
                }
                approval_status_display = status_map.get(approval_status_str, 'Unknown')
            else:
                approval_status_display = 'No Drawing'
            
            # Get stage information
            stage_display = 'Not Started'
            if project.stage:
                if project.stage == 10:
                    stage_display = 'PDF Generated'
                elif project.stage == 9:
                    stage_display = 'Drawing Complete'
                elif project.stage < 9:
                    stage_display = f'Stage {project.stage}/9'
            
            project_summaries.append({
                'project': project,
                'latest_pdf': latest_pdf,
                'junction_count': junction_count,
                'cable_count': cable_count,
                'terminal_count': terminal_count,
                'approval_status': approval_status_display,
                'stage': stage_display,
                'has_drawing': latest_pdf is not None,
                'last_updated': project.updated_date
            })
        
        # Calculate overall statistics
        total_junctions = JunctionBox.query.count()
        total_cables = Cable.query.count()
        total_terminals = Terminal.query.count()
        
        # Get approval status distribution
        status_counts = {
            'approved': GeneratedPDF.query.filter_by(level3_status='approved').count(),
            'rejected': GeneratedPDF.query.filter(
                db.or_(
                    GeneratedPDF.level1_status == 'rejected',
                    GeneratedPDF.level2_status == 'rejected',
                    GeneratedPDF.level3_status == 'rejected'
                )
            ).count(),
            'pending': GeneratedPDF.query.filter(
                db.and_(
                    GeneratedPDF.level3_status == 'pending',
                    GeneratedPDF.level2_status != 'rejected',
                    GeneratedPDF.level1_status != 'rejected'
                )
            ).count(),
            'no_drawing': Project.query.filter(~Project.generated_pdfs.any()).count()
        }
        
        # Pagination numbers for template
        pagination_numbers = []
        if total_pages <= 7:
            pagination_numbers = list(range(1, total_pages + 1))
        else:
            if page <= 4:
                pagination_numbers = list(range(1, 6)) + ['...', total_pages]
            elif page >= total_pages - 3:
                pagination_numbers = [1, '...'] + list(range(total_pages - 4, total_pages + 1))
            else:
                pagination_numbers = [1, '...'] + list(range(page - 1, page + 2)) + ['...', total_pages]
        
        return render_template('dashboard.html',
                            projects=project_summaries,
                            all_projects=all_projects_for_filter,
                            current_filters={
                                'project_id': project_id,
                                'approval_status': approval_status
                            },
                            pagination={
                                'page': page,
                                'total_pages': total_pages,
                                'total_records': total_projects,
                                'rows_per_page': rows_per_page,
                                'start_idx': start_idx,
                                'end_idx': end_idx,
                                'pagination_numbers': pagination_numbers
                            },
                            stats={
                                'total_junctions': total_junctions,
                                'total_cables': total_cables,
                                'total_terminals': total_terminals,
                                'total_projects': total_projects,
                                'status_counts': status_counts
                            })
        
    except Exception as e:
        print(f"Error in view_list: {e}")
        flash('Error loading view list', 'danger')
        return redirect(url_for('main.approval_tracking'))



# ==================== BEFORE REQUEST HOOK ====================

@bp.before_app_request
def check_active_role():
    """Check if current user's role is still active on every request"""
    if current_user.is_authenticated:
        # Skip for static files and specific routes
        if request.endpoint and 'static' in request.endpoint:
            return
        
        # Skip for these endpoints
        excluded_endpoints = ['main.login', 'main.logout', 'main.role_master_toggle_status']
        if request.endpoint in excluded_endpoints:
            return
        
        # Check if user's role is active
        if not check_user_role_status(current_user):
            logout_user()
            flash('Your role has been deactivated. Please contact administrator.', 'warning')
            return redirect(url_for('main.login'))

def check_user_role_status(user):
    """Check if user's role is active"""
    role = RoleMaster.query.filter_by(role_name=user.role_name).first()
    return role and role.is_active

@bp.route('/check-designation-users/<int:id>')
def check_designation_users(id):
    """Check how many users have this designation"""
    if current_user.role_name != '4':
        return jsonify({'user_count': 0}), 403
    
    designation = DesignationMaster.query.get_or_404(id)
    user_count = User.query.filter_by(designation_id=id).count()
    
    return jsonify({
        'designation_name': designation.designation_name,
        'user_count': user_count
    })
    
# ==================== ROLE MANAGEMENT ====================

@bp.route('/roles')
def role_master_list():
    if current_user.role_name != '4':
        flash('Only admin users can access this page.', 'danger')
        return redirect(url_for('main.approval_tracking'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    # Get search parameters
    search = request.args.get('search', '')
    status_filter = request.args.get('status', 'all')
    
    # Base query
    query = RoleMaster.query
    
    # Apply search filter
    if search:
        query = query.filter(RoleMaster.role_name.ilike(f'%{search}%'))
    
    # Apply status filter
    if status_filter == 'active':
        query = query.filter(RoleMaster.is_active == True)
    elif status_filter == 'inactive':
        query = query.filter(RoleMaster.is_active == False)
    
    # Order and paginate
    roles = query.order_by(RoleMaster.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Build pagination info
    pagination = {
        'page': roles.page,
        'per_page': roles.per_page,
        'total': roles.total,
        'pages': roles.pages,
        'has_prev': roles.has_prev,
        'has_next': roles.has_next,
        'prev_num': roles.prev_num,
        'next_num': roles.next_num,
        'iter_pages': list(roles.iter_pages(left_edge=1, right_edge=1, left_current=2, right_current=2))
    }
    
    # Count active users per role for display
    role_user_counts = {}
    for role in roles.items:
        user_count = User.query.filter_by(
            role_name=role.role_name, 
            is_active=True
        ).count()
        role_user_counts[role.id] = user_count
    
    return render_template('roles/list.html', 
                         roles=roles.items,
                         pagination=pagination,
                         search=search,
                         status_filter=status_filter,
                         role_user_counts=role_user_counts)

@bp.route('/roles/create', methods=['GET', 'POST'])
def role_master_create(): 
    if current_user.role_name != '4':
        flash('Only admin users can create roles.', 'danger')
        return redirect(url_for('main.role_master_list'))
    
    if request.method == 'POST':
        role_name = request.form.get('role_name', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        # Validate
        if not role_name:
            flash('Role name is required.', 'danger')
            return redirect(url_for('main.role_master_create'))
        
        # Check if role already exists
        existing_role = RoleMaster.query.filter_by(role_name=role_name).first()
        if existing_role:
            flash(f'Role "{role_name}" already exists.', 'danger')
            return redirect(url_for('main.role_master_create'))
        
        # Create new role
        new_role = RoleMaster(
            role_name=role_name,
            is_active=is_active
        )
        
        try:
            db.session.add(new_role)
            db.session.commit()
            flash(f'Role "{role_name}" created successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating role: {str(e)}', 'danger')
        
        return redirect(url_for('main.role_master_list'))
    
    return render_template('roles/form.html', role=None, title='Create New Role')

@bp.route('/roles/<int:id>/edit', methods=['GET', 'POST'])
def role_master_edit(id):
    if current_user.role_name != '4':
        flash('Only admin users can edit roles.', 'danger')
        return redirect(url_for('main.role_master_list'))
    
    role = RoleMaster.query.get_or_404(id)
    
    if request.method == 'POST':
        role_name = request.form.get('role_name', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        # Validate
        if not role_name:
            flash('Role name is required.', 'danger')
            return redirect(url_for('main.role_master_edit', id=id))
        
        # Check if role name already exists (excluding current role)
        existing_role = RoleMaster.query.filter(
            RoleMaster.role_name == role_name,
            RoleMaster.id != id
        ).first()
        
        if existing_role:
            flash(f'Role "{role_name}" already exists.', 'danger')
            return redirect(url_for('main.role_master_edit', id=id))
        
        # Update role
        role.role_name = role_name
        role.is_active = is_active
        
        try:
            db.session.commit()
            flash(f'Role "{role_name}" updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating role: {str(e)}', 'danger')
        
        return redirect(url_for('main.role_master_list'))
    
    return render_template('roles/form.html', role=role, title='Edit Role')

@bp.route('/roles/<int:id>/delete')
def role_master_delete(id):
    if current_user.role_name != '4':
        flash('Only admin users can delete roles.', 'danger')
        return redirect(url_for('main.role_master_list'))
    
    role = RoleMaster.query.get_or_404(id)
    role_name = role.role_name
    
    # Check if role is being used
    users_with_role = User.query.filter_by(role_name=role.role_name).count()
    if users_with_role > 0:
        flash(f'Cannot delete role "{role_name}" because {users_with_role} user(s) have this role. Please reassign users first.', 'danger')
        return redirect(url_for('main.role_master_list'))
    
    try:
        db.session.delete(role)
        db.session.commit()
        flash(f'Role "{role_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting role: {str(e)}', 'danger')
    
    return redirect(url_for('main.role_master_list'))

@bp.route('/roles/<int:id>/toggle-status', methods=['POST'])
def role_master_toggle_status(id):
    if current_user.role_name != '4':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    role = RoleMaster.query.get_or_404(id)
    old_status = role.is_active
    role.is_active = not role.is_active
    role.updated_at = datetime.utcnow()
    
    try:
        db.session.commit()
        
        # Count active users with this role
        active_users_count = User.query.filter_by(
            role_name=role.role_name, 
            is_active=True
        ).count()
        
        response_data = {
            'success': True,
            'is_active': role.is_active,
            'role_name': role.role_name
        }
        
        # If role is being deactivated, add warning
        if not role.is_active and old_status and active_users_count > 0:
            response_data['warning'] = f'Role deactivated. {active_users_count} active user(s) with this role will be logged out.'
        
        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== DESIGNATION MANAGEMENT ====================

@bp.route('/designations')
def designation_master_list():
    if current_user.role_name != '4':
        flash('Only admin users can access this page.', 'danger')
        return redirect(url_for('main.approval_tracking'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    # Get search parameters
    search = request.args.get('search', '')
    status_filter = request.args.get('status', 'all')
    
    # Base query
    query = DesignationMaster.query
    
    # Apply search filter
    if search:
        query = query.filter(DesignationMaster.designation_name.ilike(f'%{search}%'))
    
    # Apply status filter
    if status_filter == 'active':
        query = query.filter(DesignationMaster.is_active == True)
    elif status_filter == 'inactive':
        query = query.filter(DesignationMaster.is_active == False)
    
    # Order and paginate
    designations = query.order_by(DesignationMaster.approval_level.asc(), 
                                 DesignationMaster.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Count active users per designation for display
    designation_user_counts = {}
    for designation in designations.items:
        user_count = User.query.filter_by(
            designation_id=designation.id, 
            is_active=True
        ).count()
        designation_user_counts[designation.id] = user_count
    
    # Build pagination info
    pagination = {
        'page': designations.page,
        'per_page': designations.per_page,
        'total': designations.total,
        'pages': designations.pages,
        'has_prev': designations.has_prev,
        'has_next': designations.has_next,
        'prev_num': designations.prev_num,
        'next_num': designations.next_num,
        'iter_pages': list(designations.iter_pages(left_edge=1, right_edge=1, left_current=2, right_current=2))
    }
    
    return render_template('designations/list.html', 
                         designations=designations.items,
                         pagination=pagination,
                         search=search,
                         status_filter=status_filter,
                         designation_user_counts=designation_user_counts)

@bp.route('/designations/create', methods=['GET', 'POST'])
def designation_master_create():
    if current_user.role_name != '4':
        flash('Only admin users can create designations.', 'danger')
        return redirect(url_for('main.designation_master_list'))
    
    if request.method == 'POST':
        designation_name = request.form.get('designation_name', '').strip()
        approval_level = request.form.get('approval_level', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        # Validate
        if not designation_name:
            flash('Designation name is required.', 'danger')
            return redirect(url_for('main.designation_master_create'))
        
        # Check if designation already exists
        existing_designation = DesignationMaster.query.filter_by(
            designation_name=designation_name
        ).first()
        if existing_designation:
            flash(f'Designation "{designation_name}" already exists.', 'danger')
            return redirect(url_for('main.designation_master_create'))
        
        # Parse approval level (can be empty)
        approval_level_int = None
        if approval_level and approval_level.isdigit():
            approval_level_int = int(approval_level)
        
        # Create new designation
        new_designation = DesignationMaster(
            designation_name=designation_name,
            approval_level=approval_level_int,
            is_active=is_active
        )
        
        try:
            db.session.add(new_designation)
            db.session.commit()
            flash(f'Designation "{designation_name}" created successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating designation: {str(e)}', 'danger')
        
        return redirect(url_for('main.designation_master_list'))
    
    return render_template('designations/form.html', designation=None, title='Create New Designation')

@bp.route('/designations/<int:id>/edit', methods=['GET', 'POST'])
def designation_master_edit(id):
    if current_user.role_name != '4':
        flash('Only admin users can edit designations.', 'danger')
        return redirect(url_for('main.designation_master_list'))
    
    designation = DesignationMaster.query.get_or_404(id)
    
    if request.method == 'POST':
        designation_name = request.form.get('designation_name', '').strip()
        approval_level = request.form.get('approval_level', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        # Validate
        if not designation_name:
            flash('Designation name is required.', 'danger')
            return redirect(url_for('main.designation_master_edit', id=id))
        
        # Check if designation name already exists (excluding current designation)
        existing_designation = DesignationMaster.query.filter(
            DesignationMaster.designation_name == designation_name,
            DesignationMaster.id != id
        ).first()
        
        if existing_designation:
            flash(f'Designation "{designation_name}" already exists.', 'danger')
            return redirect(url_for('main.designation_master_edit', id=id))
        
        # Parse approval level (can be empty)
        approval_level_int = None
        if approval_level and approval_level.isdigit():
            approval_level_int = int(approval_level)
        
        # Update designation
        designation.designation_name = designation_name
        designation.approval_level = approval_level_int
        designation.is_active = is_active
        
        try:
            db.session.commit()
            flash(f'Designation "{designation_name}" updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating designation: {str(e)}', 'danger')
        
        return redirect(url_for('main.designation_master_list'))
    
    return render_template('designations/form.html', designation=designation, title='Edit Designation')

@bp.route('/designations/<int:id>/delete')
def designation_master_delete(id):
    if current_user.role_name != '4':
        flash('Only admin users can delete designations.', 'danger')
        return redirect(url_for('main.designation_master_list'))
    
    designation = DesignationMaster.query.get_or_404(id)
    designation_name = designation.designation_name
    
    # Check if designation is being used
    users_with_designation = User.query.filter_by(designation_id=id).count()
    if users_with_designation > 0:
        flash(f'Cannot delete designation "{designation_name}" because {users_with_designation} user(s) have this designation. Please reassign users first.', 'danger')
        return redirect(url_for('main.designation_master_list'))
    
    try:
        db.session.delete(designation)
        db.session.commit()
        flash(f'Designation "{designation_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting designation: {str(e)}', 'danger')
    
    return redirect(url_for('main.designation_master_list'))

@bp.route('/designations/<int:id>/toggle-status', methods=['POST'])
def designation_master_toggle_status(id):
    if current_user.role_name != '4':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    designation = DesignationMaster.query.get_or_404(id)
    old_status = designation.is_active
    designation.is_active = not designation.is_active
    designation.updated_at = datetime.utcnow()
    
    try:
        db.session.commit()
        
        # Count active users with this designation
        active_users_count = User.query.filter_by(
            designation_id=designation.id, 
            is_active=True
        ).count()
        
        response_data = {
            'success': True,
            'is_active': designation.is_active,
            'designation_name': designation.designation_name
        }
        
        # If designation is being deactivated, add warning
        if not designation.is_active and old_status and active_users_count > 0:
            response_data['warning'] = f'Designation deactivated. {active_users_count} active user(s) with this designation will be logged out on their next request.'
        
        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500



# Add this route to your routes.py file
@bp.route('/project/<int:project_id>/stage/9/api-data')
@login_required
def stage9_api_data(project_id):
    """API endpoint for Stage 9 (Relay Box) data"""
    project = Project.query.get_or_404(project_id)
    
    # Get query parameters
    location_id = request.args.get('location_id', type=str)
    cable_id = request.args.get('cable_id', type=int)
    location_name = request.args.get('location_name', type=str)
    
    # Start query for cable boxes (relay boxes)
    query = CableBox.query.filter_by(project_id=project_id)
    
    # Apply location filter if provided
    if location_id and location_id != 'null':
        try:
            # Try to parse location_id as integer (it might be the JunctionBox id)
            location_id_int = int(location_id)
            
            # Get the junction box by ID first
            junction_box = JunctionBox.query.filter_by(
                project_id=project_id,
                id=location_id_int
            ).first()
            
            if junction_box:
                # Now filter cable boxes by junction_id (string field)
                query = query.filter_by(junction_box=junction_box.junction_id)
                location_name = junction_box.junction_name
            else:
                # If no junction box found by ID, try by junction_id string
                junction_box = JunctionBox.query.filter_by(
                    project_id=project_id,
                    junction_id=location_id
                ).first()
                if junction_box:
                    query = query.filter_by(junction_box=location_id)
                    location_name = junction_box.junction_name
                else:
                    # Return empty results
                    query = query.filter_by(junction_box="NON_EXISTENT_JUNCTION")
        except ValueError:
            # If location_id is not a number, use it as junction_id directly
            junction_box = JunctionBox.query.filter_by(
                project_id=project_id,
                junction_id=location_id
            ).first()
            if junction_box:
                query = query.filter_by(junction_box=location_id)
                location_name = junction_box.junction_name
            else:
                # Return empty results
                query = query.filter_by(junction_box="NON_EXISTENT_JUNCTION")
    
    # Apply cable ID filter if provided
    if cable_id:
        query = query.filter_by(cable_id=cable_id)
    
    # Execute query and order results
    cable_boxes = query.order_by(CableBox.cable_id.asc()).all()
    
    # Format data for response
    cable_box_list = [box.to_dict() for box in cable_boxes]
    
    # Get counts for different statuses if needed
    relay_box_count = len([cb for cb in cable_boxes if getattr(cb, 'cable_type', None) == 'relay_box'])
    cable_count = len(cable_boxes) - relay_box_count
    
    return jsonify({
        'success': True,
        'has_data': len(cable_box_list) > 0,
        'location_id': location_id if location_id and location_id != 'null' else None,
        'location_name': location_name,
        'data': {
            'cable_boxes': cable_box_list
        },
        'counts': {
            'cable_boxes': len(cable_box_list),
            'relay_boxes': relay_box_count,
            'cables': cable_count
        }
    })


@bp.route("/set_preview_status/")
@login_required
def set_preview_status():
    """Set project status to 'preview' (for AJAX calls)"""
    project_id = get_current_project()
    if not project_id:
        return jsonify({"success": False, "message": "Please select a project first"}), 400
   
    # Get project
    current_project = Project.query.get(project_id)
    if not current_project:
        return jsonify({"success": False, "message": "Project not found"}), 404
    
    try:
        # Update project status to 'preview'
        current_project.stage = 10
        current_project.status = 'preview'
        current_project.updated_date = get_ist_now()
        db.session.commit()
        
        print(f"✅ Project {project_id} marked as stage 10 with status 'preview' (via AJAX)")
        
        return jsonify({
            "success": True, 
            "message": "Project status updated to 'preview'",
            "project_id": project_id,
            "status": "preview"
        })
       
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error setting preview status: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500



# ============================================================================
# 📋 PREVIEW DOWNLOAD ENDPOINT
# ============================================================================

@bp.route("/download_preview/")
@login_required
def download_preview():
    """Download XLSX file for preview - sets status to 'preview'"""
    project_id = get_current_project()
    if not project_id:
        flash("Please select a project first", "error")
        return redirect(url_for("main.project_selection"))
   
    # Get project
    current_project = Project.query.get(project_id)
    if not current_project:
        flash("Project not found", "error")
        return redirect(url_for("main.index"))
    
    # Seed default station drawing
    seed_default_station_drawing(project_id)
   
    try:
        # ========== CREATE XLSX WORKBOOK ==========
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        total_records = 0
       
        # Get cable names mapping
        cable_name_map = {}
        try:
            cables = Cable.query.filter_by(project_id=project_id).all()
            cable_name_map = {str(cable.cable_id): cable.cable_name for cable in cables}
        except Exception as e:
            print(f"Warning loading cable names: {str(e)}")
       
        # Create all sheets (copy your existing sheet creation code here)
        for sheet_name, columns in SHEETS.items():
            ws = wb.create_sheet(title=sheet_name)
           
            # Handle special formatting for terminals sheet
            if sheet_name == 'terminal':
                # Use template headers for terminals (without ID column)
                template_headers = ['cable_id', 'cable_name', 'terminal_id', 'terminal_no', 'symbol',
                                  'input_left', 'input_right', 'spare', 'input_connected',
                                  'output_connected', 'input_connected_extra', 'output_connected_extra',
                                  'output_left', 'output_right']
                ws.append(template_headers)
               
                model = MODEL_MAP[sheet_name]
                records = model.query.filter_by(project_id=project_id).all()
               
                # Sort records
                sorted_records = sorted(records, key=lambda x: (
                    int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0,
                    int(x.terminal_id) if x.terminal_id and str(x.terminal_id).isdigit() else 0
                ))
               
                for record in sorted_records:
                    row = [
                        record.cable_id,
                        cable_name_map.get(str(record.cable_id), ''),
                        record.terminal_id,
                        record.terminal_no,
                        record.symbol,
                        record.input_left if record.input_left and record.input_left.strip() else '',
                        record.input_right if record.input_right and record.input_right.strip() else '',
                        record.spare,
                        record.input_connected,
                        record.output_connected,
                        record.input_connected_extra if record.input_connected_extra and record.input_connected_extra.strip() else '',
                        record.output_connected_extra if record.output_connected_extra and record.output_connected_extra.strip() else '',
                        record.output_left if record.output_left and record.output_left.strip() else '',
                        record.output_right if record.output_right and record.output_right.strip() else ''
                    ]
                    ws.append(row)
                    total_records += 1
                   
            # Handle headers sheet with cable name
            elif sheet_name == 'terminal_header':
                excel_columns = list(columns)
                cable_id_index = excel_columns.index('cable_id')
                excel_columns.insert(cable_id_index + 1, 'cable_name')
                ws.append(excel_columns)
               
                model = MODEL_MAP[sheet_name]
                records = model.query.filter_by(project_id=project_id).all()
               
                for record in records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                   
                    cable_id_index = columns.index('cable_id')
                    cable_name = cable_name_map.get(str(record.cable_id), '')
                    row.insert(cable_id_index + 1, cable_name)
                   
                    ws.append(row)
                    total_records += 1
                   
            # Handle groups sheet with cable name
            elif sheet_name == 'group':
                excel_columns = list(columns)
                cable_id_index = excel_columns.index('cable_id')
                excel_columns.insert(cable_id_index + 1, 'cable_name')
                ws.append(excel_columns)
               
                model = MODEL_MAP[sheet_name]
                records = model.query.filter_by(project_id=project_id).all()
               
                for record in records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                   
                    cable_id_index = columns.index('cable_id')
                    cable_name = cable_name_map.get(str(record.cable_id), '')
                    row.insert(cable_id_index + 1, cable_name)
                   
                    ws.append(row)
                    total_records += 1
                   
            # Handle choketable sheet with cable name
            elif sheet_name == 'choketable':
                excel_columns = list(columns)
                cable_id_index = excel_columns.index('cable_id')
                excel_columns.insert(cable_id_index + 1, 'cable_name')
                ws.append(excel_columns)
               
                model = MODEL_MAP[sheet_name]
                records = model.query.filter_by(project_id=project_id).all()
               
                for record in records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                   
                    cable_id_index = columns.index('cable_id')
                    cable_name = cable_name_map.get(str(record.cable_id), '')
                    row.insert(cable_id_index + 1, cable_name)
                   
                    ws.append(row)
                    total_records += 1
                   
            # Handle resistortable sheet with cable name
            elif sheet_name == 'resistortable':
                excel_columns = list(columns)
                cable_id_index = excel_columns.index('cable_id')
                excel_columns.insert(cable_id_index + 1, 'cable_name')
                ws.append(excel_columns)
               
                model = MODEL_MAP[sheet_name]
                records = model.query.filter_by(project_id=project_id).all()
               
                for record in records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                   
                    cable_id_index = columns.index('cable_id')
                    cable_name = cable_name_map.get(str(record.cable_id), '')
                    row.insert(cable_id_index + 1, cable_name)
                   
                    ws.append(row)
                    total_records += 1
                   
            else:
                # Standard handling for other sheets
                ws.append(columns)
                model = MODEL_MAP[sheet_name]
                records = model.query.filter_by(project_id=project_id).all()
               
                if sheet_name == 'cable':
                    sorted_records = sorted(records, key=lambda x: (
                        int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0
                    ))
                else:
                    sorted_records = records
               
                for record in sorted_records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                    ws.append(row)
                    total_records += 1
        
        # ========== UPDATE PROJECT STATUS FOR PREVIEW ==========
        # Update both stage and status for preview
        current_project.stage = 10  # Set stage to 10
        current_project.status = 'preview'  # Set status to preview
        current_project.updated_date = get_ist_now()
        db.session.commit()
        
        print(f"✅ Project {project_id} marked as stage 10 with status 'preview' (preview download)")
        
        # ========== SAVE XLSX TO DOWNLOAD FOLDER ==========
        # Create download directory
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        download_dir = os.path.join(base_dir, 'xlsx_download')
        os.makedirs(download_dir, exist_ok=True)
        
        # Generate filename with PREVIEW prefix
        timestamp = get_ist_now().strftime('%Y%m%d_%H%M%S')
        project_name_clean = re.sub(r'[^\w\-_\. ]', '', current_project.name).replace(' ', '_')
        filename = f"PREVIEW_RAILWAYPROJECT_ID{project_id}_{project_name_clean}_{timestamp}.xlsx"
        file_path = os.path.join(download_dir, filename)
        
        # Save the workbook
        wb.save(file_path)
        
        # ========== FLASH MESSAGE ==========
        flash(f"✅ PREVIEW XLSX downloaded successfully! Project marked as stage 10 with status 'preview'", "success")
        
        # ========== RETURN FILE FOR DOWNLOAD ==========
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
       
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error generating preview download: {str(e)}", "error")
        print(f"❌ Preview download error: {str(e)}")
        import traceback
        traceback.print_exc()
        return redirect(url_for("main.index"))

@bp.route('/debug/pdf_records/<int:project_id>')
@login_required
def debug_pdf_records(project_id):
    """Debug endpoint to check PDF records in database"""
    try:
        # Get all PDF records for this project
        pdf_records = GeneratedPDF.query.filter_by(project_id=project_id).order_by(GeneratedPDF.created_at.desc()).all()
        
        records_list = []
        for pdf in pdf_records:
            records_list.append({
                'id': pdf.id,
                'pdf_filename': pdf.pdf_filename,
                'project_id': pdf.project_id,
                'version': pdf.version,
                'created_at': pdf.created_at.isoformat() if pdf.created_at else None,
                'level1_status': pdf.level1_status,
                'level2_status': pdf.level2_status,
                'level3_status': pdf.level3_status
            })
        
        # NEW: Also get files from uploads directory
        import os
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        uploads_dir = os.path.join(base_dir, 'uploads')
        
        preview_pdfs = []
        final_pdfs = []
        
        if os.path.exists(uploads_dir):
            for filename in os.listdir(uploads_dir):
                if filename.lower().endswith('.pdf'):
                    if f'project_{project_id}' in filename or f'_{project_id}_' in filename:
                        if 'PREVIEW' in filename:
                            preview_pdfs.append({
                                'filename': filename,
                                'path': os.path.join(uploads_dir, filename),
                                'exists': True,
                                'type': 'PREVIEW'
                            })
                        else:
                            final_pdfs.append({
                                'filename': filename,
                                'path': os.path.join(uploads_dir, filename),
                                'exists': True,
                                'type': 'FINAL'
                            })
        
        # Get project info
        project = Project.query.get(project_id)
        project_info = {
            'id': project.id if project else None,
            'name': project.name if project else None,
            'stage': project.stage if project else None,
            'status': project.status if project else None
        }
        
        return jsonify({
            'success': True,
            'project_id': project_id,
            'project': project_info,
            'pdf_records': records_list,
            'preview_pdfs': preview_pdfs,
            'final_pdfs': final_pdfs,
            'all_pdfs': preview_pdfs + final_pdfs,
            'pdf_count': len(records_list),
            'uploads': {
                'uploads_dir': uploads_dir,
                'exists': os.path.exists(uploads_dir),
                'total_files': len(os.listdir(uploads_dir)) if os.path.exists(uploads_dir) else 0
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@bp.route('/debug/check_pdf/<filename>')
def debug_check_pdf(filename):
    """Debug endpoint to check if PDF exists"""
    import os
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    uploads_dir = os.path.join(base_dir, 'uploads')
    file_path = os.path.join(uploads_dir, filename)
    
    return jsonify({
        'filename': filename,
        'uploads_dir': uploads_dir,
        'file_path': file_path,
        'exists': os.path.exists(file_path),
        'files_in_uploads': os.listdir(uploads_dir) if os.path.exists(uploads_dir) else []
    })


@bp.route('/project/<int:project_id>/set-location-filter/<int:location_id>', methods=['POST'])
@login_required
def set_location_filter(project_id, location_id):
    """Set location filter in session for a specific project"""
    try:
        project = Project.query.get_or_404(project_id)
        
        # Get the location (junction box)
        location = JunctionBox.query.filter_by(
            project_id=project_id,
            id=location_id
        ).first()
        
        if not location:
            return jsonify({'success': False, 'error': 'Location not found'}), 404
        
        # Set location filter in session
        session['current_location_id'] = location_id
        session['current_location_name'] = location.junction_name
        session['current_project_id'] = project_id
        
        return jsonify({
            'success': True,
            'message': f'Location filter set to {location.junction_name}',
            'project_id': project_id,
            'location_id': location_id,
            'location_name': location.junction_name
        })
        
    except Exception as e:
        print(f"Error setting location filter: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500



@bp.route("/download/location/")
@login_required
def download_location():
    """Download XLSX file for a specific location (JunctionBox) only"""
    project_id = get_current_project()
    location_id = request.args.get('location_id', type=int)
    
    if not project_id:
        flash("Please select a project first", "error")
        return redirect(url_for("main.project_selection"))
    
    if not location_id:
        flash("Location ID is required", "error")
        return redirect(request.referrer or url_for("main.index"))
    
    # Get project
    current_project = Project.query.get(project_id)
    if not current_project:
        flash("Project not found", "error")
        return redirect(url_for("main.index"))
    
    # Get location (JunctionBox)
    location = JunctionBox.query.filter_by(
        id=location_id, 
        project_id=project_id
    ).first()
    
    if not location:
        flash("Location (Junction Box) not found in this project", "error")
        return redirect(request.referrer or url_for("main.index"))
    
    try:
        # ========== CREATE XLSX WORKBOOK ==========
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        total_records = 0
        
        # Get cable names mapping
        cable_name_map = {}
        try:
            cables = Cable.query.filter_by(project_id=project_id).all()
            cable_name_map = {str(cable.cable_id): cable.cable_name for cable in cables}
        except Exception as e:
            print(f"Warning loading cable names: {str(e)}")
        
        # Create all sheets but filter by location (junction_box)
        for sheet_name, columns in SHEETS.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # Get the model for this sheet
            model = MODEL_MAP.get(sheet_name)
            if not model:
                continue
            
            # Check if model has junction_box field
            if hasattr(model, 'junction_box'):
                # Filter by junction_box
                records = model.query.filter_by(
                    project_id=project_id,
                    junction_box=location.junction_id
                ).all()
            else:
                # For models without junction_box, get all records for the project
                records = model.query.filter_by(project_id=project_id).all()
            
            # Skip if no records
            if not records:
                ws.append([f"No data available for junction box: {location.junction_name}"])
                continue
            
            # Handle special formatting for terminals sheet
            if sheet_name == 'terminal':
                # Use template headers for terminals (without ID column)
                template_headers = ['cable_id', 'cable_name', 'terminal_id', 'terminal_no', 'symbol',
                                  'input_left', 'input_right', 'spare', 'input_connected',
                                  'output_connected', 'input_connected_extra', 'output_connected_extra',
                                  'output_left', 'output_right']
                ws.append(template_headers)
                
                # For location-specific download, filter by cable_ids from this junction
                if location_id:
                    # Get cables that belong to this junction
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        # Get terminals for these cable_ids
                        filtered_records = []
                        for cable_id in cable_ids:
                            cable_terminals = model.query.filter_by(
                                project_id=project_id,
                                cable_id=cable_id
                            ).all()
                            filtered_records.extend(cable_terminals)
                        records = filtered_records
                    else:
                        records = []
                
                # Sort records
                sorted_records = sorted(records, key=lambda x: (
                    int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0,
                    int(x.terminal_id) if x.terminal_id and str(x.terminal_id).isdigit() else 0
                ))
                
                for record in sorted_records:
                    row = [
                        record.cable_id,
                        cable_name_map.get(str(record.cable_id), ''),
                        record.terminal_id,
                        record.terminal_no,
                        record.symbol,
                        record.input_left if record.input_left and record.input_left.strip() else '',
                        record.input_right if record.input_right and record.input_right.strip() else '',
                        record.spare,
                        record.input_connected,
                        record.output_connected,
                        record.input_connected_extra if record.input_connected_extra and record.input_connected_extra.strip() else '',
                        record.output_connected_extra if record.output_connected_extra and record.output_connected_extra.strip() else '',
                        record.output_left if record.output_left and record.output_left.strip() else '',
                        record.output_right if record.output_right and record.output_right.strip() else ''
                    ]
                    ws.append(row)
                    total_records += 1
                    
            # Handle headers sheet with cable name
            elif sheet_name == 'terminal_header':
                excel_columns = list(columns)
                if 'cable_id' in excel_columns:
                    cable_id_index = excel_columns.index('cable_id')
                    excel_columns.insert(cable_id_index + 1, 'cable_name')
                ws.append(excel_columns)
                
                # For location-specific download, filter by cable_ids from this junction
                if location_id and records:
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        # Filter records by cable_ids
                        records = [r for r in records if r.cable_id in cable_ids]
                    else:
                        records = []
                
                for record in records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                    
                    if 'cable_id' in columns:
                        cable_id_index = columns.index('cable_id')
                        cable_name = cable_name_map.get(str(record.cable_id), '')
                        row.insert(cable_id_index + 1, cable_name)
                    
                    ws.append(row)
                    total_records += 1
                    
            # Handle groups sheet with cable name
            elif sheet_name == 'group':
                excel_columns = list(columns)
                if 'cable_id' in excel_columns:
                    cable_id_index = excel_columns.index('cable_id')
                    excel_columns.insert(cable_id_index + 1, 'cable_name')
                ws.append(excel_columns)
                
                # For location-specific download, filter by cable_ids from this junction
                if location_id and records:
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        # Filter records by cable_ids
                        records = [r for r in records if r.cable_id in cable_ids]
                    else:
                        records = []
                
                for record in records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                    
                    if 'cable_id' in columns:
                        cable_id_index = columns.index('cable_id')
                        cable_name = cable_name_map.get(str(record.cable_id), '')
                        row.insert(cable_id_index + 1, cable_name)
                    
                    ws.append(row)
                    total_records += 1
                    
            # Handle choketable sheet with cable name
            elif sheet_name == 'choketable':
                excel_columns = list(columns)
                if 'cable_id' in excel_columns:
                    cable_id_index = excel_columns.index('cable_id')
                    excel_columns.insert(cable_id_index + 1, 'cable_name')
                ws.append(excel_columns)
                
                # For location-specific download, filter by cable_ids from this junction
                if location_id and records:
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        # Filter records by cable_ids
                        records = [r for r in records if r.cable_id in cable_ids]
                    else:
                        records = []
                
                for record in records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                    
                    if 'cable_id' in columns:
                        cable_id_index = columns.index('cable_id')
                        cable_name = cable_name_map.get(str(record.cable_id), '')
                        row.insert(cable_id_index + 1, cable_name)
                    
                    ws.append(row)
                    total_records += 1
                    
            # Handle resistortable sheet with cable name
            elif sheet_name == 'resistortable':
                excel_columns = list(columns)
                if 'cable_id' in excel_columns:
                    cable_id_index = excel_columns.index('cable_id')
                    excel_columns.insert(cable_id_index + 1, 'cable_name')
                ws.append(excel_columns)
                
                # For location-specific download, filter by cable_ids from this junction
                if location_id and records:
                    junction_cables = Cable.query.filter_by(
                        project_id=project_id,
                        junction_box=location.junction_id
                    ).all()
                    
                    if junction_cables:
                        cable_ids = [cable.cable_id for cable in junction_cables]
                        # Filter records by cable_ids
                        records = [r for r in records if r.cable_id in cable_ids]
                    else:
                        records = []
                
                for record in records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                    
                    if 'cable_id' in columns:
                        cable_id_index = columns.index('cable_id')
                        cable_name = cable_name_map.get(str(record.cable_id), '')
                        row.insert(cable_id_index + 1, cable_name)
                    
                    ws.append(row)
                    total_records += 1
                    
            else:
                # Standard handling for other sheets
                ws.append(columns)
                
                if sheet_name == 'cable':
                    # For location-specific download, already filtered by junction_box
                    sorted_records = sorted(records, key=lambda x: (
                        int(x.cable_id) if x.cable_id and str(x.cable_id).isdigit() else 0
                    ))
                else:
                    sorted_records = records
                
                for record in sorted_records:
                    row = []
                    for col in columns:
                        value = getattr(record, col, "")
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            value = ''
                        row.append(str(value) if value is not None else "")
                    ws.append(row)
                    total_records += 1
        
        # ========== SAVE XLSX TO DOWNLOAD FOLDER ==========
        # Create download directory
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        download_dir = os.path.join(base_dir, 'xlsx_download')
        os.makedirs(download_dir, exist_ok=True)
        
        # Generate filename with location (junction)
        timestamp = get_ist_now().strftime('%Y%m%d_%H%M%S')
        project_name_clean = re.sub(r'[^\w\-_\. ]', '', current_project.name).replace(' ', '_')
        location_name_clean = re.sub(r'[^\w\-_\. ]', '', location.junction_name).replace(' ', '_') if location.junction_name else f"junction_{location_id}"
        filename = f"PROJECT_{project_id}_{project_name_clean}_JUNCTION_{location.junction_id}_{location_name_clean}_{timestamp}.xlsx"
        file_path = os.path.join(download_dir, filename)
        
        # Save the workbook
        wb.save(file_path)
        
        # ========== FLASH MESSAGE ==========
        flash(f"✅ XLSX for junction box '{location.junction_name}' downloaded successfully!", "success")
        
        # ========== RETURN FILE FOR DOWNLOAD ==========
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error generating location download: {str(e)}", "error")
        print(f"❌ Location download error: {str(e)}")
        import traceback
        traceback.print_exc()
        return redirect(request.referrer or url_for("main.index"))


@bp.route('/get_location_status/<int:project_id>/<string:junction_name>')
@login_required
def get_location_status(project_id, junction_name):
    """Get the status of a specific location (junction)"""
    try:
        # Decode the junction name if it's URL encoded
        from urllib.parse import unquote
        junction_name = unquote(junction_name)
        
        # Get the project
        project = Project.query.get_or_404(project_id)
        
        # Check if project has a PDF
        latest_pdf = GeneratedPDF.query.filter_by(project_id=project_id)\
            .order_by(GeneratedPDF.created_at.desc()).first()
        
        # Check if this location exists in junction boxes
        junction_box = JunctionBox.query.filter_by(
            project_id=project_id,
            junction_name=junction_name
        ).first()
        
        if not junction_box:
            return jsonify({
                'status': 'no_data',
                'color': 'danger',
                'message': 'Location not found'
            })
        
        # Check if there are any cables for this location
        cables_count = Cable.query.filter_by(
            project_id=project_id,
            junction_name=junction_name
        ).count()
        
        # Check if there are any terminals for cables in this location
        terminals_count = 0
        if cables_count > 0:
            # Get all cable IDs for this location
            cable_ids = [c.cable_id for c in Cable.query.filter_by(
                project_id=project_id,
                junction_name=junction_name
            ).all()]
            
            if cable_ids:
                terminals_count = Terminal.query.filter(
                    Terminal.project_id == project_id,
                    Terminal.cable_id.in_(cable_ids)
                ).count()
        
        # Check project stage
        project_stage = project.stage or 1
        
        # Determine status based on data and stage
        if cables_count == 0 and terminals_count == 0:
            status = 'no_data'
            color = 'danger'
        elif project_stage < 9:
            status = 'in_progress'
            color = 'warning'
        elif project_stage >= 9:
            # Check if fully approved
            if latest_pdf and latest_pdf.level3_status == 'approved':
                status = 'completed'
                color = 'success'
            else:
                status = 'pending_approval'
                color = 'info'
        else:
            status = 'in_progress'
            color = 'warning'
        
        return jsonify({
            'status': status,
            'color': color,
            'cables_count': cables_count,
            'terminals_count': terminals_count,
            'project_stage': project_stage,
            'has_pdf': latest_pdf is not None,
            'pdf_approved': latest_pdf.level3_status == 'approved' if latest_pdf else False
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'color': 'secondary',
            'message': str(e)
        }), 500


@bp.route('/get_all_locations_status/<int:project_id>')
@login_required
def get_all_locations_status(project_id):
    """Get status for all locations in a project"""
    try:
        # Get the project
        project = Project.query.get_or_404(project_id)
        
        # Get all unique location names for this project
        location_names = []
        
        # From junction boxes
        junction_boxes = JunctionBox.query.filter_by(project_id=project_id).all()
        for jb in junction_boxes:
            if jb.junction_name and jb.junction_name not in location_names:
                location_names.append(jb.junction_name)
        
        # From cables (in case some locations are only in cables)
        cables = Cable.query.filter_by(project_id=project_id).all()
        for cable in cables:
            if cable.junction_name and cable.junction_name not in location_names:
                location_names.append(cable.junction_name)
        
        # Get latest PDF
        latest_pdf = GeneratedPDF.query.filter_by(project_id=project_id)\
            .order_by(GeneratedPDF.created_at.desc()).first()
        
        project_stage = project.stage or 1
        
        # Get status for each location
        location_statuses = {}
        
        for location_name in location_names:
            # Check cables for this location
            cables_count = Cable.query.filter_by(
                project_id=project_id,
                junction_name=location_name
            ).count()
            
            # Check terminals
            terminals_count = 0
            if cables_count > 0:
                cable_ids = [c.cable_id for c in Cable.query.filter_by(
                    project_id=project_id,
                    junction_name=location_name
                ).all()]
                
                if cable_ids:
                    terminals_count = Terminal.query.filter(
                        Terminal.project_id == project_id,
                        Terminal.cable_id.in_(cable_ids)
                    ).count()
            
            # Determine status
            if cables_count == 0 and terminals_count == 0:
                status = 'no_data'
                color = 'danger'
                badge_class = 'bg-danger'
            elif project_stage < 9:
                status = 'in_progress'
                color = 'warning'
                badge_class = 'bg-warning'
            elif project_stage >= 9:
                if latest_pdf and latest_pdf.level3_status == 'approved':
                    status = 'completed'
                    color = 'success'
                    badge_class = 'bg-success'
                else:
                    status = 'pending_approval'
                    color = 'info'
                    badge_class = 'bg-info'
            else:
                status = 'in_progress'
                color = 'warning'
                badge_class = 'bg-warning'
            
            location_statuses[location_name] = {
                'status': status,
                'color': color,
                'badge_class': badge_class,
                'cables_count': cables_count,
                'terminals_count': terminals_count
            }
        
        return jsonify({
            'project_id': project_id,
            'project_stage': project_stage,
            'location_statuses': location_statuses,
            'total_locations': len(location_names)
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500


# Add this helper function to your template context
@bp.context_processor
def utility_processor():
    def get_junction_approvals(pdf_id, junction_id):
        """Get approval status for a specific junction"""
        # This is a placeholder - you'll need to implement based on your data structure
        # For now, return empty list
        return []
    
    return dict(get_junction_approvals=get_junction_approvals)


@bp.route('/get_location_approval_status/<int:project_id>/<location_id>')
@login_required
def get_location_approval_status(project_id, location_id):
    """Get approval status for all locations or a specific location in a project"""
    try:
        project = Project.query.get_or_404(project_id)
        
        # Get all locations for this project
        if location_id == 'all':
            junctions = project.junction_boxes
        else:
            junctions = [JunctionBox.query.get_or_404(location_id)]
        
        # Get the latest PDF for this project
        latest_pdf = GeneratedPDF.query.filter_by(project_id=project_id)\
            .order_by(GeneratedPDF.created_at.desc()).first()
        
        locations_data = []
        
        for junction in junctions:
            if not junction.junction_name:
                continue
            
            # Look for junction approval records
            if latest_pdf:
                # Try to find junction approval for this junction and PDF
                junction_approval = JunctionApproval.query.filter_by(
                    project_id=project_id,
                    generated_pdf_id=latest_pdf.id,
                    junction_box_id=junction.id
                ).first()
                
                if junction_approval:
                    # Get approver names
                    level1_approver_name = junction_approval.level1_approver.username if junction_approval.level1_approver else None
                    level2_approver_name = junction_approval.level2_approver.username if junction_approval.level2_approver else None
                    level3_approver_name = junction_approval.level3_approver.username if junction_approval.level3_approver else None
                    
                    # Format dates
                    level1_date = junction_approval.level1_approval_date.strftime('%d-%m-%Y %H:%M') if junction_approval.level1_approval_date else None
                    level2_date = junction_approval.level2_approval_date.strftime('%d-%m-%Y %H:%M') if junction_approval.level2_approval_date else None
                    level3_date = junction_approval.level3_approval_date.strftime('%d-%m-%Y %H:%M') if junction_approval.level3_approval_date else None
                    
                    location_info = {
                        'id': junction.id,
                        'name': junction.junction_name,
                        'status': junction.status or '1',
                        'level1': {
                            'status': junction_approval.level1_status,
                            'approver': level1_approver_name,
                            'date': level1_date
                        },
                        'level2': {
                            'status': junction_approval.level2_status,
                            'approver': level2_approver_name,
                            'date': level2_date
                        },
                        'level3': {
                            'status': junction_approval.level3_status,
                            'approver': level3_approver_name,
                            'date': level3_date
                        },
                        'rejection_reason': junction_approval.rejection_reason
                    }
                else:
                    # No junction approval record exists yet
                    location_info = {
                        'id': junction.id,
                        'name': junction.junction_name,
                        'status': junction.status or '1',
                        'level1': {'status': 'pending', 'approver': None, 'date': None},
                        'level2': {'status': 'pending', 'approver': None, 'date': None},
                        'level3': {'status': 'pending', 'approver': None, 'date': None},
                        'rejection_reason': None
                    }
            else:
                # No PDF exists yet
                location_info = {
                    'id': junction.id,
                    'name': junction.junction_name,
                    'status': junction.status or '1',
                    'level1': {'status': 'pending', 'approver': None, 'date': None},
                    'level2': {'status': 'pending', 'approver': None, 'date': None},
                    'level3': {'status': 'pending', 'approver': None, 'date': None},
                    'rejection_reason': None
                }
            
            locations_data.append(location_info)
        
        return jsonify({
            'success': True,
            'project_id': project_id,
            'location_id': location_id,
            'locations': locations_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/projects/<int:project_id>/latest-pdf')
@login_required
def get_latest_pdf(project_id):
    """
    Get the latest PDF for a project.
    If location_id is provided, returns the PDF associated with that junction box.
    Otherwise returns the most recent project-level GeneratedPDF.
    """
    location_id = request.args.get('location_id', type=int)
    print(location_id)
    try:
        print(f"📄 DEBUG: Fetching latest PDF for project {project_id}" +
              (f", location {location_id}" if location_id else ""))

        # ---------- LOCATION-SPECIFIC PDF ----------
        if location_id:
            location = JunctionBox.query.filter_by(
                id=location_id,
                project_id=project_id
            ).first()
            print(location)
            if not location:
                return jsonify({
                    'success': False,
                    'error': 'Location not found in this project'
                }), 404

            # Find the latest JunctionApproval for this location with a linked PDF
            junction_approval = JunctionApproval.query.filter_by(
                project_id=project_id,
                junction_box_id=location.id
            ).filter(
                JunctionApproval.generated_pdf_id.isnot(None)
            ).order_by(
                JunctionApproval.created_at.desc()
            ).first()
            print (junction_approval)
            if junction_approval and junction_approval.generated_pdf_id:
                pdf = GeneratedPDF.query.get(junction_approval.generated_pdf_id)
                print(pdf)   
                if pdf:
                    project = Project.query.get(project_id)
                    project_name = project.name if project else f"Project {project_id}"

                    # ---- PDF data (with correct keys) ----
                    pdf_data = {
                        'id': pdf.id,
                        'pdf_filename': pdf.pdf_filename,
                        'project_id': pdf.project_id,
                        'version': pdf.version,
                        'created_at': pdf.created_at.isoformat() if pdf.created_at else None,
                        # Approval fields – exactly as frontend expects
                        'level1_status': pdf.level1_status,
                        'level2_status': pdf.level2_status,
                        'level3_status': pdf.level3_status,
                        'level1_approver': pdf.level1_approver.username if pdf.level1_approver else None,
                        'level2_approver': pdf.level2_approver.username if pdf.level2_approver else None,
                        'level3_approver': pdf.level3_approver.username if pdf.level3_approver else None,
                        'level1_approval_date': pdf.level1_approval_date.isoformat() if pdf.level1_approval_date else None,
                        'level2_approval_date': pdf.level2_approval_date.isoformat() if pdf.level2_approval_date else None,
                        'level3_approval_date': pdf.level3_approval_date.isoformat() if pdf.level3_approval_date else None,
                    }

                    # ---- Junction‑specific approval data (overrides PDF status) ----
                    junction_approval_data = {
                        'level1_status': junction_approval.level1_status,
                        'level2_status': junction_approval.level2_status,
                        'level3_status': junction_approval.level3_status,
                        'level1_approver': junction_approval.level1_approver.username if junction_approval.level1_approver else None,
                        'level2_approver': junction_approval.level2_approver.username if junction_approval.level2_approver else None,
                        'level3_approver': junction_approval.level3_approver.username if junction_approval.level3_approver else None,
                        'level1_approval_date': junction_approval.level1_approval_date.isoformat() if junction_approval.level1_approval_date else None,
                        'level2_approval_date': junction_approval.level2_approval_date.isoformat() if junction_approval.level2_approval_date else None,
                        'level3_approval_date': junction_approval.level3_approval_date.isoformat() if junction_approval.level3_approval_date else None,
                    }

                    print(f"✅ Found location PDF: {pdf.pdf_filename} for junction {location.junction_name}")

                    return jsonify({
                        'success': True,
                        'pdf': pdf_data,
                        'junction_approval': junction_approval_data,
                        'project_name': project_name,
                        'location_name': location.junction_name,
                        'location_id': location.id
                    })

            # No PDF linked to this location
            return jsonify({
                'success': False,
                'error': 'No PDF found for this location'
            }), 404

        # ---------- PROJECT-LEVEL PDF (original behaviour) ----------
        else:
            latest_pdf = GeneratedPDF.query.filter_by(
                project_id=project_id
            ).order_by(GeneratedPDF.created_at.desc()).first()

            if not latest_pdf:
                print(f"❌ No PDF found for project {project_id}")
                return jsonify({
                    'success': False,
                    'error': 'No PDF found for this project'
                }), 404

            project = Project.query.get(project_id)
            project_name = project.name if project else f"Project {project_id}"

            pdf_data = {
                'id': latest_pdf.id,
                'pdf_filename': latest_pdf.pdf_filename,
                'project_id': latest_pdf.project_id,
                'version': latest_pdf.version,
                'created_at': latest_pdf.created_at.isoformat() if latest_pdf.created_at else None,
                'level1_status': latest_pdf.level1_status,
                'level2_status': latest_pdf.level2_status,
                'level3_status': latest_pdf.level3_status,
                'level1_approver': latest_pdf.level1_approver.username if latest_pdf.level1_approver else None,
                'level2_approver': latest_pdf.level2_approver.username if latest_pdf.level2_approver else None,
                'level3_approver': latest_pdf.level3_approver.username if latest_pdf.level3_approver else None,
                'level1_approval_date': latest_pdf.level1_approval_date.isoformat() if latest_pdf.level1_approval_date else None,
                'level2_approval_date': latest_pdf.level2_approval_date.isoformat() if latest_pdf.level2_approval_date else None,
                'level3_approval_date': latest_pdf.level3_approval_date.isoformat() if latest_pdf.level3_approval_date else None,
            }

            print(f"✅ Found project PDF: {latest_pdf.pdf_filename} for project {project_id}")

            return jsonify({
                'success': True,
                'pdf': pdf_data,
                'project_name': project_name
            })

    except Exception as e:
        print(f"❌ Error fetching latest PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/pdfs/<int:pdf_id>/view')
@login_required
def view_pdf(pdf_id):
    """View a PDF inline"""
    try:
        pdf_record = GeneratedPDF.query.get(pdf_id)
        if not pdf_record:
            return "PDF not found", 404
        
        # Get the uploads directory path
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        uploads_dir = os.path.join(base_dir, 'uploads')
        pdf_path = os.path.join(uploads_dir, pdf_record.pdf_filename)
        
        if not os.path.exists(pdf_path):
            return "PDF file not found on server", 404
        
        # Send the PDF for inline viewing
        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=pdf_record.pdf_filename
        )
        
    except Exception as e:
        print(f"❌ Error viewing PDF: {str(e)}")
        return str(e), 500

@bp.route('/api/pdfs/<int:pdf_id>/download')
@login_required
def download_pdf_by_id(pdf_id):
    """Download a PDF by its database ID"""
    try:
        pdf_record = GeneratedPDF.query.get(pdf_id)
        if not pdf_record:
            return "PDF not found", 404
        
        # Get the uploads directory path
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        uploads_dir = os.path.join(base_dir, 'uploads')
        pdf_path = os.path.join(uploads_dir, pdf_record.pdf_filename)
        
        if not os.path.exists(pdf_path):
            return "PDF file not found on server", 404
        
        # Send the PDF for download
        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=pdf_record.pdf_filename
        )
        
    except Exception as e:
        print(f"❌ Error downloading PDF: {str(e)}")
        return str(e), 500



#######3



@bp.route('/api/pdfs/<int:pdf_id>/download', methods=['GET'])
@login_required
def download_pdf_api(pdf_id):
    """Download a PDF"""
    try:
        # Get PDF record
        pdf_record = GeneratedPDF.query.get(pdf_id)
        if not pdf_record:
            return jsonify({'success': False, 'message': 'PDF record not found'}), 404
        
        # Get file path
        import os
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        uploads_dir = os.path.join(base_dir, 'uploads')
        pdf_path = os.path.join(uploads_dir, pdf_record.pdf_filename)
        
        if not os.path.exists(pdf_path):
            return jsonify({'success': False, 'message': 'PDF file not found'}), 404
        
        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=pdf_record.pdf_filename
        )
        
    except Exception as e:
        print(f"❌ Error downloading PDF: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# Also add this helper function for the debug endpoint that already exists


@bp.route("/debug/routes")
def debug_routes():
    """Debug endpoint to see all registered routes"""
    import flask
    routes = []
    for rule in flask.current_app.url_map.iter_rules():
        routes.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'rule': str(rule)
        })
    return jsonify(routes)


#### add by yash 



@bp.route("/approve_junction/<int:junction_id>/<int:pdf_id>/<level>", methods=["POST"])
@login_required
def approve_junction(junction_id, pdf_id, level):
    """Approve a specific junction at specified level"""
    try:
        junction = JunctionBox.query.get_or_404(junction_id)
        pdf = GeneratedPDF.query.get_or_404(pdf_id)
        project = Project.query.get_or_404(junction.project_id)
        
        # Get approval remarks from form
        approval_remarks = request.form.get('approval_remarks', '')
        
        # Get user permissions based on role
        permissions = get_user_permissions(current_user)
        
        # Check if junction approval record exists, create if not
        junction_approval = JunctionApproval.query.filter_by(
            project_id=project.id,
            generated_pdf_id=pdf.id,
            junction_box_id=junction.id
        ).first()
        
        if not junction_approval:
            junction_approval = JunctionApproval(
                project_id=project.id,
                generated_pdf_id=pdf.id,
                junction_box_id=junction.id,
                level1_status='pending',
                level2_status='pending', 
                level3_status='pending',
                created_at=get_ist_now()
            )
            db.session.add(junction_approval)
        
        # Initialize status if None (for existing records)
        if junction_approval.level1_status is None:
            junction_approval.level1_status = 'pending'
        if junction_approval.level2_status is None:
            junction_approval.level2_status = 'pending'
        if junction_approval.level3_status is None:
            junction_approval.level3_status = 'pending'
        
        if level == '1':
            # ===== LEVEL 1 APPROVAL =====
            if not permissions['can_approve_level1']:
                flash("Access denied. You don't have permission to approve at Level 1.", "danger")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Check if already processed
            if junction_approval.level1_status != 'pending':
                flash(f"This junction has already been {junction_approval.level1_status} at Level 1.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Update junction approval status
            junction_approval.level1_status = 'approved'
            junction_approval.level1_approver_id = current_user.id
            junction_approval.level1_approval_date = get_ist_now()
            
            # Also update junction box status to "Under Process" (status=2)
            junction.status = 2
            
            # Record approval in Approval table
            approval = Approval(
                generated_pdf_id=pdf.id,
                level=1,
                status='approved',
                approver_id=current_user.id,
                remarks=approval_remarks,
                created_at=get_ist_now()
            )
            db.session.add(approval)
            
            flash(f"✅ Junction '{junction.junction_name}' approved at Level 1!", "success")
            
        elif level == '2':
            # ===== LEVEL 2 APPROVAL =====
            if not permissions['can_approve_level2']:
                flash("Access denied. You don't have permission to approve at Level 2.", "danger")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Check prerequisites
            if junction_approval.level1_status != 'approved':
                flash("This junction must be approved at Level 1 first.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            if junction_approval.level2_status != 'pending':
                flash(f"This junction has already been {junction_approval.level2_status} at Level 2.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Update junction approval status
            junction_approval.level2_status = 'approved'
            junction_approval.level2_approver_id = current_user.id
            junction_approval.level2_approval_date = get_ist_now()
            
            # Record approval in Approval table
            approval = Approval(
                generated_pdf_id=pdf.id,
                level=2,
                status='approved',
                approver_id=current_user.id,
                remarks=approval_remarks,
                created_at=get_ist_now()
            )
            db.session.add(approval)
            
            flash(f"✅ Junction '{junction.junction_name}' approved at Level 2!", "success")
            
        elif level == '3':
            # ===== LEVEL 3 APPROVAL (FINAL) =====
            if not permissions['can_approve_level3']:
                flash("Access denied. You don't have permission to approve at Level 3.", "danger")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Check prerequisites
            if junction_approval.level2_status != 'approved':
                flash("This junction must be approved at Level 2 first.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            if junction_approval.level3_status != 'pending':
                flash(f"This junction has already been {junction_approval.level3_status} at Level 3.", "warning")
                return redirect(request.referrer or url_for('main.approval_tracking'))
            
            # Update junction approval status
            junction_approval.level3_status = 'approved'
            junction_approval.level3_approver_id = current_user.id
            junction_approval.level3_approval_date = get_ist_now()
            
            # Update junction box status to "Completed" (status=3)
            junction.status = 3
            
            # Record approval in Approval table
            approval = Approval(
                generated_pdf_id=pdf.id,
                level=3,
                status='approved',
                approver_id=current_user.id,
                remarks=approval_remarks,
                created_at=get_ist_now()
            )
            db.session.add(approval)
            
            flash(f"✅ Junction '{junction.junction_name}' FULLY APPROVED at Level 3!", "success")
        
        else:
            flash("Invalid approval level.", "danger")
            return redirect(request.referrer or url_for('main.approval_tracking'))
        
        db.session.commit()
        
        # Check if all junctions in project are approved
        check_and_update_project_status(project.id, pdf.id)
        
        return redirect(request.referrer or url_for('main.approval_tracking'))
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error approving junction: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f"Error approving junction: {str(e)}", "danger")
        return redirect(request.referrer or url_for('main.approval_tracking'))

@bp.route("/reject_junction/<int:junction_id>/<int:pdf_id>/<level>", methods=["POST"])
@login_required
def reject_junction(junction_id, pdf_id, level):
    """Reject a specific junction at specified level"""
    try:
        junction = JunctionBox.query.get_or_404(junction_id)
        pdf = GeneratedPDF.query.get_or_404(pdf_id)
        project = Project.query.get_or_404(junction.project_id)
        rejection_reason = request.form.get('rejection_reason', 'No reason provided')
        
        # Get user permissions based on role
        permissions = get_user_permissions(current_user)
        
        # Check if junction approval record exists, create if not
        junction_approval = JunctionApproval.query.filter_by(
            project_id=project.id,
            generated_pdf_id=pdf.id,
            junction_box_id=junction.id
        ).first()
        
        if not junction_approval:
            junction_approval = JunctionApproval(
                project_id=project.id,
                generated_pdf_id=pdf.id,
                junction_box_id=junction.id,
                created_at=get_ist_now()
            )
            db.session.add(junction_approval)
        
        # Check if user has permission for this level
        level_key = f'can_approve_level{level}'
        if not permissions.get(level_key):
            flash(f"Access denied. You don't have permission to reject at Level {level}.", "danger")
            return redirect(request.referrer or url_for('main.approval_tracking'))
        
        # Update junction approval status
        if level == '1':
            junction_approval.level1_status = 'rejected'
            junction_approval.level1_approver_id = current_user.id
            junction_approval.level1_approval_date = get_ist_now()
        elif level == '2':
            junction_approval.level2_status = 'rejected'
            junction_approval.level2_approver_id = current_user.id
            junction_approval.level2_approval_date = get_ist_now()
        elif level == '3':
            junction_approval.level3_status = 'rejected'
            junction_approval.level3_approver_id = current_user.id
            junction_approval.level3_approval_date = get_ist_now()
        
        junction_approval.rejection_reason = rejection_reason
        
        # Update junction box status to "No Drawing Data" (status=1)
        junction.status = 1
        
        # Record rejection in Approval table
        approval = Approval(
            generated_pdf_id=pdf.id,
            level=int(level),
            status='rejected',
            approver_id=current_user.id,
            remarks=rejection_reason,
            created_at=get_ist_now()
        )
        db.session.add(approval)
        
        db.session.commit()
        flash(f"✅ Junction '{junction.junction_name}' rejected at Level {level}.", "success")
        return redirect(request.referrer or url_for('main.approval_tracking'))
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error rejecting junction: {str(e)}")
        flash(f"Error rejecting junction: {str(e)}", "danger")
        return redirect(request.referrer or url_for('main.approval_tracking'))

def check_and_update_project_status(project_id, pdf_id):
    """Check if all junctions in project are approved and update project status"""
    try:
        # Get all junctions for this project
        junctions = JunctionBox.query.filter_by(project_id=project_id).all()
        
        # Get all junction approvals for this PDF
        junction_approvals = JunctionApproval.query.filter_by(
            project_id=project_id,
            generated_pdf_id=pdf_id
        ).all()
        
        if not junctions or not junction_approvals:
            return False
        
        # Check status of all junctions
        all_approved = True
        any_rejected = False
        
        for junction in junctions:
            # Find approval for this junction
            approval = next((ja for ja in junction_approvals if ja.junction_box_id == junction.id), None)
            
            if not approval:
                all_approved = False
                break
            
            if not approval.is_fully_approved():
                all_approved = False
            
            if approval.get_approval_status() == 'rejected':
                any_rejected = True
        
        # Update project status based on junction approvals
        project = Project.query.get(project_id)
        if project:
            if all_approved:
                project.status = 'approved'
            elif any_rejected:
                project.status = 'rejected'
            else:
                project.status = 'in_progress'
            
            db.session.commit()
            
        return all_approved
        
    except Exception as e:
        print(f"❌ Error checking project status: {str(e)}")
        return False

@bp.route("/get_junction_approval_status/<int:project_id>/<junction_id>")
@login_required
def get_junction_approval_status(project_id, junction_id):
    """Get approval status for a specific junction or all junctions in project"""
    try:
        project = Project.query.get_or_404(project_id)
        
        if junction_id == 'all':
            # Get all junctions for this project
            junctions = JunctionBox.query.filter_by(project_id=project_id).all()
            result = []
            
            for junction in junctions:
                # Get the latest PDF for this project
                latest_pdf = GeneratedPDF.query.filter_by(
                    project_id=project_id
                ).order_by(GeneratedPDF.created_at.desc()).first()
                
                if latest_pdf:
                    junction_approval = JunctionApproval.query.filter_by(
                        project_id=project_id,
                        generated_pdf_id=latest_pdf.id,
                        junction_box_id=junction.id
                    ).first()
                    
                    if junction_approval:
                        result.append({
                            'id': junction.id,
                            'name': junction.junction_name,
                            'status': junction.status,
                            'level1': {
                                'status': junction_approval.level1_status,
                                'approver': junction_approval.level1_approver.username if junction_approval.level1_approver else None,
                                'date': junction_approval.level1_approval_date.strftime('%Y-%m-%d %H:%M:%S') if junction_approval.level1_approval_date else None
                            },
                            'level2': {
                                'status': junction_approval.level2_status,
                                'approver': junction_approval.level2_approver.username if junction_approval.level2_approver else None,
                                'date': junction_approval.level2_approval_date.strftime('%Y-%m-%d %H:%M:%S') if junction_approval.level2_approval_date else None
                            },
                            'level3': {
                                'status': junction_approval.level3_status,
                                'approver': junction_approval.level3_approver.username if junction_approval.level3_approver else None,
                                'date': junction_approval.level3_approval_date.strftime('%Y-%m-%d %H:%M:%S') if junction_approval.level3_approval_date else None
                            }
                        })
            
            return jsonify({
                'success': True,
                'project_name': project.name,
                'locations': result
            })
        
        else:
            # Get specific junction
            junction = JunctionBox.query.get_or_404(junction_id)
            
            # Get the latest PDF for this project
            latest_pdf = GeneratedPDF.query.filter_by(
                project_id=project_id
            ).order_by(GeneratedPDF.created_at.desc()).first()
            
            junction_approval = None
            if latest_pdf:
                junction_approval = JunctionApproval.query.filter_by(
                    project_id=project_id,
                    generated_pdf_id=latest_pdf.id,
                    junction_box_id=junction.id
                ).first()
            
            return jsonify({
                'success': True,
                'junction': {
                    'id': junction.id,
                    'name': junction.junction_name,
                    'status': junction.status,
                    'approval': {
                        'level1': junction_approval.level1_status if junction_approval else 'pending',
                        'level2': junction_approval.level2_status if junction_approval else 'pending',
                        'level3': junction_approval.level3_status if junction_approval else 'pending'
                    } if junction_approval else None
                }
            })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/project/<int:project_id>/all_pdfs')
@login_required
def get_all_pdfs_for_project(project_id):
    """Get all PDFs for a specific project as JSON."""
    try:
        # Check if user has access to this project
        permissions = get_user_permissions(current_user)
        
        if permissions['can_see_all']:
            # Admin can see all
            pass
        else:
            # Check if user is assigned to this project
            if current_user.projects and project_id not in [p.id for p in current_user.projects]:
                return jsonify({'success': False, 'message': 'Access denied'}), 403
        
        # Get the project
        project = Project.query.get(project_id)
        if not project:
            return jsonify({'success': False, 'message': 'Project not found'}), 404
        
        # Get all PDFs for this project, ordered by created_at (latest first)
        pdf_records = GeneratedPDF.query.filter_by(project_id=project_id)\
            .order_by(GeneratedPDF.created_at.desc())\
            .all()
        
        import json
        from flask import url_for
        from datetime import datetime
        
        all_pdfs = []
        
        # Method 1: From database records
        for pdf in pdf_records:
            # Get SPECIFIC location names for this PDF version
            specific_locations = []
            
            # Option 1: Try to get from junction_data field (primary source)
            if pdf.junction_data:
                try:
                    data = json.loads(pdf.junction_data)
                    if isinstance(data, list):
                        for jb in data:
                            if isinstance(jb, dict):
                                jb_name = jb.get('junction_name') or jb.get('name') or jb.get('junctionName') or jb.get('junction')
                                if jb_name:
                                    specific_locations.append(str(jb_name))
                            elif isinstance(jb, str):
                                specific_locations.append(jb)
                    elif isinstance(data, dict):
                        if 'junction_boxes' in data:
                            jbs = data['junction_boxes']
                            if isinstance(jbs, list):
                                for jb in jbs:
                                    if isinstance(jb, dict):
                                        jb_name = jb.get('junction_name') or jb.get('name') or jb.get('junctionName')
                                        if jb_name:
                                            specific_locations.append(str(jb_name))
                        elif 'locations' in data:
                            locs = data['locations']
                            if isinstance(locs, list):
                                specific_locations.extend([str(loc) for loc in locs if loc])
                        elif 'junction_names' in data:
                            names = data['junction_names']
                            if isinstance(names, list):
                                specific_locations.extend([str(name) for name in names if name])
                        else:
                            for key, value in data.items():
                                if isinstance(value, str) and ('jb' in key.lower() or 'junction' in key.lower() or 'box' in key.lower()):
                                    specific_locations.append(value)
                                elif isinstance(value, list):
                                    for item in value:
                                        if isinstance(item, str):
                                            specific_locations.append(item)
                except json.JSONDecodeError:
                    try:
                        if isinstance(pdf.junction_data, str):
                            cleaned = pdf.junction_data.strip()
                            import re
                            cleaned = re.sub(r'[\[\]{}"]', '', cleaned)
                            for delim in [',', ';', '|', '\n']:
                                if delim in cleaned:
                                    parts = [p.strip() for p in cleaned.split(delim) if p.strip()]
                                    if parts:
                                        specific_locations = parts
                                        break
                            if not specific_locations and cleaned:
                                specific_locations = [cleaned]
                    except Exception as e:
                        print(f"Alternative parsing failed for PDF {pdf.id}: {e}")
                except Exception as e:
                    print(f"Error parsing junction_data for PDF {pdf.id}: {e}")
                    if isinstance(pdf.junction_data, str) and pdf.junction_data.strip():
                        specific_locations = [pdf.junction_data.strip()]
            
            # Option 2: Get from JunctionApproval records (if available)
            if not specific_locations and hasattr(pdf, 'junction_approvals'):
                try:
                    for ja in pdf.junction_approvals:
                        if ja.junction_box and ja.junction_box.junction_name:
                            specific_locations.append(ja.junction_box.junction_name)
                except Exception as e:
                    print(f"Error getting locations from junction_approvals for PDF {pdf.id}: {e}")
            
            # Option 3: Try to infer from filename
            if not specific_locations and pdf.pdf_filename:
                try:
                    filename = pdf.pdf_filename
                    import re
                    patterns = [
                        r'JB[\s_]?BOX[\s_]*NO[\s_]*[\w\d]+[A-Z]?',
                        r'AC[\s_]*[\w\d]+',
                        r'JB[\s_]*[\w\d]+',
                        r'BOX[\s_]*[\w\d]+',
                        r'JN[\s_]*[\w\d]+',
                        r'LOC[\s_]*[\w\d]+',
                    ]
                    for pattern in patterns:
                        matches = re.findall(pattern, filename, re.IGNORECASE)
                        for match in matches:
                            cleaned = re.sub(r'[\s_]+', ' ', match).strip()
                            if cleaned and cleaned not in specific_locations:
                                specific_locations.append(cleaned)
                except Exception as e:
                    print(f"Error extracting locations from filename for PDF {pdf.id}: {e}")
            
            # ✅ FIX: pdf.created_at is now timezone‑aware UTC – convert to IST for display
            created_at_ist = None
            if pdf.created_at:
                created_at_ist = pdf.created_at.astimezone(IST)  # convert UTC to IST
            
            version_str = str(pdf.version) if pdf.version else "1"
            
            if specific_locations:
                specific_locations = list(set([loc.strip() for loc in specific_locations if loc and loc.strip()]))
                specific_locations.sort()
            
            # ✅ Use correct endpoint names inline_pdf and download_pdf
            all_pdfs.append({
                'id': pdf.id,
                'filename': pdf.pdf_filename,
                'version': version_str,
                'created_at': created_at_ist.isoformat() if created_at_ist else '',
                'created_at_raw': pdf.created_at.isoformat() if pdf.created_at else '',
                'specific_locations': specific_locations,
                'location_names': specific_locations[:5],
                'location_count': len(specific_locations) if specific_locations else 0,
                'level1_status': pdf.level1_status or 'pending',
                'level2_status': pdf.level2_status or 'pending',
                'level3_status': pdf.level3_status or 'pending',
                'inline_url': url_for('main.inline_pdf', filename=pdf.pdf_filename),
                'download_url': url_for('main.download_pdf', filename=pdf.pdf_filename),
                'checksum_md5': pdf.checksum_md5,
                'file_size': pdf.file_size,
                'source': 'database'
            })
        
        # Method 2: Scan filesystem for any additional PDFs (unchanged)...
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        uploads_dir = os.path.join(base_dir, 'uploads')
        
        if os.path.exists(uploads_dir):
            prefix = f"railway_project_{project_id}_"
            for filename in os.listdir(uploads_dir):
                if filename.endswith(".pdf") and filename.startswith(prefix):
                    if not any(pdf['filename'] == filename for pdf in all_pdfs):
                        try:
                            filepath = os.path.join(uploads_dir, filename)
                            stat = os.stat(filepath)
                            size_kb = max(1, stat.st_size // 1024)
                            mtime = datetime.fromtimestamp(stat.st_mtime)
                            
                            version = "1"
                            if "_v" in filename.lower():
                                import re
                                match = re.search(r'_v(\d+)', filename, re.IGNORECASE)
                                if match:
                                    version = match.group(1)
                            
                            specific_locations = []
                            try:
                                patterns = [
                                    r'JB[\s_]?BOX[\s_]*NO[\s_]*[\w\d]+[A-Z]?',
                                    r'AC[\s_]*[\w\d]+',
                                    r'JB[\s_]*[\w\d]+',
                                    r'BOX[\s_]*[\w\d]+',
                                    r'JN[\s_]*[\w\d]+',
                                    r'LOC[\s_]*[\w\d]+',
                                ]
                                for pattern in patterns:
                                    matches = re.findall(pattern, filename, re.IGNORECASE)
                                    for match in matches:
                                        cleaned = re.sub(r'[\s_]+', ' ', match).strip()
                                        if cleaned and cleaned not in specific_locations:
                                            specific_locations.append(cleaned)
                            except Exception as e:
                                print(f"Error extracting locations from filesystem filename {filename}: {e}")
                            
                            all_pdfs.append({
                                'id': None,
                                'filename': filename,
                                'version': version,
                                'created_at': mtime.strftime('%Y-%m-%d %H:%M:%S'),
                                'created_at_raw': mtime.isoformat(),
                                'specific_locations': specific_locations,
                                'location_names': specific_locations[:5],
                                'location_count': len(specific_locations) if specific_locations else 0,
                                'level1_status': 'pending',
                                'level2_status': 'pending',
                                'level3_status': 'pending',
                                'inline_url': url_for('main.inline_pdf', filename=filename),
                                'download_url': url_for('main.download_pdf', filename=filename),
                                'checksum_md5': None,
                                'file_size': size_kb,
                                'source': 'filesystem'
                            })
                        except Exception as e:
                            print(f"Error processing file {filename}: {e}")
                            continue
        
        # Sort by created date (most recent first)
        all_pdfs.sort(key=lambda x: x.get('created_at_raw', ''), reverse=True)
        
        print(f"DEBUG: Found {len(all_pdfs)} PDFs for project {project_id}")
        for pdf in all_pdfs[:3]:
            print(f"  - {pdf['filename']} (v{pdf['version']}) - Specific locations: {pdf['specific_locations']}")
        
        return jsonify({
            'success': True,
            'project_id': project_id,
            'project_name': project.name,
            'pdfs': all_pdfs,
            'total_pdfs': len(all_pdfs)
        })
        
    except Exception as e:
        print(f"Error in get_all_pdfs_for_project: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error loading PDFs: {str(e)}'
        }), 500

@bp.route('/test_pdf_route/<int:project_id>')
@login_required
def test_pdf_route(project_id):
    """Test endpoint to check if the PDF route is working."""
    return jsonify({
        'success': True,
        'message': f'Route is working for project {project_id}',
        'test_url': f'/project/{project_id}/all_pdfs'
    })


@bp.route('/sync_all_pdf_approvals')
@login_required
def sync_all_pdf_approvals():
    """Sync all PDF approvals to per-location approvals"""
    try:
        all_pdfs = GeneratedPDF.query.filter(
            (GeneratedPDF.level1_status == 'approved') | 
            (GeneratedPDF.level1_status == 'rejected') |
            (GeneratedPDF.level2_status == 'approved') | 
            (GeneratedPDF.level2_status == 'rejected') |
            (GeneratedPDF.level3_status == 'approved') | 
            (GeneratedPDF.level3_status == 'rejected')
        ).all()
        
        count = 0
        for pdf in all_pdfs:
            project = pdf.project
            junctions = project.junction_boxes
            
            for junction in junctions:
                if not junction.junction_name:
                    continue
                
                # Check if junction approval exists
                existing_approval = JunctionApproval.query.filter_by(
                    project_id=project.id,
                    generated_pdf_id=pdf.id,
                    junction_box_id=junction.id
                ).first()
                
                if not existing_approval:
                    # Create new junction approval
                    junction_approval = JunctionApproval(
                        project_id=project.id,
                        generated_pdf_id=pdf.id,
                        junction_box_id=junction.id,
                        level1_status=pdf.level1_status,
                        level2_status=pdf.level2_status,
                        level3_status=pdf.level3_status,
                        level1_approver_id=pdf.level1_approver_id,
                        level2_approver_id=pdf.level2_approver_id,
                        level3_approver_id=pdf.level3_approver_id,
                        level1_approval_date=pdf.level1_approval_date,
                        level2_approval_date=pdf.level2_approval_date,
                        level3_approval_date=pdf.level3_approval_date
                    )
                    db.session.add(junction_approval)
                    count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Synced {count} junction approvals from {len(all_pdfs)} PDFs'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



### def get_role_id_by_name(role_name):
    """Get role ID from role name string (e.g., '1', '2', '3', '4')"""
    if not role_name:
        return None
    
    # Try to find role by name in RoleMaster
    role = RoleMaster.query.filter_by(role_name=role_name).first()
    if role:
        return role.id
    
    # Fallback: create a mapping for common roles
    role_mapping = {
        '0': None,  # Viewer role might not exist in RoleMaster
        '1': None,  # Creator role
        '2': None,  # Level 2 approver
        '3': None,  # Level 3 approver
        '4': None,  # Admin
    }
    
    # If role exists in mapping, try to find it by role_name again
    if role_name in role_mapping:
        # Try to find by common names
        common_names = {
            '0': ['viewer', 'Viewer', 'VIEWER'],
            '1': ['creator', 'Creator', 'CREATOR', 'role1', 'Role1'],
            '2': ['approver_level2', 'Approver Level 2', 'level2', 'Level2'],
            '3': ['approver_level3', 'Approver Level 3', 'level3', 'Level3'],
            '4': ['admin', 'Admin', 'ADMIN', 'administrator']
        }
        
        for name in common_names.get(role_name, []):
            role = RoleMaster.query.filter_by(role_name=name).first()
            if role:
                return role.id
    
    return None

# Add this function to help get status by code
def get_status_by_code(status_code):
    """Get StatusMaster object by status_code"""
    return StatusMaster.query.filter_by(status_code=status_code).first()

def get_status_id_by_code(status_code):
    """Get status ID by status_code"""
    status = get_status_by_code(status_code)
    return status.id if status else 1  # Default to first status

def get_permissions(role_name):
    """Get permissions based on role name"""
    if not role_name:
        return {
            'can_create_drawing': False,
            'can_approve_level1': False,
            'can_approve_level2': False,
            'can_approve_level3': False,
            'can_see_all': False
        }
    
    role_name_str = str(role_name)
    
    # Admin (role 4) has all permissions
    if role_name_str == '4':
        return {
            'can_create_drawing': True,
            'can_approve_level1': True,
            'can_approve_level2': True,
            'can_approve_level3': True,
            'can_see_all': True
        }
    
    # Creator (role 1) can create drawings
    if role_name_str == '1':
        return {
            'can_create_drawing': True,
            'can_approve_level1': False,
            'can_approve_level2': False,
            'can_approve_level3': False,
            'can_see_all': False
        }
    
    # Approvers (roles 2, 3) can approve
    if role_name_str == '2':
        return {
            'can_create_drawing': True,
            'can_approve_level1': True,
            'can_approve_level2': False,  # Can only approve level 1
            'can_approve_level3': False,
            'can_see_all': False
        }
    
    if role_name_str == '3':
        return {
            'can_create_drawing': True,
            'can_approve_level1': True,
            'can_approve_level2': True,
            'can_approve_level3': False,  # Can only approve up to level 2
            'can_see_all': False
        }
    
    # Viewer (role 0) can only view
    if role_name_str == '0':
        return {
            'can_create_drawing': False,
            'can_approve_level1': False,
            'can_approve_level2': False,
            'can_approve_level3': False,
            'can_see_all': False
        }
    
    # Default: no permissions
    return {
        'can_create_drawing': False,
        'can_approve_level1': False,
        'can_approve_level2': False,
        'can_approve_level3': False,
        'can_see_all': False
    }


def get_user_permissions(user):
    """Get permissions for current user based on role"""
    role_name = user.role_name if hasattr(user, 'role_name') else '0'
    return get_permissions(role_name)

def update_ctr_status(ctr_upload, new_status_code, action_details=None):
    """Update status of CTR upload and create history record"""
    try:
        # Get new status
        new_status = get_status_by_code(new_status_code)
        if not new_status:
            print(f"DEBUG: Status code {new_status_code} not found")
            return False
        
        old_status_id = ctr_upload.status_id
        old_status = StatusMaster.query.get(old_status_id) if old_status_id else None
        
        # Update status
        ctr_upload.status_id = new_status.id
        db.session.add(ctr_upload)
        
        # Create status history record
        if hasattr(ctr_upload, 'status_history'):
            history = CTRStatusHistory(
                ctr_upload_id=ctr_upload.id,
                old_status_id=old_status_id,
                new_status_id=new_status.id,
                changed_by_user_id=current_user.id,
                changed_by_role_id=current_user.role_id,
                change_notes=action_details or f"Status changed to {new_status.status_name}",
                version_number=ctr_upload.version
            )
            db.session.add(history)
        
        db.session.commit()
        
        print(f"DEBUG: Updated status for upload {ctr_upload.id} from {old_status.status_name if old_status else 'None'} to {new_status.status_name}")
        return True
        
    except Exception as e:
        print(f"ERROR updating status: {str(e)}")
        db.session.rollback()
        return False

def get_current_status_display(ctr_upload):
    """Get the current status display based on approval progress"""
    if not ctr_upload.status_id:
        return "Unknown"
    
    status_obj = StatusMaster.query.get(ctr_upload.status_id)
    if not status_obj:
        return "Unknown"
    
    # Check approval status first
    if ctr_upload.is_fully_approved:
        return "Fully Approved"
    
    if not ctr_upload.sent_for_approval:
        # Not sent for approval yet
        if status_obj.status_code in ['uploaded', 'processed', 'generating_pdf', 'completed']:
            return status_obj.status_name
        else:
            return "Ready for Approval"
    
    # Sent for approval - check levels
    approval_summary = get_approval_summary(ctr_upload)
    
    # Check Level 2
    if approval_summary[2]['status'] == 'rejected':
        return "Rejected at Level 2"
    elif approval_summary[2]['status'] == 'changes_requested':
        return "Changes Requested at Level 2"
    elif approval_summary[2]['status'] == 'approved':
        # Level 2 approved, check Level 3
        if approval_summary[3]['status'] == 'rejected':
            return "Rejected at Level 3"
        elif approval_summary[3]['status'] == 'changes_requested':
            return "Changes Requested at Level 3"
        elif approval_summary[3]['status'] == 'approved':
            return "Approved at All Levels"
        else:
            return "Pending at Level 3"
    else:
        return "Pending at Level 2"


# Update the get_approval_summary function
def get_approval_summary(ctr_upload):
    """Get a summary of approval status for all levels"""
    approvals = {}
    
    # Level 1 - Creator (auto-approved when sent)
    approvals[1] = {
        'status': 'done' if ctr_upload.sent_for_approval else 'pending',
        'approver': ctr_upload.user.username if ctr_upload.user else None,
        'comments': 'Created and sent for approval' if ctr_upload.sent_for_approval else 'Not sent yet',
        'updated_at': ctr_upload.upload_date
    }
    
    # Levels 2 and 3
    for level in [2, 3]:
        approval = CTRApproval.query.filter_by(
            ctr_upload_id=ctr_upload.id,
            approval_level=level
        ).first()
        
        if approval:
            approvals[level] = {
                'status': approval.approval_status,
                'approver': approval.approver_user.username if approval.approver_user else None,
                'comments': approval.comments,
                'updated_at': approval.updated_at
            }
        else:
            # If not sent for approval or level not reached yet
            if not ctr_upload.sent_for_approval or (level == 3 and ctr_upload.current_approval_level < 3):
                approvals[level] = {
                    'status': 'not_started',
                    'approver': None,
                    'comments': None,
                    'updated_at': None
                }
            else:
                approvals[level] = {
                    'status': 'pending',
                    'approver': None,
                    'comments': None,
                    'updated_at': None
                }
    
    return approvals


# Add this function to help get status by code
def get_status_by_code(status_code):
    """Get StatusMaster object by status_code"""
    return StatusMaster.query.filter_by(status_code=status_code).first()

def get_status_id_by_code(status_code):
    """Get status ID by status_code"""
    status = get_status_by_code(status_code)
    return status.id if status else 1  # Default to first status

# Update permissions function to use status codes instead of strings
def get_approval_status_for_user(ctr_upload, user_role):
    """Get approval status information for a specific user role"""
    # Get status codes
    status_codes = {
        'completed': get_status_id_by_code('completed'),
        'pending_approval': get_status_id_by_code('pending_approval'),
        'approved': get_status_id_by_code('approved'),
        'rejected': get_status_id_by_code('rejected')
    }
    
    if user_role == 4:  # Admin sees everything
        return {
            'can_approve': True,
            'can_view': True,
            'current_level': ctr_upload.current_approval_level,
            'is_fully_approved': ctr_upload.is_fully_approved,
            'sent_for_approval': ctr_upload.sent_for_approval,
            'approval_history': ctr_upload.approval_history
        }
    
    if user_role == 0:  # Viewers only see fully approved
        return {
            'can_approve': False,
            'can_view': ctr_upload.is_fully_approved,
            'current_level': ctr_upload.current_approval_level,
            'is_fully_approved': ctr_upload.is_fully_approved,
            'sent_for_approval': ctr_upload.sent_for_approval,
            'approval_history': []
        }
    
    # For creators (role 1)
    if user_role == 1:
        if not hasattr(current_user, 'id'):
            return {
                'can_approve': False,
                'can_send_for_approval': False,
                'can_view': False,
                'current_level': ctr_upload.current_approval_level,
                'is_fully_approved': ctr_upload.is_fully_approved,
                'sent_for_approval': ctr_upload.sent_for_approval,
                'approval_history': []
            }
        
        # Check if user is the uploader
        if current_user.id == ctr_upload.user_id:
            completed_status_id = status_codes.get('completed')
            return {
                'can_approve': False,  # Creators cannot approve
                'can_send_for_approval': (
                    ctr_upload.status_id == completed_status_id and 
                    not ctr_upload.sent_for_approval and
                    not ctr_upload.is_fully_approved
                ),
                'can_view': True,
                'current_level': ctr_upload.current_approval_level,
                'is_fully_approved': ctr_upload.is_fully_approved,
                'sent_for_approval': ctr_upload.sent_for_approval,
                'approval_history': ctr_upload.approval_history
            }
        else:
            return {
                'can_approve': False,
                'can_send_for_approval': False,
                'can_view': False,
                'current_level': ctr_upload.current_approval_level,
                'is_fully_approved': ctr_upload.is_fully_approved,
                'sent_for_approval': ctr_upload.sent_for_approval,
                'approval_history': []
            }
    
    # For approvers (roles 2, 3)
    if user_role in [2, 3]:
        # Check if drawing has been sent for approval
        if not ctr_upload.sent_for_approval:
            return {
                'can_approve': False,
                'can_view': False,
                'current_level': ctr_upload.current_approval_level,
                'is_fully_approved': ctr_upload.is_fully_approved,
                'sent_for_approval': ctr_upload.sent_for_approval,
                'approval_history': []
            }
        
        # User can only approve if they're at the current approval level
        if user_role == ctr_upload.current_approval_level:
            approval_record = CTRApproval.query.filter_by(
                ctr_upload_id=ctr_upload.id,
                approval_level=user_role
            ).first()
            
            return {
                'can_approve': approval_record and approval_record.approval_status == 'pending',
                'can_view': True,
                'current_level': ctr_upload.current_approval_level,
                'is_fully_approved': ctr_upload.is_fully_approved,
                'sent_for_approval': ctr_upload.sent_for_approval,
                'approval_history': ctr_upload.approval_history,
                'current_approval': approval_record
            }
        else:
            # User is not at current approval level
            return {
                'can_approve': False,
                'can_view': user_role < ctr_upload.current_approval_level,  # Can view if they've already approved
                'current_level': ctr_upload.current_approval_level,
                'is_fully_approved': ctr_upload.is_fully_approved,
                'sent_for_approval': ctr_upload.sent_for_approval,
                'approval_history': [a for a in ctr_upload.approval_history if a.approval_level <= user_role]
            }
    
    # Default for any other role
    return {
        'can_approve': False,
        'can_view': False,
        'current_level': ctr_upload.current_approval_level,
        'is_fully_approved': ctr_upload.is_fully_approved,
        'sent_for_approval': ctr_upload.sent_for_approval,
        'approval_history': []
    }

def get_approval_history_for_upload(upload_id):
    """Get complete approval history for a specific upload"""
    history = CTRApprovalHistory.query.filter_by(
        ctr_upload_id=upload_id
    ).order_by(CTRApprovalHistory.action_date.desc()).all()
    
    return [{
        'id': h.id,
        'action': h.action,
        'action_level': h.action_level,
        'action_details': h.action_details,
        'action_by': h.action_by_user.username if h.action_by_user else 'Unknown',
        'action_by_role': h.action_by_role.role_name if h.action_by_role and hasattr(h.action_by_role, 'role_name') else 'Unknown',
        'action_date': h.action_date.strftime('%d-%m-%Y %H:%M') if h.action_date else None,
        'previous_status': h.previous_status.status_name if h.previous_status and hasattr(h.previous_status, 'status_name') else 'Unknown',
        'new_status': h.new_status.status_name if h.new_status and hasattr(h.new_status, 'status_name') else 'Unknown',
        'version_number': h.version_number
    } for h in history]

def create_approval_history_record(ctr_upload_id, action, action_level, action_details, 
                                  action_by_user_id, action_by_role_id, 
                                  previous_status_code=None, new_status_code=None):
    """Create a new approval history record"""
    try:
        # Get the upload to get current version
        upload = CTRUpload.query.get(ctr_upload_id)
        if not upload:
            return None
        
        # Get status IDs from codes
        previous_status_id = get_status_id_by_code(previous_status_code) if previous_status_code else None
        new_status_id = get_status_id_by_code(new_status_code) if new_status_code else None
        
        history = CTRApprovalHistory(
            ctr_upload_id=ctr_upload_id,
            action=action,
            action_level=action_level,
            action_details=action_details,
            action_by_user_id=action_by_user_id,
            action_by_role_id=action_by_role_id,
            previous_status_id=previous_status_id,
            new_status_id=new_status_id,
            version_number=upload.version
        )
        
        db.session.add(history)
        db.session.commit()
        
        return history
    except Exception as e:
        print(f"Error creating approval history: {e}")
        db.session.rollback()
        return None

def get_all_versions_approval_history(station_name, user_id=None):
    """Get approval history for all versions of a station"""
    query = CTRUpload.query.filter_by(station_name=station_name)
    
    if user_id:
        query = query.filter_by(user_id=user_id)
    
    uploads = query.order_by(CTRUpload.version.desc()).all()
    
    history_by_version = {}
    for upload in uploads:
        history = get_approval_history_for_upload(upload.id)
        if history:
            history_by_version[f"v{upload.version}"] = {
                'upload_id': upload.id,
                'upload_date': upload.upload_date.strftime('%d-%m-%Y %H:%M'),
                'status': upload.status,
                'is_latest': upload.is_latest_version,
                'is_approved': upload.is_fully_approved,
                'sent_for_approval': upload.sent_for_approval,
                'user': upload.user.username if upload.user else 'Unknown',
                'filename': upload.filename,
                'history': history
            }
    
    return history_by_version


# ---------- Routes ----------

@bp.route('/download_ctr_sample_template')
@login_required
def download_ctr_sample_template():
    """Download CTR Excel template with sample data"""
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from openpyxl.worksheet.datavalidation import DataValidation
        
        return send_file(
            r"C:\Railway\git\CTR_Sample_Template.xlsx", # "/var/www/html/git/CTR_Sample_Template.xlsx"
            as_attachment=True,
            download_name="CTR_Sample_Template.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)
        
        # 1. Summary Sheet - WITH SAMPLE DATA
        summary_sheet = wb.create_sheet(title="Summary")
        summary_headers = [
            "id", 
            "name", 
            "station_name", 
            "no_of_rows", 
            "no_of_terminal_per_row", 
            "desg1", 
            "desg2", 
            "desg3",
            "sig_play_no",
            "ver_no",
            "page_no",
            "date"
        ]
        summary_sheet.append(summary_headers)
        
        # Add sample data for Summary sheet
        # 1	WR-1	SIDDHPUR STATION	8	72	Sr.DSTE/ADI	ADSTE/ADI	SSE/SIG	SC/PL 411/14

        summary_data = [
            1, 
            "WR-1", 
            "SIDDHPUR STATION", 
            8, 
            72, 
            "Sr.DSTE/ADI", 
            "ADSTE/ADI", 
            "SSE/SIG", 
            "SC/PL 411/14"
        ]
        summary_sheet.append(summary_data)
        
        # 2. Diagram Sheet - WITH SAMPLE DATA
        diagram_sheet = wb.create_sheet(title="Diagram")
        diagram_headers = [
            "TerminalNo", 
            "positive", 
            "function", 
            "negative", 
            "is_spare"
        ]
        diagram_sheet.append(diagram_headers)
        
        # Add sample data for Diagram sheet
        diagram_data = [
            [1, "B", "Point", "N", ""],
            [2, "BX", "Signal", "NX", ""],
            [3, "BX", "LC-36", "NX", ""],
            [4, "BX", "HPR N", "NX", ""],
            [5, "BX", "TK(S)", "NX", ""],
            [6, "BX", "We", "NX", ""],
            [7, "B24V", "EXT ACCR", "N24V", ""],
            [8, "B60V", "SP", "N60V", ""],
            [9, "B24V", "LC LC36", "N24V", ""],
            [10, "B24V", "SP", "N24V", ""]
        ]
        
       

        for row in diagram_data:
            diagram_sheet.append(row)
        
        # 3. RowDetail Sheet - WITH SAMPLE DATA
        rowdetail_sheet = wb.create_sheet(title="RowDetail")
        rowdetail_headers = [
            "RowMarker", 
            "TerminalNo", 
            "Description", 
            "CableName", 
            "CableCoreStart", 
            "CableCoreEnd", 
            "BlockSize",
            "Color"
        ]
        rowdetail_sheet.append(rowdetail_headers)
        
        # Add sample data for RowDetail sheet
       
        
       
        rowdetail_data = [
            ["A", 2, "5 DPR", "2401 (JB-11)", 1, 2, 2, "red"],
            ["A", 2, "5 DG", "2401 (JB-11)", 3, 4, 2, "yellow"],
            ["A", 2, "5 RG", "2401 (JB-11)", 5, 6, 2, "green"],
            ["A", 2, "5 a UPR", "2401 (JB-11)", 7, 8, 2, "no color"],
            ["A", 4, "SP", "2401 (JB-11)", 9, 12, 2, "navy blue"],
            ["A", 12, "SP", "2401 (JB-11)", 13, 24, 2, "sky blue"],
            ["A", 2, "5 HPR", "2402 (JB-11)", 1, 2, 2, "orange"],
            ["A", 2, "5 HG", "2402 (JB-11)", 3, 4, 2, "light blue"],
            ["A", 2, "5 a UG", "2402 (JB-11)", 5, 6, 2, "red text"],
            ["A", 4, "SP", "2402 (JB-11)", 7, 8, 2, "no color"],
            ["A", 12, "SP", "2402 (JB-11)", 9, 12, 2, "no color"],
            ["A", 13, "SP", "2402 (JB-11)", 13, 24, 2, "no color"],
            ["A", 2, "B/N24V(2TPR)", "2404 (JB-11)", 1, 2, 2, "no color"],
            ["A", 3, "B/N24V(101TPR2404", "2404 (JB-11)", 3, 4, 2, "no color"],
            ["A", 5, "B/N24V(103TPR2404", "2404 (JB-11)", 5, 6, 2, "no color"],
            ["A", 7, "B/N24V(104TPR2404", "2404 (JB-11)", 7, 8, 2, "no color"],
            ["A", 9, "B/N24V(103TPR2404", "2404 (JB-11)", 9, 10, 2, "no color"],
            ["A", 11, "2SP", "2404 (JB-11)", 11, 12, 2, "no color"],
            ["A", 13, "2MN BX110V", "2404 (JB-11)", 13, 14, 2, "no color"],
            ["A", 15, "2MN BX110V", "2404 (JB-11)", 15, 16, 2, "no color"],
            ["A", 17, "2MN BX110V", "2404 (JB-11)", 17, 18, 2, "no color"],
            ["A", 19, "2MN NX110V", "2404 (JB-11)", 19, 20, 2, "no color"],
            ["A", 21, "2MN NX110V", "2404 (JB-11)", 21, 22, 2, "no color"],
            ["A", 23, "2B/N24V(COSTPR2405", "2404 (JB-11)", 1, 2, 2, "no color"],
            ["B", 1, "2B/N24V(5TPR)", "2405 (JB-11)", 3, 4, 2, "no color"],
            ["B", 3, "2B/N24V(10TPR2405", "2405 (JB-11)", 5, 6, 2, "no color"],
            ["B", 5, "2B/N24V(10TPR2405", "2405 (JB-11)", 7, 8, 2, "no color"],
            ["B", 7, "2B/N24V(10TPR2405", "2405 (JB-11)", 9, 10, 2, "no color"],
            ["B", 9, "2BX110V R/SUPP2405", "2405 (JB-11)", 11, 12, 2, "no color"],
            ["B", 11, "2BX110V R/SUPP2405", "2405 (JB-11)", 13, 14, 2, "no color"],
            ["B", 13, "2BX110V R/SUPP2405", "2405 (JB-11)", 15, 16, 2, "no color"],
            ["B", 15, "2BX110V R/SUPP2405", "2405 (JB-11)", 17, 18, 2, "no color"],
            ["B", 17, "2NX110V R/SUPP2405", "2405 (JB-11)", 19, 20, 2, "no color"],
            ["B", 19, "2NX110V R/SUPP2405", "2405 (JB-11)", 21, 22, 2, "no color"],
            ["B", 21, "2NX110V R/SUPP2405", "2405 (JB-11)", 23, 24, 2, "no color"],
            ["B", 1, "2TCMR", "2406 (JB-11)", 1, 2, 2, "no color"],
            ["B", 3, "10SP", "2406 (JB-11)", 3, 12, 2, "no color"],
            ["B", 23, "2LOC LIGHT", "2406 (JB-11)", 13, 22, 2, "no color"],
            ["B", 1, "2COST TPR", "2407 (JB-11)", 1, 2, 2, "no color"],
            ["B", 3, "2ST TPR", "2407 (JB-11)", 3, 4, 2, "no color"],
            ["B", 5, "2I2T TPR", "2407 (JB-11)", 5, 6, 2, "no color"],
            ["B", 7, "2101T TPR", "2407 (JB-11)", 7, 8, 2, "no color"],
            ["B", 9, "2103T TPR", "2407 (JB-11)", 9, 10, 2, "no color"],
            ["B", 11, "2104T TPR", "2407 (JB-11)", 11, 12, 2, "no color"],
            ["B", 13, "2 2T TPR", "2407 (JB-11)", 13, 14, 2, "no color"],
            ["B", 15, "10SP", "2407 (JB-11)", 15, 24, 2, "no color"],
            ["C", 1, "24SP", "2409 (JB-11)", 1, 24, 2, "no color"],
            ["C", 1, "2AXTVR1 (R1)(F)(UP2AXT SID-", "2409 (JB-11)", 1, 2, 2, "no color"],
            ["C", 3, "2AXTVR1 (R1)(F)(UP2AXT SID-", "2409 (JB-11)", 3, 4, 2, "no color"],
            ["C", 5, "2AXTPR (R1)(R)(UP2AXT SID-", "2409 (JB-11)", 5, 6, 2, "no color"],
            ["C", 1, "6SP", "603 (JB-11)", 1, 6, 2, "no color"],
            ["C", 3, "2COS HG", "603 (JB-11)", 1, 4, 2, "no color"],
            ["C", 1, "2COS HPR", "603 (JB-11)", 3, 6, 2, "no color"],
            ["C", 5, "2SP", "603 (JB-11)", 5, 6, 2, "no color"],
            ["C", 1, "2B/N24VPT 101/1604", "604 (JB-11)", 1, 2, 2, "no color"],
            ["C", 3, "4SP", "604 (JB-11)", 3, 6, 2, "no color"],
            ["C", 1, "6PT 101/102", "604 (JB-11)", 1, 6, 2, "no color"],
            ["C", 1, "6PT 101/102 RW606", "604 (JB-11)", 1, 6, 2, "no color"],
            ["C", 1, "6PT 101/102 CW607", "604 (JB-11)", 1, 6, 2, "no color"],
            ["C", 1, "6SP", "608 (JB-11)", 1, 6, 2, "no color"],
            ["D", 1, "2 2 DG", "2403 (JB-11)", 1, 2, 2, "no color"],
            ["D", 3, "2 2 RG", "2403 (JB-11)", 3, 4, 2, "no color"],
            ["D", 5, "6SP", "2403 (JB-11)", 5, 12, 2, "no color"],
            ["D", 7, "2SH 7 HG", "2403 (JB-11)", 13, 14, 2, "no color"],
            ["D", 9, "2SH 7 RG", "2403 (JB-11)", 15, 16, 2, "no color"],
            ["D", 11, "2SH 7 HPR", "2403 (JB-11)", 17, 18, 2, "no color"],
            ["D", 13, "4SP", "2403 (JB-11)", 19, 24, 2, "no color"],
            ["D", 1, "2102T TPR", "2408 (JB-11)", 1, 2, 2, "no color"],
            ["D", 3, "2105T TPR", "2408 (JB-11)", 3, 4, 2, "no color"],
            ["D", 5, "2106T TPR", "2408 (JB-11)", 5, 6, 2, "no color"],
            ["D", 7, "12SP", "2408 (JB-11)", 7, 24, 2, "no color"],
            ["D", 1, "6PT 103/104 NW2410", "2410 (JB-11)", 1, 6, 2, "no color"],
            ["D", 7, "6PT 103/104 RW2410", "2410 (JB-11)", 7, 12, 2, "no color"],
            ["D", 13, "6PT 103/104 CW2410", "2410 (JB-11)", 13, 18, 2, "no color"],
            ["D", 19, "6SP", "2410 (JB-11)", 19, 24, 2, "no color"],
            ["E", 1, "6PT 105/106 NW2411", "2411 (JB-11)", 1, 6, 2, "no color"],
            ["E", 7, "6PT 105/106 RW2411", "2411 (JB-11)", 7, 12, 2, "no color"],
            ["E", 13, "6PT 105/106 CW2411", "2411 (JB-11)", 13, 18, 2, "no color"],
            ["E", 19, "4B/N24VPT 103/1609", "2411 (JB-11)", 19, 24, 2, "no color"],
            ["E", 1, "4SP", "609 (JB-11)", 1, 6, 2, "no color"],
            ["E", 1, "2SH 16 HG", "2412 (JB-11)", 1, 2, 2, "no color"],
            ["E", 3, "2SH 16 RG", "2412 (JB-11)", 3, 4, 2, "no color"],
            ["E", 5, "2SH 16 HPR", "2412 (JB-11)", 5, 8, 2, "no color"],
            ["E", 9, "4SP", "2412 (JB-11)", 9, 12, 2, "no color"],
            ["E", 13, "218 HG", "2412 (JB-11)", 13, 14, 2, "no color"],
            ["E", 15, "218 RG", "2412 (JB-11)", 15, 16, 2, "no color"],
            ["E", 17, "6SP", "2412 (JB-11)", 17, 24, 2, "no color"],
            ["E", 1, "212 HG", "2413 (JB-11)", 1, 2, 2, "no color"],
            ["E", 3, "212 RG", "2413 (JB-11)", 3, 4, 2, "no color"],
            ["E", 5, "212 DPR", "2413 (JB-11)", 5, 6, 2, "no color"],
            ["E", 7, "212 HPR", "2413 (JB-11)", 7, 8, 2, "no color"],
            ["E", 9, "2SP", "2413 (JB-11)", 9, 10, 2, "no color"],
            ["E", 11, "214 HG", "2413 (JB-11)", 11, 12, 2, "no color"],
            ["E", 13, "214 RG", "2413 (JB-11)", 13, 14, 2, "no color"],
            ["E", 15, "214 HPR", "2413 (JB-11)", 15, 16, 2, "no color"],
            ["E", 17, "6SP", "2413 (JB-11)", 17, 24, 2, "no color"],
            ["F", 1, "2AXT MOD-I", "(UP2AXT SID-", 1, 2, 2, "no color"],
            ["F", 3, "2 MOD-I DISPLAY (UP2AXT SID-", 3, 4, 2, "no color"],
            ["F", 5, "2 MOD-I DISPLAY (UP2AXT SID-", 5, 6, 2, "no color"],
            ["F", 7, "2AXT (+4B VRE (UP2AXT SID-", 7, 10, 2, "no color"],
            ["F", 9, "2AXT (+4B VRE (UP2AXT SID-", 9, 10, 2, "no color"],
            ["F", 11, "8SP", "2414 (JB-11)", 11, 24, 2, "no color"],
            ["F", 1, "2AXT MOD-II", "(UP2AXT SID-", 1, 2, 2, "no color"],
            ["F", 3, "2SP", "(UP2AXT SID-", 3, 4, 2, "no color"],
            ["F", 5, "124V (+)", "(2AXT SID-KW", 1, 1, 1, "no color"],
            ["F", 2, "124V (-)", "(2AXT SID-KW", 2, 2, 1, "no color"],
            ["F", 3, "148V (+)", "(2AXT SID-KW", 3, 3, 1, "no color"],
            ["F", 4, "148V (-)", "(2AXT SID-KW", 4, 4, 1, "no color"],
            ["F", 5, "1PRR A2", "(2AXT SID-KW", 5, 5, 1, "no color"],
            ["F", 6, "1VPR A5", "(2AXT SID-KW", 6, 6, 1, "no color"],
            ["F", 7, "1VPR B2", "(2AXT SID-KW", 7, 7, 1, "no color"],
            ["F", 8, "2DISPLAY MODEN(2AXT SID-KW", 8, 9, 2, "no color"],
            ["F", 10, "2DISPLAY MODEN(2AXT SID-KW", 10, 11, 2, "no color"],
            ["F", 12, "1SP", "(2AXT SID-KW", 12, 12, 1, "no color"],
            ["F", 13, "124V (+)", "(5AXT SID-KW", 13, 13, 1, "no color"],
            ["F", 14, "124V (-)", "(5AXT SID-KW", 14, 14, 1, "no color"],
            ["F", 15, "148V (+)", "(5AXT SID-KW", 15, 15, 1, "no color"],
            ["F", 16, "148V (-)", "(5AXT SID-KW", 16, 16, 1, "no color"],
            ["F", 17, "1PRR A2", "(5AXT SID-KW", 17, 17, 1, "no color"],
            ["F", 18, "1VPR A5", "(5AXT SID-KW", 18, 18, 1, "no color"],
            ["F", 19, "1VPR B2", "(5AXT SID-KW", 19, 19, 1, "no color"],
            ["F", 20, "2DISPLAY MODEN(5AXT SID-KW", 20, 21, 2, "no color"],
            ["F", 22, "2DISPLAY MODEN(5AXT SID-KW", 22, 23, 2, "no color"],
            ["F", 24, "1SP", "(5AXT SID-KW", 24, 24, 1, "no color"],
            ["G", 1, "12SP", "2414 (JB-11)", 13, 24, 2, "no color"],
            ["G", 1, "2B/N24V(01ATPR2415", "2415 (JB-11)", 1, 2, 2, "no color"],
            ["G", 3, "2B/N24V(02ATPR2415", "2415 (JB-11)", 3, 4, 2, "no color"],
            ["G", 5, "2B/N24V(03ATPR2415", "2415 (JB-11)", 5, 6, 2, "no color"],
            ["G", 7, "4SP", "2415 (JB-11)", 7, 8, 2, "no color"],
            ["G", 9, "2B/N24V(01BTPR2415", "2415 (JB-11)", 9, 14, 2, "no color"],
            ["G", 15, "2B/N24V(02BTPR2415", "2415 (JB-11)", 15, 16, 2, "no color"],
            ["G", 17, "2B/N24V(03BTPR2415", "2415 (JB-11)", 17, 18, 2, "no color"],
            ["G", 19, "2B/N24V(04BTPR2415", "2415 (JB-11)", 19, 20, 2, "no color"],
            ["G", 21, "4SP", "2415 (JB-11)", 21, 24, 2, "no color"],
            ["G", 1, "2CHI KLCR", "24N (KLCR B", 1, 2, 2, "no color"],
            ["G", 3, "2CHI YCR", "24N (KLCR B", 3, 4, 2, "no color"],
            ["G", 5, "2CH2 KLCR", "24N (KLCR B", 5, 6, 2, "no color"],
            ["G", 7, "2CH2 YCR", "24N (KLCR B", 7, 8, 2, "no color"],
            ["G", 9, "2CH3 KLCR", "24N (KLCR B", 9, 10, 2, "no color"],
            ["G", 11, "2CH3 YCR", "24N (KLCR B", 11, 12, 2, "no color"],
            ["G", 13, "6SP", "24N (KLCR B", 13, 24, 2, "no color"],
            ["H", 1, "12SP", "245 (KLCR B", 1, 12, 2, "no color"],
            ["H", 13, "2FIRE ALARM +2-245", "245 (KLCR B", 13, 22, 2, "no color"],
            ["H", 23, "2FIRE ALARM +2-245", "245 (KLCR B", 23, 24, 2, "no color"],
            ["H", 1, "2PT 101/102 NW1209", " (JB-9B)", 1, 2, 2, "no color"],
            ["H", 3, "2PT 101/102 RW1209", " (JB-9B)", 3, 4, 2, "no color"],
            ["H", 5, "8SP", "1209 (JB-9B)", 5, 12, 2, "no color"],
            ["H", 1, "2PT 103/104 NW1210", " (JB-9B)", 1, 2, 2, "no color"],
            ["H", 3, "2PT 103/104 RW1210", " (JB-9B)", 3, 4, 2, "no color"],
            ["H", 5, "8SP", "1210 (JB-9B)", 5, 12, 2, "no color"],
            ["H", 1, "2PT 105/106 NW1211", " (JB-9B)", 1, 2, 2, "no color"],
            ["H", 3, "2PT 105/106 RW1211", " (JB-9B)", 3, 4, 2, "no color"],
            ["H", 5, "8SP", "1211 (JB-9B)", 5, 12, 2, "no color"],
            ["H", 1, "2SMRB R1", "1211 (JB-9B)", 1, 4, 2, "no color"],
            ["H", 5, "2SMRB R2", "12C (VDU TA", 5, 12, 2, "no color"],
            ["H", 1, "2SMKEYND INO12C", "12C (VDU TA", 1, 4, 2, "no color"],
            ["H", 5, "1VDU R1", "12C (VDU TA", 5, 9, 1, "no color"],
            ["H", 10, "1VDU R2", "12C (VDU TA", 10, 12, 1, "no color"],
            ["H", 1, "2D/LVDU 'A'CH", "12C (VDU TA", 1, 2, 2, "no color"],
            ["H", 3, "2D/LVDU 'B'CH", "12C (VDU TA", 3, 4, 2, "no color"],
            ["H", 5, "2D/L PC-I", "12C (VDU TA", 5, 6, 2, "no color"],
            ["H", 7, "2D/L PC-II", "12C (VDU TA", 7, 8, 2, "no color"],
            ["H", 9, "2SP", "12C (VDU TA", 9, 10, 2, "no color"],
            ["H", 11, "2EWUYN KEY", "12C (VDU TA", 11, 12, 2, "no color"]
        ]
        for row in rowdetail_data:
            rowdetail_sheet.append(row)
        # Color dropdown
        

        
        colors = ["","red" , "yellow" , "green" , "no color" , "navy blue" , "sky blue" , "orange" ,  "light gray" , "red text"]
        dv = DataValidation(type="list", formula1=f'"{",".join(colors)}"', allow_blank=True)

        rowdetail_sheet.add_data_validation(dv)
        dv.add("H2:H500")
        # Simple column width adjustment
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
        
        # Create a BytesIO buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Send the file
        return send_file(
            buffer,
            as_attachment=True,
            download_name="CTR_Sample_Template.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating sample template: {str(e)}", "danger")
        return redirect(url_for('main.ctr_drawing'))


@bp.route('/download_ctr_blank_template')
@login_required
def download_ctr_blank_template():
    """Download CTR Excel template without data (only headers)"""
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)
        
        # 1. Summary Sheet - ONLY HEADERS
        summary_sheet = wb.create_sheet(title="Summary")
        summary_headers = [
            "id", 
            "ctr_name", 
            "station_name", 
            "no_of_rows", 
            "no_of_terminal_per_row", 
            "desg1", 
            "desg2", 
            "desg3",
            "sip_no",
            "total_page_no",
            "page_no",
            "date"
        ]
        summary_sheet.append(summary_headers)
        
        # 2. Diagram Sheet - ONLY HEADERS
        diagram_sheet = wb.create_sheet(title="Diagram")
        diagram_headers = [
            "TerminalNo", 
            "positive", 
            "function", 
            "negative", 
            "is_spare"
        ]
        diagram_sheet.append(diagram_headers)
        
        # 3. RowDetail Sheet - ONLY HEADERS
        rowdetail_sheet = wb.create_sheet(title="RowDetail")
        rowdetail_headers = [
            "RowMarker", 
            "TerminalNo", 
            "Description", 
            "CableName", 
            "CableCoreStart", 
            "CableCoreEnd", 
            "BlockSize",
            "Color"
        ]
        rowdetail_sheet.append(rowdetail_headers)
        # Color dropdown
        

        
        #colors = ["","red" , "yellow" , "green" , "no color" , "navy blue" , "sky blue" , "orange" , "light blue" , "red text"]
        colors = ["","red", "pink", "yellow" , "green" , "no color" , "navy blue" , "sky blue" , "orange" ,  "light gray" , "red text"]
        dv = DataValidation(type="list", formula1=f'"{",".join(colors)}"', allow_blank=True)

        rowdetail_sheet.add_data_validation(dv)
        dv.add("H2:H500")
        # Simple column width adjustment
        for sheet in wb.worksheets:
            for column in sheet.columns:
                column_letter = column[0].column_letter
                sheet.column_dimensions[column_letter].width = 20
        
        # Create a BytesIO buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Send the file
        return send_file(
            buffer,
            as_attachment=True,
            download_name="CTR_Blank_Template.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating blank template: {str(e)}", "danger")
        return redirect(url_for('main.ctr_drawing'))


@bp.route('/ctr_drawing')
@login_required
def ctr_drawing():
    """CTR Drawing Template Page with Filtering"""
    # Get user permissions
    permissions = get_user_permissions(current_user)
    
    # Determine what user can see based on role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # Only allow specific roles to access
    if user_role not in [0, 1, 2, 3, 4]:
        flash("You don't have permission to access CTR Drawing page.", "danger")
        return redirect(url_for('main.approval_tracking'))
    
    # Get query parameters for filtering
    station_filter = request.args.get('station', '')
    station_id_filter = request.args.get('station_id', '')
    status_filter = request.args.get('status', '')
    approval_filter = request.args.get('approval', '')
    version_filter = request.args.get('version', 'latest')
    search_query = request.args.get('search', '')
    
    # Get assigned projects for current user from user_projects table
    if user_role == 4:  # Admin
        assigned_projects_query = Project.query.with_entities(Project.id, Project.name).distinct()
        base_query = CTRUpload.query
    else:
        assigned_projects_query = Project.query\
            .join(user_projects, Project.id == user_projects.c.project_id)\
            .filter(user_projects.c.user_id == current_user.id)\
            .with_entities(Project.id, Project.name).distinct()
        
        assigned_projects = [{'id': row[0], 'name': row[1]} for row in assigned_projects_query if row[1]]
        assigned_project_ids = [str(p['id']) for p in assigned_projects]
        assigned_station_names = [p['name'] for p in assigned_projects]
        
        if assigned_project_ids:
            base_query = CTRUpload.query.filter(
                and_(
                    CTRUpload.is_deleted == 0,
                    or_(
                        CTRUpload.station_id.in_(assigned_project_ids),
                        CTRUpload.station_name.in_(assigned_station_names)
                    )
                )
            )
        else:
            base_query = CTRUpload.query.filter(False)    
    assigned_projects_list = [{'id': row[0], 'name': row[1]} for row in assigned_projects_query if row[1]]
    assigned_stations = sorted([p['name'] for p in assigned_projects_list])
    
    # Start with the base query
    query = base_query
    
    # Apply additional filters based on user role
    if user_role == 0:
        query = query.filter_by(is_fully_approved=True)
    
    # Apply filters
    if station_id_filter:
        query = query.filter_by(station_id=station_id_filter)
    elif station_filter:
        query = query.filter_by(station_name=station_filter)
    
    if status_filter:
        # Get status ID from status code
        status_record = StatusMaster.query.filter_by(status_code=status_filter).first()
        if status_record:
            query = query.filter_by(status_id=status_record.id)
    
    if approval_filter:
        if approval_filter == 'fully_approved':
            query = query.filter_by(is_fully_approved=True)
        elif approval_filter == 'pending':
            subquery = CTRApproval.query.filter_by(
                approval_status='pending'
            ).with_entities(CTRApproval.ctr_upload_id).distinct()
            query = query.filter(CTRUpload.id.in_(subquery))
        elif approval_filter == 'rejected':
            subquery = CTRApproval.query.filter_by(
                approval_status='rejected'
            ).with_entities(CTRApproval.ctr_upload_id).distinct()
            query = query.filter(CTRUpload.id.in_(subquery))
    
    # Version filter
    if version_filter == 'latest':
        query = query.filter_by(is_latest_version=True)
    
    # Search query
    if search_query:
        search = f"%{search_query}%"
        query = query.filter(or_(
            CTRUpload.filename.ilike(search),
            CTRUpload.station_name.ilike(search),
            CTRUpload.checksum_md5.ilike(search),
            CTRUpload.station_id.ilike(search)
        ))
    
    # Get uploads with ordering
    uploads = query.order_by(CTRUpload.upload_date.desc()).all()
    for u in uploads:
        print(f"Upload {u.id}: pdf_generated_date = {u.pdf_generated_date}")
    
    # Get unique station names and IDs for filter dropdown
    station_names_query = query.with_entities(CTRUpload.station_name).distinct()
    station_names = [row[0] for row in station_names_query if row[0]]
    station_names = sorted(station_names)
    '''
    from .models import Project
    station_names = (
    db.session.query(func.trim(Project.name))
    .filter(Project.name.isnot(None))
    .distinct()
    .order_by(func.trim(Project.name))
    .all()
    )
    '''
    # station_names = [s[0] for s in station_names]
    
    # Get all statuses for filter from StatusMaster
    statuses = StatusMaster.query.filter_by(category='upload').order_by(StatusMaster.sequence).all()
    approval_statuses = ['pending', 'fully_approved', 'rejected']
    
    # Pass helper functions to template
    return render_template('ctr_drawing.html', 
                         permissions=permissions,
                         user_role=user_role,
                         uploads=uploads,
                         station_names=station_names,
                         assigned_stations=assigned_stations,
                         statuses=statuses,
                         StatusMaster=StatusMaster,  # Pass StatusMaster class to template
                         approval_statuses=approval_statuses,
                         station_filter=station_filter,
                         station_id_filter=station_id_filter,
                         status_filter=status_filter,
                         approval_filter=approval_filter,
                         version_filter=version_filter,
                         search_query=search_query,
                         get_approval_summary=get_approval_summary,
                         get_approval_status_for_user=get_approval_status_for_user
                         )

@bp.route('/station_ctr_drawing')
@login_required
def station_ctr_drawing():
    """Station-wise CTR Drawing PDFs"""
    from collections import defaultdict
    permissions = get_user_permissions(current_user)

    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4

    # Allow roles
    if user_role not in [0, 1, 2, 3, 4]:
        flash("You don't have permission to access CTR Drawing page.", "danger")
        return redirect(url_for('main.approval_tracking'))

   
    # Base query
    if user_role == 4:
        query = CTRUpload.query

    else:
        assigned_projects_query = Project.query\
            .join(user_projects, Project.id == user_projects.c.project_id)\
            .filter(user_projects.c.user_id == current_user.id)\
            .with_entities(Project.id, Project.name).distinct()

        assigned_projects = [{'id': row[0], 'name': row[1]} for row in assigned_projects_query if row[1]]

        assigned_project_ids = [str(p['id']) for p in assigned_projects]
        assigned_station_names = [p['name'] for p in assigned_projects]

        query = CTRUpload.query.filter(
            or_(
                CTRUpload.station_id.in_(assigned_project_ids),
                CTRUpload.station_name.in_(assigned_station_names)
            )
        )

    # Only approved PDFs for role 0
    if user_role == 0:
        query = query.filter_by(is_fully_approved=True)

    
    
    # Get uploads
    uploads = query.order_by(
        CTRUpload.station_name.asc(),
        CTRUpload.upload_date.desc()
    ).all()

    grouped_uploads = defaultdict(list)

    for upload in uploads:
        station = upload.station_name or "Unknown Station"
        grouped_uploads[station].append(upload)


    # Group station-wise
    
    return render_template(
        'station_ctr_drawing.html',
        permissions=permissions,
        uploads=uploads,
        grouped_uploads=grouped_uploads
        
    )

@bp.route('/download-station-pdfs/<station_name>')
@login_required
def download_station_pdfs(station_name):

    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4

    # Base query
    if user_role == 4:
        query = CTRUpload.query.filter_by(station_name=station_name)

    else:
        assigned_projects_query = Project.query\
            .join(user_projects, Project.id == user_projects.c.project_id)\
            .filter(user_projects.c.user_id == current_user.id)\
            .with_entities(Project.id, Project.name).distinct()

        assigned_projects = [{'id': row[0], 'name': row[1]} for row in assigned_projects_query if row[1]]

        assigned_project_ids = [str(p['id']) for p in assigned_projects]
        assigned_station_names = [p['name'] for p in assigned_projects]

        query = CTRUpload.query.filter(
            and_(
                CTRUpload.station_name == station_name,
                or_(
                    CTRUpload.station_id.in_(assigned_project_ids),
                    CTRUpload.station_name.in_(assigned_station_names)
                )
            )
        )

    # Only approved PDFs for role 0
    if user_role == 0:
        query = query.filter_by(is_fully_approved=True)

    uploads = query.order_by(CTRUpload.upload_date.desc()).all()

    UPLOAD_FOLDER = r"C:\Railway\git\Circuitbuilding\uploads_ctr"

    merger = PdfMerger()

    for upload in uploads:

        # Skip NULL or empty PDF names
        if not upload.generated_pdf:
            continue

        pdf_file = str(upload.generated_pdf).strip()

        # Skip blank values
        if pdf_file == "":
            continue

        pdf_path = os.path.join(
            UPLOAD_FOLDER,
            pdf_file
        )

        # File exists check
        if os.path.exists(pdf_path):
            merger.append(pdf_path)

    # No PDF found
    if len(merger.pages) == 0:
        flash("No valid PDFs found for this station.", "warning")
        return redirect(url_for('main.station_ctr_drawing'))

    merged_pdf = BytesIO()

    merger.write(merged_pdf)
    merger.close()

    merged_pdf.seek(0)

    return send_file(
        merged_pdf,
        as_attachment=True,
        download_name=f"{station_name}_CTR_Drawings.pdf",
        mimetype='application/pdf'
    )

@bp.route('/upload_ctr_xlsx', methods=['POST'])
@login_required
def upload_ctr_xlsx():
    """Handle CTR XLSX file uploads with station assignment validation"""
    print("DEBUG: Starting upload_ctr_xlsx function")
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # Only role 1 (creator) and admin can upload
    if user_role != 1 and user_role != 4:
        print("DEBUG: Permission denied - only role 1 and admin can upload")
        return jsonify({'success': False, 'message': 'You don\'t have permission to upload files.'}), 403
    
    # Check if file was uploaded
    if 'ctr_file' not in request.files:
        print("DEBUG: No 'ctr_file' in request.files")
        return jsonify({'success': False, 'message': 'No file selected.'}), 400
    
    file = request.files['ctr_file']
    
    # Check if file is empty
    if file.filename == '':
        print("DEBUG: File filename is empty")
        return jsonify({'success': False, 'message': 'No file selected.'}), 400
    
    # Check if file has content
    file.seek(0, os.SEEK_END)
    file_length = file.tell()
    file.seek(0, os.SEEK_SET)
    
    if file_length == 0:
        print("DEBUG: File is empty (0 bytes)")
        return jsonify({'success': False, 'message': 'File is empty.'}), 400
    
    # Get selected station from form
    selected_station = request.form.get('station_name', '').strip()
    print(f"DEBUG: Selected station from form: {selected_station}")
    
    # Validate station selection
    if not selected_station:
        return jsonify({'success': False, 'message': 'Please select a station.'}), 400
    
    # Get station ID from Project table
    project = Project.query.filter_by(name=selected_station).first()
    if not project:
        return jsonify({'success': False, 'message': f'Station "{selected_station}" not found in database.'}), 400
    
    station_id = str(project.id)
    station_name = project.name
    
    print(f"DEBUG: Found station - ID: {station_id}, Name: {station_name}")
    
    # Check if user is assigned to this station (for non-admin users)
    if current_user.role_name != '4':
        user_assigned = db.session.query(user_projects)\
            .filter_by(user_id=current_user.id, project_id=project.id)\
            .first()
        
        if not user_assigned:
            return jsonify({'success': False, 'message': f'You are not assigned to station "{selected_station}".'}), 403
    
    print(f"DEBUG: Filename received: {file.filename}")
    
    # Check file extension
    if not file.filename.endswith('.xlsx'):
        print(f"DEBUG: Invalid file extension: {file.filename}")
        return jsonify({'success': False, 'message': 'Only .xlsx files are allowed.'}), 400
    
    # Check file size (limit to 10MB)
    max_size = 10 * 1024 * 1024
    file.seek(0, os.SEEK_END)
    file_length = file.tell()
    file.seek(0, os.SEEK_SET)
    
    print(f"DEBUG: File size: {file_length} bytes")
    
    if file_length > max_size:
        print(f"DEBUG: File size {file_length} exceeds limit {max_size}")
        return jsonify({'success': False, 'message': 'File size exceeds 10MB limit.'}), 400
    
    try:
        # Get the upload folder from config
        BASE_DIR = r"C:\Railway\git\Circuitbuilding" #"/var/www/html/git/Circuitbuilding"

        upload_folder = os.path.join(BASE_DIR, "uploads_ctr")
        #upload_folder = current_app.config.get('CTR_UPLOAD_FOLDER', 'uploads_ctr')
        print(f"DEBUG: Upload folder: {upload_folder}")
        
        # Create uploads_ctr directory if it doesn't exist
        os.makedirs(upload_folder, exist_ok=True)
        print(f"DEBUG: Created upload folder if needed")
        
        # Generate unique filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        username = current_user.username.replace(' ', '_').lower()
        original_filename = secure_filename(file.filename)
        unique_filename = f"{timestamp}_{username}_{original_filename}"
        file_path = os.path.join(upload_folder, unique_filename)
        
        print(f"DEBUG: Original filename: {original_filename}")
        print(f"DEBUG: Unique filename: {unique_filename}")
        print(f"DEBUG: File path: {file_path}")
        
        # Calculate file checksum
        file_content = file.read()
        file.seek(0, os.SEEK_SET)
        checksum = hashlib.md5(file_content).hexdigest()
        print(f"DEBUG: Calculated checksum: {checksum}")
        
        # Save the file temporarily to extract station name
        temp_file_path = os.path.join(upload_folder, f"temp_{unique_filename}")
        file.save(temp_file_path)
        
        print(f"DEBUG: Using station name: {station_name}, Station ID: {station_id}")
        
        # Clean station name
        if station_name:
            station_name = re.sub(r'\s+', ' ', station_name.strip())

        excel_data_name = pd.read_excel(
            temp_file_path,
            sheet_name="Summary",
            dtype=str
        )

        # Get first data row
        first_row = excel_data_name.iloc[0]

        # Read name column
        current_name = str(first_row['ctr_name']).strip()
     
        
        # ==================== VERSION DETERMINATION (CHANGED) ====================
        # Get ALL uploads for this station (across all users), ordered by version desc, then id desc
        existing_uploads = CTRUpload.query.filter(
            CTRUpload.station_id == station_id,
            CTRUpload.name == current_name
        ).order_by(
            CTRUpload.version.desc(),
            CTRUpload.id.desc()
        ).all()
        
        # Find the latest fully approved version (if any)
        fully_approved_uploads = [u for u in existing_uploads if u.is_fully_approved]
        
        if fully_approved_uploads:
            # There is a fully approved version → new upload gets next version number
            latest_approved = fully_approved_uploads[0]  # already sorted by version desc
            version_number = latest_approved.version + 1
            is_new_version = True
            print(f"DEBUG: Found fully approved version {latest_approved.version}. New version will be {version_number}")
        else:
            # No fully approved version exists → keep the current version number
            if existing_uploads:
                latest = existing_uploads[0]   # latest upload (any status)
                version_number = latest.version
                is_new_version = False
                print(f"DEBUG: No fully approved version. Keeping same version {version_number}")
            else:
                # First upload for this station
                version_number = 1
                is_new_version = False
                print(f"DEBUG: First upload for station. Version = 1")
        
        # Determine parent version (the previous latest upload) for history
        parent_version_id = existing_uploads[0].id if existing_uploads else None
        
        # Mark ALL previous uploads as NOT latest
        for upload in existing_uploads:
            upload.is_latest_version = False
            db.session.add(upload)
        
        # Create approval history record for the parent upload (if any)
        if parent_version_id:
            if is_new_version:
                create_approval_history_record(
                    ctr_upload_id=parent_version_id,
                    action='new_version_created',
                    action_level=1,
                    action_details=f"New version {version_number} created by {current_user.username}",
                    action_by_user_id=current_user.id,
                    action_by_role_id=current_user.role_id,
                    previous_status_code='existing',
                    new_status_code='new_version'
                )
            else:
                create_approval_history_record(
                    ctr_upload_id=parent_version_id,
                    action='new_upload_same_version',
                    action_level=1,
                    action_details=f"New upload (same version {version_number}) by {current_user.username}",
                    action_by_user_id=current_user.id,
                    action_by_role_id=current_user.role_id,
                    previous_status_code='existing',
                    new_status_code='new_upload'
                )
        
        # Check if this exact same file already exists (optional, keep as is)
        existing_same_file = CTRUpload.query.filter_by(
            checksum_md5=checksum, 
            user_id=current_user.id
        ).first()
        
        if existing_same_file:
            print(f"DEBUG: Same file content found (ID: {existing_same_file.id}), but still creating new version")
        
        # Move temp file to final location
        os.rename(temp_file_path, file_path)
        print(f"DEBUG: File saved successfully: {file_path}")
        
        # Get status IDs
        uploaded_status_id = get_status_id_by_code('uploaded')
        processed_status_id = get_status_id_by_code('processed')
        
        # ============================================================
        # CREATE CTR UPLOAD RECORD - NOT SENT FOR APPROVAL YET
        # ============================================================
        ctr_upload = CTRUpload(
            user_id=current_user.id,
            filename=original_filename,
            stored_filename=unique_filename,
            file_size=file_length,
            checksum_md5=checksum,
            status_id=uploaded_status_id,
            station_name=station_name,
            station_id=station_id,
            version=version_number,
            is_latest_version=True,
            current_approval_level=0,
            is_fully_approved=False,
            sent_for_approval=False,
            sent_for_approval_at=None,
            parent_version_id=parent_version_id,
            name=current_name
        )
        
        db.session.add(ctr_upload)
        db.session.flush()
        print(f"DEBUG: Created NEW CTRUpload record with ID: {ctr_upload.id}, Version: {version_number}")
        
        # Create approval history record for initial upload
        create_approval_history_record(
            ctr_upload_id=ctr_upload.id,
            action='uploaded',
            action_level=1,
            action_details=f"Initial upload of version {version_number} for station {station_name} (ID: {station_id})",
            action_by_user_id=current_user.id,
            action_by_role_id=current_user.role_id,
            previous_status_code='new',
            new_status_code='uploaded'
        )
        
        # DO NOT CREATE APPROVAL RECORDS HERE - Wait for creator to send for approval
        print(f"DEBUG: NOT creating approval records - waiting for creator to send for approval")
        
        # Process Excel file (same as before)
        try:
            print("DEBUG: Starting Excel processing")
            # Read Excel sheets
            excel_data = pd.read_excel(file_path, sheet_name=None, dtype=str)
            sheet_names = list(excel_data.keys())
            print(f"DEBUG: Excel sheets found: {sheet_names}")
            
            # Process Summary sheet
            if 'Summary' in excel_data:
                summary_df = excel_data['Summary'].fillna('')
                print(f"DEBUG: Summary sheet shape: {summary_df.shape}")
                
                # Normalize column names
                normalized_columns = {}
                for col in summary_df.columns:
                    col_str = str(col).lower().strip()
                    normalized_columns[col] = col_str
                
                print(f"DEBUG: Normalized columns: {normalized_columns}")
                
                # Process each row in Summary sheet
                for idx, row in summary_df.iterrows():
                    summary_data = {
                        'ctr_upload_id': ctr_upload.id,
                        'station_id': station_id,
                        'station': '',
                        'project': '',
                        'designation1': '',
                        'designation2': '',
                        'designation3': '',
                        'station_name': '',
                        'junction_name': '',
                        'station_code': '',
                        'zone': '',
                        'division': ''
                    }
                    
                    # Map columns flexibly
                    for col_idx, col in enumerate(summary_df.columns):
                        col_lower = normalized_columns[col]
                        value = str(row[col]).strip()
                        
                        if 'station' in col_lower and 'name' in col_lower:
                            summary_data['station_name'] = value
                            summary_data['station'] = value
                        elif col_lower == 'station':
                            summary_data['station'] = value
                        elif 'project' in col_lower:
                            summary_data['project'] = value
                        elif 'name' in col_lower and 'station' not in col_lower:
                            summary_data['junction_name'] = value
                        elif 'desg1' in col_lower or 'designation1' in col_lower:
                            summary_data['designation1'] = value
                        elif 'desg2' in col_lower or 'designation2' in col_lower:
                            summary_data['designation2'] = value
                        elif 'desg3' in col_lower or 'designation3' in col_lower:
                            summary_data['designation3'] = value
                        elif 'code' in col_lower:
                            summary_data['station_code'] = value
                        elif 'zone' in col_lower:
                            summary_data['zone'] = value
                        elif 'division' in col_lower:
                            summary_data['division'] = value
                    
                    # Set defaults if empty
                    if not summary_data['station'] and summary_data['station_name']:
                        summary_data['station'] = summary_data['station_name']
                    if not summary_data['project'] and summary_data['junction_name']:
                        summary_data['project'] = summary_data['junction_name']
                    
                    print(f"DEBUG: Summary row {idx} data: {summary_data}")
                    
                    summary = CTRSummary(**summary_data)
                    db.session.add(summary)
            
            # Process Diagram sheet
            if 'Diagram' in excel_data:
                diagram_df = excel_data['Diagram'].fillna('')
                print(f"DEBUG: Diagram sheet shape: {diagram_df.shape}")
                
                # Normalize column names
                normalized_columns = {}
                for col in diagram_df.columns:
                    col_str = str(col).lower().strip().replace(' ', '').replace('_', '').replace('#', 'no')
                    normalized_columns[col] = col_str
                
                print(f"DEBUG: Normalized diagram columns: {normalized_columns}")
                
                for idx, row in diagram_df.iterrows():
                    # Skip completely empty rows
                    if row.isna().all():
                        continue
                    
                    diagram_data = {
                        'ctr_upload_id': ctr_upload.id,
                        'terminal_no': '',
                        'positive': '',
                        'function': '',
                        'negative': ''
                    }
                    
                    # Map columns flexibly
                    for col_idx, col in enumerate(diagram_df.columns):
                        col_lower = normalized_columns[col]
                        value = str(row[col]).strip()
                        
                        if 'terminal' in col_lower or 'term' in col_lower or 'no' in col_lower:
                            diagram_data['terminal_no'] = value
                        elif 'positive' in col_lower:
                            diagram_data['positive'] = value
                        elif 'function' in col_lower:
                            diagram_data['function'] = value
                        elif 'negative' in col_lower:
                            diagram_data['negative'] = value
                    
                    # Only add if we have terminal number
                    if diagram_data['terminal_no']:
                        diagram = CTRDiagram(**diagram_data)
                        db.session.add(diagram)
                        if idx % 10 == 0:
                            print(f"DEBUG: Added diagram record {idx}")
            
            # Process RowDetail sheet
            if 'RowDetail' in excel_data:
                rowdetail_df = excel_data['RowDetail'].fillna('')
                print(f"DEBUG: RowDetail sheet shape: {rowdetail_df.shape}")
                
                # Normalize column names
                column_mapping = {}
                for col in rowdetail_df.columns:
                    col_lower = str(col).lower().strip()
                    if 'terminal' in col_lower and 'no' in col_lower:
                        column_mapping['TerminalNo'] = col
                    elif 'row' in col_lower and 'marker' in col_lower:
                        column_mapping['RowMarker'] = col
                    elif 'description' in col_lower:
                        column_mapping['Description'] = col
                    elif 'cable' in col_lower and 'name' in col_lower:
                        column_mapping['CableName'] = col
                    elif 'cable' in col_lower and 'core' in col_lower and ('start' in col_lower or 'begin' in col_lower):
                        column_mapping['CableCoreStart'] = col
                    elif 'cable' in col_lower and 'core' in col_lower and ('end' in col_lower or 'finish' in col_lower):
                        column_mapping['CableCoreEnd'] = col
                    elif 'block' in col_lower and 'size' in col_lower:
                        column_mapping['BlockSize'] = col
                    elif 'color' in col_lower:
                        column_mapping['Color'] = col
                
                print(f"DEBUG: RowDetail column mapping: {column_mapping}")
                
                for idx, row in rowdetail_df.iterrows():
                    # Skip empty rows
                    if row.isna().all():
                        continue
                    
                    # Get values with column mapping
                    row_marker = ''
                    if 'RowMarker' in column_mapping:
                        marker_val = str(row[column_mapping['RowMarker']]).strip()
                        if marker_val:
                            row_marker = marker_val[0].upper()
                    
                    terminal_no = ''
                    if 'TerminalNo' in column_mapping:
                        term_val = str(row[column_mapping['TerminalNo']]).strip()
                        if term_val:
                            terminal_no = term_val
                    
                    # Skip if missing required fields
                    if not row_marker or not terminal_no:
                        continue
                    
                    description = ''
                    if 'Description' in column_mapping:
                        description = str(row[column_mapping['Description']]).strip()
                    
                    cable_name = ''
                    if 'CableName' in column_mapping:
                        cable_name = str(row[column_mapping['CableName']]).strip()
                    
                    cable_core_start = ''
                    if 'CableCoreStart' in column_mapping:
                        core_val = str(row[column_mapping['CableCoreStart']]).strip()
                        if core_val:
                            cable_core_start = core_val
                    
                    cable_core_end = ''
                    if 'CableCoreEnd' in column_mapping:
                        core_val = str(row[column_mapping['CableCoreEnd']]).strip()
                        if core_val:
                            cable_core_end = core_val
                    
                    block_size = ''
                    if 'BlockSize' in column_mapping:
                        block_val = str(row[column_mapping['BlockSize']]).strip()
                        if block_val:
                            block_size = block_val
                    
                    color = ''
                    if 'Color' in column_mapping:
                        color_val = str(row[column_mapping['Color']]).strip()
                        if color_val and color_val.lower() != 'nan':
                            color = color_val
                    
                    row_detail = CTRRowDetail(
                        ctr_upload_id=ctr_upload.id,
                        row_marker=row_marker,
                        terminal_no=terminal_no,
                        description=description,
                        cable_name=cable_name,
                        cable_core_start=cable_core_start,
                        cable_core_end=cable_core_end,
                        block_size=block_size,
                        color=color
                    )
                    db.session.add(row_detail)
                    
                    if idx % 10 == 0:
                        print(f"DEBUG: Added rowdetail record {idx}")
            
            # Update upload status to processed
            ctr_upload.status_id = processed_status_id
            db.session.commit()
            print(f"DEBUG: Database committed successfully")
            
            # Get statistics
            summary_count = CTRSummary.query.filter_by(ctr_upload_id=ctr_upload.id).count()
            diagram_count = CTRDiagram.query.filter_by(ctr_upload_id=ctr_upload.id).count()
            rowdetail_count = CTRRowDetail.query.filter_by(ctr_upload_id=ctr_upload.id).count()
            
            print(f"DEBUG: Records saved - Summary: {summary_count}, Diagram: {diagram_count}, RowDetail: {rowdetail_count}")
            def generate_pdf_background(upload_id, file_path, upload_data):
                try:
                    from Circuitbuilding.app.database import db
                    from Circuitbuilding.app.models import CTRUpload, CTRSummary, StatusMaster, CTRApprovalHistory
                    import os
                    from datetime import datetime

                    print(f"DEBUG: Starting PDF generation for upload ID: {upload_id}")

                    ctr_upload = db.session.get(CTRUpload, upload_id)

                    if not ctr_upload:
                        print("Upload not found")
                        return

                    # Update status
                    generating_pdf_status = StatusMaster.query.filter_by(status_code='generating_pdf').first()
                    if generating_pdf_status:
                        ctr_upload.status_id = generating_pdf_status.id

                    db.session.commit()

                    # Create PDF folder
                    #pdf_output_dir = "static/ctr_pdfs"
                    project_root = os.path.dirname(os.path.abspath(__file__))  # app folder
                    pdf_output_dir = os.path.join(project_root, "static", "ctr_pdfs")
                    os.makedirs(pdf_output_dir, exist_ok=True)

                    from Circuitbuilding.app.ctr_pdf_generator import generate_ctr_pdf_from_excel

                    # Generate PDF
                    version = ctr_upload.version

                    result = generate_ctr_pdf_from_excel(
                        excel_path=file_path,
                        output_dir=pdf_output_dir,
                        version=str(version)
                    )

                    pdf_path = result.get('pdf_path')
                    no_of_rows = result.get('no_of_rows', 0)
                    no_of_terminal_per_row = result.get('no_of_terminal_per_row', 0)

                    if pdf_path and os.path.exists(pdf_path):

                        ctr_upload.generated_pdf = os.path.basename(pdf_path)

                        completed_status = StatusMaster.query.filter_by(status_code='completed').first()
                        if completed_status:
                            ctr_upload.status_id = completed_status.id
                        
                        from datetime import datetime, timedelta

                        ctr_upload.pdf_generated_date = datetime.now()    

                        #ctr_upload.pdf_generated_date = datetime.now()
                        #local_time = request.form.get("local_time")
                        #ctr_upload.pdf_generated_date = datetime.fromisoformat(local_time.replace("Z",""))

                        # Update summary
                        summary_record = CTRSummary.query.filter_by(ctr_upload_id=upload_id).first()
                        if summary_record:
                            summary_record.no_of_rows = no_of_rows
                            summary_record.no_of_terminal_per_row = no_of_terminal_per_row

                        db.session.commit()

                        print(f"✅ PDF generated: {pdf_path}")

                        # Save history
                        
                        version_number = ctr_upload.version or 1  # fallback to 1 if None
                        history = CTRApprovalHistory(
                            ctr_upload_id=upload_id,
                            action='pdf_generated',
                            action_level=1,
                            action_details='PDF generated successfully',
                            action_by_user_id=ctr_upload.user_id,
                            previous_status_id=get_status_id_by_code('processed'),
                            new_status_id=get_status_id_by_code('completed'),
                            version_number=version_number
                        )

                        db.session.add(history)
                        db.session.commit()

                    else:
                        failed_status = StatusMaster.query.filter_by(status_code='pdf_generation_failed').first()
                        if failed_status:
                            ctr_upload.status_id = failed_status.id

                        db.session.commit()
                        print("❌ PDF generation failed")

                except Exception as e:
                    import traceback
                    print("❌ ERROR:", e)
                    traceback.print_exc()
            # Start PDF generation in background (UPDATED to pass version)
            '''
            def generate_pdf_background_old(upload_id, file_path, flask_app, upload_data):
                """Background thread for PDF generation"""
                try:
                    print(f"DEBUG: Starting PDF generation for upload ID: {upload_id}")
                    
                    with flask_app.app_context():
                        import sys
                        import os
                        
                        current_dir = os.path.dirname(os.path.abspath(__file__))
                        project_root = os.path.dirname(current_dir)
                        
                        if project_root not in sys.path:
                            sys.path.insert(0, project_root)
                        
                        from sqlalchemy import create_engine
                        from sqlalchemy.orm import sessionmaker
                        from flask import current_app
                        
                        from Circuitbuilding.app.database import db
                        from Circuitbuilding.app.models import CTRUpload, CTRSummary, StatusMaster, CTRApprovalHistory
                        
                        #db_uri = current_app.config['SQLALCHEMY_DATABASE_URI']
                        #engine = create_engine(db_uri)
                        #Session = sessionmaker(bind=engine)
                        #session = Session()
                        
                        try:
                            from app.models import CTRUpload, CTRSummary
                            
                            ctr_upload =  db.session.get(CTRUpload, upload_id) #session.query(CTRUpload).get(upload_id)
                            if not ctr_upload:
                                print(f"DEBUG: Upload record {upload_id} not found")
                                return
                            
                            # Update status to generating_pdf
                            generating_pdf_status = StatusMaster.query.filter_by(status_code='generating_pdf').first()
                            if generating_pdf_status:
                                ctr_upload.status_id = generating_pdf_status.id
                            db.session.commit()
                            
                            pdf_output_dir = current_app.config.get('CTR_PDF_FOLDER', 'static/ctr_pdfs')
                            os.makedirs(pdf_output_dir, exist_ok=True)
                            print(f"DEBUG: PDF output directory: {pdf_output_dir}")
                            
                            from app.ctr_pdf_generator import generate_ctr_pdf_from_excel
                            
                            try:
                                print(f"DEBUG: Generating PDF from: {file_path}")
                                # Get the version from upload record
                                version = ctr_upload.version
                                # Pass version to PDF generator
                                result = generate_ctr_pdf_from_excel(
                                    excel_path=file_path,
                                    output_dir=pdf_output_dir,
                                    version=str(version)   # <-- NEW: pass version
                                )
                                
                                # Check if PDF was generated successfully
                                pdf_path = result.get('pdf_path')
                                no_of_rows = result.get('no_of_rows', 0)
                                no_of_terminal_per_row = result.get('no_of_terminal_per_row', 0)
                                
                                if pdf_path and os.path.exists(pdf_path):
                                    ctr_upload.generated_pdf = os.path.basename(pdf_path)
                                    
                                    # Update status to completed
                                    completed_status = StatusMaster.query.filter_by(status_code='completed').first()
                                    if completed_status:
                                        ctr_upload.status_id = completed_status.id
                                    ctr_upload.pdf_generated_date = datetime.now()
                                    
                                    # Update CTRSummary with row and terminal counts
                                    summary_record = session.query(CTRSummary)\
                                        .filter_by(ctr_upload_id=upload_id)\
                                        .first()
                                    
                                    if summary_record:
                                        summary_record.no_of_rows = no_of_rows
                                        summary_record.no_of_terminal_per_row = no_of_terminal_per_row
                                        print(f"DEBUG: Updated CTRSummary with rows={no_of_rows}, terminals per row={no_of_terminal_per_row}")
                                    
                                    db.session.commit()
                                    print(f"✅ PDF generated successfully: {pdf_path}")
                                    print(f"✅ PDF saved as: {ctr_upload.generated_pdf}")
                                    print(f"✅ Summary updated with rows={no_of_rows}, terminals per row={no_of_terminal_per_row}")
                                    
                                    # Create approval history record for PDF generation
                                    from app.models import CTRApprovalHistory
                                    from sqlalchemy.orm import Session as ORMSession
                                    
                                    local_session = ORMSession(bind=engine)
                                    try:
                                        history = CTRApprovalHistory(
                                            ctr_upload_id=upload_id,
                                            action='pdf_generated',
                                            action_level=1,
                                            action_details='PDF generated successfully',
                                            action_by_user_id=ctr_upload.user_id,
                                            action_by_role_id=None,
                                            previous_status_id=get_status_id_by_code('processed'),
                                            new_status_id=get_status_id_by_code('completed'),
                                            version_number=ctr_upload.version
                                        )
                                        local_session.add(history)
                                        local_session.commit()
                                        print("✅ Approval history record created for PDF generation")
                                    except Exception as e:
                                        print(f"❌ Error creating approval history: {e}")
                                        #local_session.rollback()
                                    finally:
                                        #local_session.close()
                                    
                                else:
                                    # Update status to pdf_generation_failed
                                    failed_status = StatusMaster.query.filter_by(status_code='pdf_generation_failed').first()
                                    if failed_status:
                                        ctr_upload.status_id = failed_status.id
                                    session.commit()
                                    print(f"❌ PDF generation failed for upload ID: {upload_id}")
                                    print(f"❌ Result: {result}")
                                    
                            except Exception as e:
                                print(f"❌ PDF generation error: {str(e)}")
                                import traceback
                                traceback.print_exc()
                                # Update status to pdf_generation_failed
                                failed_status = StatusMaster.query.filter_by(status_code='pdf_generation_failed').first()
                                if failed_status:
                                    ctr_upload.status_id = failed_status.id
                                session.commit()
                                
                        finally:
                            session.close()
                            
                except Exception as e:
                    print(f"❌ Background thread error: {str(e)}")
                    import traceback
                    traceback.print_exc()
            
            
            # Start background thread for PDF generation
            thread = threading.Thread(
                target=generate_pdf_background,
                args=(ctr_upload.id, file_path, current_app._get_current_object(), {
                    'filename': original_filename,
                    'upload_id': ctr_upload.id,
                    'user_id': current_user.id,
                    'station_id': station_id,
                    'station_name': station_name
                })
            )
            #thread.daemon = True
            #thread.start()
            '''
            generate_pdf_background(ctr_upload.id, file_path,  {
                    'filename': original_filename,
                    'upload_id': ctr_upload.id,
                    'user_id': current_user.id,
                    'station_id': station_id,
                    'station_name': station_name
                })
            
            
            print(f"DEBUG: Started PDF generation thread")
            
            action_message = f"uploaded and processed as Version {version_number}"
            if is_new_version:
                action_message = f"uploaded as new Version {version_number} for station '{station_name}'"
            
            return jsonify({
                'success': True, 
                'message': f'File "{original_filename}" {action_message} successfully! PDF generation started in background.',
                'filename': unique_filename,
                'statistics': {
                    'summary': summary_count,
                    'diagram': diagram_count,
                    'rowdetail': rowdetail_count
                },
                'upload_id': ctr_upload.id,
                'is_new_version': is_new_version,
                'version': version_number,
                'station_name': station_name,
                'station_id': station_id
            })
            
        except Exception as e:
            # If processing fails, mark as error but keep the upload record
            print(f"DEBUG: Error processing CTR XLSX: {str(e)}")    
            import traceback
            traceback.print_exc()
            # Update status to error
            error_status = StatusMaster.query.filter_by(status_code='error').first()
            if error_status:
                ctr_upload.status_id = error_status.id
            db.session.commit()
            return jsonify({
                'success': False, 
                'message': f'Error processing file: {str(e)}',
                'upload_id': ctr_upload.id
            }), 500
        
    except Exception as e:
        print(f"DEBUG: Error uploading CTR XLSX: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error uploading file: {str(e)}'}), 500
        
@bp.route('/view_ctr_pdf/<int:upload_id>')
@login_required
def view_ctr_pdf(upload_id):
    """View generated CTR PDF with role-based multi-level approval access control"""
    # Get the upload record
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # ============================================================
    # CHECK PERMISSIONS BASED ON ROLE AND MULTI-LEVEL APPROVAL
    # ============================================================
    
    # Admin can view all
    if user_role == 4:
        pass
    # Viewers (role 0) can only see fully approved drawings
    elif user_role == 0:
        if not ctr_upload.is_fully_approved:
            flash("You can only view fully approved CTR drawings.", "danger")
            return redirect(url_for('main.ctr_drawing'))
    # Creators (role 1) can view their own uploads
    elif user_role == 1:
        pass
        '''
        if current_user.id != ctr_upload.user_id:
            flash("You can only view your own uploads.", "danger")
            return redirect(url_for('main.ctr_drawing'))
        ''' 
    # Approvers (roles 2, 3) can view if they're assigned to the station
    elif user_role in [2, 3]:
        if ctr_upload.station_id or ctr_upload.station_name:
            if ctr_upload.station_id:
                project = Project.query.filter_by(id=int(ctr_upload.station_id)).first()
            else:
                project = Project.query.filter_by(name=ctr_upload.station_name).first()
            
            if project:
                user_assigned = db.session.query(user_projects)\
                    .filter_by(user_id=current_user.id, project_id=project.id)\
                    .first()
                
                if not user_assigned:
                    flash("You don't have permission to view this PDF.", "danger")
                    return redirect(url_for('main.ctr_drawing'))
                
                if not ctr_upload.sent_for_approval:
                    flash("This drawing hasn't been sent for approval yet.", "warning")
                    return redirect(url_for('main.ctr_drawing'))
                
                if user_role > ctr_upload.current_approval_level:
                    flash("This drawing is not ready for your approval level yet.", "warning")
                    return redirect(url_for('main.ctr_drawing'))
                
                approval_record = CTRApproval.query.filter_by(
                    ctr_upload_id=upload_id,
                    approval_level=user_role
                ).first()
                
                if approval_record and approval_record.approval_status == 'approved':
                    pass
                elif approval_record and approval_record.approval_status == 'pending':
                    pass
                else:
                    if user_role < ctr_upload.current_approval_level:
                        pass
                    else:
                        flash("You don't have permission to view this PDF.", "danger")
                        return redirect(url_for('main.ctr_drawing'))
            else:
                flash("Station not found. You don't have permission to view this PDF.", "danger")
                return redirect(url_for('main.ctr_drawing'))
        else:
            flash("This PDF doesn't have a station assigned. You don't have permission to view it.", "danger")
            return redirect(url_for('main.ctr_drawing'))
    else:
        flash("You don't have permission to view CTR PDFs.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # ============================================================
    # CHECK IF PDF EXISTS AND IS READY
    # ============================================================
    
    # Check if PDF exists
    if not ctr_upload.generated_pdf:
        flash("PDF has not been generated yet. Please wait for the generation to complete.", "warning")
        return redirect(url_for('main.ctr_drawing'))
    
    # Get status codes for checking
    generating_pdf_status = get_status_by_code('generating_pdf')
    regenerating_pdf_status = get_status_by_code('regenerating_pdf')
    pdf_generation_failed_status = get_status_by_code('pdf_generation_failed')
    
    # Check if PDF is still being generated
    if ctr_upload.status_id in [generating_pdf_status.id if generating_pdf_status else None, 
                                regenerating_pdf_status.id if regenerating_pdf_status else None]:
        flash("PDF is still being generated. Please wait for completion.", "warning")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if PDF generation failed
    if ctr_upload.status_id == (pdf_generation_failed_status.id if pdf_generation_failed_status else None):
        flash("PDF generation failed. Please contact support or try re-uploading the file.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    try:
        # Get PDF path from configuration
        #pdf_folder = current_app.config.get('CTR_PDF_FOLDER', 'static/ctr_pdfs')
        #BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        #pdf_folder=os.path.join(BASE_DIR, 'static', 'ctr_pdfs')
        pdf_folder = r"C:\Railway\git\Circuitbuilding\uploads_ctr" #'/var/www/html/git/Circuitbuilding/uploads_ctr'
        pdf_path = os.path.join(pdf_folder, ctr_upload.generated_pdf)
        
        # Debug logging
        print(f"DEBUG: Looking for PDF at: {pdf_path}")
        print(f"DEBUG: PDF exists: {os.path.exists(pdf_path)}")
        
        # Check if the PDF exists in the expected location
        if not os.path.exists(pdf_path):
            # Try alternative paths
            print("DEBUG: PDF not found at primary location, trying alternative paths...")
            
            # Try relative to current directory
            current_dir = os.path.dirname(os.path.abspath(__file__))
            alt_path1 = os.path.join(current_dir, pdf_folder, ctr_upload.generated_pdf)
            print(f"DEBUG: Trying path 1: {alt_path1}")
            
            # Try from project root
            project_root = os.path.dirname(current_dir)
            alt_path2 = os.path.join(project_root, pdf_folder, ctr_upload.generated_pdf)
            print(f"DEBUG: Trying path 2: {alt_path2}")
            
            # Try with just the filename in the uploads folder
            upload_folder = current_app.config.get('CTR_UPLOAD_FOLDER', 'uploads_ctr')
            alt_path3 = os.path.join(upload_folder, ctr_upload.generated_pdf)
            print(f"DEBUG: Trying path 3: {alt_path3}")
            
            # Check which path exists
            found_path = None
            for path in [alt_path1, alt_path2, alt_path3, pdf_path]:
                if os.path.exists(path):
                    found_path = path
                    print(f"DEBUG: Found PDF at: {found_path}")
                    break
            
            if not found_path:
                flash(f"PDF file not found. Please regenerate the PDF. Searched in: {pdf_folder}", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            pdf_path = found_path
        
        # Get the absolute path
        pdf_path = os.path.abspath(pdf_path)
        print(f"DEBUG: Using absolute path: {pdf_path}")
        
        # Verify the file exists (double-check)
        if not os.path.exists(pdf_path):
            flash("PDF file not found. Please regenerate the PDF.", "danger")
            return redirect(url_for('main.ctr_drawing'))
        
        # Check file size to ensure it's a valid PDF
        if os.path.getsize(pdf_path) == 0:
            flash("PDF file is empty. Please regenerate the PDF.", "danger")
            return redirect(url_for('main.ctr_drawing'))
        
        # Generate download filename
        safe_station_name = ctr_upload.station_name or "CTR_Drawing"
        safe_station_name = re.sub(r'[^\w\-_\. ]', '_', safe_station_name)
        download_name = f"CTR_{safe_station_name}_v{ctr_upload.version}.pdf"
        
        # ============================================================
        # LOG THE VIEW ACCESS FOR AUDIT TRAIL (OPTIONAL)
        # ============================================================
        print(f"AUDIT: User {current_user.username} (Role: {user_role}) viewed PDF for upload ID: {upload_id}, "
              f"Station: {ctr_upload.station_name}, Station ID: {ctr_upload.station_id}, Version: {ctr_upload.version}")
        
        # Send the PDF file
        return send_file(
            pdf_path,
            as_attachment=False,
            download_name=download_name,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"❌ Error viewing PDF: {str(e)}")
        flash(f"Error viewing PDF: {str(e)}", "danger")
        return redirect(url_for('main.ctr_drawing'))

@bp.route('/send_for_approval/<int:upload_id>', methods=['POST'])
@login_required
def send_for_approval(upload_id):
    """Send a CTR drawing for approval (Role 1 only) with status update"""
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # Only role 1 (creator) and admin can send for approval
    if user_role != 1 and user_role != 4:
        flash("You don't have permission to send drawings for approval.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if user is the creator or admin
    if current_user.id != ctr_upload.user_id and user_role != 4:
        flash("You can only send your own drawings for approval.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if already sent for approval
    if ctr_upload.sent_for_approval:
        flash("This drawing has already been sent for approval.", "info")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if PDF is generated and completed
    completed_status = get_status_by_code('completed')
    if ctr_upload.status_id != (completed_status.id if completed_status else None):
        flash("PDF must be generated before sending for approval.", "warning")
        return redirect(url_for('main.ctr_drawing'))
    
    try:
        # Update status to pending_approval
        update_ctr_status(ctr_upload, 'pending_approval', 'Drawing sent for approval')
        
        # Mark as sent for approval
        ctr_upload.sent_for_approval = True
        ctr_upload.sent_for_approval_at = datetime.utcnow()
        ctr_upload.current_approval_level = 2
        
        # Create approval record for Level 2
        level2_approval = CTRApproval(
            ctr_upload_id=upload_id,
            approval_level=2,
            approval_status='pending'
        )
        db.session.add(level2_approval)
        
        # Create approval history record
        create_approval_history_record(
            ctr_upload_id=upload_id,
            action='sent_for_approval',
            action_level=1,
            action_details=f'Drawing sent for approval for station {ctr_upload.station_name}',
            action_by_user_id=current_user.id,
            action_by_role_id=current_user.role_id,
            previous_status_code='completed',
            new_status_code='pending_approval'
        )
        
        db.session.commit()
        
        flash(f"CTR drawing sent for approval successfully! Status: Pending at Level 2", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error sending for approval: {str(e)}", "danger")
    
    return redirect(url_for('main.ctr_drawing'))

@bp.route('/admin_send_for_approval/<int:upload_id>', methods=['POST'])
@login_required
def admin_send_for_approval(upload_id):
    """Admin sends a CTR drawing for approval on behalf of creator"""
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # Only admin can use this route
    if user_role != 4:
        flash("You don't have permission to use this function.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if already sent for approval
    if ctr_upload.sent_for_approval:
        flash("This drawing has already been sent for approval.", "info")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if PDF is generated and completed
    completed_status = get_status_by_code('completed')
    if ctr_upload.status_id != (completed_status.id if completed_status else None):
        flash("PDF must be generated before sending for approval.", "warning")
        return redirect(url_for('main.ctr_drawing'))
    
    try:
        # Mark as sent for approval
        ctr_upload.sent_for_approval = True
        ctr_upload.sent_for_approval_at = datetime.utcnow()
        
        # Set current approval level to 2 (for Level 2 approvers)
        ctr_upload.current_approval_level = 2
        
        # Create approval record for Level 2
        level2_approval = CTRApproval(
            ctr_upload_id=upload_id,
            approval_level=2,
            approval_status='pending'
        )
        db.session.add(level2_approval)
        
        # Create approval history record
        create_approval_history_record(
            ctr_upload_id=upload_id,
            action='sent_for_approval',
            action_level=4,
            action_details=f'Drawing sent for approval by admin on behalf of creator for station {ctr_upload.station_name} (ID: {ctr_upload.station_id})',
            action_by_user_id=current_user.id,
            action_by_role_id=current_user.role_id,
            previous_status_code='not_sent',
            new_status_code='sent_for_approval'
        )
        
        db.session.commit()
        
        flash(f"CTR drawing sent for approval successfully by admin! It will now be reviewed by Level 2 approvers.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error sending for approval: {str(e)}", "danger")
    
    return redirect(url_for('main.ctr_drawing'))

def admin_approve_all_levels(ctr_upload, upload_id, approval_notes):
    """Admin direct approval - bypass all levels"""
    # Update status to fully_approved
    update_ctr_status(ctr_upload, 'fully_approved', 'Approved by admin override')
    
    # Mark as fully approved
    ctr_upload.is_fully_approved = True
    ctr_upload.fully_approved_at = datetime.utcnow()
    ctr_upload.admin_approved = True
    ctr_upload.admin_approved_at = datetime.utcnow()
    ctr_upload.admin_approved_by = current_user.id
    ctr_upload.admin_approval_notes = approval_notes
    ctr_upload.sent_for_approval = True
    ctr_upload.current_approval_level = 3  # Mark all levels as done
    
    # Mark all existing approval levels as approved by admin
    approvals = CTRApproval.query.filter_by(ctr_upload_id=upload_id).all()
    for approval in approvals:
        approval.approval_status = 'approved'
        approval.approver_user_id = current_user.id
        approval.approver_role_id = current_user.role_id
        approval.comments = f"Approved by admin override: {approval_notes}"
        approval.updated_at = datetime.utcnow()
    
    # Create Level 2 and 3 approvals if they don't exist
    for level in [2, 3]:
        approval = CTRApproval.query.filter_by(
            ctr_upload_id=upload_id,
            approval_level=level
        ).first()
        
        if not approval:
            approval = CTRApproval(
                ctr_upload_id=upload_id,
                approval_level=level,
                approval_status='approved',
                approver_user_id=current_user.id,
                approver_role_id=current_user.role_id,
                comments=f"Approved by admin override: {approval_notes}",
                updated_at=datetime.utcnow()
            )
            db.session.add(approval)
    
    # Create approval history record
    create_approval_history_record(
        ctr_upload_id=upload_id,
        action='approved',
        action_level=4,
        action_details=f"Approved by admin override: {approval_notes}",
        action_by_user_id=current_user.id,
        action_by_role_id=current_user.role_id,
        previous_status_code='pending_approval' if ctr_upload.sent_for_approval else 'completed',
        new_status_code='fully_approved'
    )
    
    db.session.commit()
    flash("CTR PDF has been fully approved by admin override!", "success")
    return redirect(url_for('main.ctr_drawing'))

def admin_approve_specific_level(ctr_upload, upload_id, level, approval_notes):
    """Admin approving at specific level"""
    # Update or create approval record
    approval = CTRApproval.query.filter_by(
        ctr_upload_id=upload_id,
        approval_level=level
    ).first()
    
    if not approval:
        approval = CTRApproval(
            ctr_upload_id=upload_id,
            approval_level=level,
            approval_status='pending'
        )
        db.session.add(approval)
    
    # Update approval record
    approval.approval_status = 'approved'
    approval.approver_user_id = current_user.id
    approval.approver_role_id = current_user.role_id
    approval.comments = f"Approved by admin: {approval_notes}"
    approval.updated_at = datetime.utcnow()
    
    # Update CTR upload
    ctr_upload.current_approval_level = level
    ctr_upload.sent_for_approval = True
    
    # Update status based on level
    if level == 2:
        update_ctr_status(ctr_upload, 'approved_level_2', f'Approved at Level 2 by admin')
        
        # Create Level 3 approval if it doesn't exist
        level3_approval = CTRApproval.query.filter_by(
            ctr_upload_id=upload_id,
            approval_level=3
        ).first()
        
        if not level3_approval:
            level3_approval = CTRApproval(
                ctr_upload_id=upload_id,
                approval_level=3,
                approval_status='pending'
            )
            db.session.add(level3_approval)
        
        # Update to Level 3
        ctr_upload.current_approval_level = 3
        message = "CTR PDF approved at Level 2 by admin. Automatically sent to Level 3."
        
    elif level == 3:
        update_ctr_status(ctr_upload, 'fully_approved', f'Approved at Level 3 by admin')
        ctr_upload.is_fully_approved = True
        ctr_upload.fully_approved_at = datetime.utcnow()
        message = "CTR PDF has been fully approved at Level 3 by admin!"
    
    # Create approval history record
    create_approval_history_record(
        ctr_upload_id=upload_id,
        action='approved',
        action_level=level,
        action_details=f"Approved by admin: {approval_notes}",
        action_by_user_id=current_user.id,
        action_by_role_id=current_user.role_id,
        previous_status_code='pending',
        new_status_code='approved'
    )
    
    db.session.commit()
    flash(message, "success")
    return redirect(url_for('main.ctr_drawing'))

def user_approve_level(ctr_upload, upload_id, level, user_role, approval_notes):
    """Normal user approval (roles 2 or 3)"""
    # Check if user can approve at this level
    if user_role != level:
        flash(f"You don't have permission to approve at level {level}.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if drawing has been sent for approval
    if not ctr_upload.sent_for_approval:
        flash("This drawing hasn't been sent for approval yet.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if user is at the right approval level
    if user_role != ctr_upload.current_approval_level:
        flash(f"You cannot approve at this level. Current level is {ctr_upload.current_approval_level}.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Find the approval record for this level
    approval = CTRApproval.query.filter_by(
        ctr_upload_id=upload_id,
        approval_level=level
    ).first()
    
    if not approval:
        flash(f"No approval record found for level {level}.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if already approved
    if approval.approval_status == 'approved':
        flash(f"This drawing is already approved at level {level}.", "info")
        return redirect(url_for('main.ctr_drawing'))
    
    # Update approval record
    approval.approval_status = 'approved'
    approval.approver_user_id = current_user.id
    approval.approver_role_id = current_user.role_id
    approval.comments = approval_notes
    approval.updated_at = datetime.utcnow()
    
    # Update CTR upload
    ctr_upload.current_approval_level = level
    
    # Update status and handle next level
    if level == 2:
        update_ctr_status(ctr_upload, 'approved_level_2', f'Approved at Level 2')
        
        # Create Level 3 approval if it doesn't exist
        level3_approval = CTRApproval.query.filter_by(
            ctr_upload_id=upload_id,
            approval_level=3
        ).first()
        
        if not level3_approval:
            level3_approval = CTRApproval(
                ctr_upload_id=upload_id,
                approval_level=3,
                approval_status='pending'
            )
            db.session.add(level3_approval)
        
        # Update to Level 3
        ctr_upload.current_approval_level = 3
        message = "CTR PDF approved at Level 2. Automatically sent to Level 3 for approval."
        
    elif level == 3:
        update_ctr_status(ctr_upload, 'fully_approved', f'Approved at Level 3')
        ctr_upload.is_fully_approved = True
        ctr_upload.fully_approved_at = datetime.utcnow()
        message = "CTR PDF has been fully approved at all levels!"
    
    # Create approval history record
    create_approval_history_record(
        ctr_upload_id=upload_id,
        action='approved',
        action_level=level,
        action_details=f"{approval_notes}",
        action_by_user_id=current_user.id,
        action_by_role_id=current_user.role_id,
        previous_status_code='pending',
        new_status_code='approved'
    )
    
    db.session.commit()
    flash(message, "success")
    return redirect(url_for('main.ctr_drawing'))

@bp.route('/approve_ctr_level/<int:upload_id>/<int:level>', methods=['POST'])
@login_required
def approve_ctr_level(upload_id, level):
    """Approve a CTR at specific level and auto-advance to next level"""
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # Get approval notes
    approval_notes = request.form.get('approval_notes', '').strip()
    
    try:
        if user_role == 4 and level == 4:
            # Admin direct approval - bypass all levels
            return admin_approve_all_levels(ctr_upload, upload_id, approval_notes)
        
        elif user_role == 4:
            # Admin approving at specific level
            return admin_approve_specific_level(ctr_upload, upload_id, level, approval_notes)
        
        else:
            # Normal user approval (roles 2 or 3)
            return user_approve_level(ctr_upload, upload_id, level, user_role, approval_notes)
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error approving CTR: {str(e)}", "danger")
    
    return redirect(url_for('main.ctr_drawing'))


@bp.route('/reject_ctr_level/<int:upload_id>/<int:level>', methods=['POST'])
@login_required
def reject_ctr_level(upload_id, level):
    """Reject a CTR at specific level"""
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # Get rejection reason
    rejection_reason = request.form.get('rejection_reason', '').strip()
    
    if not rejection_reason:
        flash("Please provide a rejection reason.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    try:
        if user_role == 4 and level == 4:
            # Admin direct rejection
            update_ctr_status(ctr_upload, 'rejected', 'Rejected by admin override')
            
            ctr_upload.admin_approved = False
            ctr_upload.admin_approved_at = datetime.utcnow()
            ctr_upload.admin_approved_by = current_user.id
            ctr_upload.admin_approval_notes = f"Rejected by admin: {rejection_reason}"
            
            # Mark all existing approvals as rejected
            approvals = CTRApproval.query.filter_by(ctr_upload_id=upload_id).all()
            for approval in approvals:
                approval.approval_status = 'rejected'
                approval.approver_user_id = current_user.id
                approval.approver_role_id = current_user.role_id
                approval.comments = f"Rejected by admin: {rejection_reason}"
                approval.updated_at = datetime.utcnow()
            
            # Create approval history record
            create_approval_history_record(
                ctr_upload_id=upload_id,
                action='rejected',
                action_level=4,
                action_details=f"Rejected by admin: {rejection_reason}",
                action_by_user_id=current_user.id,
                action_by_role_id=current_user.role_id,
                previous_status_code='pending_approval',
                new_status_code='rejected'
            )
            
            flash("CTR PDF has been rejected by admin override!", "success")
            
        elif user_role == 4:
            # Admin rejecting at specific level
            approval = CTRApproval.query.filter_by(
                ctr_upload_id=upload_id,
                approval_level=level
            ).first()
            
            if not approval:
                approval = CTRApproval(
                    ctr_upload_id=upload_id,
                    approval_level=level,
                    approval_status='pending'
                )
                db.session.add(approval)
            
            # Update approval record
            approval.approval_status = 'rejected'
            approval.approver_user_id = current_user.id
            approval.approver_role_id = current_user.role_id
            approval.comments = f"Rejected by admin: {rejection_reason}"
            approval.updated_at = datetime.utcnow()
            
            # Update status
            status_code = f'rejected_level_{level}'
            update_ctr_status(ctr_upload, status_code, f'Rejected at Level {level} by admin')
            
            # Create approval history record
            create_approval_history_record(
                ctr_upload_id=upload_id,
                action='rejected',
                action_level=level,
                action_details=f"Rejected by admin: {rejection_reason}",
                action_by_user_id=current_user.id,
                action_by_role_id=current_user.role_id,
                previous_status_code='pending',
                new_status_code='rejected'
            )
            
            flash(f"CTR PDF has been rejected at level {level} by admin.", "success")
            
        else:
            # Normal user rejection
            if user_role != level:
                flash(f"You don't have permission to reject at level {level}.", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            # Check if drawing has been sent for approval
            if not ctr_upload.sent_for_approval:
                flash("This drawing hasn't been sent for approval yet.", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            # Check if user is at the right approval level
            if user_role != ctr_upload.current_approval_level:
                flash(f"You cannot reject at this level. Current level is {ctr_upload.current_approval_level}.", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            # Find the approval record for this level
            approval = CTRApproval.query.filter_by(
                ctr_upload_id=upload_id,
                approval_level=level
            ).first()
            
            if not approval:
                flash(f"No approval record found for level {level}.", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            # Update approval record
            approval.approval_status = 'rejected'
            approval.approver_user_id = current_user.id
            approval.approver_role_id = current_user.role_id
            approval.comments = rejection_reason
            approval.updated_at = datetime.utcnow()
            
            # Update status
            status_code = f'rejected_level_{level}'
            update_ctr_status(ctr_upload, status_code, f'Rejected at Level {level}')
            
            # Create approval history record
            create_approval_history_record(
                ctr_upload_id=upload_id,
                action='rejected',
                action_level=level,
                action_details=f"{rejection_reason}",
                action_by_user_id=current_user.id,
                action_by_role_id=current_user.role_id,
                previous_status_code='pending',
                new_status_code='rejected'
            )
            
            flash(f"CTR PDF has been rejected at level {level}.", "success")
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error rejecting CTR: {str(e)}", "danger")
    
    return redirect(url_for('main.ctr_drawing'))


@bp.route('/request_changes_ctr_level/<int:upload_id>/<int:level>', methods=['POST'])
@login_required
def request_changes_ctr_level(upload_id, level):
    """Request changes for a CTR at specific level"""
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # Get changes requested
    changes_requested = request.form.get('changes_requested', '').strip()
    
    if not changes_requested:
        flash("Please specify what changes are required.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    try:
        if user_role == 4 and level == 4:
            # Admin direct changes request
            update_ctr_status(ctr_upload, 'changes_requested', 'Changes requested by admin')
            
            ctr_upload.admin_approved = False
            ctr_upload.admin_approved_at = datetime.utcnow()
            ctr_upload.admin_approved_by = current_user.id
            ctr_upload.admin_approval_notes = f"Changes requested by admin: {changes_requested}"
            
            # Mark all existing approvals as changes requested
            approvals = CTRApproval.query.filter_by(ctr_upload_id=upload_id).all()
            for approval in approvals:
                approval.approval_status = 'changes_requested'
                approval.approver_user_id = current_user.id
                approval.approver_role_id = current_user.role_id
                approval.comments = f"Changes requested by admin: {changes_requested}"
                approval.updated_at = datetime.utcnow()
            
            # Create approval history record
            create_approval_history_record(
                ctr_upload_id=upload_id,
                action='changes_requested',
                action_level=4,
                action_details=f"Changes requested by admin: {changes_requested}",
                action_by_user_id=current_user.id,
                action_by_role_id=current_user.role_id,
                previous_status_code='pending_approval',
                new_status_code='changes_requested'
            )
            
            flash("Changes requested by admin override!", "success")
            
        elif user_role == 4:
            # Admin requesting changes at specific level
            approval = CTRApproval.query.filter_by(
                ctr_upload_id=upload_id,
                approval_level=level
            ).first()
            
            if not approval:
                approval = CTRApproval(
                    ctr_upload_id=upload_id,
                    approval_level=level,
                    approval_status='pending'
                )
                db.session.add(approval)
            
            # Update approval record
            approval.approval_status = 'changes_requested'
            approval.approver_user_id = current_user.id
            approval.approver_role_id = current_user.role_id
            approval.comments = f"Changes requested by admin: {changes_requested}"
            approval.updated_at = datetime.utcnow()
            
            # Update status
            status_code = f'changes_requested_level_{level}'
            update_ctr_status(ctr_upload, status_code, f'Changes requested at Level {level} by admin')
            
            # Create approval history record
            create_approval_history_record(
                ctr_upload_id=upload_id,
                action='changes_requested',
                action_level=level,
                action_details=f"Changes requested by admin: {changes_requested}",
                action_by_user_id=current_user.id,
                action_by_role_id=current_user.role_id,
                previous_status_code='pending',
                new_status_code='changes_requested'
            )
            
            flash(f"Changes requested for CTR PDF at level {level} by admin.", "success")
            
        else:
            # Normal user changes request
            if user_role != level:
                flash(f"You don't have permission to request changes at level {level}.", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            # Check if drawing has been sent for approval
            if not ctr_upload.sent_for_approval:
                flash("This drawing hasn't been sent for approval yet.", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            # Check if user is at the right approval level
            if user_role != ctr_upload.current_approval_level:
                flash(f"You cannot request changes at this level. Current level is {ctr_upload.current_approval_level}.", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            # Find the approval record for this level
            approval = CTRApproval.query.filter_by(
                ctr_upload_id=upload_id,
                approval_level=level
            ).first()
            
            if not approval:
                flash(f"No approval record found for level {level}.", "danger")
                return redirect(url_for('main.ctr_drawing'))
            
            # Update approval record
            approval.approval_status = 'changes_requested'
            approval.approver_user_id = current_user.id
            approval.approver_role_id = current_user.role_id
            approval.comments = changes_requested
            approval.updated_at = datetime.utcnow()
            
            # Update status
            status_code = f'changes_requested_level_{level}'
            update_ctr_status(ctr_upload, status_code, f'Changes requested at Level {level}')
            
            # Create approval history record
            create_approval_history_record(
                ctr_upload_id=upload_id,
                action='changes_requested',
                action_level=level,
                action_details=f"{changes_requested}",
                action_by_user_id=current_user.id,
                action_by_role_id=current_user.role_id,
                previous_status_code='pending',
                new_status_code='changes_requested'
            )
            
            flash(f"Changes requested for CTR PDF at level {level}.", "success")
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error requesting changes for CTR: {str(e)}", "danger")
    
    return redirect(url_for('main.ctr_drawing'))

@bp.route('/get_ctr_versions_data/<int:upload_id>')
@login_required
def get_ctr_versions_data(upload_id):
    """Get CTR versions data for modal - Allow approvers to view versions"""
    # Get the upload record
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # ============================================================
    # UPDATED PERMISSION LOGIC - More flexible for approvers
    # ============================================================
    
    # Admin can view any
    if user_role == 4:
        pass  # Admin has permission
    
    # Uploader can view their own uploads
    elif current_user.id == ctr_upload.user_id:
        pass  # Uploader has permission
    
    # Viewers (role 0) can only see fully approved
    elif user_role == 0:
        if not ctr_upload.is_fully_approved:
            return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    # Approvers (roles 2, 3) can view if:
    # 1. The drawing has been sent for approval AND
    # 2. They're at or above the current approval level
    elif user_role in [2, 3]:
        # Check if drawing has been sent for approval
        if not ctr_upload.sent_for_approval:
            return jsonify({'success': False, 'message': 'This drawing has not been sent for approval yet.'}), 403
        
        # Check if user's role matches or is below the current approval level
        if user_role > ctr_upload.current_approval_level:
            return jsonify({'success': False, 'message': 'You cannot view this drawing at your approval level.'}), 403
        
        # Check if user is assigned to this station (optional but good practice)
        if ctr_upload.station_id or ctr_upload.station_name:
            # Try to find project by station_id first, then by station_name
            if ctr_upload.station_id:
                project = Project.query.filter_by(id=int(ctr_upload.station_id)).first()
            else:
                project = Project.query.filter_by(name=ctr_upload.station_name).first()
            
            if project:
                user_assigned = db.session.query(user_projects)\
                    .filter_by(user_id=current_user.id, project_id=project.id)\
                    .first()
                
                # If not assigned, they can still view if they're approvers
                if not user_assigned and user_role not in [2, 3, 4]:
                    return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    # Creators (role 1) can only view their own
    elif user_role == 1:
        if current_user.id != ctr_upload.user_id:
            return jsonify({'success': False, 'message': 'You can only view your own uploads.'}), 403
    
    else:
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    # ============================================================
    # GET VERSION DATA
    # ============================================================
    
    # Get all versions for this station (all users' uploads for this station)
    if ctr_upload.station_id or ctr_upload.station_name:
        # Build query to filter by station_id OR station_name
        versions_query = CTRUpload.query.filter(
            or_(
                CTRUpload.station_id == ctr_upload.station_id,
                CTRUpload.station_name == ctr_upload.station_name
            ),
            CTRUpload.user_id == ctr_upload.user_id  # Only get versions by the same user
        ).order_by(CTRUpload.version.desc())
        
        versions = versions_query.all()
    else:
        versions = [ctr_upload]
    
    # Prepare data for JSON response
    versions_data = []
    for version in versions:
        # Get approval summary
        approval_summary = get_approval_summary(version)
        
        # Convert StatusMaster object to string for JSON serialization
        status_name = None
        if version.status:
            if hasattr(version.status, 'status_name'):
                status_name = version.status.status_name
            elif hasattr(version.status, 'status_code'):
                status_name = version.status.status_code
            else:
                status_name = str(version.status)
        else:
            # Get status by ID
            status_obj = StatusMaster.query.get(version.status_id) if hasattr(version, 'status_id') else None
            status_name = status_obj.status_name if status_obj else "Unknown"
        
        version_data = {
            'id': version.id,
            'filename': version.filename,
            'version': version.version,
            'upload_date': version.upload_date.strftime('%d-%m-%Y %H:%M') if version.upload_date else None,
            'station_name': version.station_name,
            'station_id': version.station_id,
            'file_size': f"{(version.file_size / 1024):.1f} KB" if version.file_size else "0 KB",
            'status': status_name,  # Now a string, not an object
            'status_id': version.status_id,
            'approval_summary': approval_summary,
            'is_latest_version': version.is_latest_version,
            'parent_version_id': version.parent_version_id,
            'generated_pdf': version.generated_pdf,
            'pdf_generated_date': version.pdf_generated_date.strftime('%d-%m-%Y %H:%M') if version.pdf_generated_date else None,
            'approved_by': version.admin_approver.username if version.admin_approver else None,
            'approved_date': version.admin_approved_at.strftime('%d-%m-%Y %H:%M') if version.admin_approved_at else None,
            'user': {
                'username': version.user.username if version.user else 'Unknown'
            },
            'sent_for_approval': version.sent_for_approval,
            'is_fully_approved': version.is_fully_approved,
            'current_approval_level': version.current_approval_level
        }
        versions_data.append(version_data)
    
    return jsonify({
        'success': True,
        'current_version': {
            'id': ctr_upload.id,
            'station_name': ctr_upload.station_name,
            'station_id': ctr_upload.station_id,
            'user': {
                'username': ctr_upload.user.username if ctr_upload.user else 'Unknown'
            }
        },
        'versions': versions_data
    })
@bp.route('/test444')
def get_test44():
    
    python_exe = "/usr/bin/python3"
    converter_script = "/root/srv/local/git/excel_to_pdf_converter.py"
    file_path = "/root/srv/local/git/xlsx_download/RAILWAYPROJECT_ID72_Upendra_20260312_151833.xlsx"
    pdf_path = "root/srv/local/git/uploads/RAILWAYPROJECT_ID72_Upendra_20260312_151834.pdf"

    try:
        result = subprocess.run(
            [python_exe, converter_script, file_path, pdf_path],
            capture_output=True,
            text=True,
            timeout=300
        )

        return jsonify({
            "returncode": result.returncode,
            "stdout": result.stdout,
            "stderr": result.stderr
        })
        

    except Exception as e:
        return jsonify({"error": str(e)})   
       
@bp.route('/get_approval_historys/<int:upload_id>')
@login_required
def get_approval_historys(upload_id):
    """Get approval history for a specific upload"""
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Get user role
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    # ============================================================
    # UPDATED PERMISSION LOGIC
    # ============================================================
    
    if user_role == 4:
        pass  # Admin can see all
    elif user_role == 0:  # Viewers
        if not ctr_upload.is_fully_approved:
            return jsonify({'success': False, 'message': 'Permission denied'}), 403
    elif user_role == 1:  # Creators
        if current_user.id != ctr_upload.user_id:
            return jsonify({'success': False, 'message': 'Permission denied'}), 403
    elif user_role in [2, 3]:  # Approvers
        # Check if drawing has been sent for approval
        if not ctr_upload.sent_for_approval:
            return jsonify({'success': False, 'message': 'This drawing has not been sent for approval yet.'}), 403
        
        # Check if user's role matches or is below the current approval level
        if user_role > ctr_upload.current_approval_level:
            return jsonify({'success': False, 'message': 'You cannot view this drawing at your approval level.'}), 403
    else:
        return jsonify({'success': False, 'message': 'Permission denied'}), 403
    
    history = get_approval_history_for_upload(upload_id)
    
    return jsonify({
        'success': True,
        'upload_id': upload_id,
        'station_name': ctr_upload.station_name,
        'station_id': ctr_upload.station_id,
        'version': ctr_upload.version,
        'is_latest': ctr_upload.is_latest_version,
        'history': history  
    })


@bp.route('/get_all_station_versions_history/<string:station_name>')
@login_required
def get_all_station_versions_history(station_name):
    """Get approval history for all versions of a station"""
    # Check if user has access to this station
    user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
    
    if user_role != 4:
        # Check if user is assigned to this station by name or ID
        project_by_name = Project.query.filter_by(name=station_name).first()
        
        if project_by_name:
            user_assigned = db.session.query(user_projects)\
                .filter_by(user_id=current_user.id, project_id=project_by_name.id)\
                .first()
            if not user_assigned:
                return jsonify({'success': False, 'message': 'Permission denied'}), 403
        else:
            # Try to find by station_id if station_name is actually an ID
            project_by_id = Project.query.filter_by(id=station_name).first()
            if project_by_id:
                user_assigned = db.session.query(user_projects)\
                    .filter_by(user_id=current_user.id, project_id=project_by_id.id)\
                    .first()
                if not user_assigned:
                    return jsonify({'success': False, 'message': 'Permission denied'}), 403
            else:
                return jsonify({'success': False, 'message': 'Station not found'}), 404
    
    # Get all uploads for this station by name or ID
    uploads = CTRUpload.query.filter(
        or_(
            CTRUpload.station_name == station_name,
            CTRUpload.station_id == station_name
        )
    ).order_by(CTRUpload.version.desc()).all()
    
    versions_history = []
    for upload in uploads:
        history = get_approval_history_for_upload(upload.id)
        
        # Get approval summary
        approval_summary = get_approval_summary(upload)
        
        versions_history.append({
            'version': upload.version,
            'upload_id': upload.id,
            'upload_date': upload.upload_date.strftime('%d-%m-%Y %H:%M'),
            'filename': upload.filename,
            'status': upload.status,
            'is_latest': upload.is_latest_version,
            'is_fully_approved': upload.is_fully_approved,
            'sent_for_approval': upload.sent_for_approval,
            'station_id': upload.station_id,
            'user': upload.user.username if upload.user else 'Unknown',
            'approval_summary': approval_summary,
            'history': history
        })
    
    return jsonify({
        'success': True,
        'station_name': station_name,
        'total_versions': len(versions_history),
        'versions': versions_history
    })


@bp.route('/view_previous_versions/<int:upload_id>')
@login_required
def view_previous_versions(upload_id):
    """View previous versions of a CTR upload"""
    # Get the upload record
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Check permissions
    permissions = get_user_permissions(current_user)
    if not (permissions['can_create_drawing'] or permissions['can_see_all']):
        flash("You don't have permission to view previous versions.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if user owns this upload or is admin
    if not (current_user.id == ctr_upload.user_id or current_user.role_name == '4'):
        flash("You don't have permission to view these versions.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Get all versions for this station
    if ctr_upload.station_id or ctr_upload.station_name:
        versions = CTRUpload.query.filter(
            or_(
                CTRUpload.station_id == ctr_upload.station_id,
                CTRUpload.station_name == ctr_upload.station_name
            ),
            CTRUpload.user_id == ctr_upload.user_id
        ).order_by(CTRUpload.version.desc()).all()
    else:
        versions = [ctr_upload]
    
    return render_template('ctr_versions.html',
                         permissions=permissions,
                         current_version=ctr_upload,
                         versions=versions)


@bp.route('/restore_version/<int:upload_id>')
@login_required
def restore_version(upload_id):
    """Restore a previous version as the latest"""
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Check permissions
    permissions = get_user_permissions(current_user)
    if not (permissions['can_create_drawing'] or permissions['can_see_all']):
        flash("You don't have permission to restore versions.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    # Check if user owns this upload or is admin
    if not (current_user.id == ctr_upload.user_id or current_user.role_name == '4'):
        flash("You don't have permission to restore this version.", "danger")
        return redirect(url_for('main.view_previous_versions', upload_id=upload_id))
    
    # Check if this is already the latest version
    if ctr_upload.is_latest_version:
        flash("This is already the latest version.", "info")
        return redirect(url_for('main.view_previous_versions', upload_id=upload_id))
    
    # Get all versions for this station
    if ctr_upload.station_id or ctr_upload.station_name:
        # Mark all versions as not latest
        versions = CTRUpload.query.filter(
            or_(
                CTRUpload.station_id == ctr_upload.station_id,
                CTRUpload.station_name == ctr_upload.station_name
            ),
            CTRUpload.user_id == ctr_upload.user_id
        ).all()
        
        for version in versions:
            version.is_latest_version = False
        
        # Mark this version as latest
        ctr_upload.is_latest_version = True
        ctr_upload.approval_status = 'pending'  # Reset approval status
        ctr_upload.admin_approved = None
        ctr_upload.admin_approved_at = None
        ctr_upload.admin_approval_notes = None
        
        db.session.commit()
        
        flash(f"Version {ctr_upload.version} has been restored as the latest version.", "success")
    else:
        flash("Cannot restore version: No station name found.", "danger")
    
    return redirect(url_for('main.view_previous_versions', upload_id=upload_id))


@bp.route('/delete_ctr_upload/<int:upload_id>', methods=['POST'])
@login_required
def delete_ctr_upload(upload_id):
    """Delete a CTR upload and all associated data"""
    ctr_upload = CTRUpload.query.get_or_404(upload_id)
    
    # Check permissions - only owner or admin can delete
    if not (current_user.id == ctr_upload.user_id or current_user.role_name == '4'):
        flash("You don't have permission to delete this upload.", "danger")
        return redirect(url_for('main.ctr_drawing'))
    
    try:

        if current_user.role_name != "4":
            ctr_upload.is_deleted = 1
            db.session.commit()
            flash(f"CTR upload '{ctr_upload.filename}' has been deleted successfully.", "success")
        else :
            # Delete associated records
            CTRSummary.query.filter_by(ctr_upload_id=upload_id).delete()
            CTRDiagram.query.filter_by(ctr_upload_id=upload_id).delete()
            CTRRowDetail.query.filter_by(ctr_upload_id=upload_id).delete()
            CTRApproval.query.filter_by(ctr_upload_id=upload_id).delete()
            CTRApprovalHistory.query.filter_by(ctr_upload_id=upload_id).delete()
            
            # Delete PDF file if exists
            if ctr_upload.generated_pdf:
                #pdf_folder = current_app.config.get('CTR_PDF_FOLDER', 'static/ctr_pdfs')
                BASE_DIR = os.path.dirname(os.path.abspath(__file__))
                pdf_folder=os.path.join(BASE_DIR, 'static', 'ctr_pdfs')
                pdf_path = os.path.join(pdf_folder, ctr_upload.generated_pdf)
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
            
            # Delete Excel file if exists
            upload_folder = current_app.config.get('CTR_UPLOAD_FOLDER', 'uploads_ctr')
            excel_path = os.path.join(upload_folder, ctr_upload.stored_filename)
            if os.path.exists(excel_path):
                os.remove(excel_path)
            
            # Check if there are other versions for this station
            if (ctr_upload.station_id or ctr_upload.station_name) and ctr_upload.is_latest_version:
                # Find the next latest version (by any user)
                other_versions = CTRUpload.query.filter(
                    or_(
                        CTRUpload.station_id == ctr_upload.station_id,
                        CTRUpload.station_name == ctr_upload.station_name
                    ),
                    CTRUpload.user_id == ctr_upload.user_id,
                    CTRUpload.id != upload_id
                ).order_by(CTRUpload.version.desc()).all()
                
                if other_versions:
                    # Mark the next version as latest
                    other_versions[0].is_latest_version = True
            
            # Delete the upload record
            db.session.delete(ctr_upload)
            db.session.commit()
            
            flash(f"CTR upload '{ctr_upload.filename}' has been deleted successfully.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting upload: {str(e)}", "danger")
    
    return redirect(url_for('main.ctr_drawing'))


@bp.route('/project/<int:project_id>/cable/<string:cable_id>/terminals', methods=['GET'])
def get_cable_terminals(project_id, cable_id):
    """Get terminals for a specific cable (no location filtering)"""
    try:
        terminals = Terminal.query.filter_by(
            project_id=project_id,
            cable_id=cable_id
        ).order_by(Terminal.terminal_id).all()
        
        return jsonify({
            'success': True,
            'terminals': [terminal.to_dict() for terminal in terminals],
            'count': len(terminals)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/get_station_details/<station_identifier>')
@login_required
def get_station_details(station_identifier):
    """Get station details by ID or name"""
    try:
        # Try to find by ID first
        station = Project.query.filter_by(id=station_identifier).first()
        
        # If not found by ID, try by name
        if not station:
            station = Project.query.filter_by(name=station_identifier).first()
        
        if station:
            return jsonify({
                'success': True,
                'station': {
                    'id': station.id,
                    'name': station.name,
                    'code': station.code if hasattr(station, 'code') else '',
                    'location': station.location if hasattr(station, 'location') else '',
                    'division': station.division if hasattr(station, 'division') else ''
                }
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Station not found'
            }), 404
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching station details: {str(e)}'
        }), 500


@bp.route('/get_stations_by_user')
@login_required
def get_stations_by_user():
    """Get stations assigned to current user"""
    try:
        user_role = int(current_user.role_name) if current_user.role_name.isdigit() else 4
        
        if user_role == 4:
            # Admin gets all stations
            stations = Project.query.order_by(Project.name).all()
        else:
            # Get stations assigned to current user
            stations = Project.query\
                .join(user_projects, Project.id == user_projects.c.project_id)\
                .filter(user_projects.c.user_id == current_user.id)\
                .order_by(Project.name).all()
        
        stations_list = []
        for station in stations:
            stations_list.append({
                'id': station.id,
                'name': station.name,
                'code': station.code if hasattr(station, 'code') else '',
                'location': station.location if hasattr(station, 'location') else ''
            })
        
        return jsonify({
            'success': True,
            'stations': stations_list,
            'count': len(stations_list)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching stations: {str(e)}'
        }), 500
        
#################################################
def process_pdf_background(app, project_id,file_path, pdf_path, filename,pdf_filename, upload_dir):
    with app.app_context(): 
        try:
            #python_exe = sys.executable
           
            current_project = Project.query.get(project_id)  # ✅ fetch again safely
            python_exe = "/usr/bin/python3"
            converter_script = "/root/srv/local/git/excel_to_pdf_converter.py"
           


            result = subprocess.run(
                [python_exe, converter_script, file_path, pdf_path],
                capture_output=True,
                text=True,
                encoding='utf-8',
                timeout=300
            )

            if result.returncode != 0:
                print("PDF conversion error:", result.stderr)
                 # ========== FLASH MESSAGE ==========
               

               
               
            else:
                # ===== MOVE EXCEL FILE TO uploads =====
                new_excel_path = os.path.join(upload_dir, filename)
                shutil.copy(file_path, new_excel_path)
                if result.returncode == 0 and os.path.exists(pdf_path):
                    file_md5 = _md5_of_file(pdf_path)
                    file_size = os.path.getsize(pdf_path)
                    meta = parse_converter_stdout(result.stdout)
                    db_checksum = meta.get("metadata_checksum") or meta.get("full_file_md5") or file_md5
                    max_version_record = GeneratedPDF.query.filter_by(project_id=project_id).order_by(GeneratedPDF.version.desc()).first()
                    next_version = max_version_record.version + 1 if max_version_record else 1
                    # Get junction data from the database at this moment
                    junction_boxes = JunctionBox.query.filter_by(project_id=project_id).all()
                    junction_data_list = []
                    for jb in junction_boxes:
                        junction_data_list.append({
                            'junction_id': jb.junction_id,
                            'junction_name': jb.junction_name,
                            'junction_size': jb.junction_size,
                            'station_id': jb.station_id,
                            'latitude': jb.latitude,
                            'longitude': jb.longitude,
                            'junction_row': jb.junction_row
                        })
                    
                    # Convert to JSON string
                    import json
                    remarks=''
                    junction_data_json = json.dumps(junction_data_list) if junction_data_list else None
                    record = GeneratedPDF(
                        project_id=project_id,
                        pdf_filename=pdf_filename,
                        xlsx_filename=filename,
                        checksum_md5=db_checksum,
                        file_size=file_size,
                        checksum_algo="md5",
                        metadata_checksum=meta.get("metadata_checksum"),
                        metadata_data=meta.get("metadata_data"),
                        initial_size_bytes=meta.get("initial_size_bytes"),
                        final_size_bytes=meta.get("final_size_bytes"),
                        metadata_ts_ist=meta.get("metadata_ts_ist"),
                        station_code=meta.get("station_code"),
                        source_pdf_name=meta.get("source_pdf_name"),
                        full_file_md5=meta.get("full_file_md5") or file_md5,
                        remarks=remarks if remarks else None,
                        created_at=get_ist_now(),
                        version=next_version,
                        junction_data=junction_data_json  # Store junction data
                    )
                    db.session.add(record)
                    db.session.commit()
                
                    # ✅ CRITICAL: SET PROJECT STAGE TO 10 (PDF Generated)
                    try:
                        if current_project:
                            # Update project stage to 10 (PDF Generated)
                            current_project.stage = 10
                            current_project.updated_date = get_ist_now()
                            print(f"✅ Updated project {project_id} stage to 10 (PDF generated)")
                    except Exception as stage_error:
                        print(f"⚠️ Warning: Could not update project stage: {stage_error}")
                    
                    # Update StationDrawing with latest checksum and version
                    try:
                        # Refresh the StationDrawing data
                        station_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
                        if station_drawing:
                            # Ensure project name matches station name
                            if station_drawing.station_name and current_project.name != station_drawing.station_name:
                                current_project.name = station_drawing.station_name
                                current_project.updated_date = get_ist_now()
                                db.session.commit()
                                print(f"✅ Synchronized project name with station name: {station_drawing.station_name}")
                        
                            # Debug output to verify data
                            print(f"🔍 DEBUG StationDrawing after import:")
                            print(f" - station_id: {station_drawing.station_id}")
                            print(f" - station_name: {station_drawing.station_name}")
                            print(f" - station_code: {station_drawing.station_code}")
                            print(f" - version: {station_drawing.version}")
                            print(f" - checksum: {station_drawing.checksum}")
                        
                    except Exception as e:
                        print(f"❌ Error synchronizing project and station data: {str(e)}")
                        db.session.rollback()
                    
                
                    admin_users = User.query.filter_by(role='admin').all()
                    for admin in admin_users:
                        # Check if admin is assigned to this project
                        if current_project in admin.projects or admin.role == 'admin':  # Admin might have access to all
                            notification = Notification(
                                user_id=admin.id,
                                pdf_id=record.id,
                                project_id=project_id,
                                level='New_Drawing',
                                status='pending',
                                # UPDATED MESSAGE FORMAT
                                message=f'NEW DRAWING requires admin attention: {current_project.name if current_project else "Unknown"}'
                            )
                            db.session.add(notification)

                    # Also create notification for level1 users assigned to this project
                    level1_users = User.query.filter_by(designation='level1').all()
                    for user in level1_users:
                        # Check if level1 user is assigned to this project
                        if current_project in user.projects:
                            notification = Notification(
                                user_id=user.id,
                                pdf_id=record.id,
                                project_id=project_id,
                                level='level1',
                                status='pending',
                                # UPDATED MESSAGE FORMAT
                                message=f'NEW DRAWING requires level1 approval by: {user.username}'
                            )
                            db.session.add(notification)
                    
                    db.session.commit()  # Commit the notifications
                    
                
                
                  
                
        except Exception as e:
            print("PDF conversion failed:", str(e))
            # Make sure the folder exists
          