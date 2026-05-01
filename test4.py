import os
import sys
import time
import hashlib
import subprocess
import shutil
import re
import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
import logging
import paramiko
import ftplib
from openpyxl import load_workbook
from flask import Flask

# Add the app directory to Python path if running standalone
if __name__ == '__main__':
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

try:
    from Circuitbuilding.app.database import db
    from Circuitbuilding.app.models import (
        get_ist_now, Project, StationDrawing, GeneratedPDF,
        User, Notification, Cable, JunctionBox, Terminal,
        Group, TerminalHeader, ChokeTable, ResistorTable, CableBox,
        StationMaster  # Added for station access checking
    )
    PACKAGE_MODE = True
except ImportError:
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
    from Circuitbuilding.app.database import db
    from Circuitbuilding.app.models import (
        get_ist_now, Project, StationDrawing, GeneratedPDF,
        User, Notification, Cable, JunctionBox, Terminal,
        Group, TerminalHeader, ChokeTable, ResistorTable, CableBox,
        StationMaster  # Added for station access checking
    )
    PACKAGE_MODE = False

# Define IST timezone (UTC+5:30)
IST = timezone(timedelta(hours=5, minutes=30))

# ==================== GLOBAL TRACKING VARIABLE ====================
PROCESSED_FILES_CACHE = set()
SFTP_CONNECTION_POOL = {}
LAST_CONNECTION_CLEANUP = datetime.now()

# ==================== ENVIRONMENT LOADING ====================
def load_environment_variables():
    """Load environment variables from .env file"""
    try:
        from dotenv import load_dotenv
        script_dir = Path(__file__).resolve().parent
        env_path = script_dir / 'Circuitbuilding' / '.env'
        
        if env_path.exists():
            load_dotenv(dotenv_path=env_path)
            print(f"✓ Loaded environment variables from: {env_path}")
        else:
            print(f"⚠️ .env file not found at: {env_path}")
            fallback_env = script_dir / '.env'
            if fallback_env.exists():
                load_dotenv(dotenv_path=fallback_env)
                print(f"✓ Loaded environment variables from fallback: {fallback_env}")
    except ImportError:
        print("⚠️ python-dotenv not installed. Install with: pip install python-dotenv")
        print("⚠️ Using system environment variables instead.")
    except Exception as e:
        print(f"⚠️ Error loading environment variables: {str(e)}")

load_environment_variables()

# ==================== CONFIGURATION ====================
CONFIG = {
    'SECRET_KEY': os.environ.get("SECRET_KEY", "Saltriver@123"),
    'DB_URI': "postgresql://postgres:Omhari%408899@pso.cellapps.com:5432/postgrestest",
    'FTP_ENABLED': os.environ.get("FTP_ENABLED", "True").lower() == "true",
    'FTP_HOST': os.environ.get("FTP_HOST", "45.67.216.178"),
    'FTP_PORT': int(os.environ.get("FTP_PORT", 22)),
    'FTP_USERNAME': os.environ.get("FTP_USERNAME", "root"),
    'FTP_PASSWORD': os.environ.get("FTP_PASSWORD", "Mehul88$"),
    'FTP_UPLOAD_DIR': os.environ.get("FTP_UPLOAD_DIR", "/srv/railway/frontend/uploads/"),
    'FTP_XLSX_TAKE_DIR': os.environ.get("FTP_XLSX_TAKE_DIR", "/srv/railway/frontend/xlsx_download/"),
    'FTP_USE_SFTP': os.environ.get("FTP_USE_SFTP", "True").lower() == "true",
    'FTP_TIMEOUT': int(os.environ.get("FTP_TIMEOUT", 30)),
    'AUTO_GENERATE_PDF_ON_DOWNLOAD': os.environ.get("AUTO_GENERATE_PDF_ON_DOWNLOAD", "True").lower() == "true",
    'LOCAL_TEMP_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp'),
    'LOCAL_XLSX_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_xlsx'),
    'LOCAL_PDF_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_pdf'),
    'LOCAL_DOWNLOADS_VISIBLE_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloaded_xlsx_files'),
    'LOCAL_UPLOADS_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads'),
    'CONVERTER_SCRIPT': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel_to_pdf_converter.py'),
    'CHECK_INTERVAL': int(os.environ.get("CHECK_INTERVAL", 30)),
    'ARCHIVE_REMOTE': os.environ.get("ARCHIVE_REMOTE", "True").lower() == "true",
    'REMOTE_ARCHIVE_DIR': os.environ.get("REMOTE_ARCHIVE_DIR", "/srv/railway/frontend/xlsx_download/processed/"),
    'REMOTE_FAILED_DIR': os.environ.get("REMOTE_FAILED_DIR", "/srv/railway/frontend/xlsx_download/failed/"),
    'REMOTE_BACKUP_DIR': os.environ.get("REMOTE_BACKUP_DIR", "/srv/railway/frontend/xlsx_download/backup/"),
    'XLSX_PATTERNS': ['*.xlsx', '*.XLSX'],
    'LOG_LEVEL': os.environ.get("LOG_LEVEL", "INFO"),
    'LOG_FILE': os.environ.get("LOG_FILE", "ftp_converter.log"),
    'PROCESSED_FILES_LOG': os.environ.get("PROCESSED_FILES_LOG", "processed_files.log"),
    'KEEP_DOWNLOADED_COPIES': os.environ.get("KEEP_DOWNLOADED_COPIES", "True").lower() == "true",
    # New connection settings
    'SFTP_MAX_RETRIES': 3,
    'SFTP_RETRY_DELAY': 5,
    'SFTP_KEEPALIVE_INTERVAL': 30,
    'OPERATION_DELAY': 1,
    'CONNECTION_POOL_TIMEOUT': 300,  # 5 minutes
}

# Setup logging
log_file_handler = logging.FileHandler(CONFIG['LOG_FILE'], encoding='utf-8')
log_stream_handler = logging.StreamHandler()
log_stream_handler.stream.reconfigure(encoding='utf-8')

logging.basicConfig(
    level=getattr(logging, CONFIG['LOG_LEVEL']),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        log_file_handler,
        log_stream_handler
    ]
)
logger = logging.getLogger(__name__)

# Initialize Flask app for db context
app = Flask(__name__)
app.config['SECRET_KEY'] = CONFIG['SECRET_KEY']
app.config['SQLALCHEMY_DATABASE_URI'] = CONFIG['DB_URI']
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

with app.app_context():
    db.create_all()

# ==================== PROCESSED FILES TRACKING ====================
def load_processed_files():
    """Load previously processed files from log to avoid reprocessing"""
    global PROCESSED_FILES_CACHE
    try:
        if os.path.exists(CONFIG['PROCESSED_FILES_LOG']):
            with open(CONFIG['PROCESSED_FILES_LOG'], 'r', encoding='utf-8') as f:
                for line in f:
                    filename = line.strip()
                    if filename:
                        PROCESSED_FILES_CACHE.add(filename)
            logger.info(f"Loaded {len(PROCESSED_FILES_CACHE)} previously processed files from log")
    except Exception as e:
        logger.error(f"Error loading processed files log: {e}")

def save_processed_file(filename):
    """Save processed filename to log"""
    global PROCESSED_FILES_CACHE
    try:
        PROCESSED_FILES_CACHE.add(filename)
        with open(CONFIG['PROCESSED_FILES_LOG'], 'a', encoding='utf-8') as f:
            f.write(f"{filename}\n")
    except Exception as e:
        logger.error(f"Error saving processed file {filename}: {e}")

def is_file_processed(filename):
    """Check if file has already been processed"""
    global PROCESSED_FILES_CACHE
    return filename in PROCESSED_FILES_CACHE

# ==================== HELPER FUNCTIONS (EXACTLY MATCHING ROUTES.PY) ====================
def get_role_display(role):
    """Convert role number to display name - EXACTLY MATCHING ROUTES.PY"""
    role_map = {
        '0': 'Viewer',
        '1': 'Creator',
        '2': 'Approver L2',
        '3': 'Approver L3',
        '4': 'Admin'
    }
    return role_map.get(str(role), 'Unknown')

def get_user_permissions(user):
    """
    Get user permissions based on role (0,1,2,3,4) - EXACTLY MATCHING ROUTES.PY
    """
    role = str(user.role)
    
    permissions = {
        'can_view_approvals': False,
        'can_create_drawing': False,
        'can_approve_level1': False,
        'can_approve_level2': False,
        'can_approve_level3': False,
        'can_see_all': False,
        'role': role,
        'role_display': get_role_display(role)
    }
    
    if role == '0':
        permissions['can_view_approvals'] = True
    elif role == '1':
        permissions['can_view_approvals'] = True
        permissions['can_create_drawing'] = True
        permissions['can_approve_level1'] = True
    elif role == '2':
        permissions['can_view_approvals'] = True
        permissions['can_approve_level2'] = True
    elif role == '3':
        permissions['can_view_approvals'] = True
        permissions['can_approve_level3'] = True
    elif role == '4':
        permissions['can_view_approvals'] = True
        permissions['can_create_drawing'] = True
        permissions['can_approve_level1'] = True
        permissions['can_approve_level2'] = True
        permissions['can_approve_level3'] = True
        permissions['can_see_all'] = True
    
    return permissions

# ==================== USER ACCESS CHECKING FUNCTIONS ====================
def check_user_project_access(user, project):
    """Check if a user has access to a specific project"""
    try:
        # Admin users (role='4') have access to all projects
        if user.role == '4':
            return True
        
        # Check direct project assignment via user.projects relationship
        if hasattr(user, 'projects'):
            if project in user.projects:
                return True
        
        # Check if user has access via station-specific permissions
        station_drawing = StationDrawing.query.filter_by(
            project_id=project.id
        ).first()
        
        if station_drawing and station_drawing.station_code:
            station_master = StationMaster.query.filter_by(
                station_code=station_drawing.station_code,
                project_id=project.id
            ).first()
            
            if station_master:
                return True
        
        return False
        
    except Exception as e:
        logger.error(f"Error checking user project access: {e}")
        return False

# ==================== NOTIFICATION FUNCTIONS ====================
def create_pdf_notifications_ftp(generated_pdf, project, station_name, version, file_type='FINAL'):
    """
    Create notifications for FTP-generated PDF - ONLY FOR USERS WITH PROJECT ACCESS
    Called when a new PDF is generated via FTP automation
    """
    with app.app_context():
        try:
            logger.info(f"📢 Creating notifications for {file_type} PDF {generated_pdf.id} in project {project.id}")
            
            # For JUNCTION files, don't create notifications
            if file_type == 'JUNCTION':
                logger.info(f"⏭️ Skipping notifications for JUNCTION file: {generated_pdf.id}")
                return
            
            # Only create notifications for FINAL files
            if file_type != 'FINAL':
                logger.info(f"⏭️ Skipping notifications for {file_type} file: {generated_pdf.id}")
                return
            
            # Get level1 users
            level1_users = User.query.filter_by(designation='level1', is_active=True).all()
            
            if not level1_users:
                logger.info("No level1 users found with designation='level1'")
                level1_users = User.query.filter_by(role='1', is_active=True).all()
                logger.info(f"Found {len(level1_users)} users with role='1' instead")
            
            notification_count = 0
            for user in level1_users:
                user_has_access = check_user_project_access(user, project)
                
                if user_has_access:
                    notification = Notification(
                        user_id=user.id,
                        pdf_id=generated_pdf.id,
                        project_id=project.id,
                        level='level1',
                        status='pending',
                        message=f'NEW DRAWING requires level1 approval by: {user.username}'
                    )
                    db.session.add(notification)
                    notification_count += 1
                    logger.info(f"✅ Created Level 1 notification for: {user.username} (has project access)")
                else:
                    logger.debug(f"Skipping user {user.username} - no access to project {project.id}")
            
            if notification_count == 0:
                logger.warning(f"No level1 users found with access to project {project.id}")
                fallback_users = User.query.filter_by(role='4', is_active=True).all()
                for user in fallback_users:
                    notification = Notification(
                        user_id=user.id,
                        pdf_id=generated_pdf.id,
                        project_id=project.id,
                        level='level1',
                        status='pending',
                        message=f'NEW DRAWING requires level1 approval by: {user.username} (Admin fallback)'
                    )
                    db.session.add(notification)
                    notification_count += 1
                    logger.info(f"✅ Created fallback Admin notification for: {user.username}")
            
            db.session.commit()
            logger.info(f"✅ Created {notification_count} Level 1 notification(s) for PDF {generated_pdf.id}")
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"❌ Error creating notifications: {str(e)}")
            import traceback
            traceback.print_exc()

def update_pdf_notifications_on_approval(pdf_file, project, level, new_status, approver_user):
    with app.app_context():
        try:
            notif = Notification.query.filter_by(
                pdf_id=pdf_file.id,
                level=level
            ).first()
            
            if notif:
                notif.status = new_status
                notif.is_read = True
                notif.updated_at = get_ist_now()
                logger.info(f"✅ Updated notification: PDF {pdf_file.id}, Level {level} = {new_status}")
            
            if new_status == 'approved':
                if level == 'level1':
                    level2_users = User.query.filter(
                        (User.role == '2') | (User.designation == 'level2'),
                        User.is_active == True
                    ).all()
                    
                    for user in level2_users:
                        user_has_access = check_user_project_access(user, project)
                        
                        if user_has_access:
                            new_notif = Notification(
                                user_id=user.id,
                                pdf_id=pdf_file.id,
                                project_id=project.id,
                                level='level2',
                                status='pending',
                                message=f"Drawing in {project.name} approved by {approver_user.username} ({get_role_display(approver_user.role)}) at Level 1. Requires your Level 2 approval."
                            )
                            db.session.add(new_notif)
                            logger.info(f"✅ Created Level 2 notification for {user.username} (has project access)")
                        else:
                            logger.debug(f"Skipping Level 2 user {user.username} - no access to project {project.id}")
                
                elif level == 'level2':
                    level3_users = User.query.filter(
                        (User.role == '3') | (User.designation == 'level3'),
                        User.is_active == True
                    ).all()
                    
                    for user in level3_users:
                        user_has_access = check_user_project_access(user, project)
                        
                        if user_has_access:
                            new_notif = Notification(
                                user_id=user.id,
                                pdf_id=pdf_file.id,
                                project_id=project.id,
                                level='level3',
                                status='pending',
                                message=f"Drawing in {project.name} approved by {approver_user.username} ({get_role_display(approver_user.role)}) at Level 2. Requires your Level 3 approval."
                            )
                            db.session.add(new_notif)
                            logger.info(f"✅ Created Level 3 notification for {user.username} (has project access)")
                        else:
                            logger.debug(f"Skipping Level 3 user {user.username} - no access to project {project.id}")
                
                elif level == 'level3':
                    all_active_users = User.query.filter_by(is_active=True).all()
                    
                    for user in all_active_users:
                        user_has_access = check_user_project_access(user, project)
                        
                        if user_has_access:
                            final_notif = Notification(
                                user_id=user.id,
                                pdf_id=pdf_file.id,
                                project_id=project.id,
                                level='final',
                                status='approved',
                                message=f"🎉 Drawing in {project.name} has been FULLY APPROVED by {approver_user.username} ({get_role_display(approver_user.role)}) at Level 3!"
                            )
                            db.session.add(final_notif)
                            logger.info(f"✅ Created FINAL APPROVED notification for {user.username} (has project access)")
                        else:
                            logger.debug(f"Skipping user {user.username} for final approval - no access to project {project.id}")
            
            if new_status == 'rejected':
                creators = User.query.filter(
                    (User.role == '1') | (User.designation == 'level1'),
                    User.is_active == True
                ).all()
                
                for user in creators:
                    user_has_access = check_user_project_access(user, project)
                    
                    if user_has_access:
                        reject_notif = Notification(
                            user_id=user.id,
                            pdf_id=pdf_file.id,
                            project_id=project.id,
                            level=level,
                            status='rejected',
                            message=f"Drawing in {project.name} was REJECTED at Level {level.replace('level', '')} by {approver_user.username} ({get_role_display(approver_user.role)}). Reason: {pdf_file.rejection_reason if hasattr(pdf_file, 'rejection_reason') and pdf_file.rejection_reason else 'No reason provided'}"
                        )
                        db.session.add(reject_notif)
                        logger.info(f"✅ Created REJECTION notification for {user.username} (has project access)")
                    else:
                        logger.debug(f"Skipping creator {user.username} for rejection - no access to project {project.id}")
            
            db.session.commit()
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"❌ Error updating notifications: {str(e)}")

# ==================== PROJECT STATUS FUNCTIONS ====================
def check_project_status(project_id):
    """Check if project has status 'ready_for_pdf' OR 'preview' OR 'junction' - Process ALL THREE"""
    with app.app_context():
        try:
            project = Project.query.get(project_id)
            if not project:
                logger.warning(f"Project {project_id} not found in database")
                print(f"❌ Project {project_id} not found in database")
                return False
            
            if hasattr(project, 'status'):
                current_status = project.status
                current_stage = project.stage if hasattr(project, 'stage') else None
                
                logger.info(f"Project {project_id} - Status: {current_status}, Stage: {current_stage}")
                
                # Check if status is 'ready_for_pdf' OR 'preview' OR 'junction' - PROCESS ALL THREE
                if current_status in ['ready_for_pdf', 'preview', 'junction']:
                    logger.info(f"✅ Project {project_id} has status '{current_status}' - WILL PROCESS")
                    print(f"✅ Project {project_id} status: '{current_status}' - Will process")
                    return True
                
                # For any other status, don't process
                logger.info(f"⚠️ Project {project_id} has status '{current_status}' - SKIPPING")
                print(f"⏸️ Project {project_id} status: '{current_status}' - Skipping")
                return False
            else:
                logger.warning(f"Project {project_id} has no 'status' field, assuming ready")
                print(f"⚠️ Project {project_id} has no status field, assuming ready")
                return True
        except Exception as e:
            logger.error(f"Error checking project status for ID {project_id}: {e}")
            print(f"❌ Error checking project {project_id} status: {e}")
            return False

def update_project_status(project_id, new_status, new_stage=None, is_preview=False, is_junction=False):
    """Update project status and stage in database - Handle preview and junction specially"""
    with app.app_context():
        try:
            project = Project.query.get(project_id)
            if not project:
                logger.error(f"Project {project_id} not found")
                return False
            
            old_status = project.status if hasattr(project, 'status') else None
            old_stage = project.stage if hasattr(project, 'stage') else None
            
            # If it's a preview file and current status is 'preview', keep it as 'preview'
            # If it's a junction file and current status is 'junction', keep it as 'junction'
            # Otherwise update to new_status
            if is_preview and old_status == 'preview':
                # Keep as preview for preview files
                project.status = 'preview'
                new_status = 'preview'  # For logging
            elif is_junction and old_status == 'junction':
                # Keep as junction for junction files
                project.status = 'junction'
                new_status = 'junction'  # For logging
            elif hasattr(project, 'status'):
                project.status = new_status
            
            if new_stage is not None and hasattr(project, 'stage'):
                project.stage = new_stage
            
            project.updated_at = get_ist_now()
            db.session.commit()
            
            status_log = f"status: {old_status} -> {new_status}" if hasattr(project, 'status') else ""
            stage_log = f"stage: {old_stage} -> {new_stage}" if new_stage is not None and hasattr(project, 'stage') else ""
            
            logger.info(f"Updated project {project_id}: {status_log} {stage_log}")
            print(f"✓ Project {project_id} updated: {status_log} {stage_log}")
            return True
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error updating project {project_id}: {e}")
            return False

# ==================== ENHANCED SFTP CONNECTION FUNCTIONS ====================
def get_sftp_connection_with_retry(max_retries=None, retry_delay=None):
    """Establish SFTP connection with retry logic and keepalive"""
    if max_retries is None:
        max_retries = CONFIG['SFTP_MAX_RETRIES']
    if retry_delay is None:
        retry_delay = CONFIG['SFTP_RETRY_DELAY']
    
    for attempt in range(max_retries):
        try:
            logger.info(f"Connecting to SFTP server (attempt {attempt+1}/{max_retries}): {CONFIG['FTP_HOST']}:{CONFIG['FTP_PORT']}")
            
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            
            ssh.connect(
                hostname=CONFIG['FTP_HOST'],
                port=CONFIG['FTP_PORT'],
                username=CONFIG['FTP_USERNAME'],
                password=CONFIG['FTP_PASSWORD'],
                timeout=CONFIG['FTP_TIMEOUT'],
                allow_agent=False,
                look_for_keys=False,
                banner_timeout=60,  # Increased banner timeout
                auth_timeout=30,    # Auth timeout
            )
            
            # Enable keepalive
            transport = ssh.get_transport()
            if transport:
                transport.set_keepalive(CONFIG['SFTP_KEEPALIVE_INTERVAL'])
            
            sftp = ssh.open_sftp()
            logger.info("SFTP connection established successfully")
            return ssh, sftp
            
        except paramiko.ssh_exception.SSHException as e:
            logger.error(f"SSH connection failed (attempt {attempt+1}): {str(e)}")
            if "Error reading SSH protocol banner" in str(e):
                logger.warning("Server forcibly closed connection. This might be due to firewall or rate limiting.")
            
            if attempt < max_retries - 1:
                wait_time = retry_delay * (attempt + 1)  # Exponential backoff
                logger.info(f"Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                logger.error(f"Failed to connect after {max_retries} attempts")
                return None, None
        except Exception as e:
            logger.error(f"Unexpected connection error (attempt {attempt+1}): {str(e)}")
            if attempt < max_retries - 1:
                wait_time = retry_delay * (attempt + 1)
                logger.info(f"Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                return None, None
    
    return None, None

def get_sftp_connection():
    """Wrapper for backward compatibility"""
    return get_sftp_connection_with_retry()

def get_connection_from_pool():
    """Get SFTP connection from pool or create new one"""
    global SFTP_CONNECTION_POOL, LAST_CONNECTION_CLEANUP
    
    # Clean up stale connections every 5 minutes
    if (datetime.now() - LAST_CONNECTION_CLEANUP).seconds > 300:
        cleanup_connection_pool()
        LAST_CONNECTION_CLEANUP = datetime.now()
    
    connection_key = f"{CONFIG['FTP_HOST']}:{CONFIG['FTP_PORT']}"
    
    if connection_key in SFTP_CONNECTION_POOL:
        ssh, sftp, last_used = SFTP_CONNECTION_POOL[connection_key]
        
        # Check if connection is stale
        if (datetime.now() - last_used).seconds < CONFIG['CONNECTION_POOL_TIMEOUT']:
            # Test if connection is still alive
            try:
                sftp.stat('.')
                SFTP_CONNECTION_POOL[connection_key] = (ssh, sftp, datetime.now())
                logger.debug(f"Reusing connection from pool: {connection_key}")
                return ssh, sftp
            except:
                logger.debug("Connection in pool is dead, creating new one")
                try:
                    sftp.close()
                    ssh.close()
                except:
                    pass
                del SFTP_CONNECTION_POOL[connection_key]
    
    # Create new connection
    ssh, sftp = get_sftp_connection_with_retry()
    if sftp:
        SFTP_CONNECTION_POOL[connection_key] = (ssh, sftp, datetime.now())
    
    return ssh, sftp

def cleanup_connection_pool():
    """Clean up stale connections from pool"""
    global SFTP_CONNECTION_POOL
    logger.debug("Cleaning up connection pool...")
    
    for connection_key, (ssh, sftp, last_used) in list(SFTP_CONNECTION_POOL.items()):
        if (datetime.now() - last_used).seconds > CONFIG['CONNECTION_POOL_TIMEOUT']:
            try:
                sftp.close()
                ssh.close()
                logger.debug(f"Closed stale connection: {connection_key}")
            except:
                pass
            del SFTP_CONNECTION_POOL[connection_key]
    
    logger.debug(f"Connection pool size: {len(SFTP_CONNECTION_POOL)}")

def upload_file_remote_with_persistent_connection(local_path, remote_path, ssh=None, sftp=None):
    """Upload file with option to reuse existing connection"""
    close_connection = False
    if ssh is None or sftp is None:
        ssh, sftp = get_connection_from_pool()
        if not sftp:
            return False
        close_connection = False  # Don't close pooled connections
    
    verified = False
    try:
        remote_dir = os.path.dirname(remote_path)
        try:
            sftp.stat(remote_dir)
        except FileNotFoundError:
            parts = remote_dir.split('/')
            current_path = ''
            for part in parts:
                if part:
                    current_path = current_path + '/' + part if current_path else '/' + part
                    try:
                        sftp.stat(current_path)
                    except:
                        sftp.mkdir(current_path)
        
        def progress_callback(transferred, total):
            if total > 0:
                percent = (transferred / total) * 100
                if int(percent) % 25 == 0:
                    logger.info(f"Upload progress: {percent:.1f}% ({transferred}/{total} bytes)")
        
        sftp.put(local_path, remote_path, callback=progress_callback)
        
        # Verify upload
        try:
            remote_stat = sftp.stat(remote_path)
            local_size = os.path.getsize(local_path)
            
            if remote_stat.st_size == local_size:
                logger.info(f"Verified SFTP upload: {remote_path} (size: {remote_stat.st_size} bytes)")
                verified = True
            else:
                logger.error(f"Size mismatch: local={local_size}, remote={remote_stat.st_size}")
                verified = False
                
        except Exception as ve:
            logger.error(f"SFTP upload verification failed: {ve}")
            verified = False
        
        if verified:
            logger.info(f"Uploaded via SFTP: {local_path} -> {remote_path}")
        else:
            logger.warning(f"Upload completed but verification failed for {remote_path}")
            
    except Exception as e:
        logger.error(f"Upload failed: {local_path} -> {remote_path}: {str(e)}")
        verified = False
    finally:
        if close_connection and ssh:
            try:
                sftp.close()
                ssh.close()
                logger.debug("SFTP connection closed")
            except:
                pass
    
    return verified

def download_file_remote_with_persistent_connection(remote_path, local_path, ssh=None, sftp=None):
    """Download with persistent connection option"""
    close_connection = False
    if ssh is None or sftp is None:
        ssh, sftp = get_connection_from_pool()
        if not sftp:
            return False
        close_connection = False
    
    try:
        def progress_callback(transferred, total):
            if total > 0:
                percent = (transferred / total) * 100
                if int(percent) % 25 == 0:
                    logger.info(f"Download progress: {percent:.1f}%")
        
        sftp.get(remote_path, local_path, callback=progress_callback)
        
        if os.path.exists(local_path) and os.path.getsize(local_path) > 0:
            logger.info(f"Downloaded via SFTP: {remote_path} -> {local_path}")
            return True
        else:
            logger.error(f"Download verification failed: {local_path}")
            return False
            
    except Exception as e:
        logger.error(f"Download failed: {remote_path} -> {local_path}: {str(e)}")
        return False
    finally:
        if close_connection and ssh:
            try:
                sftp.close()
                ssh.close()
            except:
                pass

def bulk_remote_operations(operations):
    """
    Perform multiple remote operations with a single connection
    operations: list of tuples ('upload', local_path, remote_path) or 
                ('download', remote_path, local_path) or
                ('move', source_path, dest_path) or
                ('delete', remote_path)
    """
    ssh, sftp = get_connection_from_pool()
    if not sftp:
        return []
    
    results = []
    try:
        for op_type, *args in operations:
            success = False
            if op_type == 'upload':
                local_path, remote_path = args
                success = upload_file_remote_with_persistent_connection(
                    local_path, remote_path, ssh, sftp
                )
                results.append(('upload', local_path, remote_path, success))
                
            elif op_type == 'download':
                remote_path, local_path = args
                success = download_file_remote_with_persistent_connection(
                    remote_path, local_path, ssh, sftp
                )
                results.append(('download', remote_path, local_path, success))
                
            elif op_type == 'move':
                source_path, dest_path = args
                try:
                    sftp.rename(source_path, dest_path)
                    success = True
                    logger.info(f"Moved on remote: {source_path} -> {dest_path}")
                except Exception as e:
                    logger.error(f"Move failed: {source_path} -> {dest_path}: {str(e)}")
                    success = False
                results.append(('move', source_path, dest_path, success))
                
            elif op_type == 'delete':
                remote_path = args[0]
                try:
                    sftp.remove(remote_path)
                    success = True
                    logger.info(f"Deleted from remote: {remote_path}")
                except Exception as e:
                    logger.error(f"Delete failed: {remote_path}: {str(e)}")
                    success = False
                results.append(('delete', remote_path, success))
            
            time.sleep(CONFIG['OPERATION_DELAY'])
        
        return results
        
    except Exception as e:
        logger.error(f"Bulk operations failed: {str(e)}")
        return results
    finally:
        # Update connection timestamp but don't close it
        if ssh and sftp:
            connection_key = f"{CONFIG['FTP_HOST']}:{CONFIG['FTP_PORT']}"
            SFTP_CONNECTION_POOL[connection_key] = (ssh, sftp, datetime.now())

# ==================== ENHANCED FTP/SFTP FUNCTIONS ====================
def list_remote_files(remote_dir):
    """List files in remote directory"""
    files = []
    
    if CONFIG['FTP_USE_SFTP']:
        ssh, sftp = get_connection_from_pool()
        if sftp:
            try:
                files = sftp.listdir(remote_dir)
                logger.info(f"Listed {len(files)} files in {remote_dir}")
            except Exception as e:
                logger.error(f"Error listing SFTP directory {remote_dir}: {e}")
                parent_dir = os.path.dirname(remote_dir)
                if parent_dir != remote_dir:
                    try:
                        parent_files = sftp.listdir(parent_dir)
                        logger.info(f"Fallback: Listed parent {parent_dir}: {parent_files[:10]}...")
                    except:
                        pass
    else:
        ftp = get_ftp_connection()
        if ftp:
            try:
                ftp.cwd(remote_dir)
                files = ftp.nlst()
                logger.info(f"Listed {len(files)} files in {remote_dir}")
                ftp.quit()
            except Exception as e:
                logger.error(f"Error listing FTP directory {remote_dir}: {e}")
                ftp.quit()
    
    return files

def download_file_remote(remote_path, local_path):
    """Download file from remote server"""
    return download_file_remote_with_persistent_connection(remote_path, local_path)

def upload_file_remote(local_path, remote_path):
    """Upload file to remote server"""
    return upload_file_remote_with_persistent_connection(local_path, remote_path)

def move_file_remote(source_path, dest_path):
    """Move file on remote server"""
    operations = [('move', source_path, dest_path)]
    results = bulk_remote_operations(operations)
    if results:
        return results[0][-1]
    return False

def delete_file_remote(remote_path):
    """Delete file from remote server"""
    operations = [('delete', remote_path)]
    results = bulk_remote_operations(operations)
    if results:
        return results[0][-1]
    return False

def get_ftp_connection():
    """Establish FTP connection"""
    try:
        logger.info(f"Connecting to FTP server: {CONFIG['FTP_HOST']}:21")
        
        ftp = ftplib.FTP()
        ftp.connect(CONFIG['FTP_HOST'], 21, timeout=CONFIG['FTP_TIMEOUT'])
        ftp.login(CONFIG['FTP_USERNAME'], CONFIG['FTP_PASSWORD'])
        ftp.set_pasv(True)
        
        logger.info("FTP connection established successfully")
        return ftp
        
    except Exception as e:
        logger.error(f"FTP connection failed: {str(e)}")
        return None

# ==================== ENHANCED PROCESSING FUNCTIONS ====================
def get_project_id_from_filename(filename):
    """Extract project ID from filename"""
    patterns = [
        r'PROJECT_(\d+)_',
        r'RAILWAYPROJECT_ID(\d+)_',
        r'railway_project_(\d+)_',
        r'project_(\d+)_',
        r'_ID(\d+)_',
        r'ID(\d+)[_-]',
        r'_P(\d+)[_-]'
    ]
    
    filename_upper = filename.upper()
    
    for pattern in patterns:
        match = re.search(pattern, filename_upper)
        if match:
            return int(match.group(1))
    
    numbers = re.findall(r'\b(\d{2,})\b', filename)
    if numbers:
        for num in numbers:
            if 10 <= int(num) <= 999:
                return int(num)
    
    return None

def parse_converter_stdout(stdout_text):
    """Parse converter script output for metadata"""
    META_PATTERNS = {
        "metadata_checksum": re.compile(
            r"(?:Metadata\s+Checksum|Initial\s+checksum\s+with\s+initial\s+file\s+size)\s*:\s*([0-9a-fA-F]{32})",
            re.IGNORECASE,
        ),
        "metadata_data": re.compile(r"(?:Metadata\s+Data\s+string|Checksum\s+data\s+string)\s*:\s*(.+)", re.IGNORECASE),
        "initial_size_bytes": re.compile(r"Initial\s+file\s+size\s*:\s*(\d+)\s*bytes", re.IGNORECASE),
        "final_size_bytes": re.compile(r"(?:Final\s+file\s+size|Final\s+file\s+size\s+after\s+enhancement)\s*:\s*(\d+)\s*bytes", re.IGNORECASE),
        "metadata_ts_ist": re.compile(r"Timestamp\s*\(IST\)\s*:\s*([0-9:\-\s]+)", re.IGNORECASE),
        "station_code": re.compile(r"Station\s+code\s*:\s*([A-Za-z0-9\-_]+)", re.IGNORECASE),
        "source_pdf_name": re.compile(r"File\s+name\s*:\s*(.+?\.pdf)\s*$", re.IGNORECASE | re.MULTILINE),
        "full_file_md5": re.compile(r"Full\s+file\s+MD5(?:\s*hash)?\s*:\s*([0-9a-fA-F]{32})", re.IGNORECASE),
        "generated_pdf_name": re.compile(r"Multi-page PDF saved as '([^']+)'", re.IGNORECASE),
    }
    
    out = {}
    text = stdout_text or ""
    for key, pat in META_PATTERNS.items():
        m = pat.search(text)
        out[key] = m.group(1).strip() if m else None
    
    return out

def _md5_of_file(path):
    """Calculate MD5 hash of file"""
    hasher = hashlib.md5()
    try:
        with open(path, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b''):
                hasher.update(chunk)
        return hasher.hexdigest()
    except Exception as e:
        logger.error(f"Error calculating MD5 for {path}: {e}")
        return None

def extract_xlsx_metadata(xlsx_path):
    """Extract metadata from XLSX file"""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        metadata = {
            'station_name': 'Unknown Station',
            'station_id': None,
            'station_code': None,
            'drawn_by': 'Auto Converter',
            'version': '1',
            'checksum': None,
            'diagram_name': 'railways',
            'checked_by': 'supervisor',
            'division': 'Ahemdabad',
            'zone': 'WRLY',
            'total_sheet': '17',
            'designation1': 'DY.CSTE/C-II/ADI',
            'designation2': 'DSTE/C/ADI',
            'designation3': 'SSE/SIG/C/ADI'
        }
        
        sheet_names_lower = [s.lower() for s in wb.sheetnames]
        
        if 'stationdrawing' in sheet_names_lower:
            sheet_idx = sheet_names_lower.index('stationdrawing')
            ws = wb[wb.sheetnames[sheet_idx]]
            
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
                else:
                    headers.append('')
            
            row_data = {}
            for i, cell in enumerate(ws[2]):
                if i < len(headers):
                    value = cell.value
                    if value is not None:
                        row_data[headers[i]] = str(value).strip()
            
            field_mapping = {
                'station_name': 'station_name',
                'station_id': 'station_id',
                'station_code': 'station_code',
                'drawn_by': 'drawn_by',
                'version': 'version',
                'checksum': 'checksum',
                'diagram_name': 'diagram_name',
                'checked_by': 'checked_by',
                'division': 'division',
                'zone': 'zone',
                'total_sheet': 'total_sheet',
                'designation1': 'designation1',
                'designation2': 'designation2',
                'designation3': 'designation3',
                'date': 'date'
            }
            
            for db_field, excel_field in field_mapping.items():
                if excel_field in row_data and row_data[excel_field]:
                    metadata[db_field] = row_data[excel_field]
        
        if not metadata.get('date'):
            metadata['date'] = datetime.now().strftime("%d-%m-%Y")
        
        return metadata
        
    except Exception as e:
        logger.error(f"Error extracting metadata from {xlsx_path}: {e}")
        return {
            'station_name': 'Unknown Station',
            'station_id': None,
            'station_code': None,
            'drawn_by': 'Auto Converter',
            'version': '1',
            'checksum': None,
            'diagram_name': 'railways',
            'checked_by': 'supervisor',
            'division': 'Ahemdabad',
            'zone': 'WRLY',
            'total_sheet': '17',
            'designation1': 'DY.CSTE/C-II/ADI',
            'designation2': 'DSTE/C/ADI',
            'designation3': 'SSE/SIG/C/ADI',
            'date': datetime.now().strftime("%d-%m-%Y")
        }

def get_next_version(project_id, station_code):
    """Get next version number for project and station using database models"""
    with app.app_context():
        try:
            max_version_record = GeneratedPDF.query.filter_by(
                project_id=project_id,
                station_code=station_code
            ).order_by(GeneratedPDF.version.desc()).first()
            
            if max_version_record and max_version_record.version:
                return max_version_record.version + 1
            return 1
        except Exception as e:
            logger.error(f"Error getting next version for project {project_id}, station {station_code}: {e}")
            return 1

def update_database_with_models(project_id, xlsx_filename, pdf_filename, checksum, metadata, station_info, version, is_preview=False, is_junction=False):
    """Update database using SQLAlchemy models - INCLUDES NOTIFICATIONS (for FINAL only)"""
    with app.app_context():
        try:
            station_code = station_info.get('station_code') or 'DEFAULT_STATION'
            station_name = station_info.get('station_name', f"Project_{project_id}")
            
            project = Project.query.get(project_id)
            if not project:
                logger.error(f"Project {project_id} not found in database")
                return False
            
            local_xlsx = os.path.join(CONFIG['LOCAL_XLSX_DIR'], xlsx_filename)
            local_pdf = os.path.join(CONFIG['LOCAL_PDF_DIR'], pdf_filename)
            
            xlsx_size = os.path.getsize(local_xlsx) if os.path.exists(local_xlsx) else 0
            pdf_size = os.path.getsize(local_pdf) if os.path.exists(local_pdf) else 0
            
            # Determine file type based on flags
            if is_junction:
                pdf_type = 'JUNCTION'
                remarks = 'Auto-generated JUNCTION PDF from XLSX'
            elif is_preview:
                pdf_type = 'PREVIEW'
                remarks = 'Auto-generated PREVIEW PDF from XLSX'
            else:
                pdf_type = 'FINAL'
                remarks = 'Auto-generated FINAL PDF from XLSX'
            
            generated_pdf = GeneratedPDF(
                project_id=project_id,
                pdf_filename=pdf_filename,
                xlsx_filename=xlsx_filename,
                checksum_md5=checksum,
                file_size=pdf_size,
                checksum_algo="md5",
                metadata_checksum=metadata.get('metadata_checksum'),
                metadata_data=metadata.get('metadata_data'),
                initial_size_bytes=metadata.get('initial_size_bytes'),
                final_size_bytes=metadata.get('final_size_bytes'),
                metadata_ts_ist=metadata.get('metadata_ts_ist'),
                station_code=station_code,
                source_pdf_name=metadata.get('source_pdf_name'),
                full_file_md5=metadata.get('full_file_md5') or checksum,
                remarks=remarks,
                created_at=get_ist_now(),
                version=version,
                level1_status='pending',
                level2_status='pending',
                level3_status='pending',
                junction_data=None 
            )
            
            db.session.add(generated_pdf)
            db.session.flush()
            
            station_drawing = StationDrawing.query.filter_by(
                project_id=project_id, 
                station_code=station_code
            ).first()
            
            if station_drawing:
                station_drawing.version = str(version)
                station_drawing.checksum = checksum
                station_drawing.station_name = station_name
                station_drawing.station_id = station_info.get('station_id', str(project_id))
                station_drawing.drawn_by = station_info.get('drawn_by', 'Auto Converter')
                station_drawing.diagram_name = station_info.get('diagram_name', 'railways')
                station_drawing.checked_by = station_info.get('checked_by', 'supervisor')
                station_drawing.division = station_info.get('division', 'Ahemdabad')
                station_drawing.zone = station_info.get('zone', 'WRLY')
                station_drawing.total_sheet = station_info.get('total_sheet', '17')
                station_drawing.designation1 = station_info.get('designation1', 'DY.CSTE/C-II/ADI')
                station_drawing.designation2 = station_info.get('designation2', 'DSTE/C/ADI')
                station_drawing.designation3 = station_info.get('designation3', 'SSE/SIG/C/ADI')
                station_drawing.date = station_info.get('date', datetime.now().strftime("%d-%m-%Y"))
            else:
                station_drawing = StationDrawing(
                    project_id=project_id,
                    station_id=station_info.get('station_id', str(project_id)),
                    station_name=station_name,
                    station_code=station_code,
                    version=str(version),
                    checksum=checksum,
                    drawn_by=station_info.get('drawn_by', 'Auto Converter'),
                    diagram_name=station_info.get('diagram_name', 'railways'),
                    checked_by=station_info.get('checked_by', 'supervisor'),
                    division=station_info.get('division', 'Ahemdabad'),
                    zone=station_info.get('zone', 'WRLY'),
                    total_sheet=station_info.get('total_sheet', '17'),
                    designation1=station_info.get('designation1', 'DY.CSTE/C-II/ADI'),
                    designation2=station_info.get('designation2', 'DSTE/C/ADI'),
                    designation3=station_info.get('designation3', 'SSE/SIG/C/ADI'),
                    date=station_info.get('date', datetime.now().strftime("%d-%m-%Y")),
                    created_date=get_ist_now()
                )
                db.session.add(station_drawing)
            
            db.session.commit()
            
            # Create notifications based on file type
            logger.info(f"Creating notifications for new {pdf_type} PDF...")
            create_pdf_notifications_ftp(generated_pdf, project, station_name, version, pdf_type)
            
            logger.info(f"Database updated: Project {project_id}, Station {station_code}, Version {version}, PDF ID: {generated_pdf.id}, Type: {pdf_type}")
            
            return True
            
        except Exception as e:
            logger.error(f"Database update error: {e}")
            import traceback
            traceback.print_exc()
            db.session.rollback()
            return False

def convert_xlsx_to_pdf(xlsx_path, pdf_path):
    """Convert XLSX to PDF using converter script"""
    try:
        if not os.path.exists(CONFIG['CONVERTER_SCRIPT']):
            logger.error(f"Converter script not found: {CONFIG['CONVERTER_SCRIPT']}")
            return None, "Converter script not found", None
        
        python_exe = sys.executable
        converter_dir = os.path.dirname(CONFIG['CONVERTER_SCRIPT'])
        result = subprocess.run(
            [python_exe, CONFIG['CONVERTER_SCRIPT'], xlsx_path, pdf_path],
            capture_output=True,
            text=True,
            timeout=300,
            cwd=converter_dir
        )
        
        print("=== CONVERTER STDOUT ===")
        print(result.stdout)
        print("=== CONVERTER STDERR ===")
        print(result.stderr)
        print("=== END CONVERTER LOGS ===")
        
        print(f"Converter returncode: {result.returncode}")
        print(f"Expected PDF exists: {os.path.exists(pdf_path)}")
        if os.path.exists(pdf_path):
            print(f"Expected PDF size: {os.path.getsize(pdf_path)} bytes")
        
        metadata = parse_converter_stdout(result.stdout)
        generated_pdf_name = metadata.get('generated_pdf_name')
        
        actual_pdf_path = None
        if generated_pdf_name:
            candidate_path = os.path.join(converter_dir, generated_pdf_name)
            if os.path.exists(candidate_path) and os.path.getsize(candidate_path) > 0:
                actual_pdf_path = candidate_path
                print(f"Found generated PDF: {candidate_path} (size: {os.path.getsize(candidate_path)})")
            else:
                print(f"Generated PDF name parsed but file not found: {candidate_path}")
                generated_pdf_name = None
        
        if not actual_pdf_path and os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
            actual_pdf_path = pdf_path
            print(f"Using expected PDF: {pdf_path}")
        
        if actual_pdf_path:
            logger.info(f"PDF created successfully: {os.path.basename(actual_pdf_path)} (returncode: {result.returncode})")
            print(f"Conversion successful (even with returncode {result.returncode})")
            return metadata, None, actual_pdf_path
        else:
            logger.error(f"PDF conversion failed: returncode={result.returncode}")
            error_msg = result.stderr.strip() or result.stdout.strip() or "No output from converter (check if libreoffice or dependencies are installed)"
            logger.error(f"PDF conversion failed: {error_msg}")
            return None, error_msg, None
            
    except subprocess.TimeoutExpired:
        error_msg = "Conversion timed out (300 seconds)"
        logger.error(error_msg)
        return None, error_msg, None
    except Exception as e:
        error_msg = f"Conversion error: {str(e)}"
        logger.error(error_msg)
        return None, error_msg, None

def import_xlsx_data_to_database(xlsx_path, project_id):
    """Import XLSX data into database tables (optional)"""
    with app.app_context():
        try:
            logger.info(f"Importing XLSX data for project {project_id}")
            
            wb = load_workbook(xlsx_path, data_only=True)
            
            imported_count = 0
            
            try:
                ResistorTable.query.filter_by(project_id=project_id).delete()
                ChokeTable.query.filter_by(project_id=project_id).delete()
                TerminalHeader.query.filter_by(project_id=project_id).delete()
                Group.query.filter_by(project_id=project_id).delete()
                Terminal.query.filter_by(project_id=project_id).delete()
                CableBox.query.filter_by(project_id=project_id).delete()
                Cable.query.filter_by(project_id=project_id).delete()
                JunctionBox.query.filter_by(project_id=project_id).delete()
                db.session.commit()
                logger.info(f"Deleted existing records for project {project_id}")
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error deleting existing records: {e}")
            
            if 'StationDrawing' in wb.sheetnames:
                try:
                    ws = wb['StationDrawing']
                    headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
                    
                    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                        data = {'project_id': project_id}
                        for i, cell_value in enumerate(row):
                            if i < len(headers) and cell_value is not None:
                                value_str = str(cell_value).strip()
                                if value_str:
                                    data[headers[i]] = value_str
                        
                        if data:
                            station_drawing = StationDrawing.query.filter_by(
                                project_id=project_id,
                                station_code=data.get('station_code', '')
                            ).first()
                            
                            if station_drawing:
                                for key, value in data.items():
                                    if hasattr(station_drawing, key):
                                        setattr(station_drawing, key, value)
                            else:
                                station_drawing = StationDrawing(**data)
                                db.session.add(station_drawing)
                            
                            db.session.commit()
                            imported_count += 1
                            logger.info(f"Updated StationDrawing for project {project_id}")
                except Exception as e:
                    logger.error(f"Error importing StationDrawing: {e}")
                    db.session.rollback()
            
            tables_to_import = [
                ('junction_box', JunctionBox),
                ('cable', Cable),
                ('cable_box', CableBox),
                ('terminal', Terminal),
                ('group', Group),
                ('terminal_header', TerminalHeader),
                ('choketable', ChokeTable),
                ('resistortable', ResistorTable),
            ]
            
            for sheet_name, model_class in tables_to_import:
                if sheet_name in wb.sheetnames:
                    try:
                        ws = wb[sheet_name]
                        headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
                        
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            data = {'project_id': project_id}
                            has_data = False
                            
                            for i, cell_value in enumerate(row):
                                if i < len(headers) and cell_value is not None:
                                    value_str = str(cell_value).strip()
                                    if value_str:
                                        header = headers[i].lower().replace(' ', '_').replace('-', '_')
                                        if hasattr(model_class, header):
                                            data[header] = value_str
                                            has_data = True
                            
                            if has_data:
                                try:
                                    model_instance = model_class(**data)
                                    db.session.add(model_instance)
                                    imported_count += 1
                                except Exception as e:
                                    logger.error(f"Error creating {sheet_name} instance: {e}")
                                    continue
                        
                        db.session.commit()
                        logger.info(f"Imported {sheet_name} data for project {project_id}")
                        
                    except Exception as e:
                        logger.error(f"Error importing {sheet_name}: {e}")
                        db.session.rollback()
                        continue
            
            logger.info(f"Total imported records: {imported_count}")
            return imported_count > 0
            
        except Exception as e:
            logger.error(f"Error importing XLSX data: {e}")
            db.session.rollback()
            return False

def copy_to_local_uploads_no_remove(xlsx_source, pdf_source, xlsx_dest, pdf_dest):
    """Copy files to local uploads WITHOUT removing sources"""
    try:
        os.makedirs(os.path.dirname(xlsx_dest), exist_ok=True)
        print(f"Ensuring uploads dir: {os.path.dirname(xlsx_dest)}")

        print(f"Copying XLSX: {xlsx_source} -> {xlsx_dest}")
        shutil.copy2(xlsx_source, xlsx_dest)
        logger.info(f"Copied XLSX: {xlsx_source} -> {xlsx_dest}")
        if os.path.exists(xlsx_dest):
            print(f"XLSX copied successfully: {xlsx_dest}")
        else:
            print(f"ERROR: XLSX copy failed: {xlsx_dest} does not exist")
            return False

        print(f"Copying PDF: {pdf_source} -> {pdf_dest}")
        shutil.copy2(pdf_source, pdf_dest)
        logger.info(f"Copied PDF: {pdf_source} -> {pdf_dest}")
        if os.path.exists(pdf_dest):
            print(f"PDF copied successfully: {pdf_dest}")
        else:
            print(f"ERROR: PDF copy failed: {pdf_dest} does not exist")
            return False

        logger.info(f"Files copied to local uploads (no remove): {os.path.basename(xlsx_dest)}, {os.path.basename(pdf_dest)}")
        return True

    except Exception as e:
        logger.error(f"Error in copy_to_local_uploads_no_remove: {e}")
        print(f"ERROR in copy_to_local_uploads_no_remove: {e}")
        return False

def get_junction_info_from_filename(filename):
    """Extract junction information from filename"""
    # Pattern for junction filenames: PROJECT_<id>_<station>_JUNCTION_<junction_number>_<b>_<f>_<timestamp>
    # Example: PROJECT_45_KHEDBRAHMA11_JUNCTION_1_b1_F_20260119_115304
    
    patterns = [
        r'_JUNCTION_(\d+)_',  # Extract junction number
        r'JUNCTION[_-](\d+)',  # Alternative pattern
        r'_J_(\d+)_',  # Short form
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            junction_number = match.group(1)
            return {
                'junction_number': junction_number,
                'is_junction': True
            }
    
    # Also check if "junction" appears in the filename (case insensitive)
    if 'junction' in filename.lower():
        return {
            'junction_number': '1',  # Default junction number
            'is_junction': True
        }
    
    return {
        'junction_number': None,
        'is_junction': False
    }

def process_remote_xlsx_file_enhanced(remote_filename):
    """Process a single XLSX file with enhanced connection handling"""
    remote_xlsx_path = os.path.join(CONFIG['FTP_XLSX_TAKE_DIR'], remote_filename)
    process_success = False
    try:
        logger.info(f"Processing remote file: {remote_filename}")
        print(f"\n" + "="*70)
        print(f"STARTING PROCESSING: {remote_filename}")
        print("="*70)
        
        if is_file_processed(remote_filename):
            logger.info(f"Skipping already processed file: {remote_filename}")
            print(f"⚠️  Skipping (already processed): {remote_filename}")
            return True
        
        project_id = get_project_id_from_filename(remote_filename)
        if not project_id:
            logger.error(f"Could not extract project ID from: {remote_filename}")
            print(f"❌ ERROR: No project ID in {remote_filename}")
            return False
        
        logger.info(f"Detected Project ID: {project_id}")
        print(f"📋 Project ID: {project_id}")
        
        # Check for junction info in filename
        junction_info = get_junction_info_from_filename(remote_filename)
        is_junction_file = junction_info['is_junction']
        
        # Check project status - now accepts 'ready_for_pdf', 'preview', and 'junction'
        project_ready = check_project_status(project_id)
        
        with app.app_context():
            project = Project.query.get(project_id)
            current_status = project.status if project and hasattr(project, 'status') else None
            is_preview = current_status == 'preview'
            is_junction_status = current_status == 'junction'
        
        # If file has junction in name but project status is not junction, update it
        if is_junction_file and not is_junction_status:
            with app.app_context():
                project.status = 'junction'
                project.updated_at = get_ist_now()
                db.session.commit()
                print(f"🔄 Updated project {project_id} status to 'junction' for junction file")
                is_junction_status = True
        
        if not project_ready:
            logger.info(f"Project {project_id} is not 'ready_for_pdf', 'preview', or 'junction'. Skipping file: {remote_filename}")
            print(f"⏸️  Project {project_id} not ready_for_pdf, preview, or junction. Skipping: {remote_filename}")
            return False
        
        print(f"✅ Project {project_id} status is '{current_status}' - Proceeding...")
        
        # Determine file type
        if is_junction_file or is_junction_status:
            file_type = 'JUNCTION'
        elif is_preview:
            file_type = 'PREVIEW'
        else:
            file_type = 'FINAL'
        
        print(f"📊 Processing as: {file_type}")
        
        # Use bulk operations for download
        local_xlsx_temp = os.path.join(CONFIG['LOCAL_XLSX_DIR'], remote_filename)
        operations = [('download', remote_xlsx_path, local_xlsx_temp)]
        download_results = bulk_remote_operations(operations)
        
        if not download_results or not download_results[0][-1]:
            logger.error(f"Failed to download: {remote_filename}")
            print(f"❌ ERROR: Download failed for {remote_filename}")
            return False
        
        print(f"📥 Downloaded to temp: {local_xlsx_temp}")
        
        visible_copy_path = os.path.join(CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR'], remote_filename)
        shutil.copy2(local_xlsx_temp, visible_copy_path)
        print(f"💾 Downloaded XLSX saved to: {visible_copy_path}")
        logger.info(f"Visible copy saved to: {visible_copy_path}")
        
        station_info = extract_xlsx_metadata(local_xlsx_temp)
        station_name = station_info.get('station_name', f"Project_{project_id}")
        station_code = station_info.get('station_code', 'DEFAULT_STATION')
        
        logger.info(f"Station info: {station_name} (code: {station_code})")
        print(f"🏢 Station: {station_name} (code: {station_code})")
        
        version = get_next_version(project_id, station_code)
        print(f"📊 Next version for {station_name}: v{version}")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_station_name = re.sub(r'[^\w\-_\. ]', '', station_name).replace(' ', '_')
        
        # Determine naming prefix based on file type
        if file_type == 'JUNCTION':
            # Extract junction number for naming
            junction_num = junction_info.get('junction_number', '1')
            new_xlsx_name = f"JUNCTION_{junction_num}_railway_project_{project_id}_{timestamp}_{safe_station_name}.xlsx"
            new_pdf_name = new_xlsx_name.replace('.xlsx', '.pdf')
        elif file_type == 'PREVIEW':
            new_xlsx_name = f"PREVIEW_railway_project_{project_id}_{timestamp}_{safe_station_name}.xlsx"
            new_pdf_name = new_xlsx_name.replace('.xlsx', '.pdf')
        else:  # FINAL
            new_xlsx_name = f"railway_project_{project_id}_{timestamp}_{safe_station_name}.xlsx"
            new_pdf_name = new_xlsx_name.replace('.xlsx', '.pdf')
        
        print(f"📝 New names: {new_xlsx_name}, {new_pdf_name}")
        
        local_xlsx_new = os.path.join(CONFIG['LOCAL_XLSX_DIR'], new_xlsx_name)
        local_pdf_new = os.path.join(CONFIG['LOCAL_PDF_DIR'], new_pdf_name)
        
        shutil.move(local_xlsx_temp, local_xlsx_new)
        print(f"🔄 Renamed to: {local_xlsx_new}")
        
        logger.info(f"Converting to PDF: {new_pdf_name}")
        print(f"🔄 Converting: {local_xlsx_new} -> {local_pdf_new}")
        converter_metadata, error, actual_pdf_path = convert_xlsx_to_pdf(local_xlsx_new, local_pdf_new)
        
        if error or not actual_pdf_path:
            logger.error(f"Conversion failed: {error}")
            print(f"❌ ERROR: Conversion failed - {error}")
            if os.path.exists(local_xlsx_new):
                os.remove(local_xlsx_new)
            process_success = False
        else:
            print(f"✅ Conversion successful: {actual_pdf_path}")
            if actual_pdf_path != local_pdf_new:
                shutil.move(actual_pdf_path, local_pdf_new)
                print(f"🔄 Renamed generated PDF to: {local_pdf_new}")
                actual_pdf_path = local_pdf_new
            
            checksum = _md5_of_file(local_pdf_new)
            if not checksum:
                logger.error("Failed to calculate PDF checksum")
                print("❌ ERROR: Checksum failed")
                process_success = False
            else:
                print(f"🔐 Checksum: {checksum}")
                local_uploads_xlsx = os.path.join(CONFIG['LOCAL_UPLOADS_DIR'], new_xlsx_name)
                local_uploads_pdf = os.path.join(CONFIG['LOCAL_UPLOADS_DIR'], new_pdf_name)
                copy_success = copy_to_local_uploads_no_remove(local_xlsx_new, local_pdf_new, local_uploads_xlsx, local_uploads_pdf)
                process_success = copy_success
                
                if process_success:
                    print(f"✅ Copied to local uploads: {CONFIG['LOCAL_UPLOADS_DIR']}")
                    logger.info(f"Local copy success: {new_xlsx_name}, {new_pdf_name}")
                    
                    # Prepare bulk upload operations
                    remote_xlsx_new = os.path.join(CONFIG['FTP_UPLOAD_DIR'], new_xlsx_name)
                    remote_pdf_new = os.path.join(CONFIG['FTP_UPLOAD_DIR'], new_pdf_name)
                    
                    upload_operations = [
                        ('upload', local_xlsx_new, remote_xlsx_new),
                        ('upload', local_pdf_new, remote_pdf_new)
                    ]
                    
                    logger.info("Starting bulk upload operations...")
                    print("📤 Starting bulk upload to remote...")
                    upload_results = bulk_remote_operations(upload_operations)
                    
                    upload_success = all(result[-1] for result in upload_results) if upload_results else False
                    
                    if upload_success:
                        print("✅ Bulk upload successful")
                        logger.info("Updating database...")
                        print("💾 Updating database...")
                        db_success = update_database_with_models(
                            project_id,
                            new_xlsx_name,
                            new_pdf_name,
                            checksum,
                            converter_metadata or {},
                            station_info,
                            version,
                            is_preview=(file_type == 'PREVIEW'),
                            is_junction=(file_type == 'JUNCTION')
                        )
                        
                        if db_success:
                            # Update project status based on file type
                            if file_type == 'JUNCTION':
                                status_updated = update_project_status(project_id, 'junction', 12, is_junction=True)
                                if status_updated:
                                    print(f"✅ Project status kept as 'junction' for JUNCTION file")
                            elif file_type == 'PREVIEW':
                                status_updated = update_project_status(project_id, 'preview', 10, is_preview=True)
                                if status_updated:
                                    print(f"✅ Project status kept as 'preview' for PREVIEW file")
                            else:  # FINAL
                                status_updated = update_project_status(project_id, 'waiting_for_level1', 11)
                                if status_updated:
                                    print(f"✅ Project status updated to 'waiting_for_level1' and stage to 11")
                                else:
                                    print(f"⚠️  Could not update project status")
                            
                            print(f"✅ Database updated successfully for v{version} of {station_name}")
                            
                            # Show appropriate message based on file type
                            if file_type == 'JUNCTION':
                                print(f"🔌 No notifications sent for JUNCTION PDF (special processing)")
                            elif file_type == 'PREVIEW':
                                print(f"📭 No notifications sent for PREVIEW PDF")
                            else:  # FINAL
                                print(f"📢 Notifications sent to level1 approvers WITH PROJECT ACCESS ONLY")
                            
                            logger.info("Importing XLSX data to database tables...")
                            print("💿 Importing data to DB...")
                            import_xlsx_data_to_database(local_xlsx_new, project_id)
                            print("✅ Data import complete")
                            
                            # For FINAL files only, manually reset project status
                            if file_type == 'FINAL':
                                with app.app_context():
                                    try:
                                        project = Project.query.get(project_id)
                                        if project:
                                            project.status = 'waiting_for_level1'
                                            project.stage = 11
                                            project.updated_date = get_ist_now()
                                            db.session.commit()
                                            print(f"🔒 MANUAL RESET: Project {project_id} locked to status='waiting_for_level1', stage=11")
                                            logger.info(f"Manually reset project {project_id} to waiting_for_level1 after data import")
                                    except Exception as e:
                                        logger.error(f"Error manually resetting project status: {e}")
                                        print(f"⚠️  Warning: Could not manually reset project status: {e}")
                            
                            remote_success = True
                        else:
                            print("❌ ERROR: Database update failed")
                            logger.error("Database update failed")
                            remote_success = False
                    else:
                        print("❌ ERROR: Bulk upload failed")
                        remote_success = False
                    
                    # Archive original file
                    if remote_success and CONFIG['ARCHIVE_REMOTE']:
                        remote_dest_path = os.path.join(CONFIG['REMOTE_ARCHIVE_DIR'], remote_filename)
                        logger.info(f"Archiving original to: {remote_dest_path}")
                        print(f"🗄️ Archiving remote to: {remote_dest_path}")
                        archive_operations = [('move', remote_xlsx_path, remote_dest_path)]
                        archive_results = bulk_remote_operations(operations)
                        
                        if archive_results and archive_results[0][-1]:
                            print(f"✅ Archived original to: {remote_dest_path}")
                        else:
                            logger.warning(f"Failed to archive {remote_filename} (but marked processed)")
                            print(f"⚠️ WARNING: Archive failed for {remote_filename}")
                    
                    # Mark as processed
                    save_processed_file(remote_filename)
                    print(f"✅ Marked as processed: {remote_filename}")
                    
                    # Cleanup temp files
                    if remote_success:
                        if os.path.exists(local_xlsx_new):
                            os.remove(local_xlsx_new)
                            print(f"🗑️ Removed temp XLSX: {local_xlsx_new}")
                        if os.path.exists(local_pdf_new):
                            os.remove(local_pdf_new)
                            print(f"🗑️ Removed temp PDF: {local_pdf_new}")
                        print("✅ Temp cleaned (full success)")
                    else:
                        print(f"⚠️ WARNING: Remote upload/DB failed. Temp files kept in {CONFIG['LOCAL_XLSX_DIR']} and {CONFIG['LOCAL_PDF_DIR']} for manual upload.")
                        print(f"  Manual upload to WinSCP path: {CONFIG['FTP_UPLOAD_DIR']}")
                    
                    logger.info(f"Processing complete for {remote_filename} (local: {process_success}, remote: {remote_success})")
                    print(f"Processing complete: {remote_filename} (local: {process_success}, remote: {remote_success})")
                    print("="*70)
                    print(f"PROCESSING COMPLETE: {remote_filename}")
                    print(f"Status: {'✅ SUCCESS' if remote_success else '❌ FAILED'}")
                    print(f"Type: {file_type}")
                    print("="*70)
                    return remote_success
        
        # If not process_success, move to failed
        remote_failed_path = os.path.join(CONFIG['REMOTE_FAILED_DIR'], remote_filename)
        logger.info(f"Moving failed file to: {remote_failed_path}")
        print(f"📦 Moving remote to failed: {remote_failed_path}")
        move_operations = [('move', remote_xlsx_path, remote_failed_path)]
        move_results = bulk_remote_operations(move_operations)
        
        if move_results and move_results[0][-1]:
            print(f"❌ Moved to failed: {remote_filename}")
            logger.info(f"Moved failed file to: {remote_failed_path}")
        else:
            logger.warning(f"Failed to move {remote_filename} to failed dir (may retry)")
            print(f"⚠️ WARNING: Failed move to {remote_filename}")
        
        if not CONFIG['KEEP_DOWNLOADED_COPIES']:
            if os.path.exists(local_xlsx_new):
                os.remove(local_xlsx_new)
                print(f"🗑️ Removed temp XLSX (failure): {local_xlsx_new}")
            if os.path.exists(local_pdf_new):
                os.remove(local_pdf_new)
                print(f"🗑️ Removed temp PDF (failure): {local_pdf_new}")
        
        logger.info(f"Processing complete for {remote_filename} (process_success: {process_success})")
        print(f"Processing complete: {remote_filename} (process_success: {process_success})")
        print("="*70)
        print(f"PROCESSING FAILED: {remote_filename}")
        print("="*70)
        return process_success
        
    except Exception as e:
        logger.error(f"Error processing {remote_filename}: {str(e)}")
        print(f"❌ EXCEPTION in processing {remote_filename}: {e}")
        import traceback
        traceback.print_exc()
        remote_failed_path = os.path.join(CONFIG['REMOTE_FAILED_DIR'], remote_filename)
        bulk_remote_operations([('move', remote_xlsx_path, remote_failed_path)])
        print(f"📦 Moved remote to failed due to exception: {remote_filename}")
        print("="*70)
        print(f"EXCEPTION OCCURRED: {remote_filename}")
        print("="*70)
        return False

def scan_and_process_remote():
    """Scan remote xlsx_download directory and process all XLSX files"""
    logger.info("Scanning remote directory for XLSX files...")
    print("🔍 Scanning remote for XLSX files...")
    
    try:
        remote_files = list_remote_files(CONFIG['FTP_XLSX_TAKE_DIR'])
        
        if not remote_files:
            logger.info("No files found in remote directory")
            print("📭 No files in remote directory")
            return []
        
        print(f"📁 Found {len(remote_files)} files in remote directory")
        
        xlsx_files = []
        for filename in remote_files:
            if filename.lower().endswith('.xlsx'):
                if all(sub not in filename.lower() for sub in ['processed', 'backup', 'failed']):
                    if not is_file_processed(filename):
                        project_id = get_project_id_from_filename(filename)
                        if project_id and check_project_status(project_id):
                            xlsx_files.append(filename)
                            with app.app_context():
                                project = Project.query.get(project_id)
                                status = project.status if project and hasattr(project, 'status') else 'unknown'
                            print(f"✅ {filename} - Project {project_id} is '{status}' (will process)")
                        elif project_id:
                            with app.app_context():
                                project = Project.query.get(project_id)
                                status = project.status if project and hasattr(project, 'status') else 'unknown'
                            print(f"⏸️ {filename} - Project {project_id} is '{status}' (skipping)")
                        else:
                            print(f"❌ {filename} - Could not extract project ID (skipping)")
        
        logger.info(f"Found {len(xlsx_files)} new XLSX file(s) to process (ready_for_pdf OR preview OR junction)")
        print(f"📊 New XLSX to process (ready_for_pdf OR preview OR junction): {len(xlsx_files)}")
        
        results = []
        for filename in xlsx_files:
            result = {
                'filename': filename,
                'success': False,
                'message': ''
            }
            
            try:
                success = process_remote_xlsx_file_enhanced(filename)
                result['success'] = success
                result['message'] = 'Processed successfully' if success else 'Processing failed'
            except Exception as e:
                result['message'] = f"Error: {str(e)}"
                print(f"❌ Scan error for {filename}: {e}")
            
            results.append(result)
            
            time.sleep(2)
        
        return results
        
    except Exception as e:
        logger.error(f"Error scanning remote directory: {e}")
        print(f"❌ Scan error: {e}")
        return []

def setup_local_directories():
    """Create local directories if they don't exist"""
    dirs = [
        CONFIG['LOCAL_TEMP_DIR'],
        CONFIG['LOCAL_XLSX_DIR'],
        CONFIG['LOCAL_PDF_DIR'],
        CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR'],
        CONFIG['LOCAL_UPLOADS_DIR'],
    ]
    
    for directory in dirs:
        os.makedirs(directory, exist_ok=True)
        logger.info(f"Ensured directory exists: {directory}")
        print(f"Created/Ensured dir: {directory}")
    
    readme_path = os.path.join(CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR'], 'README.txt')
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write("This folder contains XLSX files downloaded from the FTP/SFTP server.\n")
        f.write(f"Files are automatically processed and converted to PDF.\n")
        f.write(f"Script location: {os.path.abspath(__file__)}\n")
        f.write(f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    print(f"Created README in downloads: {readme_path}")

    uploads_readme_path = os.path.join(CONFIG['LOCAL_UPLOADS_DIR'], 'README.txt')
    with open(uploads_readme_path, 'w', encoding='utf-8') as f:
        f.write("This folder contains processed XLSX and PDF files after conversion.\n")
        f.write(f"Uploaded to remote: {CONFIG['FTP_UPLOAD_DIR']}\n")
        f.write(f"File types:\n")
        f.write(f"  - PREVIEW files are prefixed with 'PREVIEW_'\n")
        f.write(f"  - JUNCTION files are prefixed with 'JUNCTION_<number>_'\n")
        f.write(f"  - FINAL files have no special prefix\n")
        f.write(f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    print(f"Created README in uploads: {uploads_readme_path}")

def discover_remote_paths():
    """Discover available directories on startup to help debug paths"""
    ssh, sftp = get_sftp_connection()
    if not sftp:
        logger.error("Cannot discover paths: SFTP connection failed")
        return
    
    try:
        root_files = sftp.listdir('/')
        logger.info(f"Root (/) contents: {root_files}")
        print(f"DEBUG: Root (/) contents: {root_files}")
        
        common_paths = ['/srv', '/home', '/var', '/opt', '/usr', '/root', '/tmp']
        for path in common_paths:
            try:
                contents = sftp.listdir(path)
                logger.info(f"{path} exists, contents: {contents[:5]}...")
                print(f"DEBUG: {path} exists (sample: {contents[:3]})")
                
                if path == '/srv' and 'railway' in str(contents):
                    try:
                        railway = sftp.listdir('/srv/railway')
                        logger.info(f"/srv/railway contents: {railway}")
                        print(f"DEBUG: /srv/railway exists (sample: {railway[:3]})")
                    except:
                        pass
                        
            except Exception as e:
                logger.debug(f"{path} does not exist or inaccessible: {e}")
        
        stdin, stdout, stderr = ssh.exec_command('pwd')
        home = stdout.read().decode().strip()
        logger.info(f"User home directory: {home}")
        print(f"DEBUG: Your home: {home}")
        
        suggested_xlsx = f"{home}/xlsx_download" if home else CONFIG['FTP_XLSX_TAKE_DIR']
        print(f"SUGGESTION: If XLSX files are in a subfolder of {home}, update CONFIG['FTP_XLSX_TAKE_DIR'] to '{suggested_xlsx}'")
        
    except Exception as e:
        logger.error(f"Path discovery error: {e}")
    finally:
        sftp.close()
        ssh.close()

# ==================== MAIN MONITORING FUNCTION ====================
def monitor_xlsx_download_folder(flask_app=None):
    """Continuous monitoring function for XLSX download folder"""
    if flask_app:
        global app
        app = flask_app
    
    load_processed_files()
    setup_local_directories()
    
    print("\n" + "="*80)
    print("🚂 ENHANCED AUTOMATED PDF CONVERTER - 24/7 OPERATION")
    print("="*80)
    print("📋 ENHANCEMENTS:")
    print("  1. Connection pooling with keepalive")
    print("  2. Automatic retry logic (max 3 retries)")
    print("  3. Bulk operations to minimize connections")
    print("  4. Connection verification before reuse")
    print("  5. Processes 'ready_for_pdf', 'preview', AND 'junction' status files")
    print("  6. JUNCTION file support with special naming and processing")
    print("="*80)
    print("\n📂 DOWNLOAD LOCATIONS:")
    print("-"*40)
    print(f"📁 Visible downloads folder: {CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR']}")
    print(f"📁 Temporary processing folder: {CONFIG['LOCAL_XLSX_DIR']}")
    print(f"📁 Local uploads folder: {CONFIG['LOCAL_UPLOADS_DIR']}")
    print(f"📋 Log file: {CONFIG['LOG_FILE']}")
    print(f"📋 Processed files log: {CONFIG['PROCESSED_FILES_LOG']}")
    print("="*80)
    
    print("\n🔍 DEBUG: Discovering remote paths...")
    discover_remote_paths()
    
    logger.info("Starting enhanced continuous remote monitoring...")
    logger.info(f"Monitoring: {CONFIG['FTP_XLSX_TAKE_DIR']}")
    logger.info(f"Uploads to: {CONFIG['FTP_UPLOAD_DIR']}")
    logger.info(f"Check interval: {CONFIG['CHECK_INTERVAL']} seconds")
    print("\n🚀 Starting 24/7 monitoring with enhanced connection handling...")
    
    print("\n" + "="*80)
    print("🔍 Performing initial scan for existing XLSX files...")
    print("="*80)
    initial_results = scan_and_process_remote()
    if initial_results:
        success_count = sum(1 for r in initial_results if r['success'])
        print(f"\n📊 Initial scan complete: {success_count}/{len(initial_results)} processed successfully")
        for result in initial_results:
            status = "✅" if result['success'] else "❌"
            print(f"  {status} {result['filename']}: {result['message']}")
    else:
        print("📭 Initial scan: No new files found or path issue (check debug above)")
    
    processed_count = len(PROCESSED_FILES_CACHE)
    print(f"\n📊 Total processed files in history: {processed_count}")
    print(f"🔌 Connection pool size: {len(SFTP_CONNECTION_POOL)}")
    
    print("\n" + "="*80)
    print("🚦 ENHANCED CONTINUOUS MONITORING STARTED (24/7)")
    print("="*80)
    print("📋 Monitoring for new XLSX files where project status = 'ready_for_pdf' OR 'preview' OR 'junction'...")
    print("📋 File types and prefixes:")
    print("   • PREVIEW files: 'PREVIEW_' prefix")
    print("   • JUNCTION files: 'JUNCTION_<number>_' prefix")
    print("   • FINAL files: No special prefix")
    print("📋 Notifications:")
    print("   • FINAL files: Send notifications to level1 approvers")
    print("   • PREVIEW files: No notifications")
    print("   • JUNCTION files: No notifications (special processing)")
    print("📋 Connection pooling enabled - connections will be reused")
    print("📋 Press Ctrl+C to stop.")
    print("="*80 + "\n")
    
    try:
        while True:
            current_files = list_remote_files(CONFIG['FTP_XLSX_TAKE_DIR'])
            
            if current_files:
                current_xlsx = []
                for filename in current_files:
                    if (filename.lower().endswith('.xlsx') and
                        all(sub not in filename.lower() for sub in ['processed', 'backup', 'failed']) and
                        not is_file_processed(filename)):
                        
                        project_id = get_project_id_from_filename(filename)
                        if project_id and check_project_status(project_id):
                            current_xlsx.append(filename)
                            with app.app_context():
                                project = Project.query.get(project_id)
                                status = project.status if project and hasattr(project, 'status') else 'unknown'
                            print(f"✅ {filename} - Project {project_id} is '{status}' (will process)")
                        elif project_id:
                            with app.app_context():
                                project = Project.query.get(project_id)
                                status = project.status if project and hasattr(project, 'status') else 'unknown'
                            print(f"⏸️ {filename} - Project {project_id} is '{status}' (skipping)")
                        else:
                            print(f"❌ {filename} - Could not extract project ID (skipping)")
                
                if current_xlsx:
                    logger.info(f"Found {len(current_xlsx)} new file(s) with ready_for_pdf, preview, or junction status")
                    print(f"\n📊 New files detected (ready_for_pdf OR preview OR junction): {len(current_xlsx)}")
                    print(f"🔌 Active connections in pool: {len(SFTP_CONNECTION_POOL)}")
                    
                    for filename in current_xlsx:
                        logger.info(f"Processing new file: {filename}")
                        print(f"\n" + "="*60)
                        print(f"🔄 Processing new file: {filename}")
                        print(f"🔌 Using connection pool: {len(SFTP_CONNECTION_POOL)} connections")
                        print("="*60)
                        
                        success = process_remote_xlsx_file_enhanced(filename)
                        
                        if success:
                            logger.info(f"✅ Successfully processed: {filename}")
                            print(f"✅ Successfully processed: {filename}")
                            print(f"📤 Uploaded to: {CONFIG['FTP_UPLOAD_DIR']}")
                            print(f"💾 Database updated with new version")
                            
                            # Get project status to show appropriate message
                            project_id = get_project_id_from_filename(filename)
                            with app.app_context():
                                project = Project.query.get(project_id)
                                status = project.status if project and hasattr(project, 'status') else 'unknown'
                            
                            # Determine file type for message
                            junction_info = get_junction_info_from_filename(filename)
                            if junction_info['is_junction']:
                                file_type = 'JUNCTION'
                            elif status == 'preview':
                                file_type = 'PREVIEW'
                            else:
                                file_type = 'FINAL'
                            
                            if file_type == 'JUNCTION':
                                print(f"🔄 Project status kept as 'junction' for JUNCTION file")
                                print(f"🔌 No notifications sent for JUNCTION (special processing)")
                            elif file_type == 'PREVIEW':
                                print(f"🔄 Project status kept as 'preview' for PREVIEW file")
                                print(f"📭 No notifications sent for PREVIEW")
                            else:
                                print(f"🔄 Project status updated to 'waiting_for_level1' and stage to 11")
                                print(f"📢 Notifications sent to level1 approvers WITH PROJECT ACCESS ONLY")
                        else:
                            logger.error(f"❌ Failed to process: {filename}")
                            print(f"❌ Failed to process: {filename}")
                        print("="*60)
                        
                        print(f"🔌 Connection pool after processing: {len(SFTP_CONNECTION_POOL)} connections")
            
            # Clean up connection pool periodically
            cleanup_connection_pool()
            
            print(f"\n⏰ Sleeping for {CONFIG['CHECK_INTERVAL']} seconds...")
            time.sleep(CONFIG['CHECK_INTERVAL'])
            
    except KeyboardInterrupt:
        logger.info("Monitoring stopped by user (Ctrl+C)")
        print("\n" + "="*80)
        print("🛑 MONITORING STOPPED BY USER")
        print("="*80)
        print(f"📊 Total files processed in this session: {len(PROCESSED_FILES_CACHE)}")
        print(f"🔌 Closing {len(SFTP_CONNECTION_POOL)} connections in pool...")
        
        for connection_key, (ssh, sftp, _) in list(SFTP_CONNECTION_POOL.items()):
            try:
                sftp.close()
                ssh.close()
                logger.debug(f"Closed connection: {connection_key}")
            except:
                pass
        
        SFTP_CONNECTION_POOL.clear()
        
        print(f"📁 Check downloaded files in: {CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR']}")
        print(f"📁 Check processed files in: {CONFIG['LOCAL_UPLOADS_DIR']}")
        print(f"📋 Check logs in: {CONFIG['LOG_FILE']}")
        print("="*80)
    except Exception as e:
        logger.error(f"Monitoring error: {e}")
        print(f"\n❌ Monitoring error: {e}")
        print(f"🔄 Auto-restarting in 60 seconds...")
        time.sleep(60)
        print("🚀 Auto-restarting...")
        monitor_xlsx_download_folder()

# ==================== MAIN EXECUTION ====================
if __name__ == "__main__":
    print("\n" + "="*100)
    print("🚂 ENHANCED FTP/SFTP AUTOMATED CONVERTER (24/7 MODE)")
    print("="*100)
    print("📋 ENHANCEMENTS:")
    print("  • Connection pooling with keepalive (reduces connection drops)")
    print("  • Automatic retry logic (3 retries with exponential backoff)")
    print("  • Bulk operations to minimize connection overhead")
    print("  • Processes projects with status = 'ready_for_pdf' OR 'preview' OR 'junction'")
    print("  • JUNCTION file support with special naming and processing")
    print("  • Preview files are marked with 'PREVIEW_' prefix")
    print("  • Junction files are marked with 'JUNCTION_<number>_' prefix")
    print("  • Only FINAL files trigger notifications")
    print("  • Auto-restarts on errors for 24/7 operation")
    print("="*100)
    print(f"🌐 Host: {CONFIG['FTP_HOST']}")
    print(f"🔧 Mode: {'SFTP' if CONFIG['FTP_USE_SFTP'] else 'FTP'}")
    print(f"📥 XLSX Source: {CONFIG['FTP_XLSX_TAKE_DIR']}")
    print(f"📤 PDF Destination: {CONFIG['FTP_UPLOAD_DIR']}")
    print(f"💾 Local Uploads: {CONFIG['LOCAL_UPLOADS_DIR']}")
    print(f"🗃️ Database: PostgreSQL")
    print(f"⏰ Check Interval: {CONFIG['CHECK_INTERVAL']}s")
    print(f"🔌 Connection Pool Timeout: {CONFIG['CONNECTION_POOL_TIMEOUT']}s")
    print("="*100)
    
    restart_count = 0
    max_restarts = 10
    
    while restart_count < max_restarts:
        try:
            restart_count += 1
            if restart_count > 1:
                print(f"\n🔄 RESTART ATTEMPT {restart_count}/{max_restarts}")
                print("="*80)
            
            monitor_xlsx_download_folder()
            
        except Exception as e:
            print(f"\n💥 CRITICAL ERROR in main loop: {e}")
            print("🔄 Restarting in 30 seconds...")
            time.sleep(30)
            if restart_count >= max_restarts:
                print(f"\n❌ MAXIMUM RESTARTS REACHED ({max_restarts})")
                print("🛑 SHUTTING DOWN...")
                break
    
    print("\n" + "="*100)
    print("🛑 SCRIPT TERMINATED")
    print("="*100)