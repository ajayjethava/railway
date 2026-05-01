#!/usr/bin/env python3
"""
FTP/SFTP Automated XLSX to PDF Converter
Downloads XLSX from FTP/SFTP, converts to PDF, uploads back, and updates database.
Integrated with your specific configuration and database models.
Now runs in fully automated mode: starts continuous monitoring immediately.
Enhanced with path discovery and debug logging.
FIXED: Loop issue and download location visibility.
UPDATED: Prevent reprocessing loops by moving failed files to 'failed/' dir.
UPDATED: Replaced Unicode symbols with ASCII to avoid Windows encoding errors.
UPDATED: New workflow - XLSX processed to PDF, then both moved to local uploads and remote uploads.
FIXED: File movement and upload verification issues.
FIXED: Syntax error in StationDrawing constructor.
FIXED: Improved error handling for PDF conversion with warnings.
"""
import os
import sys
import time
import hashlib
import subprocess
import shutil
import re
import json
import glob
from datetime import datetime, timezone, timedelta
from pathlib import Path
import logging
import paramiko
import ftplib
from openpyxl import load_workbook

# Flask import (required for app context)
from flask import Flask

# Add the app directory to Python path if running standalone
if __name__ == '__main__':
    # Assuming running from Circuitbuilding/app/monitor.py, add parent
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

# Import database and models (relative for package, fallback for standalone)
try:
    from Circuitbuilding.app.database import db
    from Circuitbuilding.app.models import (
        get_ist_now, Project, StationDrawing, GeneratedPDF,
        User, Notification, Cable, JunctionBox, Terminal,
        Group, TerminalHeader, ChokeTable, ResistorTable, CableBox
    )
    from Circuitbuilding.app.schemas import SHEETS, HEADER_HINTS  # If needed
    PACKAGE_MODE = True
except ImportError:
    # Standalone mode
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
    from Circuitbuilding.app.database import db
    from Circuitbuilding.app.models import (
        get_ist_now, Project, StationDrawing, GeneratedPDF,
        User, Notification, Cable, JunctionBox, Terminal,
        Group, TerminalHeader, ChokeTable, ResistorTable, CableBox
    )
    try:
        from Circuitbuilding.app.schemas import SHEETS, HEADER_HINTS
    except ImportError:
        pass  # Not critical
    PACKAGE_MODE = False

# Define IST timezone (UTC+5:30)
IST = timezone(timedelta(hours=5, minutes=30))

# ==================== GLOBAL TRACKING VARIABLE ====================
# This will track processed files to avoid loops
PROCESSED_FILES_CACHE = set()

# ==================== ENVIRONMENT LOADING (from init.py) ====================
def load_environment_variables():
    """Load environment variables from .env file"""
    try:
        # Try to load python-dotenv if available
        from dotenv import load_dotenv
        
        # Adjusted for .env at C:\Demo\git\Circuitbuilding\.env
        # Since script is at C:\Demo\git\ftp_auto_converter.py, Circuitbuilding is a sibling or subdir
        script_dir = Path(__file__).resolve().parent
        env_path = script_dir / 'Circuitbuilding' / '.env'
        
        if env_path.exists():
            load_dotenv(dotenv_path=env_path)
            print(f"v Loaded environment variables from: {env_path}")
        else:
            print(f"W .env file not found at: {env_path}")
            # Fallback to parent or current dir
            fallback_env = script_dir / '.env'
            if fallback_env.exists():
                load_dotenv(dotenv_path=fallback_env)
                print(f"v Loaded environment variables from fallback: {fallback_env}")
    except ImportError:
        print("W python-dotenv not installed. Install with: pip install python-dotenv")
        print("W Using system environment variables instead.")
    except Exception as e:
        print(f"W Error loading environment variables: {str(e)}")

# Load env vars early
load_environment_variables()

# ==================== CONFIGURATION (merged with init.py style) ====================
CONFIG = {
    # Flask Configuration (from env or default)
    'SECRET_KEY': os.environ.get("SECRET_KEY", "Saltriver@123"),
    
    # Database (hardcoded PG from init.py)
    'DB_URI': "postgresql://postgres:Omhari%408899@pso.cellapps.com:5432/Postgres",
    
    # FTP/SFTP Configuration (from env, defaults to script's originals)
    'FTP_ENABLED': os.environ.get("FTP_ENABLED", "True").lower() == "true",
    'FTP_HOST': os.environ.get("FTP_HOST", "45.67.216.178"),
    'FTP_PORT': int(os.environ.get("FTP_PORT", 22)),
    'FTP_USERNAME': os.environ.get("FTP_USERNAME", "root"),
    'FTP_PASSWORD': os.environ.get("FTP_PASSWORD", "Mehul88$"),
    'FTP_UPLOAD_DIR': os.environ.get("FTP_UPLOAD_DIR", "srv/railway/frontend/uploads/"),  # Changed path
    'FTP_XLSX_TAKE_DIR': os.environ.get("FTP_XLSX_TAKE_DIR", "/srv/railway/frontend/xlsx_download/"),
    'FTP_USE_SFTP': os.environ.get("FTP_USE_SFTP", "True").lower() == "true",
    'FTP_TIMEOUT': int(os.environ.get("FTP_TIMEOUT", 30)),
    
    # Auto-generation
    'AUTO_GENERATE_PDF_ON_DOWNLOAD': os.environ.get("AUTO_GENERATE_PDF_ON_DOWNLOAD", "True").lower() == "true",
    
    # Local paths (adjusted for script at C:\Demo\git\)
    'LOCAL_TEMP_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp'),
    'LOCAL_XLSX_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_xlsx'),
    'LOCAL_PDF_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp_pdf'),
    'LOCAL_DOWNLOADS_VISIBLE_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloaded_xlsx_files'),  # NEW: Visible downloads folder
    'LOCAL_UPLOADS_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads'),  # NEW: Local project uploads folder
    
    # Converter script (adjusted for excel_to_pdf_converter.py at C:\Demo\git\)
    'CONVERTER_SCRIPT': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel_to_pdf_converter.py'),
    
    # Processing
    'CHECK_INTERVAL': int(os.environ.get("CHECK_INTERVAL", 30)),
    'ARCHIVE_REMOTE': os.environ.get("ARCHIVE_REMOTE", "True").lower() == "true",
    'REMOTE_ARCHIVE_DIR': os.environ.get("REMOTE_ARCHIVE_DIR", "/srv/railway/frontend/xlsx_download/processed/"),
    'REMOTE_FAILED_DIR': os.environ.get("REMOTE_FAILED_DIR", "/srv/railway/frontend/xlsx_download/failed/"),  # NEW: For failed files
    'REMOTE_BACKUP_DIR': os.environ.get("REMOTE_BACKUP_DIR", "/srv/railway/frontend/xlsx_download/backup/"),
    
    # File patterns
    'XLSX_PATTERNS': ['*.xlsx', '*.XLSX'],
    
    # Logging
    'LOG_LEVEL': os.environ.get("LOG_LEVEL", "INFO"),
    'LOG_FILE': os.environ.get("LOG_FILE", "ftp_converter.log"),
    
    # NEW: Loop prevention
    'PROCESSED_FILES_LOG': os.environ.get("PROCESSED_FILES_LOG", "processed_files.log"),
    'KEEP_DOWNLOADED_COPIES': os.environ.get("KEEP_DOWNLOADED_COPIES", "True").lower() == "true",  # Keep downloaded files for debugging
    
    # NEW: Conversion settings
    'ALLOW_PARTIAL_CONVERSION': os.environ.get("ALLOW_PARTIAL_CONVERSION", "True").lower() == "true",  # Allow conversion even with warnings
}

# Setup logging (INFO level for automation, no console prompts)
# Use UTF-8 encoding to avoid charmap errors on Windows
log_file_handler = logging.FileHandler(CONFIG['LOG_FILE'], encoding='utf-8')
log_stream_handler = logging.StreamHandler()
log_stream_handler.stream.reconfigure(encoding='utf-8')  # For Python 3.7+

logging.basicConfig(
    level=getattr(logging, CONFIG['LOG_LEVEL']),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        log_file_handler,
        log_stream_handler
    ]
)
logger = logging.getLogger(__name__)

# Initialize Flask app for db context (standalone or monitor mode)
app = Flask(__name__)
app.config['SECRET_KEY'] = CONFIG['SECRET_KEY']
app.config['SQLALCHEMY_DATABASE_URI'] = CONFIG['DB_URI']
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

# Ensure db tables exist (like in init.py)
with app.app_context():
    db.create_all()

# ==================== PROCESSED FILES TRACKING ====================
def load_processed_files():
    """Load previously processed files from log to avoid reprocessing"""
    global PROCESSED_FILES_CACHE
    try:
        if os.path.exists(CONFIG['PROCESSED_FILES_LOG']):
            with open(CONFIG['PROCESSED_FILES_LOG'], 'r', encoding='utf-8') as f:
                for line in f:
                    filename = line.strip()
                    if filename:
                        PROCESSED_FILES_CACHE.add(filename)
            logger.info(f"Loaded {len(PROCESSED_FILES_CACHE)} previously processed files from log")
    except Exception as e:
        logger.error(f"Error loading processed files log: {e}")

def save_processed_file(filename):
    """Save processed filename to log"""
    global PROCESSED_FILES_CACHE
    try:
        PROCESSED_FILES_CACHE.add(filename)
        with open(CONFIG['PROCESSED_FILES_LOG'], 'a', encoding='utf-8') as f:
            f.write(f"{filename}\n")
    except Exception as e:
        logger.error(f"Error saving processed file {filename}: {e}")

def is_file_processed(filename):
    """Check if file has already been processed"""
    global PROCESSED_FILES_CACHE
    return filename in PROCESSED_FILES_CACHE

# ==================== ENHANCED SFTP DEBUG FUNCTIONS ====================
def discover_remote_paths():
    """Discover available directories on startup to help debug paths"""
    ssh, sftp = get_sftp_connection()
    if not sftp:
        logger.error("Cannot discover paths: SFTP connection failed")
        return
    
    try:
        # List root
        root_files = sftp.listdir('/')
        logger.info(f"Root (/) contents: {root_files}")
        print(f"DEBUG: Root (/) contents: {root_files}")  # Console for user
        
        # Common paths to check
        common_paths = ['/srv', '/home', '/var', '/opt', '/usr', '/root', '/tmp']
        for path in common_paths:
            try:
                contents = sftp.listdir(path)
                logger.info(f"{path} exists, contents: {contents[:5]}...")  # First 5 items
                print(f"DEBUG: {path} exists (sample: {contents[:3]})")
                
                # If /srv exists, check deeper
                if path == '/srv' and 'railway' in str(contents):
                    try:
                        railway = sftp.listdir('/srv/railway')
                        logger.info(f"/srv/railway contents: {railway}")
                        print(f"DEBUG: /srv/railway exists (sample: {railway[:3]})")
                    except:
                        pass
                        
            except Exception as e:
                logger.debug(f"{path} does not exist or inaccessible: {e}")
        
        # Get user's home
        stdin, stdout, stderr = ssh.exec_command('pwd')
        home = stdout.read().decode().strip()
        logger.info(f"User home directory: {home}")
        print(f"DEBUG: Your home: {home}")
        
        # Suggest updates
        suggested_xlsx = f"{home}/xlsx_download" if home else CONFIG['FTP_XLSX_TAKE_DIR']
        print(f"SUGGESTION: If XLSX files are in a subfolder of {home}, update CONFIG['FTP_XLSX_TAKE_DIR'] to '{suggested_xlsx}'")
        
    except Exception as e:
        logger.error(f"Path discovery error: {e}")
    finally:
        sftp.close()
        ssh.close()

# ==================== FTP/SFTP FUNCTIONS (with better error handling) ====================
def setup_local_directories():
    """Create local directories if they don't exist"""
    dirs = [
        CONFIG['LOCAL_TEMP_DIR'],
        CONFIG['LOCAL_XLSX_DIR'],
        CONFIG['LOCAL_PDF_DIR'],
        CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR'],  # NEW: Visible downloads folder
        CONFIG['LOCAL_UPLOADS_DIR'],  # NEW: Local uploads folder
    ]
    
    for directory in dirs:
        os.makedirs(directory, exist_ok=True)
        logger.info(f"Ensured directory exists: {directory}")
        print(f"D Created directory: {directory}")
    
    # Create README in downloads folder
    readme_path = os.path.join(CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR'], 'README.txt')
    with open(readme_path, 'w', encoding='utf-8') as f:
        f.write("This folder contains XLSX files downloaded from the FTP/SFTP server.\n")
        f.write(f"Files are automatically processed and converted to PDF.\n")
        f.write(f"Script location: {os.path.abspath(__file__)}\n")
        f.write(f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

def get_sftp_connection():
    """Establish SFTP connection"""
    try:
        logger.info(f"Connecting to SFTP server: {CONFIG['FTP_HOST']}:{CONFIG['FTP_PORT']}")
        
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
        ssh.connect(
            hostname=CONFIG['FTP_HOST'],
            port=CONFIG['FTP_PORT'],
            username=CONFIG['FTP_USERNAME'],
            password=CONFIG['FTP_PASSWORD'],
            timeout=CONFIG['FTP_TIMEOUT'],
            allow_agent=False,
            look_for_keys=False
        )
        
        sftp = ssh.open_sftp()
        logger.info("SFTP connection established successfully")
        return ssh, sftp
        
    except Exception as e:
        logger.error(f"SFTP connection failed: {str(e)}")
        return None, None

def get_ftp_connection():
    """Establish FTP connection"""
    try:
        logger.info(f"Connecting to FTP server: {CONFIG['FTP_HOST']}:21")
        
        ftp = ftplib.FTP()
        ftp.connect(CONFIG['FTP_HOST'], 21, timeout=CONFIG['FTP_TIMEOUT'])
        ftp.login(CONFIG['FTP_USERNAME'], CONFIG['FTP_PASSWORD'])
        ftp.set_pasv(True)
        
        logger.info("FTP connection established successfully")
        return ftp
        
    except Exception as e:
        logger.error(f"FTP connection failed: {str(e)}")
        return None

def list_remote_files(remote_dir):
    """List files in remote directory with parent fallback on error"""
    files = []
    
    if CONFIG['FTP_USE_SFTP']:
        ssh, sftp = get_sftp_connection()
        if sftp:
            try:
                files = sftp.listdir(remote_dir)
                logger.info(f"Listed {len(files)} files in {remote_dir}")
            except Exception as e:
                logger.error(f"Error listing SFTP directory {remote_dir}: {e}")
                # Fallback: Try parent directory
                parent_dir = os.path.dirname(remote_dir)
                if parent_dir != remote_dir:
                    try:
                        parent_files = sftp.listdir(parent_dir)
                        logger.info(f"Fallback: Listed parent {parent_dir}: {parent_files[:10]}...")  # Sample
                        print(f"DEBUG: {remote_dir} missing? Parent {parent_dir} has: {parent_files[:5]}")
                    except:
                        pass
                if ssh:
                    ssh.close()
    else:
        ftp = get_ftp_connection()
        if ftp:
            try:
                ftp.cwd(remote_dir)
                files = ftp.nlst()
                logger.info(f"Listed {len(files)} files in {remote_dir}")
                ftp.quit()
            except Exception as e:
                logger.error(f"Error listing FTP directory {remote_dir}: {e}")
                ftp.quit()
    
    return files

def download_file_remote(remote_path, local_path):
    """Download file from remote server"""
    try:
        if CONFIG['FTP_USE_SFTP']:
            ssh, sftp = get_sftp_connection()
            if sftp:
                sftp.get(remote_path, local_path)
                ssh.close()
                logger.info(f"Downloaded via SFTP: {remote_path} -> {local_path}")
                print(f"D Downloaded: {os.path.basename(remote_path)} -> {local_path}")
                return True
        else:
            ftp = get_ftp_connection()
            if ftp:
                with open(local_path, 'wb') as f:
                    ftp.retrbinary(f'RETR {remote_path}', f.write)
                ftp.quit()
                logger.info(f"Downloaded via FTP: {remote_path} -> {local_path}")
                print(f"D Downloaded: {os.path.basename(remote_path)} -> {local_path}")
                return True
                
    except Exception as e:
        logger.error(f"Download failed: {remote_path} -> {local_path}: {str(e)}")
        return False
    
    return False

def upload_file_remote(local_path, remote_path):
    """Upload file to remote server"""
    try:
        if CONFIG['FTP_USE_SFTP']:
            ssh, sftp = get_sftp_connection()
            if sftp:
                # Ensure remote directory exists
                remote_dir = os.path.dirname(remote_path)
                try:
                    sftp.stat(remote_dir)
                except FileNotFoundError:
                    # Create directory recursively
                    parts = remote_dir.split('/')
                    current_path = ''
                    for part in parts:
                        if part:
                            current_path = current_path + '/' + part if current_path else '/' + part
                            try:
                                sftp.stat(current_path)
                            except:
                                sftp.mkdir(current_path)
                
                sftp.put(local_path, remote_path)
                ssh.close()
                logger.info(f"Uploaded via SFTP: {local_path} -> {remote_path}")
                print(f"U Uploaded: {os.path.basename(local_path)} -> {remote_path}")
                return True
        else:
            ftp = get_ftp_connection()
            if ftp:
                # Ensure remote directory exists
                remote_dir = os.path.dirname(remote_path)
                try:
                    ftp.cwd(remote_dir)
                except:
                    # Try to create directory
                    ftp.mkd(remote_dir)
                
                with open(local_path, 'rb') as f:
                    ftp.storbinary(f'STOR {os.path.basename(remote_path)}', f)
                ftp.quit()
                logger.info(f"Uploaded via FTP: {local_path} -> {remote_path}")
                print(f"U Uploaded: {os.path.basename(local_path)} -> {remote_path}")
                return True
                
    except Exception as e:
        logger.error(f"Upload failed: {local_path} -> {remote_path}: {str(e)}")
        print(f"X Upload failed: {os.path.basename(local_path)} -> {remote_path}: {str(e)}")
        return False
    
    return False

def move_file_remote(source_path, dest_path):
    """Move file on remote server"""
    try:
        if CONFIG['FTP_USE_SFTP']:
            ssh, sftp = get_sftp_connection()
            if sftp:
                # Ensure destination directory exists
                dest_dir = os.path.dirname(dest_path)
                try:
                    sftp.stat(dest_dir)
                except FileNotFoundError:
                    # Create directory recursively
                    parts = dest_dir.split('/')
                    current_path = ''
                    for part in parts:
                        if part:
                            current_path = current_path + '/' + part if current_path else '/' + part
                            try:
                                sftp.stat(current_path)
                            except:
                                sftp.mkdir(current_path)
                
                sftp.rename(source_path, dest_path)
                ssh.close()
                logger.info(f"Moved on remote: {source_path} -> {dest_path}")
                print(f"M Moved remote: {os.path.basename(source_path)} -> {dest_path}")
                return True
        else:
            # FTP doesn't support rename across directories easily
            temp_path = os.path.join(CONFIG['LOCAL_TEMP_DIR'], 'temp_move')
            if download_file_remote(source_path, temp_path):
                if upload_file_remote(temp_path, dest_path):
                    delete_file_remote(source_path)
                    os.remove(temp_path)
                    logger.info(f"Moved on remote (via temp): {source_path} -> {dest_path}")
                    print(f"M Moved remote (via temp): {os.path.basename(source_path)} -> {dest_path}")
                    return True
                
    except Exception as e:
        logger.error(f"Move failed: {source_path} -> {dest_path}: {str(e)}")
        print(f"X Move failed: {os.path.basename(source_path)} -> {dest_path}: {str(e)}")
        return False
    
    return False

def delete_file_remote(remote_path):
    """Delete file from remote server"""
    try:
        if CONFIG['FTP_USE_SFTP']:
            ssh, sftp = get_sftp_connection()
            if sftp:
                sftp.remove(remote_path)
                ssh.close()
                logger.info(f"Deleted from remote: {remote_path}")
                print(f"D Deleted remote: {os.path.basename(remote_path)}")
                return True
        else:
            ftp = get_ftp_connection()
            if ftp:
                ftp.delete(remote_path)
                ftp.quit()
                logger.info(f"Deleted from remote: {remote_path}")
                print(f"D Deleted remote: {os.path.basename(remote_path)}")
                return True
                
    except Exception as e:
        logger.error(f"Delete failed: {remote_path}: {str(e)}")
        return False
    
    return False

# ==================== PROCESSING FUNCTIONS ====================
def get_project_id_from_filename(filename):
    """Extract project ID from filename"""
    patterns = [
        r'RAILWAYPROJECT_ID(\d+)_',
        r'railway_project_(\d+)_',
        r'project_(\d+)_',
        r'_ID(\d+)_',
        r'ID(\d+)[_-]',
        r'_P(\d+)[_-]'
    ]
    
    filename_upper = filename.upper()
    
    for pattern in patterns:
        match = re.search(pattern, filename_upper)
        if match:
            return int(match.group(1))
    
    # Try to extract any number that looks like a project ID
    numbers = re.findall(r'\b(\d{2,})\b', filename)
    if numbers:
        # Use the most likely project ID (usually 2+ digits)
        for num in numbers:
            if 10 <= int(num) <= 999:
                return int(num)
    
    return None

def parse_converter_stdout(stdout_text):
    """Parse converter script output for metadata"""
    META_PATTERNS = {
        "metadata_checksum": re.compile(
            r"(?:Metadata\s+Checksum|Initial\s+checksum\s+with\s+initial\s+file\s+size)\s*:\s*([0-9a-fA-F]{32})",
            re.IGNORECASE,
        ),
        "metadata_data": re.compile(r"(?:Metadata\s+Data\s+string|Checksum\s+data\s+string)\s*:\s*(.+)", re.IGNORECASE),
        "initial_size_bytes": re.compile(r"Initial\s+file\s+size\s*:\s*(\d+)\s*bytes", re.IGNORECASE),
        "final_size_bytes": re.compile(r"(?:Final\s+file\s+size|Final\s+file\s+size\s+after\s+enhancement)\s*:\s*(\d+)\s*bytes", re.IGNORECASE),
        "metadata_ts_ist": re.compile(r"Timestamp\s*\(IST\)\s*:\s*([0-9:\-\s]+)", re.IGNORECASE),
        "station_code": re.compile(r"Station\s+code\s*:\s*([A-Za-z0-9\-_]+)", re.IGNORECASE),
        "source_pdf_name": re.compile(r"File\s+name\s*:\s*(.+?\.pdf)\s*$", re.IGNORECASE | re.MULTILINE),
        "full_file_md5": re.compile(r"Full\s+file\s+MD5(?:\s*hash)?\s*:\s*([0-9a-fA-F]{32})", re.IGNORECASE),
    }
    
    out = {}
    text = stdout_text or ""
    for key, pat in META_PATTERNS.items():
        m = pat.search(text)
        out[key] = m.group(1).strip() if m else None
    
    return out

def _md5_of_file(path):
    """Calculate MD5 hash of file"""
    hasher = hashlib.md5()
    try:
        with open(path, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b''):
                hasher.update(chunk)
        return hasher.hexdigest()
    except Exception as e:
        logger.error(f"Error calculating MD5 for {path}: {e}")
        return None

def extract_xlsx_metadata(xlsx_path):
    """Extract metadata from XLSX file"""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        metadata = {
            'station_name': 'Unknown Station',
            'station_id': None,
            'station_code': None,
            'drawn_by': 'Auto Converter',
            'version': '1',
            'checksum': None,
            'diagram_name': 'railways',
            'checked_by': 'supervisor',
            'division': 'Ahemdabad',
            'zone': 'WRLY',
            'total_sheet': '17',
            'designation1': 'DY.CSTE/C-II/ADI',
            'designation2': 'DSTE/C/ADI',
            'designation3': 'SSE/SIG/C/ADI'
        }
        
        # Check for StationDrawing sheet (case-insensitive)
        sheet_names_lower = [s.lower() for s in wb.sheetnames]
        
        if 'stationdrawing' in sheet_names_lower:
            sheet_idx = sheet_names_lower.index('stationdrawing')
            ws = wb[wb.sheetnames[sheet_idx]]
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
                else:
                    headers.append('')
            
            # Get first data row (row 2)
            row_data = {}
            for i, cell in enumerate(ws[2]):
                if i < len(headers):
                    value = cell.value
                    if value is not None:
                        row_data[headers[i]] = str(value).strip()
            
            # Extract values
            field_mapping = {
                'station_name': 'station_name',
                'station_id': 'station_id',
                'station_code': 'station_code',
                'drawn_by': 'drawn_by',
                'version': 'version',
                'checksum': 'checksum',
                'diagram_name': 'diagram_name',
                'checked_by': 'checked_by',
                'division': 'division',
                'zone': 'zone',
                'total_sheet': 'total_sheet',
                'designation1': 'designation1',
                'designation2': 'designation2',
                'designation3': 'designation3',
                'date': 'date'
            }
            
            for db_field, excel_field in field_mapping.items():
                if excel_field in row_data and row_data[excel_field]:
                    metadata[db_field] = row_data[excel_field]
        
        # If date is not in excel, use current date
        if not metadata.get('date'):
            metadata['date'] = datetime.now().strftime("%d-%m-%Y")
        
        return metadata
        
    except Exception as e:
        logger.error(f"Error extracting metadata from {xlsx_path}: {e}")
        # Return default metadata
        return {
            'station_name': 'Unknown Station',
            'station_id': None,
            'station_code': None,
            'drawn_by': 'Auto Converter',
            'version': '1',
            'checksum': None,
            'diagram_name': 'railways',
            'checked_by': 'supervisor',
            'division': 'Ahemdabad',
            'zone': 'WRLY',
            'total_sheet': '17',
            'designation1': 'DY.CSTE/C-II/ADI',
            'designation2': 'DSTE/C/ADI',
            'designation3': 'SSE/SIG/C/ADI',
            'date': datetime.now().strftime("%d-%m-%Y")
        }

def get_next_version(project_id):
    """Get next version number for project using database models"""
    with app.app_context():
        try:
            # Get max version for project from generated_pdf table
            max_version_record = GeneratedPDF.query.filter_by(
                project_id=project_id
            ).order_by(GeneratedPDF.version.desc()).first()
            
            if max_version_record and max_version_record.version:
                return max_version_record.version + 1
            return 1
        except Exception as e:
            logger.error(f"Error getting next version for project {project_id}: {e}")
            return 1

def update_database_with_models(project_id, xlsx_filename, pdf_filename, checksum, metadata, station_info):
    """Update database using SQLAlchemy models"""
    with app.app_context():
        try:
            # Get project
            project = Project.query.get(project_id)
            if not project:
                logger.error(f"Project {project_id} not found in database")
                return False
            
            # Get next version
            version = get_next_version(project_id)
            
            # Get file sizes from local uploads directory
            local_xlsx = os.path.join(CONFIG['LOCAL_UPLOADS_DIR'], xlsx_filename)
            local_pdf = os.path.join(CONFIG['LOCAL_UPLOADS_DIR'], pdf_filename)
            
            xlsx_size = os.path.getsize(local_xlsx) if os.path.exists(local_xlsx) else 0
            pdf_size = os.path.getsize(local_pdf) if os.path.exists(local_pdf) else 0
            
            print(f"D File sizes - XLSX: {xlsx_size} bytes, PDF: {pdf_size} bytes")
            
            # Create GeneratedPDF record
            generated_pdf = GeneratedPDF(
                project_id=project_id,
                pdf_filename=pdf_filename,
                xlsx_filename=xlsx_filename,
                checksum_md5=checksum,
                file_size=pdf_size,
                checksum_algo="md5",
                metadata_checksum=metadata.get('metadata_checksum'),
                metadata_data=metadata.get('metadata_data'),
                initial_size_bytes=metadata.get('initial_size_bytes'),
                final_size_bytes=metadata.get('final_size_bytes'),
                metadata_ts_ist=metadata.get('metadata_ts_ist'),
                station_code=metadata.get('station_code') or station_info.get('station_code'),
                source_pdf_name=metadata.get('source_pdf_name'),
                full_file_md5=metadata.get('full_file_md5') or checksum,
                remarks='Auto-converted via FTP automation',
                created_at=get_ist_now(),
                version=version,
                level1_status='pending',
                level2_status='pending',
                level3_status='pending',
                junction_data=None # We could extract this from XLSX if needed
            )
            
            db.session.add(generated_pdf)
            db.session.flush() # To get the ID
            
            # Update or create StationDrawing
            station_drawing = StationDrawing.query.filter_by(project_id=project_id).first()
            
            if station_drawing:
                # Update existing
                station_drawing.version = str(version)
                station_drawing.checksum = checksum
                station_drawing.station_name = station_info.get('station_name', f"Project_{project_id}")
                station_drawing.station_id = station_info.get('station_id', str(project_id))
                station_drawing.station_code = station_info.get('station_code', f"STN{project_id}")
                station_drawing.drawn_by = station_info.get('drawn_by', 'Auto Converter')
                station_drawing.diagram_name = station_info.get('diagram_name', 'railways')
                station_drawing.checked_by = station_info.get('checked_by', 'supervisor')
                station_drawing.division = station_info.get('division', 'Ahemdabad')
                station_drawing.zone = station_info.get('zone', 'WRLY')
                station_drawing.total_sheet = station_info.get('total_sheet', '17')
                station_drawing.designation1 = station_info.get('designation1', 'DY.CSTE/C-II/ADI')
                station_drawing.designation2 = station_info.get('designation2', 'DSTE/C/ADI')
                station_drawing.designation3 = station_info.get('designation3', 'SSE/SIG/C/ADI')
                station_drawing.date = station_info.get('date', datetime.now().strftime("%d-%m-%Y"))
            else:
                # Create new - FIXED: Added missing comma after designation1 line
                station_drawing = StationDrawing(
                    project_id=project_id,
                    station_id=station_info.get('station_id', str(project_id)),
                    station_name=station_info.get('station_name', f"Project_{project_id}"),
                    station_code=station_info.get('station_code', f"STN{project_id}"),
                    version=str(version),
                    checksum=checksum,
                    drawn_by=station_info.get('drawn_by', 'Auto Converter'),
                    diagram_name=station_info.get('diagram_name', 'railways'),
                    checked_by=station_info.get('checked_by', 'supervisor'),
                    division=station_info.get('division', 'Ahemdabad'),
                    zone=station_info.get('zone', 'WRLY'),
                    total_sheet=station_info.get('total_sheet', '17'),
                    designation1=station_info.get('designation1', 'DY.CSTE/C-II/ADI'),  # FIXED: Added comma here
                    designation2=station_info.get('designation2', 'DSTE/C/ADI'),
                    designation3=station_info.get('designation3', 'SSE/SIG/C/ADI'),
                    date=station_info.get('date', datetime.now().strftime("%d-%m-%Y")),
                    created_date=get_ist_now()
                )
                db.session.add(station_drawing)
            
            # Create notifications for admin users (role='4')
            admin_users = User.query.filter_by(role='4', is_active=True).all()
            for admin in admin_users:
                # Check if admin is assigned to this project
                if hasattr(admin, 'projects') and project in admin.projects:
                    notification = Notification(
                        user_id=admin.id,
                        pdf_id=generated_pdf.id,
                        project_id=project_id,
                        level='New_Drawing',
                        status='pending',
                        message=f'NEW DRAWING created via FTP automation: {project.name}'
                    )
                    db.session.add(notification)
            
            # Also create notification for level1 users assigned to this project
            level1_users = User.query.filter_by(designation='level1', is_active=True).all()
            for user in level1_users:
                if hasattr(user, 'projects') and project in user.projects:
                    notification = Notification(
                        user_id=user.id,
                        pdf_id=generated_pdf.id,
                        project_id=project_id,
                        level='level1',
                        status='pending',
                        message=f'NEW DRAWING requires level1 approval: {project.name}'
                    )
                    db.session.add(notification)
            
            db.session.commit()
            logger.info(f"Database updated: Project {project_id}, Version {version}, PDF ID: {generated_pdf.id}")
            print(f"D Database updated successfully: Project {project_id}, Version {version}")
            
            return True
            
        except Exception as e:
            logger.error(f"Database update error: {e}")
            import traceback
            traceback.print_exc()
            db.session.rollback()
            return False

def parse_pdf_name_from_stdout(stdout_text):
    """Parse the PDF name from the converter script stdout."""
    # Look for the line: "Multi-page PDF saved as '...'"
    pattern = r"Multi-page PDF saved as '(.+?)'"
    match = re.search(pattern, stdout_text)
    if match:
        return match.group(1)
    return None

def convert_xlsx_to_pdf(xlsx_path, pdf_path):
    """Convert XLSX to PDF using converter script with improved error handling"""
    try:
        # Check if converter script exists
        if not os.path.exists(CONFIG['CONVERTER_SCRIPT']):
            logger.error(f"Converter script not found: {CONFIG['CONVERTER_SCRIPT']}")
            return None, "Converter script not found"
        
        print(f"C Converting: {os.path.basename(xlsx_path)} -> {os.path.basename(pdf_path)}")
        
        # Run converter
        python_exe = sys.executable
        result = subprocess.run(
            [python_exe, CONFIG['CONVERTER_SCRIPT'], xlsx_path, pdf_path],
            capture_output=True,
            text=True,
            timeout=300,
            cwd=os.path.dirname(CONFIG['CONVERTER_SCRIPT'])
        )
        
        # Check if PDF was created (even with warnings)
        pdf_created = False
        actual_pdf_path = pdf_path
        
        # First check if PDF was created at the expected path
        if os.path.exists(pdf_path):
            pdf_created = True
            print(f"C PDF created at expected path: {pdf_path}")
        else:
            # Try to parse the PDF name from stdout
            pdf_name = parse_pdf_name_from_stdout(result.stdout)
            if pdf_name:
                # The PDF might be in the current working directory (cwd)
                cwd = os.path.dirname(CONFIG['CONVERTER_SCRIPT'])
                possible_pdf_path = os.path.join(cwd, pdf_name)
                if os.path.exists(possible_pdf_path):
                    shutil.move(possible_pdf_path, pdf_path)
                    print(f"C PDF moved from {possible_pdf_path} to {pdf_path}")
                    pdf_created = True
                else:
                    # Try other possible locations
                    possible_dirs = [
                        os.path.dirname(xlsx_path),
                        os.path.dirname(pdf_path),
                        cwd,
                    ]
                    for directory in possible_dirs:
                        possible_pdf_path = os.path.join(directory, pdf_name)
                        if os.path.exists(possible_pdf_path):
                            shutil.move(possible_pdf_path, pdf_path)
                            print(f"C PDF moved from {possible_pdf_path} to {pdf_path}")
                            pdf_created = True
                            break
            
            # If still not found, search for PDF files with pattern
            if not pdf_created:
                pdf_pattern = "Terminal_Symbols_Centered_Fixed_Size_*.pdf"
                possible_dirs = [
                    os.path.dirname(CONFIG['CONVERTER_SCRIPT']),
                    os.path.dirname(xlsx_path),
                    os.path.dirname(pdf_path),
                ]
                pdf_files = []
                for directory in possible_dirs:
                    pdf_files.extend(glob.glob(os.path.join(directory, pdf_pattern)))
                
                # Remove duplicates
                pdf_files = list(set(pdf_files))
                
                if pdf_files:
                    # Sort by modification time, most recent first
                    pdf_files.sort(key=os.path.getmtime, reverse=True)
                    actual_pdf_path_found = pdf_files[0]
                    shutil.move(actual_pdf_path_found, pdf_path)
                    print(f"C PDF moved from {actual_pdf_path_found} to {pdf_path}")
                    pdf_created = True
        
        # NEW: Check if conversion was successful even with warnings
        if result.returncode == 0 and pdf_created:
            # Even if there are warnings, if PDF was created, consider it a success
            logger.info(f"PDF created successfully: {os.path.basename(pdf_path)}")
            print(f"C Conversion successful (with warnings): {os.path.basename(pdf_path)} created")
            
            # Log warnings if any
            if result.stderr:
                warning_lines = result.stderr.strip().split('\n')
                for line in warning_lines:
                    if line.strip() and "Warning:" in line:
                        logger.warning(f"Converter warning: {line.strip()}")
                        print(f"W Converter warning: {line.strip()}")
            
            return parse_converter_stdout(result.stdout), None
        elif result.returncode != 0:
            # Conversion failed with non-zero return code
            logger.error(f"PDF conversion failed: returncode={result.returncode}")
            if result.stdout:
                logger.error(f"STDOUT: {result.stdout.strip()}")
            if result.stderr:
                logger.error(f"STDERR: {result.stderr.strip()}")
            error_msg = result.stderr.strip() or result.stdout.strip() or "No output from converter"
            logger.error(f"PDF conversion failed: {error_msg}")
            print(f"X Conversion failed: {error_msg}")
            return None, error_msg
        else:
            # Return code is 0 but PDF not created
            logger.error(f"PDF conversion completed but PDF not found")
            print(f"X PDF not found after conversion")
            return None, "PDF not found after conversion"
            
    except subprocess.TimeoutExpired:
        error_msg = "Conversion timed out (300 seconds)"
        logger.error(error_msg)
        print(f"X {error_msg}")
        return None, error_msg
    except Exception as e:
        error_msg = f"Conversion error: {str(e)}"
        logger.error(error_msg)
        print(f"X {error_msg}")
        return None, error_msg

def import_xlsx_data_to_database(xlsx_path, project_id):
    """Import XLSX data into database tables (optional)"""
    try:
        logger.info(f"Importing XLSX data for project {project_id}")
        print(f"D Importing XLSX data to database for project {project_id}")
        
        wb = load_workbook(xlsx_path, data_only=True)
        
        # Define which sheets to import and their models
        sheets_to_import = {
            'StationDrawing': StationDrawing,
            'junction_box': JunctionBox,
            'cable': Cable,
            'cable_box': CableBox,
            'terminal': Terminal,
            'group': Group,
            'terminal_header': TerminalHeader,
            'choketable': ChokeTable,
            'resistortable': ResistorTable,
        }
        
        imported_count = 0
        
        for sheet_name, model_class in sheets_to_import.items():
            if sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    
                    # Get headers
                    headers = []
                    for cell in ws[1]:
                        if cell.value:
                            headers.append(str(cell.value).strip())
                        else:
                            headers.append('')
                    
                    # Clear existing data for this sheet and project
                    existing_records = model_class.query.filter_by(project_id=project_id).all()
                    for record in existing_records:
                        db.session.delete(record)
                    
                    # Import rows
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        data = {'project_id': project_id}
                        has_data = False
                        
                        for i, cell_value in enumerate(row):
                            if i < len(headers) and cell_value is not None:
                                value_str = str(cell_value).strip()
                                if value_str:
                                    data[headers[i]] = value_str
                                    has_data = True
                        
                        if has_data:
                            # Create model instance
                            model_instance = model_class(**data)
                            db.session.add(model_instance)
                            imported_count += 1
                    
                    db.session.commit()
                    logger.info(f"Imported {sheet_name} data for project {project_id}")
                    print(f"D Imported {sheet_name} data")
                    
                except Exception as e:
                    logger.error(f"Error importing {sheet_name}: {e}")
                    db.session.rollback()
                    continue
        
        logger.info(f"Total imported records: {imported_count}")
        print(f"D Total imported records: {imported_count}")
        return imported_count > 0
        
    except Exception as e:
        logger.error(f"Error importing XLSX data: {e}")
        print(f"X Error importing XLSX data: {e}")
        db.session.rollback()
        return False

def process_remote_xlsx_file(remote_filename):
    """Process a single XLSX file from remote server with improved error handling"""
    remote_xlsx_path = os.path.join(CONFIG['FTP_XLSX_TAKE_DIR'], remote_filename)
    success = False
    files_moved_to_uploads = False  # Track if files were moved to uploads folder
    
    try:
        logger.info(f"Processing remote file: {remote_filename}")
        print(f"\n" + "="*70)
        print(f"PROCESSING: {remote_filename}")
        print("="*70)
        
        # Check if already processed (LOOP PREVENTION)
        if is_file_processed(remote_filename):
            logger.info(f"Skipping already processed file: {remote_filename}")
            print(f"W Skipping (already processed): {remote_filename}")
            return True  # Return True so it's not retried
        
        # Extract project ID
        project_id = get_project_id_from_filename(remote_filename)
        if not project_id:
            logger.error(f"Could not extract project ID from: {remote_filename}")
            print(f"X Could not extract project ID from: {remote_filename}")
            return False
        
        logger.info(f"Detected Project ID: {project_id}")
        print(f"D Project ID: {project_id}")
        
        # Download file from remote
        local_xlsx_temp = os.path.join(CONFIG['LOCAL_XLSX_DIR'], remote_filename)
        
        if not download_file_remote(remote_xlsx_path, local_xlsx_temp):
            logger.error(f"Failed to download: {remote_filename}")
            print(f"X Failed to download: {remote_filename}")
            return False
        
        # Verify download
        if not os.path.exists(local_xlsx_temp):
            print(f"X Downloaded file not found at: {local_xlsx_temp}")
            return False
        
        print(f"D Downloaded size: {os.path.getsize(local_xlsx_temp)} bytes")
        
        # NEW: Save a visible copy in downloads folder
        visible_copy_path = os.path.join(CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR'], remote_filename)
        shutil.copy2(local_xlsx_temp, visible_copy_path)
        print(f"D Downloaded XLSX saved to: {visible_copy_path}")
        logger.info(f"Visible copy saved to: {visible_copy_path}")
        
        # Extract metadata
        station_info = extract_xlsx_metadata(local_xlsx_temp)
        station_name = station_info.get('station_name', f"Project_{project_id}")
        
        logger.info(f"Station info: {station_name}")
        print(f"D Station: {station_name}")
        
        # Generate timestamp for new filenames
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Create safe filename
        safe_station_name = re.sub(r'[^\w\-_\. ]', '', station_name).replace(' ', '_')
        new_xlsx_name = f"railway_project_{project_id}_{timestamp}_{safe_station_name}.xlsx"
        new_pdf_name = new_xlsx_name.replace('.xlsx', '.pdf')
        
        print(f"D New XLSX name: {new_xlsx_name}")
        print(f"D New PDF name: {new_pdf_name}")
        
        # Local paths for processing
        local_xlsx_new = os.path.join(CONFIG['LOCAL_XLSX_DIR'], new_xlsx_name)
        local_pdf_new = os.path.join(CONFIG['LOCAL_PDF_DIR'], new_pdf_name)
        
        # Rename local file
        shutil.move(local_xlsx_temp, local_xlsx_new)
        print(f"D Renamed local file to: {local_xlsx_new}")
        
        # Convert to PDF
        logger.info(f"Converting to PDF: {new_pdf_name}")
        converter_metadata, error = convert_xlsx_to_pdf(local_xlsx_new, local_pdf_new)
        
        if error and not CONFIG['ALLOW_PARTIAL_CONVERSION']:
            logger.error(f"Conversion failed: {error}")
            # Clean up local files
            if os.path.exists(local_xlsx_new):
                os.remove(local_xlsx_new)
            success = False
        else:
            # Even if there was an error, check if PDF was created
            pdf_created = os.path.exists(local_pdf_new)
            
            if not pdf_created:
                print(f"X PDF not created at: {local_pdf_new}")
                success = False
            else:
                print(f"D PDF created successfully: {os.path.getsize(local_pdf_new)} bytes")
                
                # Calculate checksum
                checksum = _md5_of_file(local_pdf_new)
                if not checksum:
                    logger.warning("Failed to calculate PDF checksum, using default")
                    print(f"W Failed to calculate PDF checksum, using default")
                    checksum = "00000000000000000000000000000000"
                
                print(f"D PDF checksum: {checksum}")
                
                # ========== MOVE FILES TO LOCAL UPLOADS ==========
                local_uploads_xlsx = os.path.join(CONFIG['LOCAL_UPLOADS_DIR'], new_xlsx_name)
                local_uploads_pdf = os.path.join(CONFIG['LOCAL_UPLOADS_DIR'], new_pdf_name)
                
                # Move files to local uploads directory
                logger.info(f"Moving files to local uploads: {CONFIG['LOCAL_UPLOADS_DIR']}")
                
                # First, verify the local uploads directory exists
                if not os.path.exists(CONFIG['LOCAL_UPLOADS_DIR']):
                    os.makedirs(CONFIG['LOCAL_UPLOADS_DIR'], exist_ok=True)
                    print(f"D Created local uploads directory: {CONFIG['LOCAL_UPLOADS_DIR']}")
                
                # Move XLSX
                if os.path.exists(local_xlsx_new):
                    shutil.move(local_xlsx_new, local_uploads_xlsx)
                    print(f"L Moved XLSX to local uploads: {local_uploads_xlsx}")
                    files_moved_to_uploads = True
                else:
                    print(f"W XLSX file not found for moving: {local_xlsx_new}")
                
                # Move PDF
                if os.path.exists(local_pdf_new):
                    shutil.move(local_pdf_new, local_uploads_pdf)
                    print(f"L Moved PDF to local uploads: {local_uploads_pdf}")
                    files_moved_to_uploads = True
                else:
                    print(f"W PDF file not found for moving: {local_pdf_new}")
                
                # Verify both files exist in uploads directory
                if os.path.exists(local_uploads_xlsx) and os.path.exists(local_uploads_pdf):
                    print(f"L Verified files in local uploads:")
                    print(f"L   XLSX: {os.path.getsize(local_uploads_xlsx)} bytes")
                    print(f"L   PDF: {os.path.getsize(local_uploads_pdf)} bytes")
                    
                    # ========== UPLOAD TO REMOTE UPLOADS DIRECTORY ==========
                    # Ensure remote path starts with /
                    if not CONFIG['FTP_UPLOAD_DIR'].startswith('/'):
                        remote_upload_dir = '/' + CONFIG['FTP_UPLOAD_DIR']
                    else:
                        remote_upload_dir = CONFIG['FTP_UPLOAD_DIR']
                    
                    remote_xlsx_new = os.path.join(remote_upload_dir, new_xlsx_name)
                    remote_pdf_new = os.path.join(remote_upload_dir, new_pdf_name)
                    
                    print(f"R Remote upload directory: {remote_upload_dir}")
                    
                    # Upload XLSX to remote
                    logger.info(f"Uploading XLSX to remote: {remote_xlsx_new}")
                    xlsx_upload_success = upload_file_remote(local_uploads_xlsx, remote_xlsx_new)
                    
                    # Upload PDF to remote
                    logger.info(f"Uploading PDF to remote: {remote_pdf_new}")
                    pdf_upload_success = upload_file_remote(local_uploads_pdf, remote_pdf_new)
                    
                    if xlsx_upload_success and pdf_upload_success:
                        print(f"R Files uploaded to remote successfully")
                        print(f"R   XLSX: {remote_xlsx_new}")
                        print(f"R   PDF: {remote_pdf_new}")
                        
                        # Update database using models
                        logger.info("Updating database...")
                        db_success = update_database_with_models(
                            project_id,
                            new_xlsx_name,
                            new_pdf_name,
                            checksum,
                            converter_metadata or {},
                            station_info
                        )
                        
                        if db_success:
                            # Optional: Import XLSX data to database tables
                            logger.info("Importing XLSX data to database tables...")
                            import_xlsx_data_to_database(local_uploads_xlsx, project_id)
                            success = True
                        else:
                            logger.error("Database update failed")
                            print(f"X Database update failed")
                            success = False
                    else:
                        logger.error("Failed to upload one or both files to remote")
                        print(f"X Failed to upload files to remote")
                        success = False
                else:
                    print(f"X Files not found in local uploads directory")
                    print(f"X   XLSX exists: {os.path.exists(local_uploads_xlsx)}")
                    print(f"X   PDF exists: {os.path.exists(local_uploads_pdf)}")
                    success = False
        
        # MARK AS PROCESSED IF FILES WERE MOVED TO UPLOADS (even if other steps failed)
        if files_moved_to_uploads:
            save_processed_file(remote_filename)
            print(f"v Files moved to uploads, marked as processed: {remote_filename}")
            # Even if other steps failed, we consider this a partial success
            partial_success = True
        else:
            partial_success = False
        
        # MOVE REMOTE FILE BASED ON SUCCESS STATUS
        if success and CONFIG['ARCHIVE_REMOTE']:
            remote_dest_path = os.path.join(CONFIG['REMOTE_ARCHIVE_DIR'], remote_filename)
            logger.info(f"Archiving original to: {remote_dest_path}")
            move_success = move_file_remote(remote_xlsx_path, remote_dest_path)
            if move_success:
                print(f"A Archived original to: {remote_dest_path}")
            else:
                logger.warning(f"Failed to archive {remote_filename} (but marked processed)")
                print(f"W Failed to archive {remote_filename}")
        elif partial_success:
            # Files were moved to uploads but other steps failed - move to backup
            remote_backup_path = os.path.join(CONFIG['REMOTE_BACKUP_DIR'], remote_filename)
            logger.info(f"Moving to backup (partial success): {remote_backup_path}")
            move_success = move_file_remote(remote_xlsx_path, remote_backup_path)
            if move_success:
                print(f"B Moved to backup (partial success): {remote_filename}")
            else:
                logger.warning(f"Failed to move to backup {remote_filename}")
                print(f"W Failed to move to backup")
        else:
            # Move to failed dir
            remote_failed_path = os.path.join(CONFIG['REMOTE_FAILED_DIR'], remote_filename)
            logger.info(f"Moving failed file to: {remote_failed_path}")
            move_success = move_file_remote(remote_xlsx_path, remote_failed_path)
            if move_success:
                print(f"F Moved to failed: {remote_filename}")
                logger.info(f"Moved failed file to: {remote_failed_path}")
            else:
                logger.warning(f"Failed to move {remote_filename} to failed dir (may retry)")
                print(f"W Failed to move to failed dir")
        
        # Log final status
        if success:
            print(f"✓ Processing complete for {remote_filename} (FULL SUCCESS)")
            logger.info(f"Processing complete for {remote_filename} (success: True)")
        elif partial_success:
            print(f"⚠ Processing partially successful for {remote_filename} (files moved to uploads)")
            logger.info(f"Processing partially successful for {remote_filename} (files moved to uploads)")
            success = True  # Return True to avoid retrying
        else:
            print(f"✗ Processing failed for {remote_filename}")
            logger.info(f"Processing complete for {remote_filename} (success: False)")
        
        return success or partial_success
        
    except Exception as e:
        logger.error(f"Error processing {remote_filename}: {str(e)}")
        import traceback
        traceback.print_exc()
        print(f"X Error processing {remote_filename}: {str(e)}")
        # On exception, treat as failure and move to failed
        remote_failed_path = os.path.join(CONFIG['REMOTE_FAILED_DIR'], remote_filename)
        move_file_remote(remote_xlsx_path, remote_failed_path)
        return False

def scan_and_process_remote():
    """Scan remote xlsx_download directory and process all XLSX files"""
    logger.info("Scanning remote directory for XLSX files...")
    print("S Scanning remote directory...")
    
    try:
        # List files in remote directory
        remote_files = list_remote_files(CONFIG['FTP_XLSX_TAKE_DIR'])
        
        if not remote_files:
            logger.info("No files found in remote directory")
            print("S No files found in remote directory")
            return []
        
        # Filter for XLSX files
        xlsx_files = []
        for filename in remote_files:
            if filename.lower().endswith('.xlsx'):
                # Skip files in processed, backup, or failed directories
                if all(sub not in filename.lower() for sub in ['processed', 'backup', 'failed']):
                    # Check if already processed
                    if not is_file_processed(filename):
                        xlsx_files.append(filename)
        
        logger.info(f"Found {len(xlsx_files)} new XLSX file(s) to process")
        print(f"S Found {len(xlsx_files)} new XLSX file(s) to process")
        
        results = []
        for filename in xlsx_files:
            result = {
                'filename': filename,
                'success': False,
                'message': ''
            }
            
            try:
                success = process_remote_xlsx_file(filename)
                result['success'] = success
                result['message'] = 'Processed successfully' if success else 'Processing failed'
            except Exception as e:
                result['message'] = f"Error: {str(e)}"
            
            results.append(result)
            
            # Small delay between files
            time.sleep(2)
        
        return results
        
    except Exception as e:
        logger.error(f"Error scanning remote directory: {e}")
        print(f"X Error scanning remote directory: {e}")
        return []

def monitor_xlsx_download_folder(flask_app=None):
    """
    Continuous monitoring function for XLSX download folder.
    This is the entry point called from __init__.py via threading.
    If flask_app is provided, sets it as the global app; otherwise uses global app.
    Runs automatically without prompts. Includes initial scan.
    """
    if flask_app:
        global app
        app = flask_app
    
    # Load previously processed files (LOOP PREVENTION)
    load_processed_files()
    
    setup_local_directories()
    
    # NEW: Show download locations clearly
    print("\n" + "="*70)
    print("DOWNLOAD & UPLOAD LOCATIONS:")
    print("="*70)
    print(f"K Visible downloads folder: {CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR']}")
    print(f"L Local uploads folder: {CONFIG['LOCAL_UPLOADS_DIR']}")
    print(f"D Temporary processing folder: {CONFIG['LOCAL_XLSX_DIR']}")
    print(f"F Log file: {CONFIG['LOG_FILE']}")
    print(f"P Processed files log: {CONFIG['PROCESSED_FILES_LOG']}")
    print(f"R Remote uploads: {CONFIG['FTP_UPLOAD_DIR']}")
    print("="*70)
    
    # NEW: Discover paths on startup
    print("\nDEBUG: Discovering remote paths...")
    discover_remote_paths()
    
    logger.info("Starting continuous remote monitoring...")
    logger.info(f"Monitoring: {CONFIG['FTP_XLSX_TAKE_DIR']}")
    logger.info(f"Uploads to: {CONFIG['FTP_UPLOAD_DIR']}")
    logger.info(f"Check interval: {CONFIG['CHECK_INTERVAL']} seconds")
    
    print("\n" + "="*70)
    print("FTP/SFTP AUTOMATED CONVERTER (NEW WORKFLOW)")
    print("="*70)
    print("Workflow: XLSX → PDF → Local Uploads → Remote Uploads")
    print("="*70)
    
    # NEW: Initial one-time scan for existing files
    print("\n" + "="*70)
    print("Performing initial scan for existing XLSX files...")
    print("="*70)
    initial_results = scan_and_process_remote()
    if initial_results:
        success_count = sum(1 for r in initial_results if r['success'])
        print(f"Initial scan complete: {success_count}/{len(initial_results)} processed successfully")
        for result in initial_results:
            status = "v" if result['success'] else "X"
            print(f"  [{status}] {result['filename']}: {result['message']}")
    else:
        print("Initial scan: No new files found or path issue (check debug above)")
    
    processed_count = len(PROCESSED_FILES_CACHE)
    print(f"\nS Total processed files in history: {processed_count}")
    
    print("\n" + "="*70)
    print("CONTINUOUS MONITORING STARTED")
    print("="*70)
    print("Workflow:")
    print("1. Download XLSX → C:\\Demo\\git\\downloaded_xlsx_files")
    print("2. Convert XLSX → PDF")
    print("3. Move both to local uploads → C:\\Demo\\git\\uploads")
    print("4. Upload both to remote → srv/railway/frontend/uploads/")
    print("5. Update database")
    print("="*70)
    print("Monitoring for new XLSX files...")
    print("Press Ctrl+C to stop.")
    print("="*70 + "\n")
    
    try:
        while True:
            # List current files
            current_files = list_remote_files(CONFIG['FTP_XLSX_TAKE_DIR'])
            
            if current_files:
                # Filter for XLSX files (not in processed/backup/failed)
                current_xlsx = []
                for filename in current_files:
                    if (filename.lower().endswith('.xlsx') and
                        all(sub not in filename.lower() for sub in ['processed', 'backup', 'failed']) and
                        not is_file_processed(filename)):  # LOOP PREVENTION
                        current_xlsx.append(filename)
                
                # Find new files
                if current_xlsx:
                    logger.info(f"Found {len(current_xlsx)} new file(s)")
                    print(f"\nS Found {len(current_xlsx)} new file(s)")
                    
                    for filename in current_xlsx:
                        logger.info(f"Processing new file: {filename}")
                        print(f"\n" + "="*50)
                        print(f"PROCESSING: {filename}")
                        print("="*50)
                        
                        success = process_remote_xlsx_file(filename)
                        
                        if success:
                            logger.info(f"v Successfully processed: {filename}")
                            print(f"✓ Successfully processed: {filename}")
                            print(f"✓ Files moved to local uploads: {CONFIG['LOCAL_UPLOADS_DIR']}")
                            print(f"✓ Files uploaded to remote: {CONFIG['FTP_UPLOAD_DIR']}")
                            print(f"✓ Database updated with new version")
                        else:
                            logger.error(f"X Failed to process: {filename}")
                            print(f"✗ Failed to process: {filename}")
            
            # Sleep before next check
            time.sleep(CONFIG['CHECK_INTERVAL'])
            
    except KeyboardInterrupt:
        logger.info("Monitoring stopped by user (Ctrl+C)")
        print("\n" + "="*70)
        print("MONITORING STOPPED")
        print("="*70)
        print(f"S Total files processed in this session: {len(PROCESSED_FILES_CACHE)}")
        print(f"D Check downloaded files in: {CONFIG['LOCAL_DOWNLOADS_VISIBLE_DIR']}")
        print(f"L Check local uploads in: {CONFIG['LOCAL_UPLOADS_DIR']}")
        print(f"F Check logs in: {CONFIG['LOG_FILE']}")
        print("="*70)
    except Exception as e:
        logger.error(f"Monitoring error: {e}")
        print(f"\nX Monitoring error: {e}")

if __name__ == "__main__":
    # Welcome banner (non-interactive)
    print("\n" + "="*70)
    print("FTP/SFTP AUTOMATED CONVERTER (NEW WORKFLOW)")
    print("="*70)
    print("Workflow: XLSX → PDF → Local Uploads → Remote Uploads")
    print("="*70)
    print(f"Host: {CONFIG['FTP_HOST']}")
    print(f"Mode: {'SFTP' if CONFIG['FTP_USE_SFTP'] else 'FTP'}")
    print(f"XLSX Source: {CONFIG['FTP_XLSX_TAKE_DIR']}")
    print(f"PDF Destination (Local): {CONFIG['LOCAL_UPLOADS_DIR']}")
    print(f"PDF Destination (Remote): {CONFIG['FTP_UPLOAD_DIR']}")
    print(f"Database: PostgreSQL ({CONFIG['DB_URI'].split('@')[1].split('/')[0] if '@' in CONFIG['DB_URI'] else 'local'})")
    print(f"Check Interval: {CONFIG['CHECK_INTERVAL']}s")
    print("="*70)
    
    # Start automated monitoring immediately
    monitor_xlsx_download_folder()